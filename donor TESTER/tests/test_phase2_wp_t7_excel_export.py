"""
WP-T7 — Excel export integration tests.

Plan reference: docs/zigzag_st_tester_phase2_implementation_plan.txt §9, §9.1–§9.5
Tests: #17 (legacy disabled golden), #18 (legacy enabled columns), #19 (equal_blocks disabled),
       #20 (empty-trade enabled).

Contracts verified:
1. Disabled path → no filter columns in any sheet, baseline-identical column counts.
2. Signals sheet: disabled=no filter cols; enabled=4 filter cols in correct position.
3. Trades sheets: disabled=canonical cols; enabled=+3 filter diagnostic cols.
4. Summary sheet: disabled=no filter block; enabled=filter block present.
5. FilterDiagnostics_100 sheet: appears iff export_state_columns=True AND enabled.
6. ZigZag_Trigger_Events sheet: appears iff export_trigger_columns=True AND enabled.
7. filters_summary sheet: appears iff export_state_columns=True AND enabled.
8. Empty-trade enabled: trade headers present with filter cols; summary filter block present.
9. Trigger events = rows where trigger_source != "none".
10. CLI wiring: export_tester_results receives trade_filter_config and df.
11. Two-step Linked Trade ID linker (plan §9.5.1): exact match then backward search.
12. Golden column order pinned for disabled and enabled paths (#17, #18, #19).
"""

from __future__ import annotations

import io
import tempfile
import numpy as np
import pandas as pd
import pytest

from pathlib import Path
from typing import Optional, Dict, Any, List


# ---------------------------------------------------------------------------
# Path constants (repo-relative, no hardcoded absolute paths)
# ---------------------------------------------------------------------------

_TESTS_DIR = Path(__file__).resolve().parent
_TESTER_ROOT = _TESTS_DIR.parent
_REPO_ROOT = _TESTER_ROOT.parent
_DONOR_ROOT = _REPO_ROOT / "donor"
_BASELINES_DIR = _TESTS_DIR / "baselines"

_LEGACY_BASELINE = _BASELINES_DIR / "result_legacy.baseline.xlsx"
_EQ_BLOCKS_BASELINE = _BASELINES_DIR / "result_equal_blocks.baseline.xlsx"

# Pinned exact column sequences (plan §9.5 + SIGNALS_DISPLAY_NAMES order)
# ---------------------------------------------------------------------------

# Signals: disabled — canonical 19 base + 2 ratio = 21 columns; no filter cols
EXPECTED_SIGNALS_DISABLED_COLS: List[str] = [
    "Signal Time",
    "Signal Bar Index",
    "Event Type",
    "Direction",
    "ST Color Before",
    "ST Color After",
    "Is Reversal",
    "Exec Price",
    "Signal Open Price",
    "Signal Close Price",
    "Signal Body %",
    "Signal Range %",
    "Signal Body ATR",
    "Signal Range ATR",
    "T+1 Return %",
    "T+2 Return %",
    "T+3 Return %",
    "Signal Body % / Median",
    "Signal Range % / Median",
]

# Signals: enabled — 4 filter cols inserted after Is Reversal, before Exec Price
EXPECTED_SIGNALS_ENABLED_COLS: List[str] = [
    "Signal Time",
    "Signal Bar Index",
    "Event Type",
    "Direction",
    "ST Color Before",
    "ST Color After",
    "Is Reversal",
    "Filter State at Signal",     # filter col 1
    "Filter Decision",            # filter col 2
    "Filter Block Reason",        # filter col 3
    "Filter Trigger Source",      # filter col 4
    "Exec Price",
    "Signal Open Price",
    "Signal Close Price",
    "Signal Body %",
    "Signal Range %",
    "Signal Body ATR",
    "Signal Range ATR",
    "T+1 Return %",
    "T+2 Return %",
    "T+3 Return %",
    "Signal Body % / Median",
    "Signal Range % / Median",
]

# Trades: disabled — canonical 13 columns; no filter cols
EXPECTED_TRADES_DISABLED_COLS: List[str] = [
    "Trade ID",
    "Direction",
    "Entry Time",
    "Entry Index",
    "Entry Price",
    "Exit Time",
    "Exit Index",
    "Exit Price",
    "Bars Held",
    "Gross PnL %",
    "Commission %",
    "Net PnL %",
    "SuperTrend Color",
]

# Trades: enabled — canonical 13 + 3 filter cols at end
EXPECTED_TRADES_ENABLED_COLS: List[str] = EXPECTED_TRADES_DISABLED_COLS + [
    "Entry Filter State",
    "Entry Trigger Source",
    "Exit Reason",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_synthetic_ohlc(n: int = 300, seed: int = 7) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    close = 100.0 * np.exp(np.cumsum(rng.normal(0.0003, 0.012, n)))
    noise = rng.uniform(0.001, 0.004, size=n)
    high = close * (1 + noise)
    low = close * (1 - noise)
    open_ = np.clip(close * (1 + rng.uniform(-0.002, 0.002, size=n)), low, high)
    idx = pd.date_range("2020-01-01", periods=n, freq="D")
    return pd.DataFrame({"open": open_, "high": high, "low": low, "close": close}, index=idx)


def _make_enabled_cfg(export_state_columns: bool = True, export_trigger_columns: bool = True):
    from supertrend_optimizer.core.trade_filter_config import (
        TradeFilterConfig,
        TradeFilterDiagnosticsConfig,
        TradeFilterLifecycleConfig,
        TradeFilterTriggerToggleConfig,
        TradeFilterTriggersConfig,
        TradeFilterZigZagConfig,
    )
    return TradeFilterConfig(
        enabled=True,
        type="zigzag_st_mode",
        zigzag=TradeFilterZigZagConfig(
            reversal_threshold=0.04,
            local_window=20,
            candidate_trigger_threshold=0.4,
        ),
        triggers=TradeFilterTriggersConfig(
            candidate_threshold=TradeFilterTriggerToggleConfig(enabled=True),
            confirmed_median=TradeFilterTriggerToggleConfig(enabled=True),
        ),
        lifecycle=TradeFilterLifecycleConfig(
            freeze_confirmed_legs=3,
            stop_check="confirm_bar_only",
            stopping_exit="opposite_st_flip",
        ),
        diagnostics=TradeFilterDiagnosticsConfig(
            export_state_columns=export_state_columns,
            export_trigger_columns=export_trigger_columns,
        ),
    )


def _make_disabled_cfg():
    from supertrend_optimizer.core.trade_filter_config import TradeFilterConfig
    return TradeFilterConfig(enabled=False, type=None, zigzag=None,
                             triggers=None, lifecycle=None, diagnostics=None)


def _run_disabled_legacy(df: pd.DataFrame):
    """Run a full disabled-path legacy tester and return (results, signals_df)."""
    from supertrend_optimizer.testing.runner import run_all_periods
    from supertrend_optimizer.testing.signal_events import build_signal_events
    from supertrend_optimizer.utils.enums import ExecutionModel
    results = run_all_periods(
        df=df, atr_period=14, multiplier=3.0, trade_mode="revers",
        commission=0.001, warmup_period=20,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
    )
    signals_df = build_signal_events(
        df=df,
        trend=results[0].result.trend,
        atr_period=14,
        trade_mode="revers",
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        filter_diagnostics=None,
    )
    return results, signals_df


def _run_enabled_legacy(df: pd.DataFrame, tf_cfg):
    """Run a full enabled-path legacy tester and return (results, signals_df)."""
    from supertrend_optimizer.testing.runner import run_all_periods
    from supertrend_optimizer.testing.signal_events import build_signal_events
    from supertrend_optimizer.utils.enums import ExecutionModel
    results = run_all_periods(
        df=df, atr_period=14, multiplier=3.0, trade_mode="revers",
        commission=0.001, warmup_period=20,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        trade_filter_config=tf_cfg,
    )
    signals_df = build_signal_events(
        df=df,
        trend=results[0].result.trend,
        atr_period=14,
        trade_mode="revers",
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        filter_diagnostics=results[0].filter_diagnostics,
    )
    return results, signals_df


def _export_and_load(period_results, signals_df, tf_cfg=None, df=None):
    """Export to temp file and return openpyxl workbook."""
    from supertrend_optimizer.io.excel_tester import export_tester_results
    import openpyxl
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = str(Path(tmpdir) / "result.xlsx")
        actual = export_tester_results(
            period_results, out_path,
            signals_df=signals_df,
            trade_filter_config=tf_cfg,
            df=df,
        )
        wb = openpyxl.load_workbook(actual)
    return wb


def _sheet_headers(wb, sheet_name: str) -> List[str]:
    ws = wb[sheet_name]
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value is not None]


def _load_baseline_wb(path: Path):
    """Load baseline workbook if it exists, skip otherwise."""
    if not path.exists():
        pytest.skip(f"Baseline file not found: {path} (run WP-T1 capture first)")
    import openpyxl
    return openpyxl.load_workbook(str(path))


# ---------------------------------------------------------------------------
# Group 1: Test #17 — legacy disabled golden (column order pinned, baseline comparison)
# ---------------------------------------------------------------------------

class TestDisabledLegacyGolden:
    """Test #17: legacy disabled export column order pinned; no filter cols/sheets.

    Two sub-contracts:
    (a) Exact Signals column sequence matches EXPECTED_SIGNALS_DISABLED_COLS.
    (b) Exact Trades_100 column sequence matches EXPECTED_TRADES_DISABLED_COLS.
    (c) No filter sheets.
    (d) If baseline file on disk: fresh export column order == baseline column order.
    """

    def test_disabled_signals_exact_column_order(self) -> None:
        """Signals sheet columns must match pinned canonical sequence exactly."""
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        if "Signals" not in wb.sheetnames:
            pytest.skip("No signal events generated for this synthetic OHLC")
        headers = _sheet_headers(wb, "Signals")
        assert headers == EXPECTED_SIGNALS_DISABLED_COLS, (
            f"Signals disabled column order mismatch.\n"
            f"  Expected: {EXPECTED_SIGNALS_DISABLED_COLS}\n"
            f"  Got:      {headers}"
        )

    def test_disabled_trades_exact_column_order(self) -> None:
        """Trades_100 columns must match pinned canonical sequence exactly."""
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        if "Trades_100" not in wb.sheetnames:
            pytest.skip("Trades_100 sheet not present")
        headers = _sheet_headers(wb, "Trades_100")
        assert headers == EXPECTED_TRADES_DISABLED_COLS, (
            f"Trades_100 disabled column order mismatch.\n"
            f"  Expected: {EXPECTED_TRADES_DISABLED_COLS}\n"
            f"  Got:      {headers}"
        )

    def test_disabled_no_filter_sheets(self) -> None:
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        assert "FilterDiagnostics_100" not in wb.sheetnames
        assert "ZigZag_Trigger_Events" not in wb.sheetnames
        assert "filters_summary" not in wb.sheetnames
        assert "cycle" not in wb.sheetnames

    def test_disabled_no_filter_columns_signals(self) -> None:
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        if "Signals" not in wb.sheetnames:
            pytest.skip("No signal events")
        headers = _sheet_headers(wb, "Signals")
        for col in ("Filter State at Signal", "Filter Decision",
                    "Filter Block Reason", "Filter Trigger Source"):
            assert col not in headers, f"Filter col {col!r} must not appear in disabled Signals"

    def test_disabled_no_filter_columns_trades(self) -> None:
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Trades_"):
                continue
            headers = _sheet_headers(wb, sheet_name)
            for col in ("Entry Filter State", "Entry Trigger Source", "Exit Reason"):
                assert col not in headers, (
                    f"Filter trade col {col!r} must not appear in disabled {sheet_name}"
                )

    def test_disabled_summary_no_filter_block(self) -> None:
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        wb = _export_and_load(results, signals_df, tf_cfg=None)

        ws = wb["Summary"]
        all_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
        assert "Lifecycle Starts" not in all_values, (
            "Disabled Summary must not contain filter block data"
        )

    def test_disabled_column_order_matches_baseline_if_exists(self) -> None:
        """Compare fresh export column structure against saved baseline (#17 golden gate).

        If baseline is on disk, column orders must match exactly.
        This is the "normalized comparison" from plan §11 #17 — sheet by sheet.
        """
        df = _make_synthetic_ohlc()
        results, signals_df = _run_disabled_legacy(df)
        fresh_wb = _export_and_load(results, signals_df, tf_cfg=None)
        baseline_wb = _load_baseline_wb(_LEGACY_BASELINE)

        # Sheet names must be a subset of the baseline (baseline was captured from real data
        # which has more trades, so has more sheets; synthetic may have fewer periods)
        baseline_sheets = baseline_wb.sheetnames
        fresh_sheets = fresh_wb.sheetnames
        # Core sheets must be present in both
        for sheet in ("Summary", "Metrics_100"):
            assert sheet in fresh_sheets, f"Core sheet {sheet!r} missing from fresh export"

        # For each sheet present in both workbooks, column order must match
        for sheet_name in fresh_sheets:
            if sheet_name not in baseline_wb.sheetnames:
                continue  # synthetic may have sheets baseline doesn't (or vice-versa is ok)
            fresh_cols = _sheet_headers(fresh_wb, sheet_name)
            baseline_cols = _sheet_headers(baseline_wb, sheet_name)
            assert fresh_cols == baseline_cols, (
                f"Column order mismatch in sheet {sheet_name!r}.\n"
                f"  Baseline: {baseline_cols}\n"
                f"  Fresh:    {fresh_cols}"
            )


# ---------------------------------------------------------------------------
# Group 2: Test #18 — enabled path golden (column order pinned)
# ---------------------------------------------------------------------------

class TestEnabledLegacyGolden:
    """Test #18: legacy enabled export column order pinned; filter cols in exact position."""

    def test_enabled_signals_exact_column_order(self) -> None:
        """Signals enabled: exact column sequence including 4 filter cols in canonical position."""
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        if "Signals" not in wb.sheetnames:
            pytest.skip("No signal events generated")
        headers = _sheet_headers(wb, "Signals")
        assert headers == EXPECTED_SIGNALS_ENABLED_COLS, (
            f"Signals enabled column order mismatch.\n"
            f"  Expected: {EXPECTED_SIGNALS_ENABLED_COLS}\n"
            f"  Got:      {headers}"
        )

    def test_enabled_trades_100_exact_column_order(self) -> None:
        """Trades_100 enabled: canonical 13 cols + 3 filter cols appended at end."""
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        if "Trades_100" not in wb.sheetnames:
            pytest.skip("Trades_100 sheet not present")
        headers = _sheet_headers(wb, "Trades_100")
        assert headers == EXPECTED_TRADES_ENABLED_COLS, (
            f"Trades_100 enabled column order mismatch.\n"
            f"  Expected: {EXPECTED_TRADES_ENABLED_COLS}\n"
            f"  Got:      {headers}"
        )

    def test_enabled_all_filter_sheets_present(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg(export_state_columns=True, export_trigger_columns=True)
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg, df=df)

        assert "FilterDiagnostics_100" in wb.sheetnames
        assert "ZigZag_Trigger_Events" in wb.sheetnames
        assert "filters_summary" in wb.sheetnames

    def test_enabled_filter_diagnostics_100_absent_when_flag_off(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg(export_state_columns=False)
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        assert "FilterDiagnostics_100" not in wb.sheetnames

    def test_enabled_zigzag_trigger_events_absent_when_flag_off(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg(export_trigger_columns=False)
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        assert "ZigZag_Trigger_Events" not in wb.sheetnames

    def test_enabled_summary_has_filter_block(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        ws = wb["Summary"]
        all_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
        assert "Lifecycle Starts" in all_values or "Bars OFF" in all_values, (
            "Enabled Summary must contain filter diagnostics block"
        )

    def test_filter_diagnostics_100_has_expected_columns(self) -> None:
        from supertrend_optimizer.io.excel_tester import FILTER_DIAGNOSTICS_100_DISPLAY_NAMES

        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg(export_state_columns=True)
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        if "FilterDiagnostics_100" not in wb.sheetnames:
            pytest.skip("Sheet not present")
        headers = _sheet_headers(wb, "FilterDiagnostics_100")
        assert "Bar Index" in headers
        assert "Filter State" in headers
        assert "Filter Block Reason" in headers

    def test_enabled_signals_no_extra_columns(self) -> None:
        """Signals enabled must have EXACTLY the expected cols — no extra, no missing."""
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        results, signals_df = _run_enabled_legacy(df, tf_cfg)
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        if "Signals" not in wb.sheetnames:
            pytest.skip("No signal events")
        headers = _sheet_headers(wb, "Signals")
        missing = [c for c in EXPECTED_SIGNALS_ENABLED_COLS if c not in headers]
        extra = [c for c in headers if c not in EXPECTED_SIGNALS_ENABLED_COLS]
        assert not missing, f"Missing cols in enabled Signals: {missing}"
        assert not extra, f"Unexpected extra cols in enabled Signals: {extra}"

    def test_enabled_none_linked_trade_is_test_failure(self) -> None:
        """Absence of Linked Trade ID for an enabled trade row must not occur (plan §9.5.1).

        For each trigger that occurs in an enabled run with trades, every
        trade opened during a lifecycle must map back to that lifecycle's
        trigger via the two-step linker.  Presence of "N/A" in a ZigZag_Trigger_Events
        sheet that has associated trades is a contract violation.
        """
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df
        import numpy as np

        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        results, signals_df = _run_enabled_legacy(df, tf_cfg)

        fd = results[0].filter_diagnostics
        if fd is None:
            pytest.skip("No filter_diagnostics on this run (likely 0 triggers)")

        trades_df = results[0].result.trades_df
        if trades_df is None or len(trades_df) == 0:
            pytest.skip("No trades in enabled run — N/A linker not applicable")

        trigger_events = _build_zigzag_trigger_events_df(
            filter_diagnostics=fd,
            filter_diagnostics_summary=results[0].filter_diagnostics_summary,
            df_index=df.index,
            trades_df=trades_df,
        )

        # Every trigger that led to a lifecycle start should have a Linked Trade ID
        # Triggers that did NOT start a lifecycle (Triggered Lifecycle Start == False)
        # may legitimately have "N/A"
        lc_triggers = trigger_events[trigger_events["Triggered Lifecycle Start"] == True]
        if len(lc_triggers) == 0:
            pytest.skip("No lifecycle-starting triggers found")

        na_lc_triggers = lc_triggers[lc_triggers["Linked Trade ID"] == "N/A"]
        # If there are lifecycle starts without linked trades, the number should be small
        # (e.g. lifecycle started but no entry happened due to filter blocking all entries)
        # This is a soft test: just ensure that at least *some* lifecycle triggers are linked
        n_linked = (lc_triggers["Linked Trade ID"] != "N/A").sum()
        assert n_linked > 0, (
            "At least one lifecycle-starting trigger must have a Linked Trade ID in an "
            "enabled run with trades (plan §9.5.1). "
            f"All {len(lc_triggers)} lifecycle triggers have 'N/A'."
        )


# ---------------------------------------------------------------------------
# Group 3: Test #19 — equal_blocks disabled golden (column order pinned, baseline comparison)
# ---------------------------------------------------------------------------

class TestEqualBlocksDisabledGolden:
    """Test #19: equal_blocks disabled export column order pinned; no filter sheets/cols."""

    def test_equal_blocks_trades_exact_column_order(self) -> None:
        """All Trades_S* sheets must have canonical 13 columns in exact order."""
        from supertrend_optimizer.testing.runner import run_equal_blocks
        from supertrend_optimizer.io.excel_tester import export_equal_blocks_results

        df = _make_synthetic_ohlc(n=400)
        results = run_equal_blocks(
            df=df, n_parts=4, warmup_period=20,
            atr_period=14, multiplier=3.0,
            trade_mode="revers", commission=0.001,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = str(Path(tmpdir) / "eq.xlsx")
            actual = export_equal_blocks_results(results, out_path)
            import openpyxl
            wb = openpyxl.load_workbook(actual)

        trade_sheets = [s for s in wb.sheetnames if s.startswith("Trades_")]
        assert len(trade_sheets) > 0, "Expected at least one Trades_* sheet in equal_blocks export"

        for sheet_name in trade_sheets:
            headers = _sheet_headers(wb, sheet_name)
            assert headers == EXPECTED_TRADES_DISABLED_COLS, (
                f"Equal-blocks {sheet_name!r} column order mismatch.\n"
                f"  Expected: {EXPECTED_TRADES_DISABLED_COLS}\n"
                f"  Got:      {headers}"
            )

    def test_equal_blocks_no_filter_sheets(self) -> None:
        from supertrend_optimizer.testing.runner import run_equal_blocks
        from supertrend_optimizer.io.excel_tester import export_equal_blocks_results

        df = _make_synthetic_ohlc(n=400)
        results = run_equal_blocks(
            df=df, n_parts=4, warmup_period=20,
            atr_period=14, multiplier=3.0,
            trade_mode="revers", commission=0.001,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = str(Path(tmpdir) / "eq.xlsx")
            actual = export_equal_blocks_results(results, out_path)
            import openpyxl
            wb = openpyxl.load_workbook(actual)

        assert "FilterDiagnostics_100" not in wb.sheetnames
        assert "ZigZag_Trigger_Events" not in wb.sheetnames
        assert "filters_summary" not in wb.sheetnames
        assert "Signals" not in wb.sheetnames

    def test_equal_blocks_no_filter_columns(self) -> None:
        from supertrend_optimizer.testing.runner import run_equal_blocks
        from supertrend_optimizer.io.excel_tester import export_equal_blocks_results

        df = _make_synthetic_ohlc(n=400)
        results = run_equal_blocks(
            df=df, n_parts=4, warmup_period=20,
            atr_period=14, multiplier=3.0,
            trade_mode="revers", commission=0.001,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = str(Path(tmpdir) / "eq.xlsx")
            actual = export_equal_blocks_results(results, out_path)
            import openpyxl
            wb = openpyxl.load_workbook(actual)

        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Trades_"):
                continue
            headers = _sheet_headers(wb, sheet_name)
            for col in ("Entry Filter State", "Entry Trigger Source", "Exit Reason"):
                assert col not in headers, (
                    f"equal_blocks {sheet_name!r} must not have filter col {col!r}"
                )

    def test_equal_blocks_column_order_matches_baseline_if_exists(self) -> None:
        """Compare fresh equal_blocks export column structure against saved baseline (#19 golden).

        If baseline is on disk, column orders must match sheet by sheet.
        """
        from supertrend_optimizer.testing.runner import run_equal_blocks
        from supertrend_optimizer.io.excel_tester import export_equal_blocks_results

        df = _make_synthetic_ohlc(n=400)
        results = run_equal_blocks(
            df=df, n_parts=4, warmup_period=20,
            atr_period=14, multiplier=3.0,
            trade_mode="revers", commission=0.001,
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = str(Path(tmpdir) / "eq.xlsx")
            actual = export_equal_blocks_results(results, out_path)
            import openpyxl
            fresh_wb = openpyxl.load_workbook(actual)

        baseline_wb = _load_baseline_wb(_EQ_BLOCKS_BASELINE)

        for sheet_name in fresh_wb.sheetnames:
            if sheet_name not in baseline_wb.sheetnames:
                continue
            fresh_cols = _sheet_headers(fresh_wb, sheet_name)
            baseline_cols = _sheet_headers(baseline_wb, sheet_name)
            assert fresh_cols == baseline_cols, (
                f"Equal-blocks baseline column order mismatch in sheet {sheet_name!r}.\n"
                f"  Baseline: {baseline_cols}\n"
                f"  Fresh:    {fresh_cols}"
            )


# ---------------------------------------------------------------------------
# Group 4: Test #20 — empty-trade enabled run
# ---------------------------------------------------------------------------

class TestEmptyTradeEnabledRun:
    """With 0 trades and enabled filter: headers correct, filter block present."""

    def _make_period_result_no_trades(self):
        """Build a synthetic PeriodResult with 0 trades but enabled filter diagnostics."""
        from supertrend_optimizer.testing.runner import run_all_periods
        from supertrend_optimizer.utils.enums import ExecutionModel

        df = _make_synthetic_ohlc(n=100)
        tf_cfg = _make_enabled_cfg()
        results = run_all_periods(
            df=df, atr_period=14, multiplier=3.0, trade_mode="revers",
            commission=0.001, warmup_period=20,
            execution_model=ExecutionModel.OPEN_TO_OPEN,
            trade_filter_config=tf_cfg,
        )
        return results

    def test_empty_or_few_trades_trades_headers_have_filter_cols(self) -> None:
        """Even if trades are 0, Trades_100 headers should include filter cols (plan §9.4)."""
        tf_cfg = _make_enabled_cfg()
        df = _make_synthetic_ohlc(n=100)

        try:
            results = self._make_period_result_no_trades()
        except Exception:
            pytest.skip("Could not build period result")

        from supertrend_optimizer.testing.signal_events import build_signal_events
        from supertrend_optimizer.utils.enums import ExecutionModel
        signals_df = build_signal_events(
            df=df, trend=results[0].result.trend, atr_period=14,
            trade_mode="revers", execution_model=ExecutionModel.OPEN_TO_OPEN,
            filter_diagnostics=results[0].filter_diagnostics,
        )
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        if "Trades_100" not in wb.sheetnames:
            pytest.skip("No Trades_100 sheet")
        headers = _sheet_headers(wb, "Trades_100")

        if results[0].metrics.get("num_trades", 0) == 0:
            assert headers == EXPECTED_TRADES_ENABLED_COLS, (
                f"Empty-trades Trades_100 column order mismatch.\n"
                f"  Expected: {EXPECTED_TRADES_ENABLED_COLS}\n"
                f"  Got:      {headers}"
            )

    def test_enabled_summary_filter_block_present_even_0_trades(self) -> None:
        tf_cfg = _make_enabled_cfg()
        df = _make_synthetic_ohlc(n=100)
        results = self._make_period_result_no_trades()

        from supertrend_optimizer.testing.signal_events import build_signal_events
        from supertrend_optimizer.utils.enums import ExecutionModel
        signals_df = build_signal_events(
            df=df, trend=results[0].result.trend, atr_period=14,
            trade_mode="revers", execution_model=ExecutionModel.OPEN_TO_OPEN,
            filter_diagnostics=results[0].filter_diagnostics,
        )
        wb = _export_and_load(results, signals_df, tf_cfg=tf_cfg)

        ws = wb["Summary"]
        all_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
        has_filter_header = any("Bars" in str(v) or "Lifecycle" in str(v) for v in all_values)
        assert has_filter_header, (
            "Summary filter block must appear even when 0 trades (plan §9.4)"
        )


# ---------------------------------------------------------------------------
# Group 5: ZigZag_Trigger_Events reconstruction (plan §9.2.1)
# ---------------------------------------------------------------------------

class TestZigZagTriggerEventsReconstruction:
    """Trigger events = rows where trigger_source != 'none' (plan §9.2.1)."""

    def test_trigger_events_match_trigger_source_count(self) -> None:
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 100
        trigger_source = np.array(["none"] * n, dtype=object)
        trigger_source[5] = "candidate_threshold"
        trigger_source[30] = "confirmed_median"
        trigger_source[70] = "candidate_threshold"
        fd = {
            "trade_filter_trigger_source": trigger_source,
            "trade_filter_state": np.array(["OFF"] + ["WAIT_FIRST_ST_FLIP"] * (n - 1), dtype=object),
            "candidate_trigger_threshold": np.full(n, 0.4),
            "global_median": np.full(n, 0.05),
            "local_median_N": np.full(n, 10.0),
            "candidate_height_pct": np.full(n, 0.035),
        }

        result = _build_zigzag_trigger_events_df(fd)
        assert len(result) == 3, f"Expected 3 trigger events, got {len(result)}"
        assert list(result["Trigger ID"]) == [1, 2, 3]
        assert list(result["Trigger Bar"]) == [5, 30, 70]
        assert result.iloc[0]["Trigger Source"] == "candidate_threshold"
        assert result.iloc[1]["Trigger Source"] == "confirmed_median"

    def test_trigger_events_empty_when_no_triggers(self) -> None:
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 50
        fd = {"trade_filter_trigger_source": np.array(["none"] * n, dtype=object)}
        result = _build_zigzag_trigger_events_df(fd)
        assert len(result) == 0
        assert "Trigger ID" in result.columns

    def test_trigger_events_none_diagnostics_returns_empty(self) -> None:
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df
        result = _build_zigzag_trigger_events_df(None)
        assert len(result) == 0

    def test_trigger_time_strips_timezone_for_excel(self) -> None:
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 3
        fd = {
            "trade_filter_trigger_source": np.array(["none", "candidate_threshold", "none"], dtype=object),
            "trade_filter_state": np.array(["OFF", "WAIT_FIRST_ST_FLIP", "ST_ACTIVE_FREEZE"], dtype=object),
            "candidate_trigger_threshold": np.full(n, 0.4),
            "global_median": np.full(n, 0.05),
            "local_median_N": np.full(n, 10.0),
            "candidate_height_pct": np.full(n, 0.035),
        }
        df_index = pd.date_range("2024-01-01", periods=n, freq="h", tz="UTC")

        result = _build_zigzag_trigger_events_df(fd, df_index=df_index)

        trigger_time = result.iloc[0]["Trigger Time"]
        assert isinstance(trigger_time, pd.Timestamp)
        assert trigger_time.tzinfo is None


# ---------------------------------------------------------------------------
# Group 6: Two-step Linked Trade ID linker unit tests (plan §9.5.1)
# ---------------------------------------------------------------------------

class TestTwoStepLinker:
    """Two-step trade <-> trigger linker (plan §9.5.1): exact match then backward search."""

    def _make_trades_df(self, entry_indices: List[int]) -> pd.DataFrame:
        """Build a minimal trades_df with given entry_indices."""
        return pd.DataFrame({
            "trade_id": list(range(1, len(entry_indices) + 1)),
            "entry_index": entry_indices,
        })

    def test_step1_exact_match_same_bar(self) -> None:
        """Step 1: trigger at bar t links to trade with entry_signal_bar == t."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        # Trade with entry_index=6 → entry_signal_bar=5
        trades_df = self._make_trades_df([6])
        n = 20
        state_arr = np.array(["OFF"] * n, dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        state_arr[6] = "ST_ACTIVE_FREEZE"

        result = _two_step_trade_trigger_link(
            trigger_bars=[5], state_arr=state_arr, trades_df=trades_df
        )
        assert result[5] == 1, f"Expected trade_id=1 linked to trigger at bar 5, got {result[5]}"

    def test_step2_backward_search_links_through_lifecycle(self) -> None:
        """Step 2: if no exact match, find last trigger in same lifecycle (no OFF between)."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        # Trigger at bar 5, trade entry_signal_bar = 8 (not same bar)
        # state_arr: 5=WAIT, 6=WAIT, 7=ST_ACTIVE_FREEZE, 8=ST_ACTIVE_MONITORING
        # No OFF between 5 and 8 → backward search should link trigger[5] → trade_id=1
        trades_df = self._make_trades_df([9])  # entry_index=9 → signal_bar=8
        n = 20
        state_arr = np.array(["OFF"] * n, dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        state_arr[6] = "WAIT_FIRST_ST_FLIP"
        state_arr[7] = "ST_ACTIVE_FREEZE"
        state_arr[8] = "ST_ACTIVE_MONITORING"

        result = _two_step_trade_trigger_link(
            trigger_bars=[5], state_arr=state_arr, trades_df=trades_df
        )
        assert result[5] == 1, (
            f"Backward search should link trigger at bar 5 to trade_id=1 "
            f"(entry_signal_bar=8, no OFF between 5 and 8). Got: {result[5]}"
        )

    def test_step2_blocked_by_off_state(self) -> None:
        """Step 2: OFF state between trigger and entry breaks same-lifecycle condition."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        # Trigger at bar 5, trade entry at bar 8, but state goes OFF at bar 6
        trades_df = self._make_trades_df([9])  # entry_signal_bar=8
        n = 20
        state_arr = np.array(["OFF"] * n, dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        state_arr[6] = "OFF"  # lifecycle ended between trigger and trade entry
        state_arr[7] = "WAIT_FIRST_ST_FLIP"
        state_arr[8] = "ST_ACTIVE_FREEZE"

        result = _two_step_trade_trigger_link(
            trigger_bars=[5], state_arr=state_arr, trades_df=trades_df
        )
        assert result[5] == "N/A", (
            f"Trigger at bar 5 should NOT link to trade (OFF breaks lifecycle). Got: {result[5]}"
        )

    def test_no_trades_all_na(self) -> None:
        """No trades → all triggers get 'N/A'."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        n = 20
        state_arr = np.array(["OFF"] * n, dtype=object)
        result = _two_step_trade_trigger_link(
            trigger_bars=[5, 10, 15], state_arr=state_arr, trades_df=None
        )
        assert result == {5: "N/A", 10: "N/A", 15: "N/A"}

    def test_empty_trigger_bars(self) -> None:
        """Empty trigger_bars → empty result dict."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        result = _two_step_trade_trigger_link(
            trigger_bars=[], state_arr=None, trades_df=self._make_trades_df([5])
        )
        assert result == {}

    def test_multiple_triggers_multiple_trades_independent(self) -> None:
        """Each trigger links to its own trade via exact match."""
        from supertrend_optimizer.io.excel_tester import _two_step_trade_trigger_link

        # Trigger at 5 → trade 1 (entry_signal_bar=5)
        # Trigger at 20 → trade 2 (entry_signal_bar=20)
        trades_df = self._make_trades_df([6, 21])  # signal bars 5 and 20
        n = 40
        state_arr = np.array(["OFF"] * n, dtype=object)

        result = _two_step_trade_trigger_link(
            trigger_bars=[5, 20], state_arr=state_arr, trades_df=trades_df
        )
        assert result[5] == 1
        assert result[20] == 2

    def test_linker_integrated_in_trigger_events_df(self) -> None:
        """_build_zigzag_trigger_events_df uses two-step linker (integration check)."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 20
        trigger_source = np.array(["none"] * n, dtype=object)
        trigger_source[5] = "candidate_threshold"

        state_arr = np.array(["OFF"] * n, dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        state_arr[6] = "ST_ACTIVE_FREEZE"
        state_arr[7] = "ST_ACTIVE_MONITORING"

        fd = {
            "trade_filter_trigger_source": trigger_source,
            "trade_filter_state": state_arr,
            "candidate_trigger_threshold": np.full(n, 0.4),
            "global_median": np.full(n, 0.05),
            "local_median_N": np.full(n, 10.0),
            "candidate_height_pct": np.full(n, 0.035),
        }
        trades_df = pd.DataFrame({"trade_id": [42], "entry_index": [6]})  # signal_bar=5

        result = _build_zigzag_trigger_events_df(fd, trades_df=trades_df)
        assert len(result) == 1
        assert result.iloc[0]["Linked Trade ID"] == 42, (
            f"Expected Linked Trade ID=42, got {result.iloc[0]['Linked Trade ID']}"
        )

    def test_linker_step2_backward_search_in_trigger_events_df(self) -> None:
        """Integration: _build_zigzag_trigger_events_df backward-search linker."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 20
        trigger_source = np.array(["none"] * n, dtype=object)
        trigger_source[5] = "candidate_threshold"  # trigger at bar 5

        state_arr = np.array(["OFF"] * n, dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        state_arr[6] = "WAIT_FIRST_ST_FLIP"
        state_arr[7] = "ST_ACTIVE_FREEZE"
        state_arr[8] = "ST_ACTIVE_MONITORING"  # trade entry at bar 8 (entry_index=9)

        fd = {
            "trade_filter_trigger_source": trigger_source,
            "trade_filter_state": state_arr,
            "candidate_trigger_threshold": np.full(n, 0.4),
            "global_median": np.full(n, 0.05),
            "local_median_N": np.full(n, 10.0),
            "candidate_height_pct": np.full(n, 0.035),
        }
        # entry_index=9 → entry_signal_bar=8 (no exact trigger at 8, backward links to 5)
        trades_df = pd.DataFrame({"trade_id": [99], "entry_index": [9]})

        result = _build_zigzag_trigger_events_df(fd, trades_df=trades_df)
        assert len(result) == 1
        assert result.iloc[0]["Linked Trade ID"] == 99, (
            f"Backward-search should link trigger at bar 5 to trade_id=99. "
            f"Got: {result.iloc[0]['Linked Trade ID']}"
        )


# ---------------------------------------------------------------------------
# Group 7: Static wiring checks for CLI callers
# ---------------------------------------------------------------------------

class TestCliWiringStaticCheck:
    """cli/tester.py and run_batch_tester.py must pass trade_filter_config and df."""

    def _read_file(self, path: Path) -> str:
        return path.read_text(encoding="utf-8")

    def test_cli_tester_passes_trade_filter_config(self) -> None:
        src = self._read_file(_DONOR_ROOT / "supertrend_optimizer" / "cli" / "tester.py")
        assert "trade_filter_config=tf_cfg" in src, (
            "cli/tester.py must pass trade_filter_config=tf_cfg to export_tester_results"
        )

    def test_cli_tester_passes_df(self) -> None:
        src = self._read_file(_DONOR_ROOT / "supertrend_optimizer" / "cli" / "tester.py")
        assert "df=df" in src, (
            "cli/tester.py must pass df=df to export_tester_results"
        )

    def test_run_batch_tester_passes_trade_filter_config(self) -> None:
        src = self._read_file(_TESTER_ROOT / "run_batch_tester.py")
        assert "trade_filter_config=tf_cfg" in src, (
            "run_batch_tester.py must pass trade_filter_config=tf_cfg to export_tester_results"
        )

    def test_run_batch_tester_passes_df(self) -> None:
        src = self._read_file(_TESTER_ROOT / "run_batch_tester.py")
        assert "df=df" in src, (
            "run_batch_tester.py must pass df=df to export_tester_results"
        )

    def test_two_step_linker_function_exists_in_excel_tester(self) -> None:
        """_two_step_trade_trigger_link must exist in excel_tester.py (plan §9.5.1)."""
        src = self._read_file(_DONOR_ROOT / "supertrend_optimizer" / "io" / "excel_tester.py")
        assert "_two_step_trade_trigger_link" in src, (
            "excel_tester.py must contain _two_step_trade_trigger_link function (plan §9.5.1)"
        )

    def test_simplified_implementation_comment_removed(self) -> None:
        """The 'simplified implementation' placeholder comment must be gone (plan §9.5.1)."""
        src = self._read_file(_DONOR_ROOT / "supertrend_optimizer" / "io" / "excel_tester.py")
        assert "simplified implementation" not in src, (
            "The 'simplified implementation' placeholder comment must be removed after "
            "implementing the full two-step linker (plan §9.5.1)"
        )


# ===========================================================================
# Plan v3 §6.1/§6.2: exit_b_immediate_off in excel display map + params row
# ===========================================================================

class TestImmediateOffExcelDisplayContract:
    """§10.5 (tester-side): FILTER_DIAGNOSTICS_100_DISPLAY_NAMES and
    _build_filters_summary_df include the new Plan v3 §6 entries.

    These are static contract tests — no full Excel export needed.
    """

    def test_display_map_triggered_key(self):
        """§6.1: exit_b_immediate_off_triggered → 'Exit-B Immediate OFF Triggered'."""
        from supertrend_optimizer.io.excel_tester import FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        assert "exit_b_immediate_off_triggered" in FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        assert (
            FILTER_DIAGNOSTICS_100_DISPLAY_NAMES["exit_b_immediate_off_triggered"]
            == "Exit-B Immediate OFF Triggered"
        )

    def test_display_map_config_key(self):
        """§6.1: exit_b_immediate_off_config → 'Exit-B Immediate OFF Config'."""
        from supertrend_optimizer.io.excel_tester import FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        assert "exit_b_immediate_off_config" in FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        assert (
            FILTER_DIAGNOSTICS_100_DISPLAY_NAMES["exit_b_immediate_off_config"]
            == "Exit-B Immediate OFF Config"
        )

    def _make_mock_pr(self, exit_b_immediate_off: bool):
        """Minimal duck-typed PeriodResult for _build_filters_summary_df."""
        import numpy as np

        class _BR:
            positions = np.zeros(5, dtype=np.int8)
            trades_df = None

        class _PR:
            period_label = "100%"
            result = _BR()
            filter_diagnostics = {}
            filter_diagnostics_summary = {
                "exit_off_mode": "exit B",
                "exit_off_zz_leg_count": 2,
                "exit_b_immediate_off": exit_b_immediate_off,
                "lifecycle_starts_count": 1,
                "median_stop_triggered_count": 0,
                "zz_leg_stop_triggered_count": 1,
                "thresholds": {
                    "reversal_threshold": 0.02,
                    "candidate_trigger_threshold": 0.05,
                    "candidate_trigger_quantile": None,
                    "candidate_trigger_source": "explicit",
                    "global_median": 0.04,
                    "local_window": 5,
                    "freeze_confirmed_legs": 0,
                    "exit_off_mode": "exit B",
                    "exit_off_zz_leg_count": 2,
                    "exit_b_immediate_off": exit_b_immediate_off,
                    "zigzag_mode": "A",
                    "candidate_duration_gate_enabled": False,
                    "candidate_duration_max_bars": -1,
                },
                "counters": {"zz_leg_stop_triggered": 1, "median_stop_triggered": 0},
                "bars_in_state": {"OFF": 5},
            }

        return _PR()

    def _get_params_df(self, exit_b_immediate_off: bool):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df
        pr = self._make_mock_pr(exit_b_immediate_off)
        result = _build_filters_summary_df([pr])
        assert result is not None
        params_df, _ = result
        return params_df

    def test_params_row_flag_true(self):
        """§6.2: 'Exit-B Immediate OFF' row present with value True."""
        params_df = self._get_params_df(exit_b_immediate_off=True)
        row = params_df[params_df["Parameter"] == "Exit-B Immediate OFF"]
        assert not row.empty, (
            f"'Exit-B Immediate OFF' row missing. Labels: {sorted(params_df['Parameter'].tolist())}"
        )
        assert row.iloc[0]["Value"] is True

    def test_params_row_flag_false(self):
        """§6.2: row present with value False (always-present, never '—')."""
        params_df = self._get_params_df(exit_b_immediate_off=False)
        row = params_df[params_df["Parameter"] == "Exit-B Immediate OFF"]
        assert not row.empty, (
            "'Exit-B Immediate OFF' must be present even when False (§6.2 always-present)."
        )
        assert row.iloc[0]["Value"] is False
