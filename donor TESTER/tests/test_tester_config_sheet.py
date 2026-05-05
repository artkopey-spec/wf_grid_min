"""Regression: Tester_Config sheet (first sheet, YAML snapshot + run metadata)."""

from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from supertrend_optimizer.engine.result import BacktestResult
from supertrend_optimizer.io import excel_tester
from supertrend_optimizer.testing.runner import PERIOD_SPLITS, PeriodResult, SegmentResult


def _minimal_period_results(*, effective_warmup: int = 1):
    results = []
    for label, frac in PERIOD_SPLITS:
        n = max(30, int(100 * frac))
        result = BacktestResult(
            atr_period=14,
            multiplier=3.0,
            trade_mode="revers",
            commission=0.0,
            warmup=5,
            returns=np.zeros(n - 1, dtype=np.float64),
            equity_curve=np.ones(n, dtype=np.float64),
            positions=np.zeros(n, dtype=np.int8),
            trend=np.zeros(n, dtype=np.int8),
            metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
            early_exit=False,
            exit_bar=None,
            exit_drawdown=None,
            trades_df=pd.DataFrame(),
            n_bars_original=n,
            period_label=label,
            effective_warmup=effective_warmup,
        )
        results.append(
            PeriodResult(
                period_label=label,
                n_bars=n,
                result=result,
                filter_diagnostics=None,
                filter_diagnostics_summary=None,
            )
        )
    return results


def test_tester_config_legacy_first_sheet_early_exit_and_warmup(tmp_path):
    out = tmp_path / "out.xlsx"
    raw = {"early_exit": False, "periods_per_year": 200, "commission": 0.01}
    run_meta = {
        "resolved_periods_per_year": 252.5,
        "warmup_period_resolved": 99,
        "warmup_period_effective": 7,
    }
    actual = excel_tester.export_tester_results(
        _minimal_period_results(effective_warmup=7),
        str(out),
        signals_df=None,
        config_yaml_snapshot=raw,
        run_metadata=run_meta,
    )

    wb = load_workbook(actual, read_only=True)
    try:
        assert wb.sheetnames[0] == excel_tester.TESTER_CONFIG_SHEET_NAME
        ws = wb[excel_tester.TESTER_CONFIG_SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("Section", "Parameter", "Value")
        flat = {(r[0], r[1]): r[2] for r in rows[1:] if r[0]}
        assert flat.get(("config_file", "early_exit")) == "false"
        assert flat.get(("config_file", "periods_per_year")) == "200"
        assert flat.get(("run", "resolved_periods_per_year")) == "252.5"
        assert flat.get(("run", "warmup_period_resolved")) == "99"
        assert flat.get(("run", "warmup_period_effective")) == "7"
    finally:
        wb.close()


def test_tester_config_empty_yaml_root_mapping_note(tmp_path):
    out = tmp_path / "empty_root.xlsx"
    actual = excel_tester.export_tester_results(
        _minimal_period_results(),
        str(out),
        signals_df=None,
        config_yaml_snapshot={},
        run_metadata={"k": 1},
    )
    wb = load_workbook(actual, read_only=True)
    try:
        ws = wb[excel_tester.TESTER_CONFIG_SHEET_NAME]
        text = " ".join(
            str(c) for row in ws.iter_rows(values_only=True) for c in row if c
        )
        assert "YAML root mapping is empty" in text
    finally:
        wb.close()


def test_tester_config_equal_blocks_first_sheet(tmp_path):
    seg = SegmentResult(
        segment_label="S1",
        segment_index=0,
        n_parts=3,
        range_label="0–33%",
        start_bar=0,
        end_bar=10,
        n_bars=10,
        prepend_bars=0,
        atr_period=14,
        multiplier=3.0,
        trade_mode="revers",
        commission=0.0,
        segment_metrics={"sum_pnl_pct": 0.0, "num_trades": 0},
        segment_trades_df=None,
        start_date=None,
        end_date=None,
        ext_slice_effective_warmup=4,
    )

    out = tmp_path / "eq.xlsx"
    actual = excel_tester.export_equal_blocks_results(
        [seg],
        str(out),
        config_yaml_snapshot={"periods_per_year": 100},
        run_metadata={
            "resolved_periods_per_year": 365.0,
            "warmup_period_resolved": 14,
            "warmup_period_effective": seg.ext_slice_effective_warmup,
        },
    )
    wb = load_workbook(actual, read_only=True)
    try:
        assert wb.sheetnames[0] == excel_tester.TESTER_CONFIG_SHEET_NAME
        rows = list(wb[excel_tester.TESTER_CONFIG_SHEET_NAME].iter_rows(values_only=True))
        flat = {(r[0], r[1]): r[2] for r in rows[1:] if r[0]}
        assert flat[("config_file", "periods_per_year")] == "100"
        assert flat[("run", "resolved_periods_per_year")] == "365.0"
        assert flat[("run", "warmup_period_effective")] == "4"
    finally:
        wb.close()
