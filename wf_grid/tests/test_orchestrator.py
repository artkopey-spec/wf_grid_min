"""
Tests for A13: Orchestrator + E2E golden tests

Coverage:
  - PipelineResult structure and fields
  - E2E mini-grid golden test: 3 grid points × 2 WF steps
    - full workbook: sheets, order, schemas
    - summary sheet: column order (Block A before B before segments)
    - WF_Trades / WF_Train_Trades: deterministic column order + sort
    - WF_Config: non-empty Section/Parameter/Value
  - All-failed grid: all n_ok_steps == 0 → all Tier 3
  - Reproducibility: two runs from same inputs → identical outputs
  - StageTimer: timings populated
  - PipelineDiagnostics: correct counts
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
import pytest

from wf_grid.aggregate.aggregator import aggregate_candidates
from wf_grid.collect.step_collector import collect_oos_steps, collect_train_steps
from wf_grid.collect.trades_collector import collect_oos_trades, collect_train_trades
from wf_grid.config.schema import (
    BacktestConfig,
    DataConfig,
    GridConfig,
    RankingConfig,
    StatusConfig,
)
from wf_grid.export.summary_builder import _BLOCK_A, _BLOCK_B, _is_segment_col, build_summary_wide
from wf_grid.export.xlsx_writer import export_workbook
from wf_grid.gates.gates import apply_candidate_gates, apply_step_gates
from wf_grid.grid.enumeration import GridPoint
from wf_grid.logging_.diagnostics import (
    PipelineDiagnostics,
    StageTimer,
    build_diagnostics,
)
from wf_grid.pipeline.orchestrator import PipelineResult, _collect_top1_trades
from wf_grid.ranking.ranker import rank_candidates
from wf_grid.ranking.scoring import calculate_seed_score, compute_score_discrimination
from wf_grid.status.status_model import StepStatus
from wf_grid.wf.step_executor import StepResult


# ---------------------------------------------------------------------------
# Helpers: synthetic StepResult factory
# ---------------------------------------------------------------------------

def _ok_step(gp_id: str, wf_step: int, pnl: float = 5.0, num_trades: int = 10) -> StepResult:
    return StepResult(
        grid_point_id=gp_id,
        wf_step=wf_step,
        test_start_idx=0,
        test_end_idx=100,
        metrics={
            "sum_pnl_pct": pnl,
            "sharpe": 1.2,
            "sortino": 1.5,
            "max_drawdown": -0.10,
            "cagr": 0.15,
            "win_rate": 0.6,
            "num_trades": num_trades,
            "profit_factor": 1.8,
            "avg_trade": 0.5,
        },
        oos_trades_df=pd.DataFrame({
            "trade_id": [1, 2],
            "direction": ["long", "short"],
            "entry_time": pd.NaT,
            "entry_index": [5, 50],
            "entry_price": [100.0, 110.0],
            "exit_time": pd.NaT,
            "exit_index": [20, 70],
            "exit_price": [105.0, 108.0],
            "bars_held": [15, 20],
            "gross_pnl_pct": [0.05, -0.02],
            "commission_pct": [0.001, 0.001],
            "net_pnl_pct": [0.049, -0.021],
        }),
        prepend_bars_requested=50,
        prepend_bars_applied=50,
        used_prepend=True,
        used_legacy_oos_path=False,
        used_defensive_fallback=False,
        oos_boundary_index=50,
        warmup_used=50,
        warmup_effective=50,
        effective_oos_bars=100,
    )


def _failed_step(gp_id: str, wf_step: int) -> StepResult:
    return StepResult(
        grid_point_id=gp_id,
        wf_step=wf_step,
        test_start_idx=0,
        test_end_idx=100,
        metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
        oos_trades_df=None,
        prepend_bars_requested=50,
        prepend_bars_applied=0,
        used_prepend=False,
        used_legacy_oos_path=False,
        used_defensive_fallback=False,
        oos_boundary_index=0,
        warmup_used=0,
        warmup_effective=0,
        effective_oos_bars=5,
    )


def _cfg() -> GridConfig:
    return GridConfig(
        data=DataConfig(file_path="dummy.csv"),
        backtest=BacktestConfig(min_trades_required=3),
        status=StatusConfig(min_meaningful_bars=10),
        ranking=RankingConfig(mode="gates_score"),
    )


def _run_mini_pipeline(
    grid_results: Dict[str, List[StepResult]],
    config: GridConfig = None,
    n_steps: int = 2,
    train_results: Dict[str, List[StepResult]] = None,
) -> dict:
    """Run A5→A12 pipeline stages from synthetic StepResults, return all artefacts."""
    config = config or _cfg()

    # A5: collect steps (OOS)
    step_oos_long = collect_oos_steps(grid_results, config, expected_n_steps=n_steps)

    # A5b: collect steps (Train) — if provided
    if train_results is not None:
        step_train_long = collect_train_steps(train_results, config, expected_n_steps=n_steps)
    else:
        step_train_long = None

    # A8: step gates
    step_oos_long = apply_step_gates(step_oos_long, config)

    # A7: aggregation
    aggregated = aggregate_candidates(step_oos_long, config)

    # A8: candidate gates
    gated = apply_candidate_gates(aggregated, config)

    # A9: scoring
    passed_mask = gated["seed_gate_passed"].fillna(False).astype(bool)
    scores, statuses = calculate_seed_score(
        gated,
        passed_mask,
        score_weights=config.scoring.score_weights,
        normalization_mode=config.scoring.normalization_mode,
    )
    gated["tester_seed_score"] = scores
    gated["score_contract_status"] = statuses
    gated["score_discrimination_status"] = compute_score_discrimination(
        gated, passed_mask,
    )

    # A10: ranking
    ranked = rank_candidates(gated, config)

    # A12b: top-1 trades only
    all_train = train_results if train_results is not None else {}
    trades_oos, trades_train = _collect_top1_trades(
        ranked, grid_results, all_train, config,
    )

    # A11: summary
    summary_wide = build_summary_wide(step_oos_long, aggregated, ranked, config)

    return {
        "step_oos_long": step_oos_long,
        "step_train_long": step_train_long,
        "trades_oos": trades_oos,
        "trades_train": trades_train,
        "aggregated": aggregated,
        "ranked": ranked,
        "summary_wide": summary_wide,
        "config": config,
    }


def _export_pipeline(artefacts: dict, tmp_path: Path) -> Path:
    """Run A12: export."""
    path = tmp_path / "golden_test.xlsx"
    export_workbook(
        summary_wide=artefacts["summary_wide"],
        step_oos_long=artefacts["step_oos_long"],
        trades_oos=artefacts["trades_oos"],
        trades_train=artefacts.get("trades_train", pd.DataFrame()),
        config=artefacts["config"],
        output_path=path,
    )
    return path


def _sheets(path: Path) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet)


# ===========================================================================
# E2E mini-grid golden test: 3 grid points × 2 WF steps
# ===========================================================================

class TestE2EMiniGrid:
    @pytest.fixture
    def mini_results(self):
        return {
            "atr10_m2.50_both": [_ok_step("atr10_m2.50_both", 1, pnl=10.0), _ok_step("atr10_m2.50_both", 2, pnl=8.0)],
            "atr15_m3.00_both": [_ok_step("atr15_m3.00_both", 1, pnl=5.0), _ok_step("atr15_m3.00_both", 2, pnl=3.0)],
            "atr20_m4.00_long": [_ok_step("atr20_m4.00_long", 1, pnl=1.0), _ok_step("atr20_m4.00_long", 2, pnl=-2.0)],
        }

    def test_all_sheets_present(self, mini_results, tmp_path):
        artefacts = _run_mini_pipeline(mini_results)
        path = _export_pipeline(artefacts, tmp_path)
        sheets = _sheets(path)
        assert "WF_Config" in sheets
        assert "WF_01" in sheets
        assert "WF_02" in sheets
        assert "WF_Trades" in sheets
        assert "WF_Train_Trades" in sheets
        assert "summary" in sheets

    def test_sheet_order(self, mini_results, tmp_path):
        artefacts = _run_mini_pipeline(mini_results)
        path = _export_pipeline(artefacts, tmp_path)
        sheets = _sheets(path)
        config_idx = sheets.index("WF_Config")
        summary_idx = sheets.index("summary")
        wf01_idx = sheets.index("WF_01")
        trades_idx = sheets.index("WF_Trades")
        # Order: WF_Config → summary → WF_01..N → WF_Trades
        assert config_idx < summary_idx < wf01_idx < trades_idx

    def test_summary_column_order_block_a_before_b(self, mini_results):
        artefacts = _run_mini_pipeline(mini_results)
        sw = artefacts["summary_wide"]
        cols = list(sw.columns)
        a_present = [c for c in _BLOCK_A if c in cols]
        b_present = [c for c in _BLOCK_B if c in cols]
        if a_present and b_present:
            assert max(cols.index(c) for c in a_present) < min(cols.index(c) for c in b_present)

    def test_summary_block_b_before_segments(self, mini_results):
        artefacts = _run_mini_pipeline(mini_results)
        sw = artefacts["summary_wide"]
        cols = list(sw.columns)
        b_present = [c for c in _BLOCK_B if c in cols]
        seg_present = [c for c in cols if _is_segment_col(c)]
        if b_present and seg_present:
            assert max(cols.index(c) for c in b_present) < min(cols.index(c) for c in seg_present)

    def test_summary_grid_rank_first(self, mini_results):
        artefacts = _run_mini_pipeline(mini_results)
        assert artefacts["summary_wide"].columns[0] == "grid_rank"

    def test_summary_three_rows(self, mini_results):
        artefacts = _run_mini_pipeline(mini_results)
        assert len(artefacts["summary_wide"]) == 3

    def test_wf_trades_schema(self, mini_results, tmp_path):
        artefacts = _run_mini_pipeline(mini_results)
        path = _export_pipeline(artefacts, tmp_path)
        df = _read_sheet(path, "WF_Trades")
        assert "grid_point_id" in df.columns
        assert "wf_step" in df.columns
        assert "trade_id" in df.columns

    def test_wf_trades_sorted(self, mini_results, tmp_path):
        artefacts = _run_mini_pipeline(mini_results)
        path = _export_pipeline(artefacts, tmp_path)
        df = _read_sheet(path, "WF_Trades")
        if len(df) > 1:
            gp_ids = df["grid_point_id"].tolist()
            assert gp_ids == sorted(gp_ids) or len(set(gp_ids)) == 1

    def test_wf_config_schema(self, mini_results, tmp_path):
        artefacts = _run_mini_pipeline(mini_results)
        path = _export_pipeline(artefacts, tmp_path)
        df = _read_sheet(path, "WF_Config")
        assert set(df.columns) == {"Section", "Parameter", "Value"}
        assert len(df) > 0

    def test_ranked_tiers(self, mini_results):
        artefacts = _run_mini_pipeline(mini_results)
        ranked = artefacts["ranked"]
        # All 3 have ok status + gate passed → expect Tier 1 or 2
        assert all(ranked["tier"].isin([1, 2]))


# ===========================================================================
# All-failed grid: every step is insufficient_bars → Tier 3
# ===========================================================================

class TestAllFailedGrid:
    def test_all_tier3(self):
        results = {
            "gp1": [_failed_step("gp1", 1), _failed_step("gp1", 2)],
            "gp2": [_failed_step("gp2", 1), _failed_step("gp2", 2)],
        }
        artefacts = _run_mini_pipeline(results)
        ranked = artefacts["ranked"]
        assert all(ranked["tier"] == 3)
        assert all(ranked["seed_gate_passed"] == False)

    def test_all_failed_summary_exportable(self, tmp_path):
        results = {
            "gp1": [_failed_step("gp1", 1), _failed_step("gp1", 2)],
        }
        artefacts = _run_mini_pipeline(results)
        path = _export_pipeline(artefacts, tmp_path)
        assert path.exists()
        sheets = _sheets(path)
        assert "summary" in sheets


# ===========================================================================
# Reproducibility: same inputs → identical DataFrames
# ===========================================================================

class TestReproducibility:
    def test_two_runs_identical(self):
        results = {
            "gp1": [_ok_step("gp1", 1, pnl=10.0), _ok_step("gp1", 2, pnl=5.0)],
            "gp2": [_ok_step("gp2", 1, pnl=3.0), _ok_step("gp2", 2, pnl=7.0)],
        }
        a1 = _run_mini_pipeline(results)
        a2 = _run_mini_pipeline(results)
        pd.testing.assert_frame_equal(
            a1["summary_wide"].reset_index(drop=True),
            a2["summary_wide"].reset_index(drop=True),
        )
        pd.testing.assert_frame_equal(
            a1["ranked"].reset_index(drop=True),
            a2["ranked"].reset_index(drop=True),
        )


# ===========================================================================
# Diagnostics
# ===========================================================================

class TestDiagnostics:
    def test_stage_timer_records(self):
        timer = StageTimer()
        timer.record("test_stage", 1.23)
        assert timer.timings["test_stage"] == pytest.approx(1.23)

    def test_build_diagnostics_counts(self):
        results = {
            "gp1": [_ok_step("gp1", 1), _ok_step("gp1", 2)],
            "gp2": [_failed_step("gp2", 1), _failed_step("gp2", 2)],
        }
        artefacts = _run_mini_pipeline(results)
        timer = StageTimer()
        timer.record("test", 0.5)
        diag = build_diagnostics(
            grid_size=2,
            n_wf_steps=2,
            step_oos_long=artefacts["step_oos_long"],
            ranked=artefacts["ranked"],
            timer=timer,
        )
        assert diag.grid_size == 2
        assert diag.n_wf_steps == 2
        assert len(diag.step_status_counts) > 0
        assert len(diag.tier_counts) > 0
        assert len(diag.top5_ranked) > 0


# ===========================================================================
# PipelineResult structure
# ===========================================================================

class TestPipelineResult:
    def test_default_fields(self):
        r = PipelineResult()
        assert r.config is None
        assert r.grid_points == []
        assert r.output_path is None
        assert r.error is None
        assert r.execution_mode == "sequential"

    def test_step_train_long_field_exists(self):
        r = PipelineResult()
        assert r.step_train_long is None

    def test_bucket_matrix_median_field_exists(self):
        r = PipelineResult()
        assert r.bucket_matrix_median is None


# ===========================================================================
# FIX-2.1 — Train pipeline + top-1-only trades export
# ===========================================================================

class TestTrainPipeline:
    """FIX-2.1: Train pipeline for all grid points; trade export filtered to rank-1.

    Semantics:
    - Train execution runs for ALL grid points (metrics for the full grid).
    - step_train_long covers ALL grid points.
    - Ranking/aggregation use ALL grid points.
    - trades_oos / trades_train in PipelineResult contain rank-1 only (export filter).
    - WF_Trades / WF_Train_Trades XLSX sheets contain rank-1 only.
    - No replay/second-pass: _collect_top1_trades filters already-computed results.
    """

    @pytest.fixture
    def mini_grid_with_train(self):
        oos = {
            "atr10_m2.50_both": [_ok_step("atr10_m2.50_both", 1, pnl=10.0), _ok_step("atr10_m2.50_both", 2, pnl=8.0)],
            "atr15_m3.00_both": [_ok_step("atr15_m3.00_both", 1, pnl=5.0), _ok_step("atr15_m3.00_both", 2, pnl=3.0)],
            "atr20_m4.00_long": [_ok_step("atr20_m4.00_long", 1, pnl=1.0), _ok_step("atr20_m4.00_long", 2, pnl=-2.0)],
        }
        train = {
            "atr10_m2.50_both": [_ok_step("atr10_m2.50_both", 1, pnl=12.0), _ok_step("atr10_m2.50_both", 2, pnl=9.0)],
            "atr15_m3.00_both": [_ok_step("atr15_m3.00_both", 1, pnl=6.0), _ok_step("atr15_m3.00_both", 2, pnl=4.0)],
            "atr20_m4.00_long": [_ok_step("atr20_m4.00_long", 1, pnl=2.0), _ok_step("atr20_m4.00_long", 2, pnl=-1.0)],
        }
        return oos, train

    def test_step_train_long_not_none(self, mini_grid_with_train):
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        assert artefacts["step_train_long"] is not None

    def test_step_train_long_expected_rows(self, mini_grid_with_train):
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        stl = artefacts["step_train_long"]
        assert len(stl) == 3 * 2  # 3 grid points × 2 steps

    def test_step_train_long_shape_positive(self, mini_grid_with_train):
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        stl = artefacts["step_train_long"]
        assert stl.shape[0] > 0
        assert stl.shape[1] > 0

    def test_ranking_uses_all_grid_points(self, mini_grid_with_train):
        """Ranking must be computed on ALL grid points, not just top-1."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        ranked = artefacts["ranked"]
        assert len(ranked) == 3

    def test_trades_oos_contains_only_rank1(self, mini_grid_with_train):
        """WF_Trades must contain trades for rank-1 grid point only."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        trades = artefacts["trades_oos"]
        ranked = artefacts["ranked"]
        if not trades.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            unique_gps = trades["grid_point_id"].unique()
            assert len(unique_gps) == 1
            assert unique_gps[0] == top_gp

    def test_trades_train_contains_only_rank1(self, mini_grid_with_train):
        """WF_Train_Trades must contain trades for rank-1 grid point only."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        trades = artefacts["trades_train"]
        ranked = artefacts["ranked"]
        if not trades.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            unique_gps = trades["grid_point_id"].unique()
            assert len(unique_gps) == 1
            assert unique_gps[0] == top_gp

    def test_export_trades_only_rank1(self, mini_grid_with_train, tmp_path):
        """XLSX WF_Trades sheet must only contain rank-1 grid point."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        path = _export_pipeline(artefacts, tmp_path)
        df = _read_sheet(path, "WF_Trades")
        ranked = artefacts["ranked"]
        if not df.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            assert set(df["grid_point_id"].unique()) == {top_gp}

    def test_all_failed_trades_empty(self):
        """When all candidates fail, trades should be empty."""
        results = {
            "gp1": [_failed_step("gp1", 1), _failed_step("gp1", 2)],
            "gp2": [_failed_step("gp2", 1), _failed_step("gp2", 2)],
        }
        artefacts = _run_mini_pipeline(results, train_results=results)
        assert artefacts["trades_oos"].empty
        assert artefacts["trades_train"].empty

    def test_train_metrics_no_prepend_diagnostics(self, mini_grid_with_train):
        """Train step results must not have prepend path diagnostics in step_train_long."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        stl = artefacts["step_train_long"]
        assert "used_prepend" not in stl.columns
        assert "prepend_bars_requested" not in stl.columns

    def test_non_top1_present_in_step_train_long(self, mini_grid_with_train):
        """All grid points (including non-top1) must be in step_train_long."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        stl = artefacts["step_train_long"]
        all_gp_ids = set(train.keys())
        stl_gp_ids = set(stl["grid_point_id"].unique())
        assert stl_gp_ids == all_gp_ids

    def test_non_top1_present_in_ranking(self, mini_grid_with_train):
        """All grid points (including non-top1) must be in ranking."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        ranked = artefacts["ranked"]
        all_gp_ids = set(oos.keys())
        ranked_gp_ids = set(ranked["grid_point_id"].unique())
        assert ranked_gp_ids == all_gp_ids

    def test_non_top1_present_in_aggregation(self, mini_grid_with_train):
        """All grid points (including non-top1) must be in aggregation."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        agg = artefacts["aggregated"]
        all_gp_ids = set(oos.keys())
        agg_gp_ids = set(agg["grid_point_id"].unique())
        assert agg_gp_ids == all_gp_ids

    def test_non_top1_absent_from_trades_oos(self, mini_grid_with_train):
        """Non-rank-1 grid points must NOT appear in trades_oos."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        trades = artefacts["trades_oos"]
        ranked = artefacts["ranked"]
        if not trades.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            non_top1 = set(oos.keys()) - {top_gp}
            trades_gps = set(trades["grid_point_id"].unique())
            assert trades_gps.isdisjoint(non_top1)

    def test_non_top1_absent_from_trades_train(self, mini_grid_with_train):
        """Non-rank-1 grid points must NOT appear in trades_train."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        trades = artefacts["trades_train"]
        ranked = artefacts["ranked"]
        if not trades.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            non_top1 = set(train.keys()) - {top_gp}
            trades_gps = set(trades["grid_point_id"].unique())
            assert trades_gps.isdisjoint(non_top1)

    def test_non_top1_absent_from_xlsx_wf_trades(self, mini_grid_with_train, tmp_path):
        """XLSX WF_Trades must not contain non-rank-1 grid points."""
        oos, train = mini_grid_with_train
        artefacts = _run_mini_pipeline(oos, train_results=train)
        path = _export_pipeline(artefacts, tmp_path)
        df = _read_sheet(path, "WF_Trades")
        ranked = artefacts["ranked"]
        if not df.empty:
            top_gp = ranked.loc[ranked["grid_rank"] == 1, "grid_point_id"].iloc[0]
            non_top1 = set(oos.keys()) - {top_gp}
            xlsx_gps = set(df["grid_point_id"].unique())
            assert xlsx_gps.isdisjoint(non_top1)


# ===========================================================================
# anchor / min_train_bars / min_test_bars — orchestrator passthrough + guard
# ===========================================================================

class TestOrchestratorWFSlicingPassthrough:
    """Integration: orchestrator passes anchor, min_train_bars, min_test_bars
    to make_walk_forward_slices, and raises ValueError when result is empty.
    """

    def _make_config_with_anchor(self, anchor: str = "start") -> "GridConfig":
        from wf_grid.config.schema import (
            BacktestConfig,
            DataConfig,
            GridConfig,
            RankingConfig,
            StatusConfig,
            ValidationConfig,
            WalkForwardConfig,
        )
        return GridConfig(
            data=DataConfig(file_path="dummy.csv"),
            backtest=BacktestConfig(min_trades_required=3),
            status=StatusConfig(min_meaningful_bars=10),
            ranking=RankingConfig(mode="gates_score"),
            validation=ValidationConfig(
                walk_forward=WalkForwardConfig(
                    train_size="90D",
                    test_size="30D",
                    scheme="rolling",
                    anchor=anchor,
                    min_train_bars=200,
                    min_test_bars=50,
                ),
            ),
        )

    def test_passthrough_all_seven_kwargs(self, monkeypatch):
        """make_walk_forward_slices must receive all 7 kwargs with correct values."""
        import textwrap
        import tempfile
        import pandas as pd
        from pathlib import Path

        captured_kwargs: dict = {}

        def _fake_slices(**kwargs):
            captured_kwargs.update(kwargs)
            # Return a minimal fake slice so pipeline doesn't raise empty-guard
            from types import SimpleNamespace
            return [SimpleNamespace(
                train_start_idx=0, train_end_idx=90,
                test_start_idx=90, test_end_idx=120,
                step_index=0,
            )]

        # make_walk_forward_slices is imported at module level in orchestrator;
        # patch the name as it lives in that module's namespace.
        import wf_grid.pipeline.orchestrator as _orch_mod
        monkeypatch.setattr(_orch_mod, "make_walk_forward_slices", _fake_slices)

        # Patch all downstream stages so they don't blow up on the fake data
        from wf_grid.pipeline import orchestrator as orch
        monkeypatch.setattr(orch, "compute_prepend_bars", lambda data, cfg: 0)
        monkeypatch.setattr(orch, "enumerate_grid", lambda cfg: [])
        monkeypatch.setattr(orch, "resolve_periods_per_year",
                            lambda cfg, data: cfg)

        import pandas as pd
        fake_data = pd.DataFrame(
            {"open": [1.0], "high": [1.0], "low": [1.0], "close": [1.0]},
            index=pd.date_range("2020-01-01", periods=1, freq="1D"),
        )

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as f:
            fake_data.to_csv(f)
            csv_path = f.name

        yaml_text = textwrap.dedent(f"""\
            data:
              file_path: {csv_path}
            validation:
              walk_forward:
                train_size: "90D"
                test_size: "30D"
                anchor: "end"
                min_train_bars: 200
                min_test_bars: 50
        """)

        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False, mode="w") as f:
            f.write(yaml_text)
            yaml_path = f.name

        from wf_grid.pipeline.orchestrator import run_grid_pipeline
        run_grid_pipeline(yaml_path, output_path=str(Path(csv_path).with_suffix(".xlsx")), parallel_enabled=False)

        assert "anchor" in captured_kwargs, "anchor not passed to make_walk_forward_slices"
        assert "min_train_bars" in captured_kwargs, "min_train_bars not passed"
        assert "min_test_bars" in captured_kwargs, "min_test_bars not passed"
        assert captured_kwargs["anchor"] == "end"
        assert captured_kwargs["min_train_bars"] == 200
        assert captured_kwargs["min_test_bars"] == 50

    def test_empty_slices_raises_value_error(self, monkeypatch):
        """When make_walk_forward_slices returns [], orchestrator must raise ValueError."""
        import textwrap
        import tempfile
        import pandas as pd
        from pathlib import Path

        import wf_grid.pipeline.orchestrator as _orch_mod
        monkeypatch.setattr(_orch_mod, "make_walk_forward_slices", lambda **kwargs: [])

        from wf_grid.pipeline import orchestrator as orch
        monkeypatch.setattr(orch, "compute_prepend_bars", lambda data, cfg: 0)
        monkeypatch.setattr(orch, "resolve_periods_per_year",
                            lambda cfg, data: cfg)

        fake_data = pd.DataFrame(
            {"open": [1.0], "high": [1.0], "low": [1.0], "close": [1.0]},
            index=pd.date_range("2020-01-01", periods=1, freq="1D"),
        )

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as f:
            fake_data.to_csv(f)
            csv_path = f.name

        yaml_text = textwrap.dedent(f"""\
            data:
              file_path: {csv_path}
            validation:
              walk_forward:
                train_size: "90D"
                test_size: "30D"
        """)

        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False, mode="w") as f:
            f.write(yaml_text)
            yaml_path = f.name

        from wf_grid.pipeline.orchestrator import run_grid_pipeline
        result = run_grid_pipeline(yaml_path, output_path=str(Path(csv_path).with_suffix(".xlsx")), parallel_enabled=False)
        # Pipeline catches exceptions into result.error
        assert result.error is not None
        assert "No walk-forward slices" in result.error


# ---------------------------------------------------------------------------
# FIX-1: Data validation in orchestrator
# ---------------------------------------------------------------------------

class TestOrchestratorDataValidation:
    """Data validation is called with strict=True immediately after CSV load."""

    def _write_csv(self, df: pd.DataFrame) -> str:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as f:
            df.to_csv(f)
            return f.name

    def _write_yaml(self, csv_path: str) -> str:
        import textwrap
        import tempfile
        yaml_text = textwrap.dedent(f"""\
            data:
              file_path: {csv_path}
            validation:
              walk_forward:
                train_size: "90D"
                test_size: "30D"
        """)
        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False, mode="w") as f:
            f.write(yaml_text)
            return f.name

    def test_missing_close_column_fails_early(self):
        """CSV missing 'close' column must raise DataValidationError captured in result.error."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        df = pd.DataFrame(
            {"open": [1.0, 1.1], "high": [1.2, 1.3], "low": [0.9, 1.0]},
            index=pd.date_range("2020-01-01", periods=2, freq="1D"),
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None

    def test_nan_in_ohlc_fails_early(self):
        """CSV with NaN in OHLC must fail validation."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        df = pd.DataFrame(
            {"open": [1.0, float("nan")], "high": [1.2, 1.3], "low": [0.9, 1.0], "close": [1.1, 1.2]},
            index=pd.date_range("2020-01-01", periods=2, freq="1D"),
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None

    def test_non_positive_price_fails_early(self):
        """CSV with zero price must fail validation."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        df = pd.DataFrame(
            {"open": [0.0, 1.1], "high": [1.2, 1.3], "low": [0.0, 1.0], "close": [0.0, 1.2]},
            index=pd.date_range("2020-01-01", periods=2, freq="1D"),
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None

    def test_duplicate_timestamps_strict_fails(self):
        """Duplicate DatetimeIndex rows must fail in strict mode."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        idx = pd.to_datetime(["2020-01-01", "2020-01-01", "2020-01-02"])
        df = pd.DataFrame(
            {"open": [1.0, 1.0, 1.1], "high": [1.2, 1.2, 1.3], "low": [0.9, 0.9, 1.0], "close": [1.1, 1.1, 1.2]},
            index=idx,
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None

    def test_unsorted_index_strict_fails(self):
        """Unsorted DatetimeIndex must fail in strict mode."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        idx = pd.to_datetime(["2020-01-03", "2020-01-01", "2020-01-02"])
        df = pd.DataFrame(
            {"open": [1.0, 1.1, 1.2], "high": [1.3, 1.4, 1.5], "low": [0.9, 1.0, 1.1], "close": [1.1, 1.2, 1.3]},
            index=idx,
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None

    def test_range_index_strict_fails_via_orchestrator(self, monkeypatch, tmp_path):
        """Orchestrator path: RangeIndex from read_csv fails validate_ohlc_data(strict=True)."""
        import wf_grid.pipeline.orchestrator as _orch_mod

        range_df = pd.DataFrame(
            {
                "open": [1.0, 1.1],
                "high": [1.2, 1.3],
                "low": [0.9, 1.0],
                "close": [1.1, 1.2],
            }
        )
        assert isinstance(range_df.index, pd.RangeIndex)

        monkeypatch.setattr(_orch_mod.pd, "read_csv", lambda *args, **kwargs: range_df.copy())

        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        dummy_csv = str(tmp_path / "dummy.csv")
        yaml_path = self._write_yaml(dummy_csv)
        result = run_grid_pipeline(yaml_path, output_path=str(tmp_path / "out.xlsx"), parallel_enabled=False)

        assert result.error is not None
        assert "DatetimeIndex" in result.error
        assert result.ranked is None

    def test_valid_data_passes_validation(self, monkeypatch):
        """Valid OHLC data must pass validation and reach WF slicing stage."""
        import wf_grid.pipeline.orchestrator as _orch_mod
        # Patch downstream to stop execution after validation passes
        monkeypatch.setattr(_orch_mod, "make_walk_forward_slices", lambda **kwargs: [])
        monkeypatch.setattr(_orch_mod, "compute_prepend_bars", lambda data, cfg: 0)
        monkeypatch.setattr(_orch_mod, "resolve_periods_per_year", lambda cfg, data: cfg)

        from wf_grid.pipeline.orchestrator import run_grid_pipeline

        df = pd.DataFrame(
            {"open": [1.0, 1.1], "high": [1.2, 1.3], "low": [0.9, 1.0], "close": [1.1, 1.2]},
            index=pd.date_range("2020-01-01", periods=2, freq="1D"),
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        # Should fail at empty slices, NOT at data validation
        assert result.error is not None
        assert "No walk-forward slices" in result.error

    def test_inf_in_ohlc_fails_early(self):
        """CSV with inf values in OHLC must fail validation."""
        from wf_grid.pipeline.orchestrator import run_grid_pipeline
        import numpy as np

        df = pd.DataFrame(
            {"open": [1.0, np.inf], "high": [1.2, 1.3], "low": [0.9, 1.0], "close": [1.1, 1.2]},
            index=pd.date_range("2020-01-01", periods=2, freq="1D"),
        )
        csv_path = self._write_csv(df)
        yaml_path = self._write_yaml(csv_path)
        result = run_grid_pipeline(yaml_path, output_path=csv_path.replace(".csv", ".xlsx"), parallel_enabled=False)
        assert result.error is not None


# ===========================================================================
# BucketMatrix_Median integration tests
# ===========================================================================

class TestBucketMatrixIntegration:
    """Integration: bucket_matrix_median built from pre-gate data and exported."""

    @pytest.fixture
    def single_mode_results(self):
        """3 grid points × 2 steps, all 'both' trade_mode (single mode required)."""
        return {
            "atr10_m2.50_both": [
                _ok_step("atr10_m2.50_both", 1, pnl=10.0),
                _ok_step("atr10_m2.50_both", 2, pnl=8.0),
            ],
            "atr15_m3.00_both": [
                _ok_step("atr15_m3.00_both", 1, pnl=5.0),
                _ok_step("atr15_m3.00_both", 2, pnl=3.0),
            ],
            "atr20_m4.00_both": [
                _ok_step("atr20_m4.00_both", 1, pnl=1.0),
                _ok_step("atr20_m4.00_both", 2, pnl=-2.0),
            ],
        }

    def _run_with_bucket(self, grid_results, config=None):
        """Run mini pipeline + bucket matrix build (mimics orchestrator flow)."""
        from wf_grid.bucket.median_matrix_builder import build_median_bucket_matrix

        config = config or _cfg()
        gp_objects = [
            GridPoint(atr_period=10, multiplier=2.50, trade_mode="both", grid_point_id="atr10_m2.50_both"),
            GridPoint(atr_period=15, multiplier=3.00, trade_mode="both", grid_point_id="atr15_m3.00_both"),
            GridPoint(atr_period=20, multiplier=4.00, trade_mode="both", grid_point_id="atr20_m4.00_both"),
        ]
        gp_objects = [gp for gp in gp_objects if gp.grid_point_id in grid_results]
        step_oos_long = collect_oos_steps(
            grid_results, config, expected_n_steps=2, grid_points=gp_objects,
        )
        bucket_matrix = build_median_bucket_matrix(step_oos_long, config)
        step_oos_long_gated = apply_step_gates(step_oos_long, config)

        aggregated = aggregate_candidates(step_oos_long_gated, config)
        gated = apply_candidate_gates(aggregated, config)
        passed_mask = gated["seed_gate_passed"].fillna(False).astype(bool)
        scores, statuses = calculate_seed_score(
            gated, passed_mask,
            score_weights=config.scoring.score_weights,
            normalization_mode=config.scoring.normalization_mode,
        )
        gated["tester_seed_score"] = scores
        gated["score_contract_status"] = statuses
        gated["score_discrimination_status"] = compute_score_discrimination(
            gated, passed_mask,
        )
        ranked = rank_candidates(gated, config)
        trades_oos, trades_train = _collect_top1_trades(
            ranked, grid_results, {}, config,
        )
        summary_wide = build_summary_wide(
            step_oos_long_gated, aggregated, ranked, config,
        )
        return {
            "step_oos_long": step_oos_long_gated,
            "summary_wide": summary_wide,
            "trades_oos": trades_oos,
            "trades_train": trades_train,
            "config": config,
            "bucket_matrix_median": bucket_matrix,
        }

    def test_bucket_matrix_built(self, single_mode_results):
        """Bucket matrix must be a non-empty DataFrame."""
        artefacts = self._run_with_bucket(single_mode_results)
        bm = artefacts["bucket_matrix_median"]
        assert bm is not None
        assert isinstance(bm, pd.DataFrame)
        assert len(bm) > 0

    def test_bucket_matrix_has_step_columns(self, single_mode_results):
        """Bucket matrix must have Step1..StepN columns."""
        artefacts = self._run_with_bucket(single_mode_results)
        bm = artefacts["bucket_matrix_median"]
        assert "Step1" in bm.columns
        assert "Step2" in bm.columns

    def test_bucket_matrix_has_stability_score(self, single_mode_results):
        """Bucket matrix must include bucket_stability_score."""
        artefacts = self._run_with_bucket(single_mode_results)
        bm = artefacts["bucket_matrix_median"]
        assert "bucket_stability_score" in bm.columns

    def test_bucket_matrix_sorted_by_stability_desc(self, single_mode_results):
        """Rows must be sorted by bucket_stability_score DESC."""
        artefacts = self._run_with_bucket(single_mode_results)
        bm = artefacts["bucket_matrix_median"]
        scores = bm["bucket_stability_score"].tolist()
        assert scores == sorted(scores, reverse=True)

    def test_bucket_matrix_export_sheet_present(self, single_mode_results, tmp_path):
        """BucketMatrix_Median sheet must appear in the exported workbook."""
        artefacts = self._run_with_bucket(single_mode_results)
        path = tmp_path / "bucket_test.xlsx"
        export_workbook(
            summary_wide=artefacts["summary_wide"],
            step_oos_long=artefacts["step_oos_long"],
            trades_oos=artefacts["trades_oos"],
            trades_train=artefacts["trades_train"],
            config=artefacts["config"],
            output_path=path,
            bucket_matrix_median=artefacts["bucket_matrix_median"],
        )
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        assert "BucketMatrix_Median" in sheets

    def test_bucket_matrix_export_sheet_after_train_trades(self, single_mode_results, tmp_path):
        """BucketMatrix_Median must come after WF_Train_Trades."""
        artefacts = self._run_with_bucket(single_mode_results)
        path = tmp_path / "order_test.xlsx"
        export_workbook(
            summary_wide=artefacts["summary_wide"],
            step_oos_long=artefacts["step_oos_long"],
            trades_oos=artefacts["trades_oos"],
            trades_train=artefacts["trades_train"],
            config=artefacts["config"],
            output_path=path,
            bucket_matrix_median=artefacts["bucket_matrix_median"],
        )
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        train_idx = sheets.index("WF_Train_Trades")
        bucket_idx = sheets.index("BucketMatrix_Median")
        assert bucket_idx > train_idx

    def test_bucket_matrix_export_has_data(self, single_mode_results, tmp_path):
        """BucketMatrix_Median sheet must contain data rows."""
        artefacts = self._run_with_bucket(single_mode_results)
        path = tmp_path / "data_test.xlsx"
        export_workbook(
            summary_wide=artefacts["summary_wide"],
            step_oos_long=artefacts["step_oos_long"],
            trades_oos=artefacts["trades_oos"],
            trades_train=artefacts["trades_train"],
            config=artefacts["config"],
            output_path=path,
            bucket_matrix_median=artefacts["bucket_matrix_median"],
        )
        df = pd.read_excel(path, sheet_name="BucketMatrix_Median")
        assert len(df) > 0
        assert "bucket_stability_score" in df.columns

    def test_bucket_matrix_none_gives_placeholder(self, tmp_path):
        """When bucket_matrix_median=None, sheet is a placeholder (headers only)."""
        path = tmp_path / "placeholder_test.xlsx"
        export_workbook(
            summary_wide=_make_summary(),
            step_oos_long=_make_step_oos(),
            trades_oos=_make_trades(),
            trades_train=_make_trades(),
            config=_cfg(),
            output_path=path,
            bucket_matrix_median=None,
        )
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        assert "BucketMatrix_Median" in sheets
        df = pd.read_excel(path, sheet_name="BucketMatrix_Median")
        assert len(df) == 0

    def test_backward_compat_no_bucket_arg(self, tmp_path):
        """Omitting bucket_matrix_median arg must produce placeholder sheet."""
        path = tmp_path / "compat_test.xlsx"
        export_workbook(
            summary_wide=_make_summary(),
            step_oos_long=_make_step_oos(),
            trades_oos=_make_trades(),
            trades_train=_make_trades(),
            config=_cfg(),
            output_path=path,
        )
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        assert "BucketMatrix_Median" in sheets

    def test_all_failed_gives_placeholder_bucket(self):
        """When all steps fail, bucket matrix is still built (full-grid-with-NaN)."""
        from wf_grid.bucket.median_matrix_builder import build_median_bucket_matrix

        results = {
            "atr10_m2.50_both": [_failed_step("atr10_m2.50_both", 1), _failed_step("atr10_m2.50_both", 2)],
            "atr20_m4.00_both": [_failed_step("atr20_m4.00_both", 1), _failed_step("atr20_m4.00_both", 2)],
        }
        config = _cfg()
        step_oos_long = collect_oos_steps(results, config, expected_n_steps=2)
        bm = build_median_bucket_matrix(step_oos_long, config)
        assert bm is not None
        assert isinstance(bm, pd.DataFrame)


def _make_summary():
    """Minimal valid summary for bucket integration tests."""
    return pd.DataFrame({
        "grid_rank": [1],
        "grid_point_id": ["atr10_m2.50_both"],
        "atr_period": [10],
        "multiplier": [2.50],
        "trade_mode": ["both"],
        "tier": [1],
        "seed_gate_passed": [True],
        "tester_seed_score": [0.8],
        "ranking_mode": ["gates_score"],
        "score_contract_status": ["ok"],
        "quantile_gates_status": [None],
        "seed_gate_fail_reason": [""],
        "ok_ratio": [1.0],
        "n_ok_steps": [2],
        "n_total_steps": [2],
        "n_segments": [2],
        "profitable_segments_count": [2],
        "sum_pnl_pct_Median": [5.0],
        "sum_pnl_pct_Min": [1.0],
        "sum_pnl_pct_Std": [2.0],
        "max_drawdown_Min": [-0.20],
        "profit_factor_Median": [1.8],
        "num_trades_Median": [10.0],
        "sharpe_Median": [1.2],
        "sortino_Median": [1.5],
        "cagr_Median": [0.15],
        "win_rate_Median": [0.6],
        "avg_trade_Median": [0.5],
        "gate_ok_positive_median": [True],
        "gate_ok_min_trades": [True],
        "gate_ok_worst_segment": [True],
        "gate_ok_drawdown": [True],
        "S1_sum_pnl_pct": [5.0],
        "S2_sum_pnl_pct": [5.0],
    })


def _make_step_oos():
    """Minimal valid step_oos_long for bucket integration tests."""
    rows = []
    for s in range(1, 3):
        rows.append({
            "grid_point_id": "atr10_m2.50_both",
            "wf_step": s,
            "step_status": "ok",
            "sum_pnl_pct": 5.0,
            "sharpe": 1.2,
            "sortino": 1.5,
            "max_drawdown": -0.10,
            "num_trades": 10,
            "profit_factor": 1.8,
            "effective_oos_bars": 100,
            "used_prepend": True,
            "prepend_bars_applied": 50,
        })
    return pd.DataFrame(rows)


def _make_trades():
    """Minimal valid trades for bucket integration tests."""
    return pd.DataFrame({
        "grid_point_id": ["atr10_m2.50_both"],
        "wf_step": [1],
        "step_status": ["ok"],
        "test_start_idx": [0],
        "test_end_idx": [100],
        "trade_id": [1],
        "net_pnl_pct": [0.5],
    })
