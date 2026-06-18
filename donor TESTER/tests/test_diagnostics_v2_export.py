from __future__ import annotations

from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from supertrend_optimizer.engine.result import BacktestResult
from supertrend_optimizer.io import diagnostics_v2, excel_tester
from supertrend_optimizer.testing.runner import PeriodResult


def _period_result(
    *,
    label: str = "100%",
    n: int = 8,
    filter_diagnostics: dict[str, np.ndarray] | None = None,
) -> PeriodResult:
    result = BacktestResult(
        atr_period=14,
        multiplier=3.0,
        trade_mode="revers",
        commission=0.0,
        warmup=0,
        returns=np.zeros(n - 1, dtype=np.float64),
        equity_curve=np.ones(n, dtype=np.float64),
        positions=np.zeros(n, dtype=np.int8),
        trend=np.zeros(n, dtype=np.int8),
        metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
        early_exit=False,
        exit_bar=None,
        exit_drawdown=None,
        trades_df=pd.DataFrame(),
        n_bars_original=n,
        period_label=label,
        effective_warmup=0,
        filter_diagnostics=filter_diagnostics,
    )
    return PeriodResult(
        period_label=label,
        n_bars=n,
        result=result,
        filter_diagnostics=filter_diagnostics,
        filter_diagnostics_summary=None,
    )


def _enabled_zigzag_cfg() -> SimpleNamespace:
    return SimpleNamespace(
        enabled=True,
        zigzag=SimpleNamespace(enabled=True),
        volume=None,
        diagnostics=SimpleNamespace(
            export_state_columns=False,
            export_trigger_columns=False,
        ),
    )


def _df(n: int = 8) -> pd.DataFrame:
    close = np.arange(100.0, 100.0 + n)
    return pd.DataFrame(
        {
            "open": close,
            "high": close + 1.0,
            "low": close - 1.0,
            "close": close,
        },
        index=pd.date_range("2026-01-01", periods=n, freq="h"),
    )


def _sheetnames(path: str) -> list[str]:
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def test_export_diagnostics_v2_default_writes_no_v2_sheets(tmp_path):
    out = tmp_path / "default.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    assert not any(name in diagnostics_v2.V2_SHEET_ORDER for name in names)


def test_export_diagnostics_v2_disabled_path_does_not_build_context(
    tmp_path,
    monkeypatch,
):
    out = tmp_path / "no_context.xlsx"

    def fail_build(*args, **kwargs):
        raise AssertionError("diagnostics v2 context should not be built")

    monkeypatch.setattr(diagnostics_v2, "build_diagnostics_v2_context", fail_build)
    monkeypatch.setattr(diagnostics_v2, "build_enabled_v2_sheets", fail_build)

    excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_diagnostics_v2=False,
        add_timestamp=False,
    )


def test_export_diagnostics_v2_filter_disabled_writes_no_v2_sheets(tmp_path):
    out = tmp_path / "filter_disabled.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=SimpleNamespace(enabled=False),
        export_diagnostics_v2=True,
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    assert not any(name in diagnostics_v2.V2_SHEET_ORDER for name in names)


def test_export_diagnostics_v2_export_diagnostics_false_writes_no_v2_sheets(tmp_path):
    out = tmp_path / "diagnostics_false.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_diagnostics=False,
        export_diagnostics_v2=True,
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    assert not any(name in diagnostics_v2.V2_SHEET_ORDER for name in names)


def test_export_diagnostics_v2_missing_pr_100_writes_no_v2_sheets(tmp_path):
    out = tmp_path / "missing_pr100.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result(label="75%")],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_diagnostics_v2=True,
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    assert not any(name in diagnostics_v2.V2_SHEET_ORDER for name in names)


def test_export_diagnostics_v2_enabled_appends_after_legacy_sheets(tmp_path):
    out = tmp_path / "enabled.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_signals=False,
        export_false_start=False,
        export_cycle=False,
        export_diagnostics_v2=True,
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    expected_v2 = list(diagnostics_v2.V2_SHEET_ORDER)
    assert names[-len(expected_v2) :] == expected_v2
    assert names[: -len(expected_v2)] == [
        excel_tester.TESTER_CONFIG_SHEET_NAME,
        "Summary",
        "Metrics_100",
        "Trades_100",
    ]


def test_export_diagnostics_v2_child_false_removes_only_that_sheet(tmp_path):
    out = tmp_path / "child_false.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_signals=False,
        export_false_start=False,
        export_cycle=False,
        export_diagnostics_v2=True,
        diagnostics_v2_flags={"run_health": False},
        add_timestamp=False,
    )

    names = _sheetnames(actual)
    assert "Run_Health" not in names
    assert "Reproducibility" in names
    assert "Trade_Analytics" in names


def test_export_diagnostics_v2_applies_generic_sheet_formatting(tmp_path):
    out = tmp_path / "formatted.xlsx"

    actual = excel_tester.export_tester_results(
        [_period_result()],
        str(out),
        trade_filter_config=_enabled_zigzag_cfg(),
        df=_df(),
        export_signals=False,
        export_false_start=False,
        export_cycle=False,
        export_diagnostics_v2=True,
        add_timestamp=False,
    )

    wb = load_workbook(actual)
    try:
        ws = wb["Index"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1:F")
        assert ws.column_dimensions["A"].width >= 8
    finally:
        wb.close()
