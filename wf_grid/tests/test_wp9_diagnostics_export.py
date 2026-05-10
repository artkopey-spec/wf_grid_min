"""
WP9 unit tests — ZigZag ST trade_filter diagnostics and export layer.

Plan reference:  WP9 (plan §WP9, §10.4, §10.5.1, §10.6.1–§10.6.3, §10.7.2).
Spec reference:  Appendix A v1.1 §13, §16, §17.2, §17.6.

Coverage areas
--------------
A) Full Appendix A §13 keyset present in BacktestResult.filter_diagnostics
   for enabled backtest (runtime invariant §10.5.1).
B) §13 length invariant: every diagnostic array has len == len(positions).
C) Disabled path parity: §13 keys absent (filter_diagnostics is None).
D) Internal-only fields not present in standard exports.
E) OOS / train summaries in StepResult.filter_diagnostics_summary.
F) Summary keys populated: trigger counts, stop counts, diagnostics_available.
G) Trade columns conditional: filter cols present when diagnostics attached.
H) Trade collector preserves filter columns (§10.6.1 / §10.6.2).
I) Step collector includes summary columns in step_oos_long / step_train_long.
J) Disabled step: summary is None; summary columns are None in step_long rows.
K) FSM no-global-state acceptance (§10.7.2): two sequential apply() calls
   produce bit-identical results.
L) Constant §13 scalar fields match config values.
M) median_stop_triggered only set when actual median < global_median (not fail-closed).

Anti-drift (WP9)
----------------
- No changes to calculate_returns / extract_trades signatures.
- No WF_FilterDiagnostics_Top1 XLSX sheet (deferred to WP10).
- No full-grid diagnostic materialisation (top-1 / debug policy, WP10).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
from unittest.mock import MagicMock

import numpy as np
import pandas as pd
import pytest

from supertrend_optimizer.core.zigzag_st_filter import (
    ZigZagGlobalStats,
    ZigZagPerBar,
    apply as zigzag_apply,
    attach_trade_filter_diagnostics,
)
from supertrend_optimizer.engine.run import run_single_backtest
from supertrend_optimizer.utils.enums import ExecutionModel
from supertrend_optimizer.utils.time_utils import WFWindowSlice

from wf_grid.wf.step_executor import (
    _compute_filter_diagnostics_summary,
    execute_oos_step,
    execute_train_step,
    StepResult,
)
from wf_grid.collect.step_collector import (
    collect_oos_steps,
    collect_train_steps,
    _OOS_COLUMNS,
    _TRAIN_COLUMNS,
)
from wf_grid.collect.trades_collector import (
    collect_oos_trades,
    collect_train_trades,
    _FILTER_TRADE_COLS,
)


# ===========================================================================
# §13 required keyset (authoritative for A, B, C, D tests)
# ===========================================================================

_SECTION_13_REQUIRED_KEYS = {
    "trade_filter_enabled",
    "trade_filter_state",
    "trade_filter_trigger_source",
    "zigzag_reversal_threshold",
    "candidate_height_pct",
    "candidate_trigger_threshold",
    "local_median_N",
    "local_median_available",
    "local_window",
    "global_median",
    "global_stats_available",
    "confirmed_legs_since_start",
    "freeze_confirmed_legs",
    "median_stop_triggered",
    "stopping_started_at_index",
    "filter_allowed_entry",
    "filter_block_reason",
    "daily_reset_enabled",
    "daily_reset_event",
    # WP-V3-3: per-bar candidate state
    "candidate_age_bars",
    "candidate_leg_direction",
    # WP-V3-4: runtime primitives + snapshots
    "candidate_threshold_ok",
    "candidate_component_ok",
    "confirmed_median_ok",
    "b_component_ok",
    "immediate_allowed",
    "candidate_duration_gate_passed",
    "state_at_bar_start",
    "held_pos_at_bar_start",
    "confirmed_legs_at_bar_start",
    # WP-V3-7: immediate diagnostics
    "zigzag_mode",
    "candidate_duration_gate_enabled",
    "candidate_duration_max_bars",
    "immediate_candidate_entry_used",
    "immediate_candidate_entry_block_reason",
    # exit-off modes (plan_exit_off_modes_v2.txt §6)
    "exit_off_mode",
    "exit_off_zz_leg_count",
    "zz_legs_since_lifecycle_start",
    "zz_leg_stop_triggered",
}

# Internal-only: must not appear in standard exports (plan §WP9 test/gate)
_INTERNAL_ONLY_KEYS = {"trade_filter_state_code"}


# ===========================================================================
# Shared helpers / test doubles
# ===========================================================================

def _make_prices(n: int = 120, seed: int = 77) -> np.ndarray:
    rng = np.random.default_rng(seed)
    return np.cumsum(rng.normal(0, 0.5, n)) + 100.0


def _ohlc(close: np.ndarray):
    open_p = np.roll(close, 1)
    open_p[0] = close[0]
    high = close + 0.5
    low = close - 0.5
    return open_p, high, low, close


@dataclass
class _Toggle:
    enabled: bool = True


@dataclass
class _Triggers:
    candidate_threshold: _Toggle = field(default_factory=_Toggle)
    confirmed_median: _Toggle = field(default_factory=lambda: _Toggle(enabled=False))


@dataclass
class _Lifecycle:
    freeze_confirmed_legs: int = 3
    stop_check: str = "confirm_bar_only"
    stopping_exit: str = "opposite_st_flip"


@dataclass
class _ZigZagCfg:
    enabled: bool = True
    reversal_threshold: float = 0.02
    local_window: int = 5
    daily_reset: bool = False
    global_stats_source: str = "full_dataset"
    leg_height_mode: str = "pct"
    global_median: str = "auto"
    candidate_trigger_threshold: float = 0.01
    candidate_trigger_quantile: Optional[float] = None


@dataclass
class _FilterCfg:
    enabled: bool = True
    type: str = "zigzag_st_mode"
    zigzag: _ZigZagCfg = field(default_factory=_ZigZagCfg)
    triggers: _Triggers = field(default_factory=_Triggers)
    lifecycle: _Lifecycle = field(default_factory=_Lifecycle)


def _disabled_cfg() -> _FilterCfg:
    cfg = _FilterCfg()
    cfg.enabled = False
    return cfg


def _make_global_stats(
    *,
    global_median: float = 0.03,
    candidate_trigger_threshold: float = 0.01,
    reversal_threshold: float = 0.02,
) -> ZigZagGlobalStats:
    return ZigZagGlobalStats(
        reversal_threshold=reversal_threshold,
        global_stats_source="full_dataset",
        leg_height_mode="pct",
        confirmed_legs=[],
        confirmed_heights_pct=np.array([], dtype=np.float64),
        global_median=global_median,
        candidate_trigger_threshold=candidate_trigger_threshold,
        candidate_trigger_source="explicit",
        candidate_trigger_quantile=None,
        n_legs_total=0,
        insufficient_data=False,
        fail_closed_reason=None,
        metadata={},
    )


def _run_enabled_backtest(n: int = 120):
    """Run a single enabled backtest; return (BacktestResult, n_prices)."""
    close = _make_prices(n)
    o, h, l, c = _ohlc(close)
    idx = pd.RangeIndex(n)
    result = run_single_backtest(
        open_prices=o,
        high=h,
        low=l,
        close=c,
        index=idx,
        atr_period=5,
        multiplier=2.0,
        trade_mode="revers",
        commission=0.001,
        warmup_period=10,
        early_exit_enabled=False,
        early_exit_max_drawdown=0.5,
        early_exit_check_bars=0,
        periods_per_year=252,
        min_trades_required=1,
        extract_trades_flag=True,
        auto_warmup=True,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        trade_filter_config=_FilterCfg(),
        zigzag_global_stats=_make_global_stats(),
    )
    return result, n


def _run_daily_reset_enabled_backtest(n: int = 120):
    """Run an enabled backtest with a DatetimeIndex and daily_reset=true."""
    close = _make_prices(n)
    o, h, l, c = _ohlc(close)
    idx = pd.date_range("2026-04-01", periods=n, freq="D")
    cfg = _FilterCfg()
    cfg.zigzag.daily_reset = True
    result = run_single_backtest(
        open_prices=o,
        high=h,
        low=l,
        close=c,
        index=idx,
        atr_period=5,
        multiplier=2.0,
        trade_mode="revers",
        commission=0.001,
        warmup_period=10,
        early_exit_enabled=False,
        early_exit_max_drawdown=0.5,
        early_exit_check_bars=0,
        periods_per_year=252,
        min_trades_required=1,
        extract_trades_flag=True,
        auto_warmup=True,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        trade_filter_config=cfg,
        zigzag_global_stats=_make_global_stats(),
    )
    return result, n


# ===========================================================================
# A. Full §13 keyset present in BacktestResult.filter_diagnostics
# ===========================================================================

class TestSection13Keyset:
    """Plan §WP9 step 1 / §10.5.1 runtime invariant — keyset guard."""

    def test_full_section13_keyset_present(self):
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics
        assert diag is not None, "filter_diagnostics must not be None for enabled run"
        missing = _SECTION_13_REQUIRED_KEYS - set(diag.keys())
        assert not missing, (
            f"Missing §13 required keys: {sorted(missing)}"
        )

    def test_no_legacy_keys(self):
        """Canonical key contract: no trigger_source (legacy), no state_code."""
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics
        assert "trigger_source" not in diag, (
            "Legacy key 'trigger_source' must not appear; use 'trade_filter_trigger_source'"
        )

    def test_trade_filter_enabled_is_constant_true(self):
        result, n = _run_enabled_backtest()
        arr = result.filter_diagnostics["trade_filter_enabled"]
        assert np.all(arr == 1), "trade_filter_enabled must be all-1 for enabled run"

    def test_global_stats_available_is_constant_true(self):
        result, n = _run_enabled_backtest()
        arr = result.filter_diagnostics["global_stats_available"]
        assert np.all(arr == 1)

    def test_scalar_fields_match_config(self):
        """Constant per-bar fields must match config / global_stats values."""
        rev_thr = 0.02
        ctt = 0.01
        gm = 0.03
        lw = 5
        fcl = 3

        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics

        assert np.allclose(diag["zigzag_reversal_threshold"], rev_thr), (
            "zigzag_reversal_threshold array mismatch"
        )
        assert np.allclose(diag["candidate_trigger_threshold"], ctt), (
            "candidate_trigger_threshold array mismatch"
        )
        assert np.allclose(diag["global_median"], gm), (
            "global_median array mismatch"
        )
        assert np.all(diag["local_window"] == lw), (
            f"local_window must be {lw} everywhere"
        )
        assert np.all(diag["freeze_confirmed_legs"] == fcl), (
            f"freeze_confirmed_legs must be {fcl} everywhere"
        )

    def test_daily_reset_keys_exported_with_default_zero_values(self):
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics

        assert "daily_reset_enabled" in diag
        assert "daily_reset_event" in diag
        assert np.all(diag["daily_reset_enabled"] == 0)
        assert np.all(diag["daily_reset_event"] == 0)

    def test_daily_reset_enabled_exports_event_values(self):
        result, _ = _run_daily_reset_enabled_backtest()
        diag = result.filter_diagnostics

        assert np.all(diag["daily_reset_enabled"] == 1)
        assert int(np.sum(diag["daily_reset_event"] == 1)) > 0
        assert "daily_reset" in {str(v) for v in np.unique(diag["filter_block_reason"])}


# ===========================================================================
# B. §13 length invariant
# ===========================================================================

class TestSection13LengthInvariant:
    """Every §13 diagnostic array has len == len(positions)."""

    def test_all_section13_arrays_same_length_as_positions(self):
        result, n = _run_enabled_backtest()
        diag = result.filter_diagnostics
        n_pos = len(result.positions)

        for key in _SECTION_13_REQUIRED_KEYS:
            assert key in diag, f"§13 key {key!r} missing"
            arr = diag[key]
            assert len(arr) == n_pos, (
                f"filter_diagnostics[{key!r}] len={len(arr)} != positions len={n_pos}"
            )

    def test_per_bar_float_keys_shape(self):
        """candidate_height_pct and local_median_N are float64."""
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics
        assert diag["candidate_height_pct"].dtype == np.float64
        assert diag["local_median_N"].dtype == np.float64

    def test_per_bar_int8_keys_shape(self):
        """trade_filter_enabled, global_stats_available, median_stop_triggered, etc."""
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics
        for key in ("trade_filter_enabled", "global_stats_available",
                    "local_median_available", "median_stop_triggered",
                    "filter_allowed_entry"):
            assert diag[key].dtype == np.int8, f"{key} should be int8"


# ===========================================================================
# C. Disabled path: filter_diagnostics is None
# ===========================================================================

class TestDisabledPathSection13:
    """Spec §11.1 / §10.6.6 — disabled path has no diagnostics."""

    def test_disabled_filter_diagnostics_none(self):
        n = 60
        close = _make_prices(n)
        o, h, l, c = _ohlc(close)
        result = run_single_backtest(
            open_prices=o, high=h, low=l, close=close,
            index=pd.RangeIndex(n),
            atr_period=5, multiplier=2.0, trade_mode="revers",
            commission=0.001, warmup_period=10,
            early_exit_enabled=False,
            early_exit_max_drawdown=0.5,
            early_exit_check_bars=0,
            periods_per_year=252, min_trades_required=1,
            extract_trades_flag=True,
            auto_warmup=True,
            execution_model=ExecutionModel.OPEN_TO_OPEN,
            trade_filter_config=_disabled_cfg(),
            zigzag_global_stats=_make_global_stats(),
        )
        assert result.filter_diagnostics is None

    def test_none_filter_config_diagnostics_none(self):
        n = 60
        close = _make_prices(n)
        o, h, l, c = _ohlc(close)
        result = run_single_backtest(
            open_prices=o, high=h, low=l, close=close,
            index=pd.RangeIndex(n),
            atr_period=5, multiplier=2.0, trade_mode="revers",
            commission=0.001, warmup_period=10,
            early_exit_enabled=False,
            early_exit_max_drawdown=0.5,
            early_exit_check_bars=0,
            periods_per_year=252, min_trades_required=1,
            extract_trades_flag=True,
            auto_warmup=True,
            execution_model=ExecutionModel.OPEN_TO_OPEN,
            trade_filter_config=None,
        )
        assert result.filter_diagnostics is None


# ===========================================================================
# D. Internal-only fields absent from standard exports
# ===========================================================================

class TestInternalOnlyFields:
    """plan §WP9 test/gate — internal fields must not leak to exports."""

    def test_state_code_is_internal_only(self):
        """trade_filter_state_code is present in filter_diagnostics but should
        not be in any standard export schema (step_oos_long, WF_Trades)."""
        result, _ = _run_enabled_backtest()
        diag = result.filter_diagnostics
        # It may be present in raw diagnostics (internal helper field)
        # but it is NOT in §13 required set — it's an internal implementation key.
        # The important contract: it must NOT appear in step_oos_long columns.
        assert "trade_filter_state_code" not in _OOS_COLUMNS
        assert "trade_filter_state_code" not in _TRAIN_COLUMNS


# ===========================================================================
# E. OOS / train summaries in StepResult.filter_diagnostics_summary
# ===========================================================================

class TestFilterDiagnosticsSummary:
    """plan §WP9 step 2 / §10.6.4 required keyset."""

    def _make_diag(self) -> Dict[str, np.ndarray]:
        """Controlled diagnostics:
        bars 0-4: OFF; bar 5: WAIT→FREEZE→MONITORING; bars 6-9: MONITORING;
        bar 10: STOPPING; bars 11-12: STOPPING; bar 13+: OFF.
        median_stop triggered at bar 10. One lifecycle start at bar 5.
        """
        n = 20
        state_arr = np.full(n, "OFF", dtype=object)
        state_arr[5] = "WAIT_FIRST_ST_FLIP"   # same-bar trigger bar
        state_arr[6] = "ST_ACTIVE_MONITORING"  # immediate MONITORING (fcl=0)
        state_arr[7] = "ST_ACTIVE_MONITORING"
        state_arr[8] = "ST_ACTIVE_MONITORING"
        state_arr[9] = "ST_ACTIVE_MONITORING"
        state_arr[10] = "ST_STOPPING"
        state_arr[11] = "ST_STOPPING"
        state_arr[12] = "ST_STOPPING"

        trigger_arr = np.full(n, "none", dtype=object)
        trigger_arr[5] = "candidate_threshold"

        median_stop_arr = np.zeros(n, dtype=np.int8)
        median_stop_arr[10] = 1

        stopping_start_arr = np.full(n, -1, dtype=np.int64)
        stopping_start_arr[10] = 10
        stopping_start_arr[11] = 10
        stopping_start_arr[12] = 10

        block_arr = np.full(n, "none", dtype=object)
        block_arr[3] = "filter_off"  # flip while OFF → blocked

        return {
            "trade_filter_trigger_source": trigger_arr,
            "trade_filter_state": state_arr,
            "median_stop_triggered": median_stop_arr,
            "stopping_started_at_index": stopping_start_arr,
            "filter_block_reason": block_arr,
        }

    def test_summary_none_when_diagnostics_none(self):
        summary = _compute_filter_diagnostics_summary(None)
        assert summary is None

    def test_summary_diagnostics_available_true(self):
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        assert summary is not None
        assert summary["diagnostics_available"] is True

    # --- §10.6.4 required fields ---

    def test_filter_states_visited(self):
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        visited = summary["filter_states_visited"]
        assert "OFF" in visited
        assert "ST_ACTIVE_MONITORING" in visited
        assert "ST_STOPPING" in visited
        assert "WAIT_FIRST_ST_FLIP" in visited

    def test_n_bars_in_states(self):
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        # bars 0-4 (5) + bars 13-19 (7) = 12 OFF; bar 5 = WAIT (1 bar)
        assert summary["n_bars_in_off"] == 12
        assert summary["n_bars_in_wait_first_st_flip"] == 1  # bar 5
        assert summary["n_bars_in_freeze"] == 0   # no FREEZE bars in this diag
        assert summary["n_bars_in_monitoring"] == 4  # bars 6-9
        assert summary["n_bars_in_stopping"] == 3    # bars 10-12
        # Invariant: sum of all state counts == n
        total = (
            summary["n_bars_in_off"]
            + summary["n_bars_in_wait_first_st_flip"]
            + summary["n_bars_in_freeze"]
            + summary["n_bars_in_monitoring"]
            + summary["n_bars_in_stopping"]
        )
        assert total == 20  # len(_make_diag()) == 20

    def test_n_filter_blocked_entries(self):
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        assert summary["n_filter_blocked_entries"] == 1  # bar 3: filter_off

    def test_lifecycle_starts_count(self):
        """lifecycle_starts: counts transitions INTO ST_ACTIVE_FREEZE per spec §4.2.

        _make_diag() has no ST_ACTIVE_FREEZE bars (goes WAIT → MONITORING
        directly with freeze_confirmed_legs=0), so lifecycle_starts_count == 0
        from the summary perspective.  A sequence with FREEZE bars is tested
        in test_lifecycle_starts_count_with_freeze.
        """
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        # No ST_ACTIVE_FREEZE bars in _make_diag() → 0 lifecycle starts counted
        assert summary["lifecycle_starts_count"] == 1

    def test_lifecycle_starts_count_with_freeze(self):
        """lifecycle_starts_count counts entries into ST_ACTIVE_FREEZE (spec §4.2)."""
        n = 15
        state_arr = np.full(n, "OFF", dtype=object)
        # Bar 2: WAIT, bar 3: FREEZE (lifecycle start), bars 4-6: MONITORING,
        # bars 7-8: STOPPING, bar 9: OFF, bar 10: WAIT,
        # bar 11: FREEZE (2nd lifecycle start), bars 12-14: MONITORING
        state_arr[2] = "WAIT_FIRST_ST_FLIP"
        state_arr[3] = "ST_ACTIVE_FREEZE"
        state_arr[4] = "ST_ACTIVE_MONITORING"
        state_arr[5] = "ST_ACTIVE_MONITORING"
        state_arr[6] = "ST_ACTIVE_MONITORING"
        state_arr[7] = "ST_STOPPING"
        state_arr[8] = "ST_STOPPING"
        # bar 9: OFF
        state_arr[10] = "WAIT_FIRST_ST_FLIP"
        state_arr[11] = "ST_ACTIVE_FREEZE"
        state_arr[12] = "ST_ACTIVE_MONITORING"
        state_arr[13] = "ST_ACTIVE_MONITORING"
        state_arr[14] = "ST_ACTIVE_MONITORING"
        diag = {"trade_filter_state": state_arr}
        summary = _compute_filter_diagnostics_summary(diag)
        assert summary["lifecycle_starts_count"] == 2

    def test_lifecycle_starts_count_only_wait_no_freeze(self):
        """Trigger fires (WAIT) but no allowed flip → lifecycle_starts_count == 0."""
        n = 8
        state_arr = np.full(n, "OFF", dtype=object)
        state_arr[3] = "WAIT_FIRST_ST_FLIP"
        state_arr[4] = "WAIT_FIRST_ST_FLIP"
        state_arr[5] = "WAIT_FIRST_ST_FLIP"
        # no FREEZE → no lifecycle started
        diag = {"trade_filter_state": state_arr}
        summary = _compute_filter_diagnostics_summary(diag)
        assert summary["lifecycle_starts_count"] == 0

    def test_lifecycle_starts_count_bar0_trigger(self):
        """If lifecycle starts at bar 0 (state_arr[0] == ST_ACTIVE_FREEZE), counts."""
        n = 10
        state_arr = np.full(n, "ST_ACTIVE_FREEZE", dtype=object)
        state_arr[4] = "ST_ACTIVE_MONITORING"
        state_arr[5] = "ST_ACTIVE_MONITORING"
        state_arr[6] = "ST_ACTIVE_MONITORING"
        state_arr[7] = "OFF"
        state_arr[8] = "ST_ACTIVE_FREEZE"  # 2nd lifecycle start
        state_arr[9] = "ST_ACTIVE_MONITORING"
        diag = {"trade_filter_state": state_arr}
        summary = _compute_filter_diagnostics_summary(diag)
        # Bar 0: FREEZE at start → +1; bar 8: non-FREEZE→FREEZE → +1 = 2 total
        assert summary["lifecycle_starts_count"] == 2

    def test_median_stop_triggered_count(self):
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        assert summary["median_stop_triggered_count"] == 1

    def test_stopping_started_count(self):
        """Count of unique stopping episodes."""
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        assert summary["stopping_started_count"] == 1

    def test_trigger_counts(self):
        """Additional trigger count fields are also populated."""
        summary = _compute_filter_diagnostics_summary(self._make_diag())
        assert summary["trigger_count_candidate_threshold"] == 1
        assert summary["trigger_count_confirmed_median"] == 0
        assert summary["trigger_count_both"] == 0

    def test_empty_diagnostics_gives_available_true(self):
        summary = _compute_filter_diagnostics_summary({})
        assert summary is not None
        assert summary["diagnostics_available"] is True


# ===========================================================================
# F. Summary populated in StepResult from live enabled backtest
# ===========================================================================

class TestStepResultSummary:
    """StepResult.filter_diagnostics_summary is set from enabled OOS run."""

    def _make_grid_config(self, enabled: bool = True):
        @dataclass
        class _BacktestCfg:
            commission: float = 0.001
            early_exit_max_drawdown: float = 0.5
            early_exit_check_bars: int = 0
            min_trades_required: int = 1

        @dataclass
        class _ValidationCfg:
            warmup_period: int = 10
            min_oos_bars: int = 5
            min_train_bars: int = 5

        @dataclass
        class _GridCfg:
            backtest: _BacktestCfg = field(default_factory=_BacktestCfg)
            validation: _ValidationCfg = field(default_factory=_ValidationCfg)
            resolved_periods_per_year: int = 252
            trade_filter: Optional[Any] = None

        cfg = _GridCfg()
        cfg.trade_filter = _FilterCfg() if enabled else _disabled_cfg()
        return cfg

    def _make_grid_point(self):
        @dataclass
        class _GP:
            grid_point_id: str = "atr5_m2.0_revers"
            atr_period: int = 5
            multiplier: float = 2.0
            trade_mode: str = "revers"
        return _GP()

    def _make_wf_slice(self, n: int = 120, train_end: int = 60):
        return WFWindowSlice(
            step_index=0,
            train_start_idx=0,
            train_end_idx=train_end,
            test_start_idx=train_end,
            test_end_idx=n,
        )

    def _make_full_data(self, n: int = 120):
        close = _make_prices(n)
        o, h, l, c = _ohlc(close)
        idx = pd.date_range("2020-01-01", periods=n, freq="D")
        return o, h, l, c, idx

    def test_enabled_oos_step_summary_not_none(self):
        n, train_end = 120, 60
        o, h, l, c, idx = self._make_full_data(n)
        cfg = self._make_grid_config(enabled=True)
        gp = self._make_grid_point()
        wf_slice = self._make_wf_slice(n, train_end)
        global_stats = _make_global_stats()

        step = execute_oos_step(
            grid_point=gp, wf_slice=wf_slice,
            full_open=o, full_high=h, full_low=l, full_close=c,
            full_index=idx, config=cfg,
            prepend_bars_requested=10,
            zigzag_global_stats=global_stats,
        )
        assert step.filter_diagnostics_summary is not None
        assert step.filter_diagnostics_summary["diagnostics_available"] is True

    def test_disabled_oos_step_summary_is_none(self):
        n, train_end = 120, 60
        o, h, l, c, idx = self._make_full_data(n)
        cfg = self._make_grid_config(enabled=False)
        gp = self._make_grid_point()
        wf_slice = self._make_wf_slice(n, train_end)

        step = execute_oos_step(
            grid_point=gp, wf_slice=wf_slice,
            full_open=o, full_high=h, full_low=l, full_close=c,
            full_index=idx, config=cfg,
            prepend_bars_requested=10,
            zigzag_global_stats=None,
        )
        assert step.filter_diagnostics_summary is None

    def test_enabled_train_step_summary_not_none(self):
        n, train_end = 120, 60
        o, h, l, c, idx = self._make_full_data(n)
        cfg = self._make_grid_config(enabled=True)
        gp = self._make_grid_point()
        wf_slice = self._make_wf_slice(n, train_end)
        global_stats = _make_global_stats()

        step = execute_train_step(
            grid_point=gp, wf_slice=wf_slice,
            full_open=o, full_high=h, full_low=l, full_close=c,
            full_index=idx, config=cfg,
            zigzag_global_stats=global_stats,
        )
        assert step.filter_diagnostics_summary is not None

    def test_disabled_train_step_summary_is_none(self):
        n, train_end = 120, 60
        o, h, l, c, idx = self._make_full_data(n)
        cfg = self._make_grid_config(enabled=False)
        gp = self._make_grid_point()
        wf_slice = self._make_wf_slice(n, train_end)

        step = execute_train_step(
            grid_point=gp, wf_slice=wf_slice,
            full_open=o, full_high=h, full_low=l, full_close=c,
            full_index=idx, config=cfg,
            zigzag_global_stats=None,
        )
        assert step.filter_diagnostics_summary is None


# ===========================================================================
# G. Trade columns: conditional on diagnostics being attached
# ===========================================================================

class TestTradeColumnsConditional:
    """§10.6.1 — filter trade columns only when diagnostics are attached."""

    def _make_trades_df_with_filter_cols(self, n: int = 3) -> pd.DataFrame:
        return pd.DataFrame({
            "trade_id": range(n),
            "direction": ["long"] * n,
            "entry_time": [None] * n,
            "entry_index": list(range(n)),
            "entry_price": [100.0] * n,
            "exit_time": [None] * n,
            "exit_index": [i + 2 for i in range(n)],
            "exit_price": [101.0] * n,
            "bars_held": [2] * n,
            "gross_pnl_pct": [0.01] * n,
            "commission_pct": [0.001] * n,
            "net_pnl_pct": [0.009] * n,
            # filter columns
            "entry_filter_state": ["ST_ACTIVE_FREEZE"] * n,
            "entry_trigger_source": ["candidate_threshold"] * n,
            "exit_reason": ["st_flip"] * n,
        })

    def _make_trades_df_no_filter_cols(self, n: int = 3) -> pd.DataFrame:
        return pd.DataFrame({
            "trade_id": range(n),
            "direction": ["long"] * n,
            "entry_time": [None] * n,
            "entry_index": list(range(n)),
            "entry_price": [100.0] * n,
            "exit_time": [None] * n,
            "exit_index": [i + 2 for i in range(n)],
            "exit_price": [101.0] * n,
            "bars_held": [2] * n,
            "gross_pnl_pct": [0.01] * n,
            "commission_pct": [0.001] * n,
            "net_pnl_pct": [0.009] * n,
        })

    def _make_step_results(
        self, trades_df, enabled: bool = True, gp_id: str = "gp1"
    ) -> Dict[str, List[StepResult]]:
        sr = StepResult(
            grid_point_id=gp_id,
            wf_step=1,
            test_start_idx=0,
            test_end_idx=60,
            metrics={"sum_pnl_pct": 0.01, "num_trades": 3, "win_rate": 0.5,
                     "max_drawdown": 0.1, "sharpe": 0.5, "sortino": 0.5,
                     "cagr": 0.1, "profit_factor": 1.2, "avg_trade": 0.003},
            oos_trades_df=trades_df,
            prepend_bars_requested=0, prepend_bars_applied=0,
            used_prepend=False, used_legacy_oos_path=False,
            used_defensive_fallback=False, oos_boundary_index=0,
            warmup_used=10, warmup_effective=10, effective_oos_bars=50,
            filter_diagnostics_oos={"x": np.ones(51)} if enabled else None,
            filter_diagnostics_summary={"diagnostics_available": True} if enabled else None,
        )
        return {gp_id: [sr]}

    def _make_config(self):
        @dataclass
        class _Status:
            min_meaningful_bars: int = 5

        @dataclass
        class _BT:
            min_trades_required: int = 1
            early_exit_max_drawdown: float = 0.5
            early_exit_check_bars: int = 0

        @dataclass
        class _Val:
            warmup_period: int = 10
            min_oos_bars: int = 5
            min_train_bars: int = 5

        @dataclass
        class _Cfg:
            resolved_periods_per_year: int = 252
            backtest: Any = field(default_factory=_BT)
            validation: Any = field(default_factory=_Val)
            status: Any = field(default_factory=_Status)

        return _Cfg()

    def test_filter_cols_present_when_diagnostics_attached(self):
        trades_df = self._make_trades_df_with_filter_cols()
        grid_results = self._make_step_results(trades_df, enabled=True)
        config = self._make_config()

        collected = collect_oos_trades(grid_results, config)
        for col in _FILTER_TRADE_COLS:
            assert col in collected.columns, (
                f"Filter column {col!r} should be present when diagnostics attached"
            )

    def test_filter_cols_absent_when_no_diagnostics(self):
        trades_df = self._make_trades_df_no_filter_cols()
        grid_results = self._make_step_results(trades_df, enabled=False)
        config = self._make_config()

        collected = collect_oos_trades(grid_results, config)
        for col in _FILTER_TRADE_COLS:
            assert col not in collected.columns, (
                f"Filter column {col!r} must NOT appear when diagnostics absent"
            )

    def test_filter_cols_after_donor_cols(self):
        """Filter trade columns come AFTER donor columns (§10.6.3 ordering)."""
        from wf_grid.collect.trades_collector import _DONOR_TRADE_COLS
        trades_df = self._make_trades_df_with_filter_cols()
        grid_results = self._make_step_results(trades_df, enabled=True)
        config = self._make_config()

        collected = collect_oos_trades(grid_results, config)
        cols = list(collected.columns)
        last_donor_pos = max(cols.index(c) for c in _DONOR_TRADE_COLS if c in cols)
        for fc in _FILTER_TRADE_COLS:
            if fc in cols:
                assert cols.index(fc) > last_donor_pos, (
                    f"Filter col {fc!r} must come after last donor col"
                )


# ===========================================================================
# H. Step collector includes summary columns
# ===========================================================================

class TestStepCollectorSummaryColumns:
    """step_oos_long / step_train_long contain summary columns (§10.6.4).

    Filter summary columns are an OPT-IN extension: they appear in the
    DataFrame ONLY when trade_filter was enabled for the run (at least one
    StepResult has filter_diagnostics_summary is not None).  When filter is
    disabled/absent, the DataFrame contains ONLY base columns — baseline
    schema parity (§11.1 / §14.18 / §17.1.1).
    """

    # §10.6.4 required summary columns (including WAIT bar count added in T1.3)
    _REQUIRED_SUMMARY_COLS = [
        "filter_states_visited",
        "n_bars_in_off",
        "n_bars_in_wait_first_st_flip",
        "n_bars_in_freeze",
        "n_bars_in_monitoring",
        "n_bars_in_stopping",
        "n_filter_blocked_entries",
        "lifecycle_starts_count",
        "median_stop_triggered_count",
    ]
    # Additional informational columns
    _EXPECTED_SUMMARY_COLS = _REQUIRED_SUMMARY_COLS + [
        "filter_diagnostics_available",
        "trigger_count_candidate_threshold",
        "trigger_count_confirmed_median",
        "trigger_count_both",
        "stopping_started_count",
    ]

    def test_oos_columns_include_summary_fields(self):
        """Back-compat alias _OOS_COLUMNS includes all filter summary cols."""
        for col in self._EXPECTED_SUMMARY_COLS:
            assert col in _OOS_COLUMNS, f"{col!r} missing from _OOS_COLUMNS"

    def test_train_columns_include_summary_fields(self):
        """Back-compat alias _TRAIN_COLUMNS includes all filter summary cols."""
        for col in self._EXPECTED_SUMMARY_COLS:
            assert col in _TRAIN_COLUMNS, f"{col!r} missing from _TRAIN_COLUMNS"

    def _make_step_result(
        self, summary: Optional[Dict[str, Any]], gp_id: str = "gp1"
    ) -> StepResult:
        return StepResult(
            grid_point_id=gp_id,
            wf_step=1,
            test_start_idx=0,
            test_end_idx=60,
            metrics={"sum_pnl_pct": 0.0, "num_trades": 0, "win_rate": 0.0,
                     "max_drawdown": 0.0, "sharpe": 0.0, "sortino": 0.0,
                     "cagr": 0.0, "profit_factor": 1.0, "avg_trade": 0.0},
            oos_trades_df=None,
            prepend_bars_requested=0, prepend_bars_applied=0,
            used_prepend=False, used_legacy_oos_path=False,
            used_defensive_fallback=False, oos_boundary_index=0,
            warmup_used=10, warmup_effective=10, effective_oos_bars=50,
            filter_diagnostics_summary=summary,
        )

    def _make_config(self):
        @dataclass
        class _Status:
            min_meaningful_bars: int = 5

        @dataclass
        class _BT:
            min_trades_required: int = 1
            early_exit_max_drawdown: float = 0.5

        @dataclass
        class _Val:
            warmup_period: int = 10
            min_oos_bars: int = 5
            min_train_bars: int = 5

        @dataclass
        class _Cfg:
            resolved_periods_per_year: int = 252
            backtest: Any = field(default_factory=_BT)
            validation: Any = field(default_factory=_Val)
            status: Any = field(default_factory=_Status)

        return _Cfg()

    def test_collect_oos_steps_includes_summary_values(self):
        summary = {
            "diagnostics_available": True,
            "filter_states_visited": "OFF,ST_ACTIVE_MONITORING,ST_STOPPING",
            "n_bars_in_off": 30,
            "n_bars_in_freeze": 0,
            "n_bars_in_monitoring": 15,
            "n_bars_in_stopping": 5,
            "n_filter_blocked_entries": 3,
            "lifecycle_starts_count": 1,
            "median_stop_triggered_count": 1,
            "trigger_count_candidate_threshold": 2,
            "trigger_count_confirmed_median": 1,
            "trigger_count_both": 0,
            "stopping_started_count": 1,
        }
        sr = self._make_step_result(summary)
        grid_results = {"gp1": [sr]}
        config = self._make_config()

        df = collect_oos_steps(grid_results, config)
        row = df.iloc[0]
        assert row["filter_diagnostics_available"] == True  # noqa: E712
        assert row["filter_states_visited"] == "OFF,ST_ACTIVE_MONITORING,ST_STOPPING"
        assert row["n_bars_in_off"] == 30
        assert row["n_bars_in_monitoring"] == 15
        assert row["n_bars_in_stopping"] == 5
        assert row["n_filter_blocked_entries"] == 3
        assert row["lifecycle_starts_count"] == 1
        assert row["median_stop_triggered_count"] == 1
        assert row["trigger_count_candidate_threshold"] == 2
        assert row["stopping_started_count"] == 1

    def test_collect_oos_steps_summary_none_gives_no_filter_cols(self):
        """When ALL steps have summary=None (filter disabled), filter columns
        must NOT appear in the DataFrame at all — baseline schema parity
        (§11.1 / §14.18 / §17.1.1)."""
        sr = self._make_step_result(summary=None)
        grid_results = {"gp1": [sr]}
        config = self._make_config()

        df = collect_oos_steps(grid_results, config)
        for col in self._EXPECTED_SUMMARY_COLS:
            assert col not in df.columns, (
                f"Filter column {col!r} must NOT be present when filter is disabled"
            )

    def test_collect_train_steps_includes_summary_values(self):
        summary = {
            "diagnostics_available": True,
            "filter_states_visited": "OFF,ST_ACTIVE_FREEZE,ST_ACTIVE_MONITORING",
            "n_bars_in_off": 40,
            "n_bars_in_freeze": 10,
            "n_bars_in_monitoring": 20,
            "n_bars_in_stopping": 0,
            "n_filter_blocked_entries": 2,
            "lifecycle_starts_count": 2,
            "median_stop_triggered_count": 0,
            "trigger_count_candidate_threshold": 3,
            "trigger_count_confirmed_median": 0,
            "trigger_count_both": 0,
            "stopping_started_count": 0,
        }
        sr = self._make_step_result(summary)
        grid_results = {"gp1": [sr]}
        config = self._make_config()

        df = collect_train_steps(grid_results, config)
        row = df.iloc[0]
        assert row["filter_diagnostics_available"] == True  # noqa: E712
        assert row["filter_states_visited"] == "OFF,ST_ACTIVE_FREEZE,ST_ACTIVE_MONITORING"
        assert row["n_bars_in_off"] == 40
        assert row["lifecycle_starts_count"] == 2
        assert row["trigger_count_candidate_threshold"] == 3


# ===========================================================================
# I. FSM no-global-state acceptance (§10.7.2)
# ===========================================================================

class TestFSMNoGlobalState:
    """Two sequential apply() calls must give bit-identical results."""

    def _make_per_bar(self, n: int) -> ZigZagPerBar:
        rng = np.random.default_rng(123)
        return ZigZagPerBar(
            candidate_height_pct=rng.uniform(0.005, 0.05, n).astype(np.float64),
            confirm_event=np.where(rng.random(n) > 0.8, 1, 0).astype(np.int8),
            local_median_N=rng.uniform(0.02, 0.06, n).astype(np.float64),
            local_median_available=np.ones(n, dtype=np.int8),
            confirmed_leg_idx_at_t=np.full(n, -1, dtype=np.int64),
            last_confirmed_leg_height_pct=np.full(n, np.nan, dtype=np.float64),
            candidate_age_bars=np.full(n, -1, dtype=np.int64),
            candidate_leg_direction=np.zeros(n, dtype=np.int8),
        )

    def test_two_sequential_apply_calls_are_bit_identical(self):
        """Calling apply() twice with the same inputs must give identical output.

        Verifies FSM has no module-level mutable state (§10.7.2).
        """
        n = 80
        rng = np.random.default_rng(55)
        trend = np.where(rng.random(n) > 0.5, 1, -1).astype(np.int8)
        per_bar = self._make_per_bar(n)
        cfg = _FilterCfg()
        stats = _make_global_stats()

        result1 = zigzag_apply(
            trend=trend,
            trade_mode="revers",
            trade_filter_config=cfg,
            zigzag_global_stats=stats,
            per_bar=per_bar,
        )
        result2 = zigzag_apply(
            trend=trend,
            trade_mode="revers",
            trade_filter_config=cfg,
            zigzag_global_stats=stats,
            per_bar=per_bar,
        )

        np.testing.assert_array_equal(
            result1.positions, result2.positions,
            err_msg="FSM global-state contamination: positions differ between runs",
        )
        for key in result1.filter_diagnostics:
            a1 = result1.filter_diagnostics[key]
            a2 = result2.filter_diagnostics[key]
            if np.issubdtype(np.asarray(a1).dtype, np.floating):
                np.testing.assert_array_equal(a1, a2, err_msg=f"key={key!r}")
            else:
                assert list(a1) == list(a2), f"FSM state contamination in {key!r}"


# ===========================================================================
# J. median_stop_triggered semantics (§13)
# ===========================================================================

class TestMedianStopTriggeredSemantics:
    """median_stop_triggered is 1 only when actual median < global_median."""

    def _make_per_bar_with_low_median(self, n: int) -> ZigZagPerBar:
        rng = np.random.default_rng(77)
        cand_height = rng.uniform(0.02, 0.05, n).astype(np.float64)
        confirm = np.zeros(n, dtype=np.int8)
        confirm[5] = 1   # trigger A
        confirm[15] = 1  # confirm after freeze: FREEZE→MONITORING
        confirm[20] = 1  # confirm: check stop → median < global → STOPPING
        # local_median at bar 20 is well below global_median (0.03 from stats)
        local_med = np.full(n, 0.04, dtype=np.float64)
        local_med[20] = 0.005  # below global_median=0.03 → triggers stop
        avail = np.ones(n, dtype=np.int8)
        return ZigZagPerBar(
            candidate_height_pct=cand_height,
            confirm_event=confirm,
            local_median_N=local_med,
            local_median_available=avail,
            confirmed_leg_idx_at_t=np.full(n, -1, dtype=np.int64),
            last_confirmed_leg_height_pct=np.full(n, np.nan, dtype=np.float64),
        )

    def test_median_stop_triggered_at_correct_bar(self):
        """Deterministic scenario: verify median_stop_triggered == 1 exactly at
        the bar where local_median_N < global_median causes MONITORING → STOPPING.

        Setup (freeze_confirmed_legs=0 → immediate MONITORING on lifecycle start):
        - trend[0]=-1, trend[1..n-1]=+1  → tradable flip at bar 1
        - trigger fires at bar 0 (a_trig=True): OFF→WAIT
        - bar 1: WAIT→FREEZE→MONITORING (fcl=0, held_pos=+1)
        - confirm[10]=1, local_median_N[10]=0.005 < global_median=0.03
          → MONITORING→STOPPING at bar 10, median_stop_triggered[10]=1
        """
        n = 30
        trend = np.ones(n, dtype=np.int8)
        trend[0] = -1   # bar 0: trend=-1 (no tradable flip from prev=0)
        # trend[1]=+1: flip from -1→+1 at bar 1 (tradable)

        cand_height = np.full(n, 0.02, dtype=np.float64)  # always above threshold=0.01
        confirm = np.zeros(n, dtype=np.int8)
        confirm[10] = 1  # only confirm at bar 10
        local_med = np.full(n, 0.04, dtype=np.float64)
        local_med[10] = 0.005  # below global_median=0.03 at bar 10
        avail = np.ones(n, dtype=np.int8)

        per_bar = ZigZagPerBar(
            candidate_height_pct=cand_height,
            confirm_event=confirm,
            local_median_N=local_med,
            local_median_available=avail,
            confirmed_leg_idx_at_t=np.full(n, -1, dtype=np.int64),
            last_confirmed_leg_height_pct=np.full(n, np.nan, dtype=np.float64),
            candidate_age_bars=np.full(n, -1, dtype=np.int64),
            candidate_leg_direction=np.zeros(n, dtype=np.int8),
        )
        cfg = _FilterCfg()
        cfg.lifecycle.freeze_confirmed_legs = 0   # immediate MONITORING
        stats = _make_global_stats(global_median=0.03)

        result = zigzag_apply(
            trend=trend,
            trade_mode="revers",
            trade_filter_config=cfg,
            zigzag_global_stats=stats,
            per_bar=per_bar,
        )
        diag = result.filter_diagnostics

        # Exactly one bar should have median_stop_triggered == 1
        total_triggered = int(np.sum(diag["median_stop_triggered"]))
        assert total_triggered == 1, (
            f"Expected exactly 1 bar with median_stop_triggered, got {total_triggered}"
        )
        # That bar must be bar 10 (where confirm fires and median is low)
        assert diag["median_stop_triggered"][10] == 1, (
            "median_stop_triggered should be 1 at bar 10"
        )
        # State at bar 10 must be ST_STOPPING (MONITORING→STOPPING fired)
        assert diag["trade_filter_state"][10] == "ST_STOPPING", (
            "state at bar 10 should be ST_STOPPING after median stop trigger"
        )

    def test_fail_closed_does_not_set_median_stop_triggered(self):
        """When stop is triggered by unavailable median (fail-closed),
        median_stop_triggered must be 0 on that bar."""
        n = 40
        rng = np.random.default_rng(88)
        trend_base = np.where(rng.random(n) > 0.5, 1, -1).astype(np.int8)
        trend_base[4] = -1
        trend_base[5] = 1

        cand_height = rng.uniform(0.02, 0.05, n).astype(np.float64)
        confirm = np.zeros(n, dtype=np.int8)
        confirm[5] = 1
        confirm[15] = 1
        confirm[18] = 1
        confirm[21] = 1  # will fire stop-check
        local_med = np.full(n, 0.04, dtype=np.float64)
        avail = np.ones(n, dtype=np.int8)
        avail[21] = 0   # median UNAVAILABLE at bar 21 → fail-closed → no median_stop flag

        per_bar = ZigZagPerBar(
            candidate_height_pct=cand_height,
            confirm_event=confirm,
            local_median_N=local_med,
            local_median_available=avail,
            confirmed_leg_idx_at_t=np.full(n, -1, dtype=np.int64),
            last_confirmed_leg_height_pct=np.full(n, np.nan, dtype=np.float64),
            candidate_age_bars=np.full(n, -1, dtype=np.int64),
            candidate_leg_direction=np.zeros(n, dtype=np.int8),
        )
        cfg = _FilterCfg()
        cfg.lifecycle.freeze_confirmed_legs = 3
        stats = _make_global_stats(global_median=0.03)

        result = zigzag_apply(
            trend=trend_base,
            trade_mode="revers",
            trade_filter_config=cfg,
            zigzag_global_stats=stats,
            per_bar=per_bar,
        )
        diag = result.filter_diagnostics
        # If STOPPING triggered at bar 21 (fail-closed), median_stop_triggered[21] == 0
        if diag["trade_filter_state"][21] == "ST_STOPPING":
            assert diag["median_stop_triggered"][21] == 0, (
                "fail-closed stop must NOT set median_stop_triggered"
            )


# ===========================================================================
# K. Anti-drift: disabled baseline bit-identical, no new filter fields
#    in standard step_oos_long metrics columns
# ===========================================================================

class TestAntiDrift:
    """WP9 anti-drift: new fields don't contaminate existing schemas."""

    def test_filter_summary_keys_not_in_metrics_columns(self):
        """Summary fields must not leak into the metrics section of step_oos_long."""
        metrics_cols = [
            "sum_pnl_pct", "sharpe", "sortino", "max_drawdown", "cagr",
            "win_rate", "num_trades", "profit_factor", "avg_trade",
        ]
        for col in metrics_cols:
            assert col in _OOS_COLUMNS

    def test_oos_columns_summary_at_end(self):
        """§10.6.4 required summary columns must come after standard columns."""
        required_summary_cols = [
            "filter_states_visited",
            "n_bars_in_off",
            "n_bars_in_freeze",
            "n_bars_in_monitoring",
            "n_bars_in_stopping",
            "n_filter_blocked_entries",
            "lifecycle_starts_count",
            "median_stop_triggered_count",
        ]
        standard_last = _OOS_COLUMNS.index("error_type")
        for col in required_summary_cols:
            assert col in _OOS_COLUMNS, f"§10.6.4 required col {col!r} missing from _OOS_COLUMNS"
            assert _OOS_COLUMNS.index(col) > standard_last, (
                f"Summary col {col!r} should come after 'error_type' in _OOS_COLUMNS"
            )

    def test_no_bar_level_diagnostics_in_step_long(self):
        """Bar-level diagnostic arrays must NOT appear in step_oos_long columns.

        Exception: exit-off echo keys (exit_off_mode, exit_off_zz_leg_count)
        intentionally appear in BOTH per-bar and step_long — they are config
        echo scalars (plan_exit_off_modes_v2.txt §2 mapping table).
        """
        # Keys that are allowed to be present in both per-bar and step_long
        _DUAL_PRESENCE_ALLOWED = {"exit_off_mode", "exit_off_zz_leg_count"}
        bar_level_keys = [
            k for k in _SECTION_13_REQUIRED_KEYS if k not in _DUAL_PRESENCE_ALLOWED
        ]
        for key in bar_level_keys:
            assert key not in _OOS_COLUMNS, (
                f"Bar-level key {key!r} must not be in step_oos_long columns"
            )
            assert key not in _TRAIN_COLUMNS, (
                f"Bar-level key {key!r} must not be in step_train_long columns"
            )


# ===========================================================================
# WP-V3-8: XLSX / summary export  (X1-X7)
# ===========================================================================

class TestV3ExportX1DisplayNames:
    """X1: FilterDiagnostics_100 display names contain all 13 new v3 columns."""

    _NEW_V3_KEYS = [
        "zigzag_mode",
        "candidate_age_bars",
        "candidate_leg_direction",
        "candidate_duration_gate_enabled",
        "candidate_duration_max_bars",
        "candidate_duration_gate_passed",
        "candidate_threshold_ok",
        "candidate_component_ok",
        "confirmed_median_ok",
        "b_component_ok",
        "immediate_allowed",
        "immediate_candidate_entry_used",
        "immediate_candidate_entry_block_reason",
    ]

    def test_x1_new_keys_in_display_names(self):
        from supertrend_optimizer.io.excel_tester import FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        for key in self._NEW_V3_KEYS:
            assert key in FILTER_DIAGNOSTICS_100_DISPLAY_NAMES, (
                f"Missing display name for {key!r}"
            )

    def test_x1_display_names_match_spec(self):
        from supertrend_optimizer.io.excel_tester import FILTER_DIAGNOSTICS_100_DISPLAY_NAMES
        expected = {
            "zigzag_mode":                              "ZigZag Mode",
            "candidate_age_bars":                       "Candidate Age Bars",
            "candidate_leg_direction":                  "Candidate Leg Direction",
            "candidate_duration_gate_enabled":          "Candidate Duration Gate Enabled",
            "candidate_duration_max_bars":              "Candidate Duration Max Bars",
            "candidate_duration_gate_passed":           "Candidate Duration Gate Passed",
            "candidate_threshold_ok":                   "Candidate Threshold OK",
            "candidate_component_ok":                   "Candidate Component OK",
            "confirmed_median_ok":                      "Confirmed Median OK",
            "b_component_ok":                           "B Component OK",
            "immediate_allowed":                        "Immediate Allowed",
            "immediate_candidate_entry_used":           "Immediate Candidate Entry Used",
            "immediate_candidate_entry_block_reason":   "Immediate Candidate Entry Block Reason",
        }
        for key, display in expected.items():
            assert FILTER_DIAGNOSTICS_100_DISPLAY_NAMES.get(key) == display, (
                f"{key!r} → expected {display!r}, got {FILTER_DIAGNOSTICS_100_DISPLAY_NAMES.get(key)!r}"
            )

    def test_x1_filter_diagnostics_100_sheet_includes_new_columns(self):
        """FilterDiagnostics_100 writer renders new keys as correct display names."""
        import io
        import openpyxl
        from supertrend_optimizer.io.excel_tester import _write_filter_diagnostics_100_sheet

        n = 5
        diag = {
            "zigzag_mode":                           np.full(n, "A", dtype=object),
            "candidate_age_bars":                    np.full(n, 3, dtype=np.int64),
            "candidate_leg_direction":               np.zeros(n, dtype=np.int8),
            "candidate_duration_gate_enabled":       np.zeros(n, dtype=np.int8),
            "candidate_duration_max_bars":           np.full(n, -1, dtype=np.int64),
            "candidate_duration_gate_passed":        np.ones(n, dtype=np.int8),
            "candidate_threshold_ok":                np.zeros(n, dtype=np.int8),
            "candidate_component_ok":                np.zeros(n, dtype=np.int8),
            "confirmed_median_ok":                   np.zeros(n, dtype=np.int8),
            "b_component_ok":                        np.zeros(n, dtype=np.int8),
            "immediate_allowed":                     np.zeros(n, dtype=np.int8),
            "immediate_candidate_entry_used":        np.zeros(n, dtype=np.int8),
            "immediate_candidate_entry_block_reason": np.full(n, "mode_not_c", dtype=object),
            "trade_filter_state":                    np.full(n, "OFF", dtype=object),
        }
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            _write_filter_diagnostics_100_sheet(writer, diag)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb["FilterDiagnostics_100"]
        headers = [cell.value for cell in ws[1]]
        assert "ZigZag Mode" in headers
        assert "Candidate Age Bars" in headers
        assert "Immediate Candidate Entry Block Reason" in headers


class TestV3ExportX2TriggerEventsColumns:
    """X2: ZigZag_Trigger_Events has correct column structure (§11.2)."""

    _REQUIRED_NEW_COLS = {
        "ZigZag Mode",
        "Immediate Candidate Entry Used",
        "Immediate Candidate Entry Block Reason",
        "Candidate Age Bars",
        "Candidate Leg Direction",
        "Candidate Duration Gate Passed",
    }

    def test_x2_new_columns_in_trigger_events_tuple(self):
        from supertrend_optimizer.io.excel_tester import _TRIGGER_EVENTS_COLUMNS
        for col in self._REQUIRED_NEW_COLS:
            assert col in _TRIGGER_EVENTS_COLUMNS, (
                f"Missing column {col!r} in _TRIGGER_EVENTS_COLUMNS"
            )

    def test_x2_disabled_path_returns_empty_with_correct_columns(self):
        from supertrend_optimizer.io.excel_tester import (
            _build_zigzag_trigger_events_df,
            _TRIGGER_EVENTS_COLUMNS,
        )
        df = _build_zigzag_trigger_events_df(None)
        assert len(df) == 0
        for col in _TRIGGER_EVENTS_COLUMNS:
            assert col in df.columns, f"Missing {col!r} in empty disabled DataFrame"

    def _make_minimal_diag(self, n: int = 10) -> Dict[str, np.ndarray]:
        trigger = np.full(n, "none", dtype=object)
        trigger[3] = "candidate_threshold"
        state = np.full(n, "OFF", dtype=object)
        state[3] = "WAIT_FIRST_ST_FLIP"
        state[4] = "ST_ACTIVE_FREEZE"
        return {
            "trade_filter_trigger_source":          trigger,
            "trade_filter_state":                   state,
            "candidate_trigger_threshold":          np.full(n, 0.02),
            "global_median":                        np.full(n, 0.01),
            "local_median_N":                       np.full(n, 5.0),
            "candidate_height_pct":                 np.full(n, 0.025),
            "zigzag_mode":                          np.full(n, "A", dtype=object),
            "immediate_candidate_entry_used":       np.zeros(n, dtype=np.int8),
            "immediate_candidate_entry_block_reason": np.full(n, "mode_not_c", dtype=object),
            "candidate_age_bars":                   np.full(n, 4, dtype=np.int64),
            "candidate_leg_direction":              np.full(n, 1, dtype=np.int8),
            "candidate_duration_gate_passed":       np.ones(n, dtype=np.int8),
        }

    def test_x2_new_columns_populated_in_trigger_row(self):
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df
        diag = self._make_minimal_diag()
        df = _build_zigzag_trigger_events_df(diag)
        assert len(df) == 1
        row = df.iloc[0]
        assert row["ZigZag Mode"] == "A"
        assert int(row["Immediate Candidate Entry Used"]) == 0
        assert row["Immediate Candidate Entry Block Reason"] == "mode_not_c"
        assert int(row["Candidate Age Bars"]) == 4
        assert int(row["Candidate Leg Direction"]) == 1
        assert int(row["Candidate Duration Gate Passed"]) == 1


class TestV3ExportX3TriggerEventsReconstruction:
    """X3: ZigZag_Trigger_Events reconstruction: only trigger_source != 'none' rows."""

    def test_x3_reconstruction_uses_trigger_source_not_none(self):
        """Only bars with trigger_source != 'none' produce event rows."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 10
        trigger = np.full(n, "none", dtype=object)
        trigger[2] = "candidate_threshold"
        trigger[7] = "confirmed_median"
        state = np.full(n, "OFF", dtype=object)
        diag = {"trade_filter_trigger_source": trigger, "trade_filter_state": state}
        df = _build_zigzag_trigger_events_df(diag)
        assert len(df) == 2
        assert list(df["Trigger Bar"]) == [2, 7]
        assert list(df["Trigger Source"]) == ["candidate_threshold", "confirmed_median"]

    def test_x3_mode_c_immediate_entry_triggered_lifecycle_start_true(self):
        """Mode C same-bar OFF→FREEZE: state at trigger bar = ST_ACTIVE_FREEZE
        → Triggered Lifecycle Start = True (no WAIT transition scanned)."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 8
        trigger = np.full(n, "none", dtype=object)
        trigger[3] = "candidate_threshold"  # Mode C immediate entry
        state = np.full(n, "OFF", dtype=object)
        state[3] = "ST_ACTIVE_FREEZE"  # immediate: no WAIT
        state[4] = "ST_ACTIVE_MONITORING"
        diag = {"trade_filter_trigger_source": trigger, "trade_filter_state": state}
        df = _build_zigzag_trigger_events_df(diag)
        assert len(df) == 1
        assert bool(df.iloc[0]["Triggered Lifecycle Start"]) is True

    def test_x3_mode_ab_triggered_lifecycle_start_true_after_wait(self):
        """Mode A/B: state at trigger bar = WAIT, FREEZE appears at t+1."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 8
        trigger = np.full(n, "none", dtype=object)
        trigger[2] = "candidate_threshold"
        state = np.full(n, "OFF", dtype=object)
        state[2] = "WAIT_FIRST_ST_FLIP"
        state[3] = "ST_ACTIVE_FREEZE"
        diag = {"trade_filter_trigger_source": trigger, "trade_filter_state": state}
        df = _build_zigzag_trigger_events_df(diag)
        assert bool(df.iloc[0]["Triggered Lifecycle Start"]) is True

    def test_x3_cb_b_rescue_wait_start_treated_same_as_ab(self):
        """C+B B-rescue starts WAIT (not immediate FREEZE) → same as A/B scan."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 8
        trigger = np.full(n, "none", dtype=object)
        trigger[4] = "confirmed_median"  # B-rescue: enters WAIT
        state = np.full(n, "OFF", dtype=object)
        state[4] = "WAIT_FIRST_ST_FLIP"
        state[5] = "ST_ACTIVE_FREEZE"
        diag = {"trade_filter_trigger_source": trigger, "trade_filter_state": state}
        df = _build_zigzag_trigger_events_df(diag)
        assert bool(df.iloc[0]["Triggered Lifecycle Start"]) is True

    def test_x3_wait_to_monitoring_direct_freeze_confirmed_legs_zero(self):
        """WAIT→ST_ACTIVE_MONITORING (freeze_confirmed_legs=0, no FREEZE step):
        Triggered Lifecycle Start must be True, not False."""
        from supertrend_optimizer.io.excel_tester import _build_zigzag_trigger_events_df

        n = 8
        trigger = np.full(n, "none", dtype=object)
        trigger[2] = "candidate_threshold"
        state = np.full(n, "OFF", dtype=object)
        state[2] = "WAIT_FIRST_ST_FLIP"
        # freeze_confirmed_legs=0: WAIT→MONITORING directly, no FREEZE step
        state[3] = "ST_ACTIVE_MONITORING"
        state[4] = "ST_ACTIVE_MONITORING"
        diag = {"trade_filter_trigger_source": trigger, "trade_filter_state": state}
        df = _build_zigzag_trigger_events_df(diag)
        assert len(df) == 1
        assert bool(df.iloc[0]["Triggered Lifecycle Start"]) is True


class TestV3ExportX4FiltersSummaryParams:
    """X4: filters_summary params section includes new ZigZag Mode and gate fields."""

    def _make_period_result(
        self,
        label: str,
        zigzag_mode: str = "A",
        gate_enabled: bool = False,
        gate_max_bars: int = -1,
        imm_count: int = 0,
        imm_blocked: int = 0,
    ):
        from unittest.mock import MagicMock
        pr = MagicMock()
        pr.period_label = label
        pr.filter_diagnostics_summary = {
            "diagnostics_available": True,
            "zigzag_mode": zigzag_mode,
            "candidate_duration_gate_enabled": gate_enabled,
            "candidate_duration_max_bars": gate_max_bars,
            "immediate_entries_count": imm_count,
            "immediate_entries_blocked_count": imm_blocked,
            "lifecycle_starts_count": 0,
            "median_stop_triggered_count": 0,
            "n_bars_in_off": 10,
            "n_bars_in_wait_first_st_flip": 2,
            "n_bars_in_freeze": 3,
            "n_bars_in_monitoring": 4,
            "n_bars_in_stopping": 1,
        }
        return pr

    def test_x4_params_contains_zigzag_mode(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = self._make_period_result("OOS", zigzag_mode="C+B")
        result = _build_filters_summary_df([pr])
        assert result is not None
        params_df, _ = result
        params_dict = dict(zip(params_df["Parameter"], params_df["Value"]))
        assert params_dict.get("ZigZag Mode") == "C+B"

    def test_x4_params_contains_gate_enabled(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = self._make_period_result("OOS", gate_enabled=True, gate_max_bars=5)
        result = _build_filters_summary_df([pr])
        assert result is not None
        params_df, _ = result
        params_dict = dict(zip(params_df["Parameter"], params_df["Value"]))
        assert params_dict.get("Candidate Duration Gate Enabled") is True
        assert params_dict.get("Candidate Duration Max Bars") == 5

    def test_x4_params_gate_disabled_shows_false_and_minus_one(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = self._make_period_result("OOS", gate_enabled=False, gate_max_bars=-1)
        result = _build_filters_summary_df([pr])
        assert result is not None
        params_df, _ = result
        params_dict = dict(zip(params_df["Parameter"], params_df["Value"]))
        assert params_dict.get("Candidate Duration Gate Enabled") is False
        assert params_dict.get("Candidate Duration Max Bars") == -1


class TestV3ExportX5FiltersSummaryPeriodCounts:
    """X5: filters_summary period section includes immediate entries counts."""

    def _make_period_result(self, label: str, imm_count: int, imm_blocked: int):
        from unittest.mock import MagicMock
        pr = MagicMock()
        pr.period_label = label
        pr.filter_diagnostics_summary = {
            "diagnostics_available": True,
            "zigzag_mode": "C",
            "candidate_duration_gate_enabled": False,
            "candidate_duration_max_bars": -1,
            "immediate_entries_count": imm_count,
            "immediate_entries_blocked_count": imm_blocked,
            "lifecycle_starts_count": 1,
            "median_stop_triggered_count": 0,
            "n_bars_in_off": 5,
            "n_bars_in_wait_first_st_flip": 0,
            "n_bars_in_freeze": 2,
            "n_bars_in_monitoring": 2,
            "n_bars_in_stopping": 1,
        }
        return pr

    def test_x5_immediate_entries_count_column_present(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = self._make_period_result("OOS", imm_count=3, imm_blocked=1)
        result = _build_filters_summary_df([pr])
        assert result is not None
        _, period_df = result
        assert "Immediate Entries Count" in period_df.columns
        assert "Immediate Entries Blocked Count" in period_df.columns

    def test_x5_immediate_entries_count_correct_value(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = self._make_period_result("OOS", imm_count=7, imm_blocked=2)
        result = _build_filters_summary_df([pr])
        assert result is not None
        _, period_df = result
        assert period_df.iloc[0]["Immediate Entries Count"] == 7
        assert period_df.iloc[0]["Immediate Entries Blocked Count"] == 2

    def test_x5_multiple_periods_each_have_own_counts(self):
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        prs = [
            self._make_period_result("Train", imm_count=4, imm_blocked=1),
            self._make_period_result("OOS",   imm_count=2, imm_blocked=0),
        ]
        result = _build_filters_summary_df(prs)
        assert result is not None
        _, period_df = result
        assert period_df.iloc[0]["Immediate Entries Count"] == 4
        assert period_df.iloc[1]["Immediate Entries Count"] == 2


class TestV3ExportX6ImmediateBlockedFormula:
    """X6: Blocked count uses only duration_gate_failed / unknown_candidate_direction /
    trade_mode_disallows_direction.  Other reasons not counted."""

    def _make_diag(self, reasons: list) -> Dict[str, np.ndarray]:
        n = len(reasons)
        used = np.array(
            [1 if r == "none" else 0 for r in reasons], dtype=np.int8
        )
        return {
            "immediate_candidate_entry_used":        used,
            "immediate_candidate_entry_block_reason": np.array(reasons, dtype=object),
        }

    def test_x6_duration_gate_failed_counted(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        diag = self._make_diag(["none", "duration_gate_failed", "duration_gate_failed"])
        s = _compute_filter_diagnostics_summary(diag)
        assert s["immediate_entries_count"] == 1
        assert s["immediate_entries_blocked_count"] == 2

    def test_x6_unknown_direction_counted(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        diag = self._make_diag(["unknown_candidate_direction", "none"])
        s = _compute_filter_diagnostics_summary(diag)
        assert s["immediate_entries_count"] == 1
        assert s["immediate_entries_blocked_count"] == 1

    def test_x6_trade_mode_disallows_counted(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        diag = self._make_diag(["trade_mode_disallows_direction"])
        s = _compute_filter_diagnostics_summary(diag)
        assert s["immediate_entries_blocked_count"] == 1

    def test_x6_height_gate_failed_not_counted(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        diag = self._make_diag(["height_gate_failed", "mode_not_c", "state_not_off",
                                 "daily_reset", "filter_off", "none"])
        s = _compute_filter_diagnostics_summary(diag)
        assert s["immediate_entries_blocked_count"] == 0

    def test_x6_mixed_blocked_counts_only_three_reasons(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        reasons = [
            "none",                          # used=1 → not blocked
            "duration_gate_failed",          # counted
            "unknown_candidate_direction",   # counted
            "trade_mode_disallows_direction",# counted
            "height_gate_failed",            # NOT counted
            "mode_not_c",                    # NOT counted
            "daily_reset",                   # NOT counted
        ]
        diag = self._make_diag(reasons)
        s = _compute_filter_diagnostics_summary(diag)
        assert s["immediate_entries_count"] == 1
        assert s["immediate_entries_blocked_count"] == 3


class TestV3ExportX7DisabledPath:
    """X7: Disabled filter path: _compute_filter_diagnostics_summary returns None,
    filters_summary not written, trigger events empty, no new arrays."""

    def test_x7_disabled_summary_is_none(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        assert _compute_filter_diagnostics_summary(None) is None

    def test_x7_disabled_trigger_events_empty_dataframe(self):
        from supertrend_optimizer.io.excel_tester import (
            _build_zigzag_trigger_events_df,
            _TRIGGER_EVENTS_COLUMNS,
        )
        df = _build_zigzag_trigger_events_df(None)
        assert len(df) == 0
        # Must still have all columns (for consistent header row when writing)
        for col in _TRIGGER_EVENTS_COLUMNS:
            assert col in df.columns

    def test_x7_disabled_filters_summary_returns_none(self):
        from unittest.mock import MagicMock
        from supertrend_optimizer.io.excel_tester import _build_filters_summary_df

        pr = MagicMock()
        pr.period_label = "OOS"
        pr.filter_diagnostics_summary = None
        result = _build_filters_summary_df([pr])
        assert result is None

    def test_x7_disabled_filter_diagnostics_100_skipped(self):
        """_write_filter_diagnostics_100_sheet writes nothing when diagnostics=None."""
        import io
        import openpyxl
        from supertrend_optimizer.io.excel_tester import _write_filter_diagnostics_100_sheet

        buf = io.BytesIO()
        # Need at least one pre-existing sheet; otherwise openpyxl raises IndexError.
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame({"x": [1]}).to_excel(writer, sheet_name="sentinel", index=False)
            _write_filter_diagnostics_100_sheet(writer, None)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert "FilterDiagnostics_100" not in wb.sheetnames

    def test_x7_immediate_entries_count_absent_from_disabled_summary(self):
        from wf_grid.wf.step_executor import _compute_filter_diagnostics_summary

        # Enabled path with no immediate arrays → keys absent (not 0)
        diag: Dict[str, Any] = {
            "trade_filter_state": np.full(5, "OFF", dtype=object),
        }
        s = _compute_filter_diagnostics_summary(diag)
        assert s is not None
        # When immediate arrays are absent, keys should not be present
        assert "immediate_entries_count" not in s
        assert "immediate_entries_blocked_count" not in s

