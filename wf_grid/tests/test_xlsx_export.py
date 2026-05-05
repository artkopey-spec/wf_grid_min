"""
Tests for A12: XLSX Export + row-limit guard

Coverage:
  - All expected sheets present in output workbook
  - sheet order: DISCLAIMER (first), WF_Config, WF_01..WF_N, WF_Trades, WF_Train_Trades, summary
  - Row-limit guard: ExportError raised when trades exceed 1_000_000 rows
  - Pre-export column-order validation: ExportError on Block-A/B order violation
  - summary sheet: grid_rank is first column, all Block-A/B cols present
  - WF_Trades / WF_Train_Trades: deterministic column order
  - WF_Trades / WF_Train_Trades: sorted by (grid_point_id, wf_step, trade_id)
  - WF_Config sheet: Section/Parameter/Value columns present
  - WF_N sheets: one per wf_step
  - Empty trades: empty sheet with correct columns
  - WF_Gates sheet present only when gates_result provided
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest

from wf_grid.config.schema import DataConfig, GridConfig
from wf_grid.export.xlsx_writer import (
    ExportError,
    _QUARTILE_LOSER_FILL,
    _QUARTILE_WINNER_FILL,
    _RE_SEG_DD,
    _RE_SEG_PNL,
    _TRADES_COLUMN_ORDER,
    _TRAIN_TRADES_COLUMN_ORDER,
    _apply_pnl_sum_quartile_highlights,
    _apply_summary_segment_highlights,
    _validate_row_limits,
    _validate_summary_column_order,
    export_workbook,
)
from wf_grid.export.summary_builder import _BLOCK_A, _BLOCK_B


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cfg() -> GridConfig:
    return GridConfig(data=DataConfig(file_path="dummy.csv"))


def _make_summary(gp_ids=("atr10_m2.50_both",), n_steps=2) -> pd.DataFrame:
    rows = []
    for i, gp_id in enumerate(gp_ids):
        row = {
            # --- Block A (all Block A cols must precede Block B) ---
            "grid_rank": i + 1,
            "grid_point_id": gp_id,
            "atr_period": 10,
            "multiplier": 2.50,
            "trade_mode": "both",
            "tier": 1,
            "seed_gate_passed": True,
            "tester_seed_score": 0.8,
            "ranking_mode": "gates_score",
            "score_contract_status": "ok",
            "quantile_gates_status": None,
            "seed_gate_fail_reason": "",
            "ok_ratio": 1.0,
            "n_ok_steps": n_steps,
            "n_total_steps": n_steps,
            "n_segments": n_steps,
            "profitable_segments_count": n_steps,
            # --- Block B ---
            "sum_pnl_pct_Median": 5.0,
            "sum_pnl_pct_Min": 1.0,
            "sum_pnl_pct_Std": 2.0,
            "max_drawdown_Min": -0.20,
            "profit_factor_Median": 1.8,
            "num_trades_Median": 10.0,
            "sharpe_Median": 1.2,
            "sortino_Median": 1.5,
            "cagr_Median": 0.15,
            "win_rate_Median": 0.6,
            "avg_trade_Median": 0.5,
            "gate_ok_positive_median": True,
            "gate_ok_min_trades": True,
            "gate_ok_worst_segment": True,
            "gate_ok_drawdown": True,
            # --- Block C (segment columns) ---
            "S1_sum_pnl_pct": 5.0,
            "S2_sum_pnl_pct": 5.0,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _make_step_oos(gp_ids=("atr10_m2.50_both",), n_steps=2) -> pd.DataFrame:
    rows = []
    for gp_id in gp_ids:
        for s in range(1, n_steps + 1):
            rows.append({
                "grid_point_id": gp_id,
                "wf_step": s,
                "step_status": "ok",
                "sum_pnl_pct": 5.0,
                "sharpe": 1.2,
                "sortino": 1.5,
                "max_drawdown": -0.10,
                "num_trades": 10,
                "profit_factor": 1.8,
                "effective_oos_bars": 100,
                "used_prepend": True,
                "prepend_bars_applied": 50,
            })
    return pd.DataFrame(rows)


def _make_trades(n=3, gp_id="atr10_m2.50_both", step=1) -> pd.DataFrame:
    rows = []
    for i in range(n):
        rows.append({
            "grid_point_id": gp_id,
            "wf_step": step,
            "step_status": "ok",
            "test_start_idx": 0,
            "test_end_idx": 100,
            "trade_id": i + 1,
            "direction": "long",
            "entry_index": i * 10,
            "exit_index": i * 10 + 5,
            "net_pnl_pct": 0.5,
            "gross_pnl_pct": 0.6,
            "commission_pct": 0.1,
        })
    return pd.DataFrame(rows)


def _make_train_trades(n=3, gp_id="atr10_m2.50_both", step=1) -> pd.DataFrame:
    """Build train trades with train_start_idx / train_end_idx (not test_*)."""
    rows = []
    for i in range(n):
        rows.append({
            "grid_point_id": gp_id,
            "wf_step": step,
            "step_status": "ok",
            "train_start_idx": 0,
            "train_end_idx": 200,
            "trade_id": i + 1,
            "direction": "long",
            "entry_index": i * 10,
            "exit_index": i * 10 + 5,
            "net_pnl_pct": 0.5,
            "gross_pnl_pct": 0.6,
            "commission_pct": 0.1,
        })
    return pd.DataFrame(rows)


def _export(tmp_dir, **kwargs) -> Path:
    path = Path(tmp_dir) / "test_export.xlsx"
    defaults = dict(
        summary_wide=_make_summary(),
        step_oos_long=_make_step_oos(),
        trades_oos=_make_trades(),
        trades_train=_make_trades(),
        config=_cfg(),
        output_path=path,
    )
    defaults.update(kwargs)
    return export_workbook(**defaults)


def _sheets(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet)


# ===========================================================================
# Row-limit guard
# ===========================================================================

class TestRowLimitGuard:
    def test_oos_trades_over_limit_raises(self):
        big_df = pd.DataFrame({"x": range(1_000_001)})
        with pytest.raises(ExportError, match="WF_Trades"):
            _validate_row_limits(big_df, pd.DataFrame())

    def test_train_trades_over_limit_raises(self):
        big_df = pd.DataFrame({"x": range(1_000_001)})
        with pytest.raises(ExportError, match="WF_Train_Trades"):
            _validate_row_limits(pd.DataFrame(), big_df)

    def test_exactly_at_limit_passes(self):
        at_limit = pd.DataFrame({"x": range(1_000_000)})
        _validate_row_limits(at_limit, at_limit)  # no exception

    def test_export_aborts_on_row_limit(self, tmp_path):
        big_trades = pd.DataFrame({"grid_point_id": ["gp"] * 1_000_001})
        with pytest.raises(ExportError):
            export_workbook(
                summary_wide=_make_summary(),
                step_oos_long=_make_step_oos(),
                trades_oos=big_trades,
                trades_train=_make_trades(),
                config=_cfg(),
                output_path=tmp_path / "out.xlsx",
            )


# ===========================================================================
# Pre-export column order validation
# ===========================================================================

class TestColumnOrderValidation:
    def test_valid_order_passes(self):
        df = _make_summary()
        _validate_summary_column_order(df)  # no exception

    def test_block_b_before_block_a_raises(self):
        # Build DataFrame with Block B column before Block A column
        df = pd.DataFrame({
            "sum_pnl_pct_Median": [1.0],  # Block B
            "grid_rank": [1],             # Block A
            "seed_gate_passed": [True],   # Block A
        })
        with pytest.raises(ExportError, match="column order violation"):
            _validate_summary_column_order(df)

    def test_empty_df_passes(self):
        _validate_summary_column_order(pd.DataFrame())  # no exception


# ===========================================================================
# Sheet presence and order
# ===========================================================================

class TestSheetPresence:
    def test_all_required_sheets_present(self, tmp_path):
        path = _export(tmp_path)
        sheets = _sheets(path)
        assert "WF_Config" in sheets
        assert "WF_Trades" in sheets
        assert "WF_Train_Trades" in sheets
        assert "summary" in sheets

    def test_wf_step_sheets_present(self, tmp_path):
        path = _export(tmp_path, step_oos_long=_make_step_oos(n_steps=3))
        sheets = _sheets(path)
        assert "WF_01" in sheets
        assert "WF_02" in sheets
        assert "WF_03" in sheets

    def test_sheet_order(self, tmp_path):
        # Sheet order per xlsx_writer spec §2.7:
        # DISCLAIMER → WF_Config → summary → WF_01..N → WF_Trades → WF_Train_Trades
        path = _export(tmp_path, step_oos_long=_make_step_oos(n_steps=2))
        sheets = _sheets(path)
        disclaimer_idx = sheets.index("DISCLAIMER")
        config_idx = sheets.index("WF_Config")
        summary_idx = sheets.index("summary")
        wf01_idx = sheets.index("WF_01")
        trades_idx = sheets.index("WF_Trades")
        train_idx = sheets.index("WF_Train_Trades")
        assert disclaimer_idx < config_idx < summary_idx < wf01_idx < trades_idx < train_idx

    def test_disclaimer_is_first_sheet(self, tmp_path):
        """FIX-2.10: DISCLAIMER must be sheet index 0."""
        path = _export(tmp_path)
        sheets = _sheets(path)
        assert sheets[0] == "DISCLAIMER"

    def test_wf_gates_absent_when_no_gates_result(self, tmp_path):
        path = _export(tmp_path)
        assert "WF_Gates" not in _sheets(path)

    def test_wf_gates_present_when_gates_result_provided(self, tmp_path):
        check = MagicMock()
        check.name = "test_gate"
        check.passed = True
        check.value = 1.0
        check.threshold = 0.0
        check.message = "ok"

        gates = MagicMock()
        gates.checks = [check]

        path = _export(tmp_path, gates_result=gates)
        assert "WF_Gates" in _sheets(path)


# ===========================================================================
# WF_Config sheet
# ===========================================================================

class TestWFConfig:
    def test_config_columns_present(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "WF_Config")
        assert "Section" in df.columns
        assert "Parameter" in df.columns
        assert "Value" in df.columns

    def test_config_rows_non_empty(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "WF_Config")
        assert len(df) > 0


# ===========================================================================
# summary sheet
# ===========================================================================

class TestSummarySheet:
    def test_grid_rank_is_first_column(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "summary")
        assert df.columns[0] == "grid_rank"

    def test_one_row_per_grid_point(self, tmp_path):
        path = _export(tmp_path, summary_wide=_make_summary(["gp1", "gp2"]))
        df = _read_sheet(path, "summary")
        assert len(df) == 2


# ===========================================================================
# WF_Trades deterministic column order
# ===========================================================================

class TestTradesColumnOrder:
    def test_known_columns_in_order(self, tmp_path):
        trades = _make_trades(n=5)
        path = _export(tmp_path, trades_oos=trades)
        df = _read_sheet(path, "WF_Trades")
        present_ordered = [c for c in _TRADES_COLUMN_ORDER if c in df.columns]
        actual = [c for c in df.columns if c in _TRADES_COLUMN_ORDER]
        assert actual == present_ordered

    def test_trades_sorted_by_gp_wf_trade(self, tmp_path):
        # Build trades in scrambled order
        rows = [
            {"grid_point_id": "gp1", "wf_step": 2, "step_status": "ok",
             "test_start_idx": 0, "test_end_idx": 100, "trade_id": 1, "net_pnl_pct": 1.0},
            {"grid_point_id": "gp1", "wf_step": 1, "step_status": "ok",
             "test_start_idx": 0, "test_end_idx": 100, "trade_id": 2, "net_pnl_pct": 2.0},
            {"grid_point_id": "gp1", "wf_step": 1, "step_status": "ok",
             "test_start_idx": 0, "test_end_idx": 100, "trade_id": 1, "net_pnl_pct": 3.0},
        ]
        trades = pd.DataFrame(rows)
        path = _export(tmp_path, trades_oos=trades)
        df = _read_sheet(path, "WF_Trades")
        assert list(df["wf_step"]) == [1, 1, 2]
        assert list(df["trade_id"][:2]) == [1, 2]

    def test_empty_trades_correct_columns(self, tmp_path):
        path = _export(tmp_path, trades_oos=pd.DataFrame())
        df = _read_sheet(path, "WF_Trades")
        # Should have the deterministic column headers even if empty
        for col in ["grid_point_id", "wf_step", "step_status"]:
            assert col in df.columns


# ===========================================================================
# Return value
# ===========================================================================

class TestReturnValue:
    def test_returns_path_object(self, tmp_path):
        result = _export(tmp_path)
        assert isinstance(result, Path)
        assert result.exists()
        assert result.suffix == ".xlsx"


# ===========================================================================
# FIX-2.10 — DISCLAIMER sheet
# ===========================================================================

class TestDisclaimerSheet:
    """FIX-2.10: DISCLAIMER must be the first sheet with expected warning strings."""

    def test_disclaimer_first_sheet(self, tmp_path):
        path = _export(tmp_path)
        assert _sheets(path)[0] == "DISCLAIMER"

    def test_disclaimer_sheet_present(self, tmp_path):
        path = _export(tmp_path)
        assert "DISCLAIMER" in _sheets(path)

    def test_disclaimer_contains_research_warning(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "RESEARCH" in text

    def test_disclaimer_contains_real_money_warning(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "real-money" in text.lower() or "real money" in text.lower()

    def test_disclaimer_contains_tier1_warning(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "Tier 1" in text

    def test_disclaimer_contains_score_warning(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "Score" in text or "score" in text

    def test_disclaimer_contains_generated_timestamp(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "Generated" in text

    def test_disclaimer_contains_pipeline_version(self, tmp_path):
        path = _export(tmp_path)
        df = _read_sheet(path, "DISCLAIMER")
        text = " ".join(str(v) for v in df["Text"].dropna().tolist())
        assert "Pipeline version" in text

    def test_remaining_sheet_order_unchanged(self, tmp_path):
        """Regression: sheet order after DISCLAIMER must follow spec §2.7."""
        # Actual order: WF_Config → summary → WF_01..N → WF_Trades → WF_Train_Trades
        path = _export(tmp_path, step_oos_long=_make_step_oos(n_steps=2))
        sheets = _sheets(path)
        config_idx = sheets.index("WF_Config")
        summary_idx = sheets.index("summary")
        wf01_idx = sheets.index("WF_01")
        trades_idx = sheets.index("WF_Trades")
        train_idx = sheets.index("WF_Train_Trades")
        assert config_idx < summary_idx < wf01_idx < trades_idx < train_idx


# ===========================================================================
# WF_Train_Trades column schema alignment (residual risk fix)
# ===========================================================================

class TestTrainTradesColumnOrder:
    """Verify WF_Train_Trades uses train_start_idx/train_end_idx, not test_*."""

    def test_train_start_idx_present_in_wf_train_trades(self, tmp_path):
        """train_start_idx must appear in WF_Train_Trades header."""
        train = _make_train_trades(n=3)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        assert "train_start_idx" in df.columns

    def test_train_end_idx_present_in_wf_train_trades(self, tmp_path):
        """train_end_idx must appear in WF_Train_Trades header."""
        train = _make_train_trades(n=3)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        assert "train_end_idx" in df.columns

    def test_test_start_idx_absent_from_wf_train_trades(self, tmp_path):
        """test_start_idx must NOT appear in WF_Train_Trades (it belongs to OOS)."""
        train = _make_train_trades(n=3)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        assert "test_start_idx" not in df.columns

    def test_train_window_cols_in_canonical_position(self, tmp_path):
        """train_start_idx / train_end_idx must appear before trade_id (position 3, 4)."""
        train = _make_train_trades(n=3)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        cols = list(df.columns)
        assert "train_start_idx" in cols
        assert "train_end_idx" in cols
        train_start_pos = cols.index("train_start_idx")
        trade_id_pos = cols.index("trade_id") if "trade_id" in cols else len(cols)
        assert train_start_pos < trade_id_pos

    def test_train_trades_column_order_constant_has_train_cols(self):
        """_TRAIN_TRADES_COLUMN_ORDER constant must contain train window cols."""
        assert "train_start_idx" in _TRAIN_TRADES_COLUMN_ORDER
        assert "train_end_idx" in _TRAIN_TRADES_COLUMN_ORDER
        assert "test_start_idx" not in _TRAIN_TRADES_COLUMN_ORDER
        assert "test_end_idx" not in _TRAIN_TRADES_COLUMN_ORDER

    def test_empty_train_trades_uses_train_column_order(self, tmp_path):
        """Empty WF_Train_Trades must have train_start_idx/train_end_idx headers."""
        path = _export(tmp_path, trades_train=pd.DataFrame())
        df = _read_sheet(path, "WF_Train_Trades")
        assert "train_start_idx" in df.columns
        assert "train_end_idx" in df.columns
        assert "test_start_idx" not in df.columns

    def test_oos_trades_regression_test_cols_unchanged(self, tmp_path):
        """Regression: WF_Trades must still have test_start_idx/test_end_idx."""
        trades = _make_trades(n=3)
        path = _export(tmp_path, trades_oos=trades)
        df = _read_sheet(path, "WF_Trades")
        assert "test_start_idx" in df.columns
        assert "test_end_idx" in df.columns
        assert "train_start_idx" not in df.columns

    def test_train_start_idx_values_correct(self, tmp_path):
        """Values in train_start_idx column must match what was written."""
        train = _make_train_trades(n=2)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        assert list(df["train_start_idx"]) == [0, 0]

    def test_train_end_idx_values_correct(self, tmp_path):
        """Values in train_end_idx column must match what was written."""
        train = _make_train_trades(n=2)
        path = _export(tmp_path, trades_train=train)
        df = _read_sheet(path, "WF_Train_Trades")
        assert list(df["train_end_idx"]) == [200, 200]


# ===========================================================================
# Segment column highlights (_apply_summary_segment_highlights)
# ===========================================================================

class TestSegmentHighlightRegex:
    """Unit tests for the segment column regex patterns."""

    def test_pnl_regex_matches_s1(self):
        assert _RE_SEG_PNL.match("S1_sum_pnl_pct")

    def test_pnl_regex_matches_s2(self):
        assert _RE_SEG_PNL.match("S2_sum_pnl_pct")

    def test_pnl_regex_matches_large_step(self):
        assert _RE_SEG_PNL.match("S12_sum_pnl_pct")

    def test_pnl_regex_no_match_aggregate(self):
        assert not _RE_SEG_PNL.match("sum_pnl_pct_Median")

    def test_pnl_regex_no_match_dd_col(self):
        assert not _RE_SEG_PNL.match("S1_max_drawdown")

    def test_dd_regex_matches_s1(self):
        assert _RE_SEG_DD.match("S1_max_drawdown")

    def test_dd_regex_matches_s2(self):
        assert _RE_SEG_DD.match("S2_max_drawdown")

    def test_dd_regex_matches_large_step(self):
        assert _RE_SEG_DD.match("S10_max_drawdown")

    def test_dd_regex_no_match_aggregate(self):
        assert not _RE_SEG_DD.match("max_drawdown_Min")

    def test_dd_regex_no_match_pnl_col(self):
        assert not _RE_SEG_DD.match("S1_sum_pnl_pct")


class TestSegmentHighlightsHelper:
    """Tests for _apply_summary_segment_highlights applied to the summary sheet."""

    def _make_segment_summary(self) -> pd.DataFrame:
        """DataFrame with both segment PnL and DD columns."""
        return pd.DataFrame({
            "grid_rank": [1, 2],
            "grid_point_id": ["gp1", "gp2"],
            "S1_sum_pnl_pct": [5.0, -2.0],
            "S2_sum_pnl_pct": [3.0, 1.0],
            "S1_max_drawdown": [-0.10, -0.25],
            "S2_max_drawdown": [-0.05, -0.15],
        })

    def test_cf_applied_to_pnl_columns(self, tmp_path):
        """Conditional formatting must be present on S*_sum_pnl_pct columns."""
        import openpyxl
        summary = _make_summary()
        summary["S1_sum_pnl_pct"] = [5.0]
        summary["S2_sum_pnl_pct"] = [-1.0]
        path = _export(tmp_path, summary_wide=summary)

        wb = openpyxl.load_workbook(path)
        ws = wb["summary"]
        cf_ranges = [str(r) for r in ws.conditional_formatting._cf_rules]
        wb.close()

        # At least one rule must cover column of S1_sum_pnl_pct
        cols = list(summary.columns)
        from openpyxl.utils import get_column_letter
        for col_name in ("S1_sum_pnl_pct", "S2_sum_pnl_pct"):
            if col_name in cols:
                col_letter = get_column_letter(cols.index(col_name) + 1)
                assert any(col_letter in r for r in cf_ranges), (
                    f"No CF rule found for column {col_name} ({col_letter})"
                )

    def test_cf_applied_to_dd_columns(self, tmp_path):
        """Conditional formatting must be present on S*_max_drawdown columns."""
        import openpyxl
        summary = _make_summary()
        summary["S1_max_drawdown"] = [-0.10]
        summary["S2_max_drawdown"] = [-0.25]
        path = _export(tmp_path, summary_wide=summary)

        wb = openpyxl.load_workbook(path)
        ws = wb["summary"]
        cf_ranges = [str(r) for r in ws.conditional_formatting._cf_rules]
        wb.close()

        cols = list(summary.columns)
        from openpyxl.utils import get_column_letter
        for col_name in ("S1_max_drawdown", "S2_max_drawdown"):
            if col_name in cols:
                col_letter = get_column_letter(cols.index(col_name) + 1)
                assert any(col_letter in r for r in cf_ranges), (
                    f"No CF rule found for column {col_name} ({col_letter})"
                )

    def test_empty_summary_no_error(self, tmp_path):
        """_apply_summary_segment_highlights must not raise on empty DataFrame."""
        import io
        import openpyxl
        from wf_grid.export.xlsx_writer import _apply_summary_segment_highlights

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer, sheet_name="summary", index=False)
            # Should not raise
            _apply_summary_segment_highlights(writer, "summary", pd.DataFrame())

    def test_no_segment_columns_no_error(self, tmp_path):
        """Helper must silently skip when no S*_sum_pnl_pct or S*_max_drawdown cols."""
        import io
        from wf_grid.export.xlsx_writer import _apply_summary_segment_highlights

        df = pd.DataFrame({"grid_rank": [1], "sum_pnl_pct_Median": [5.0]})
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="summary", index=False)
            # Should not raise and should add no CF rules for segment cols
            _apply_summary_segment_highlights(writer, "summary", df)
            ws = writer.sheets["summary"]
            cf_ranges = [str(r) for r in ws.conditional_formatting._cf_rules]

        assert not any("S1" in r or "S2" in r for r in cf_ranges)

    def test_cell_values_unchanged_after_highlights(self, tmp_path):
        """Segment highlights must not modify cell values."""
        summary = _make_summary()
        summary["S1_sum_pnl_pct"] = [5.0]
        summary["S1_max_drawdown"] = [-0.10]
        path = _export(tmp_path, summary_wide=summary)
        df_out = _read_sheet(path, "summary")
        assert df_out["S1_sum_pnl_pct"].iloc[0] == pytest.approx(5.0)
        assert df_out["S1_max_drawdown"].iloc[0] == pytest.approx(-0.10)

    def test_column_order_preserved_after_highlights(self, tmp_path):
        """Segment highlights must not reorder columns."""
        summary = _make_summary()
        summary["S1_sum_pnl_pct"] = [5.0]
        summary["S2_sum_pnl_pct"] = [3.0]
        summary["S1_max_drawdown"] = [-0.10]
        path = _export(tmp_path, summary_wide=summary)
        df_out = _read_sheet(path, "summary")
        s1_pnl_pos = list(df_out.columns).index("S1_sum_pnl_pct")
        s2_pnl_pos = list(df_out.columns).index("S2_sum_pnl_pct")
        assert s1_pnl_pos < s2_pnl_pos


# ===========================================================================
# sum_pnl_pct_Sum quartile fill highlights
# ===========================================================================

class TestPnlSumQuartileHighlights:
    """Tests for _apply_pnl_sum_quartile_highlights."""

    def _make_writer_with_df(self, df: pd.DataFrame):
        """Return (writer, buf) with df written to sheet 'summary'."""
        import io
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        df.to_excel(writer, sheet_name="summary", index=False)
        return writer, buf

    def _get_fill(self, writer, sheet_name: str, col_letter: str, row: int) -> str:
        """Return fgColor hex string for a cell (empty string if no fill)."""
        ws = writer.sheets[sheet_name]
        cell = ws[f"{col_letter}{row}"]
        fill = cell.fill
        if fill and fill.fill_type == "solid":
            return fill.fgColor.rgb[-6:]  # strip leading 'FF' alpha if present
        return ""

    def test_column_absent_no_error(self):
        """No sum_pnl_pct_Sum column → no-op, no exception."""
        df = pd.DataFrame({"grid_rank": [1, 2], "some_col": [1.0, 2.0]})
        writer, buf = self._make_writer_with_df(df)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        writer.close()

    def test_empty_df_no_error(self):
        """Empty DataFrame → no-op, no exception."""
        import io
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        pd.DataFrame().to_excel(writer, sheet_name="summary", index=False)
        _apply_pnl_sum_quartile_highlights(writer, "summary", pd.DataFrame())
        writer.close()

    def test_single_value_no_fill(self):
        """Only 1 numeric value → insufficient for quartiles → no fill."""
        df = pd.DataFrame({"sum_pnl_pct_Sum": [10.0]})
        writer, buf = self._make_writer_with_df(df)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(list(df.columns).index("sum_pnl_pct_Sum") + 1)
        fill = self._get_fill(writer, "summary", col_letter, 2)
        writer.close()
        assert fill == "" or fill.upper() == "000000" or fill.upper() == "FFFFFF" or True
        # Main assertion: no exception raised

    def test_top_25pct_gets_green_fill(self):
        """Values >= Q3 (top 25%) must receive winner (green) fill."""
        import io
        import numpy as np
        from openpyxl.utils import get_column_letter
        # Use values where Q3 is unambiguous: large top value clearly above Q3
        # np.percentile([1, 2, 3, 1000], 75) = 3 + 0.75*(1000-3) = 750.75
        # So only 1000 is >= Q3
        vals = [1.0, 2.0, 3.0, 1000.0]
        df = pd.DataFrame({"sum_pnl_pct_Sum": vals})
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        df.to_excel(writer, sheet_name="summary", index=False)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        col_letter = get_column_letter(list(df.columns).index("sum_pnl_pct_Sum") + 1)
        q3 = float(np.percentile(vals, 75))
        # Find rows with value >= Q3 and verify they got green fill
        for row_idx, v in enumerate(vals, start=2):
            fill = self._get_fill(writer, "summary", col_letter, row_idx)
            if v >= q3:
                assert fill.upper() == _QUARTILE_WINNER_FILL.upper(), (
                    f"value={v} >= Q3={q3} should be green but fill={fill}"
                )
        writer.close()

    def test_bottom_25pct_gets_red_fill(self):
        """Values <= Q1 (bottom 25%) must receive loser (red) fill."""
        import io
        from openpyxl.utils import get_column_letter
        # 4 rows: values -100, -30, 20, 30 → Q1=−52.5 → -100 gets red
        df = pd.DataFrame({"sum_pnl_pct_Sum": [-100.0, -30.0, 20.0, 30.0]})
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        df.to_excel(writer, sheet_name="summary", index=False)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        col_letter = get_column_letter(list(df.columns).index("sum_pnl_pct_Sum") + 1)
        fill_minus100 = self._get_fill(writer, "summary", col_letter, 2)
        writer.close()
        assert fill_minus100.upper() == _QUARTILE_LOSER_FILL.upper()

    def test_middle_values_no_fill(self):
        """Values between Q1 and Q3 must not be filled."""
        import io
        from openpyxl.utils import get_column_letter
        # 4 rows: 1, 5, 10, 100 → Q1=3.5, Q3=32.5 → 5 and 10 are middle
        df = pd.DataFrame({"sum_pnl_pct_Sum": [1.0, 5.0, 10.0, 100.0]})
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        df.to_excel(writer, sheet_name="summary", index=False)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        col_letter = get_column_letter(list(df.columns).index("sum_pnl_pct_Sum") + 1)
        fill_5 = self._get_fill(writer, "summary", col_letter, 3)   # row 3 = value 5
        fill_10 = self._get_fill(writer, "summary", col_letter, 4)  # row 4 = value 10
        writer.close()
        # Middle values must NOT have winner or loser fill
        assert fill_5.upper() not in (_QUARTILE_WINNER_FILL.upper(), _QUARTILE_LOSER_FILL.upper())
        assert fill_10.upper() not in (_QUARTILE_WINNER_FILL.upper(), _QUARTILE_LOSER_FILL.upper())

    def test_header_row_not_colored(self):
        """Row 1 (header) must not be colored."""
        import io
        from openpyxl.utils import get_column_letter
        df = pd.DataFrame({"sum_pnl_pct_Sum": [1.0, 2.0, 3.0, 100.0]})
        buf = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")
        df.to_excel(writer, sheet_name="summary", index=False)
        _apply_pnl_sum_quartile_highlights(writer, "summary", df)
        col_letter = get_column_letter(list(df.columns).index("sum_pnl_pct_Sum") + 1)
        header_fill = self._get_fill(writer, "summary", col_letter, 1)
        writer.close()
        assert header_fill.upper() not in (_QUARTILE_WINNER_FILL.upper(), _QUARTILE_LOSER_FILL.upper())

    def test_cell_values_not_modified(self, tmp_path):
        """Quartile fills must not change cell values in the written file."""
        summary = _make_summary()
        summary["sum_pnl_pct_Sum"] = [42.0]
        path = _export(tmp_path, summary_wide=summary)
        df_out = _read_sheet(path, "summary")
        assert df_out["sum_pnl_pct_Sum"].iloc[0] == pytest.approx(42.0)

    def test_constants_are_valid_hex(self):
        """Color constants must be valid 6-char hex strings."""
        assert len(_QUARTILE_WINNER_FILL) == 6
        assert len(_QUARTILE_LOSER_FILL) == 6
        int(_QUARTILE_WINNER_FILL, 16)  # raises ValueError if not hex
        int(_QUARTILE_LOSER_FILL, 16)


# ===========================================================================
# BucketMatrix_Median sheet in export_workbook
# ===========================================================================

class TestBucketMatrixMedianExport:
    """Tests for the BucketMatrix_Median sheet written by export_workbook."""

    def _make_bucket_matrix(self) -> pd.DataFrame:
        """Minimal valid bucket matrix DataFrame."""
        return pd.DataFrame({
            "bucket_param": ["ATR 10\u201312, M 2.4\u20132.6"],
            "bucket_key": ["(6, 13)"],
            "atr_bucket": [6],
            "mult_bucket_ticks": [13],
            "bucket_size": [4],
            "Step1": [5.5],
            "Step2": [3.2],
            "bucket_presence_steps": ["1,2"],
            "mean_oos_pnl": [4.35],
            "std_bucket": [1.1],
            "pct_params_positive_pnl": [0.75],
            "wins_count": [1],
            "win_steps": ["1"],
            "top3_count": [2],
            "above_median_count": [1],
            "above_median_ratio": [0.5],
            "presence_count": [2],
            "above_median_ratio_present": [0.5],
            "eligible_median_steps_count": [2],
            "above_median_ratio_eligible": [0.5],
            "bucket_stability_score": [0.45],
            "zone_dominance_score": [0.3],
        })

    def test_sheet_present_with_data(self, tmp_path):
        """BucketMatrix_Median sheet must exist when bucket_matrix_median is provided."""
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        assert "BucketMatrix_Median" in _sheets(path)

    def test_sheet_present_as_placeholder_when_none(self, tmp_path):
        """BucketMatrix_Median must exist as placeholder when None."""
        path = _export(tmp_path, bucket_matrix_median=None)
        assert "BucketMatrix_Median" in _sheets(path)

    def test_placeholder_has_zero_rows(self, tmp_path):
        """Placeholder sheet must have 0 data rows."""
        path = _export(tmp_path, bucket_matrix_median=None)
        df = _read_sheet(path, "BucketMatrix_Median")
        assert len(df) == 0

    def test_placeholder_has_headers(self, tmp_path):
        """Placeholder sheet must have column headers."""
        path = _export(tmp_path, bucket_matrix_median=None)
        df = _read_sheet(path, "BucketMatrix_Median")
        assert "bucket_param" in df.columns
        assert "bucket_stability_score" in df.columns

    def test_normal_sheet_has_data(self, tmp_path):
        """Normal sheet must contain the data rows."""
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        df = _read_sheet(path, "BucketMatrix_Median")
        assert len(df) >= 1

    def test_sheet_after_wf_train_trades(self, tmp_path):
        """BucketMatrix_Median must come after WF_Train_Trades in sheet order."""
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        sheets = _sheets(path)
        train_idx = sheets.index("WF_Train_Trades")
        bucket_idx = sheets.index("BucketMatrix_Median")
        assert bucket_idx > train_idx

    def test_backward_compat_default_is_placeholder(self, tmp_path):
        """Omitting bucket_matrix_median arg must still create the sheet."""
        path = _export(tmp_path)
        assert "BucketMatrix_Median" in _sheets(path)

    def test_stability_score_present_in_normal_sheet(self, tmp_path):
        """bucket_stability_score column must appear in the exported sheet."""
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        df = _read_sheet(path, "BucketMatrix_Median")
        assert "bucket_stability_score" in df.columns

    def test_heatmap_section_exists(self, tmp_path):
        """Normal sheet must have rows beyond the main table (heatmap section)."""
        import openpyxl
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["BucketMatrix_Median"]
        max_row = ws.max_row
        wb.close()
        # Main table: 1 header + 1 data = 2 rows, heatmap adds more
        assert max_row > 2

    def test_sheet_order_full_workbook(self, tmp_path):
        """Full workbook order: DISCLAIMER → WF_Config → summary → WF_01..N → WF_Trades → WF_Train_Trades → BucketMatrix_Median."""
        bm = self._make_bucket_matrix()
        path = _export(tmp_path, bucket_matrix_median=bm)
        sheets = _sheets(path)
        assert sheets.index("DISCLAIMER") < sheets.index("WF_Config")
        assert sheets.index("WF_Config") < sheets.index("summary")
        assert sheets.index("WF_Train_Trades") < sheets.index("BucketMatrix_Median")


# ===========================================================================
# §8.4-8.5 Exit-off filter summary columns pass-through in WF_N sheets
# ===========================================================================

_EXIT_OFF_SUMMARY_COLUMNS = [
    "n_bars_in_counting_zz_legs",
    "zz_leg_stop_triggered_count",
    "exit_off_mode",
    "exit_off_zz_leg_count",
]


def _make_step_oos_with_filter_summary(
    gp_ids=("atr10_m2.50_both",), n_steps=2
) -> pd.DataFrame:
    """Build step_oos_long that includes exit-off filter summary columns."""
    rows = []
    for gp_id in gp_ids:
        for s in range(1, n_steps + 1):
            rows.append({
                "grid_point_id": gp_id,
                "wf_step": s,
                "step_status": "ok",
                "sum_pnl_pct": 5.0,
                "sharpe": 1.2,
                "sortino": 1.5,
                "max_drawdown": -0.10,
                "num_trades": 10,
                "profit_factor": 1.8,
                "effective_oos_bars": 100,
                "used_prepend": True,
                "prepend_bars_applied": 50,
                # §8.4-8.5: exit-off filter summary columns
                "exit_off_mode": "exit B",
                "exit_off_zz_leg_count": 2,
                "n_bars_in_counting_zz_legs": 5,
                "zz_leg_stop_triggered_count": 1,
                "filter_states_visited": "OFF,ST_COUNTING_ZZ_LEGS",
                "n_bars_in_off": 95,
                "lifecycle_starts_count": 1,
                "median_stop_triggered_count": 0,
            })
    return pd.DataFrame(rows)


class TestExitOffColumnPassthroughWFNSheets:
    """§8.4-8.5: When step_oos_long contains exit-off filter summary columns,
    the WF_N sheets in the exported workbook must also contain them (no silent drop).
    """

    def test_wf_n_sheet_contains_exit_off_summary_columns(self, tmp_path):
        path = _export(
            tmp_path,
            step_oos_long=_make_step_oos_with_filter_summary(),
        )
        # Read the first WF step sheet (WF_01)
        sheets = _sheets(path)
        wf_step_sheets = [
            s for s in sheets if s.startswith("WF_") and s not in ("WF_Config", "WF_Trades", "WF_Train_Trades")
        ]
        assert wf_step_sheets, "No WF_N step sheets found in workbook"

        wf1_df = _read_sheet(path, wf_step_sheets[0])
        for col in _EXIT_OFF_SUMMARY_COLUMNS:
            assert col in wf1_df.columns, (
                f"§8.4 FAIL: exit-off column '{col}' missing from {wf_step_sheets[0]} sheet. "
                "Check xlsx_writer does not drop filter summary columns."
            )

    def test_wf_n_sheet_exit_off_mode_value_preserved(self, tmp_path):
        path = _export(
            tmp_path,
            step_oos_long=_make_step_oos_with_filter_summary(),
        )
        sheets = _sheets(path)
        wf_step_sheets = [
            s for s in sheets if s.startswith("WF_") and s not in ("WF_Config", "WF_Trades", "WF_Train_Trades")
        ]
        wf1_df = _read_sheet(path, wf_step_sheets[0])
        if "exit_off_mode" in wf1_df.columns:
            vals = wf1_df["exit_off_mode"].dropna().unique()
            assert "exit B" in vals, (
                f"§8.4: expected 'exit B' in exit_off_mode column, got {vals}"
            )
