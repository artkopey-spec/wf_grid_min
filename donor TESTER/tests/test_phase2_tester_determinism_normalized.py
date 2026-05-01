"""
Tests #23 + #28 — Determinism: in-memory and normalized XLSX comparison.

Plan reference: docs/zigzag_st_tester_phase2_implementation_plan.txt §10 #23, #28
Spec reference: Appendix A v1.1 §16

Contract (#23):
- Two consecutive legacy batch-runs on the same data + config give bit-identical
  in-memory results: positions, equity_curve, num_trades, sum_pnl_pct, filter_diagnostics.

Contract (#28 — normalized XLSX, NOT byte-identical):
- Two consecutive exports produce equal XLSX in normalized comparison:
  - Sheet names list must match.
  - Column orders per sheet must match.
  - Cell values per sheet must match (float with tolerance, strings exact).
- Comparison is NOT byte-identical (zip metadata / openpyxl timestamps may differ).
- Covers both legacy disabled and legacy enabled paths.
"""

from __future__ import annotations

import tempfile
import numpy as np
import pandas as pd
import pytest

from pathlib import Path
from typing import Dict, List, Any


def _make_synthetic_ohlc(n: int = 500, seed: int = 71) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    close = 100.0 * np.exp(np.cumsum(rng.normal(0.0003, 0.012, n)))
    noise = rng.uniform(0.001, 0.004, size=n)
    high = close * (1 + noise)
    low = close * (1 - noise)
    open_ = np.clip(close * (1 + rng.uniform(-0.002, 0.002, size=n)), low, high)
    idx = pd.date_range("2020-01-01", periods=n, freq="D")
    return pd.DataFrame({"open": open_, "high": high, "low": low, "close": close}, index=idx)


def _make_enabled_cfg():
    from supertrend_optimizer.core.trade_filter_config import (
        TradeFilterConfig, TradeFilterZigZagConfig, TradeFilterTriggersConfig,
        TradeFilterLifecycleConfig, TradeFilterDiagnosticsConfig,
        TradeFilterTriggerToggleConfig,
    )
    return TradeFilterConfig(
        enabled=True, type="zigzag_st_mode",
        zigzag=TradeFilterZigZagConfig(
            reversal_threshold=0.03, local_window=20, candidate_trigger_threshold=0.4,
        ),
        triggers=TradeFilterTriggersConfig(
            candidate_threshold=TradeFilterTriggerToggleConfig(enabled=True),
            confirmed_median=TradeFilterTriggerToggleConfig(enabled=True),
        ),
        lifecycle=TradeFilterLifecycleConfig(
            freeze_confirmed_legs=3, stop_check="confirm_bar_only",
            stopping_exit="opposite_st_flip",
        ),
        diagnostics=TradeFilterDiagnosticsConfig(
            export_state_columns=True, export_trigger_columns=True,
        ),
    )


def _run_legacy(df: pd.DataFrame, tf_cfg=None):
    from supertrend_optimizer.testing.runner import run_all_periods
    from supertrend_optimizer.utils.enums import ExecutionModel
    return run_all_periods(
        df=df, atr_period=14, multiplier=3.0, trade_mode="revers",
        commission=0.001, warmup_period=20,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        trade_filter_config=tf_cfg,
    )


def _run_signals(df: pd.DataFrame, results, tf_cfg=None):
    from supertrend_optimizer.testing.signal_events import build_signal_events
    from supertrend_optimizer.utils.enums import ExecutionModel
    return build_signal_events(
        df=df, trend=results[0].result.trend, atr_period=14,
        trade_mode="revers", execution_model=ExecutionModel.OPEN_TO_OPEN,
        filter_diagnostics=results[0].filter_diagnostics if tf_cfg and tf_cfg.enabled else None,
    )


def _export(results, signals_df, tf_cfg=None, df=None) -> str:
    from supertrend_optimizer.io.excel_tester import export_tester_results
    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = str(Path(tmpdir) / "result.xlsx")
        actual = export_tester_results(
            results, out_path,
            signals_df=signals_df,
            trade_filter_config=tf_cfg,
            df=df,
        )
        import shutil, os
        dest = str(Path(tempfile.gettempdir()) / f"result_det_{id(results)}.xlsx")
        shutil.copy(actual, dest)
        return dest


def _load_wb_normalized(path: str) -> Dict[str, List[List[Any]]]:
    """Load workbook as normalized dict: sheet_name → list of rows (list of cell values)."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        result[sheet_name] = rows
    return result


def _compare_workbooks(wb1: Dict, wb2: Dict, rtol: float = 1e-6) -> List[str]:
    """Return list of diff messages; empty = identical (within tolerance)."""
    diffs = []

    if list(wb1.keys()) != list(wb2.keys()):
        diffs.append(
            f"Sheet names differ:\n  Run1: {list(wb1.keys())}\n  Run2: {list(wb2.keys())}"
        )
        return diffs  # no point comparing further

    for sheet_name in wb1:
        rows1 = wb1[sheet_name]
        rows2 = wb2[sheet_name]

        if len(rows1) != len(rows2):
            diffs.append(
                f"Sheet {sheet_name!r}: row count {len(rows1)} != {len(rows2)}"
            )
            continue

        for ri, (r1, r2) in enumerate(zip(rows1, rows2)):
            if len(r1) != len(r2):
                diffs.append(
                    f"Sheet {sheet_name!r} row {ri}: col count {len(r1)} != {len(r2)}"
                )
                continue
            for ci, (v1, v2) in enumerate(zip(r1, r2)):
                if v1 == v2:
                    continue
                # Both numeric: allow floating-point tolerance
                if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
                    if abs(v1 - v2) <= rtol * max(abs(v1), abs(v2), 1.0):
                        continue
                diffs.append(
                    f"Sheet {sheet_name!r} row {ri} col {ci}: "
                    f"{v1!r} != {v2!r}"
                )
                if len(diffs) > 20:
                    diffs.append("... (truncated)")
                    return diffs
    return diffs


# ---------------------------------------------------------------------------
# Test #23 — In-memory determinism
# ---------------------------------------------------------------------------

class TestInMemoryDeterminism:
    """Two consecutive runs must give bit-identical in-memory results (#23)."""

    def test_disabled_positions_deterministic(self) -> None:
        df = _make_synthetic_ohlc()
        r1 = _run_legacy(df, tf_cfg=None)
        r2 = _run_legacy(df, tf_cfg=None)

        for pr1, pr2 in zip(r1, r2):
            np.testing.assert_array_equal(pr1.result.positions, pr2.result.positions)

    def test_enabled_positions_deterministic(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        r1 = _run_legacy(df, tf_cfg=tf_cfg)
        r2 = _run_legacy(df, tf_cfg=tf_cfg)

        for pr1, pr2 in zip(r1, r2):
            np.testing.assert_array_equal(pr1.result.positions, pr2.result.positions)

    def test_enabled_filter_diagnostics_deterministic(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        r1 = _run_legacy(df, tf_cfg=tf_cfg)
        r2 = _run_legacy(df, tf_cfg=tf_cfg)

        for pr1, pr2 in zip(r1, r2):
            fd1 = pr1.filter_diagnostics
            fd2 = pr2.filter_diagnostics
            if fd1 is None and fd2 is None:
                continue
            assert fd1 is not None and fd2 is not None
            for key in fd1:
                arr1 = fd1[key]
                arr2 = fd2[key]
                np.testing.assert_array_equal(arr1, arr2,
                    err_msg=f"filter_diagnostics[{key!r}] not deterministic")

    def test_metrics_deterministic(self) -> None:
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        r1 = _run_legacy(df, tf_cfg=tf_cfg)
        r2 = _run_legacy(df, tf_cfg=tf_cfg)

        for pr1, pr2 in zip(r1, r2):
            for key in ("num_trades", "sum_pnl_pct"):
                v1 = pr1.metrics.get(key)
                v2 = pr2.metrics.get(key)
                if v1 is not None and v2 is not None:
                    assert v1 == pytest.approx(v2, rel=1e-12), (
                        f"Metric {key!r} not deterministic: {v1} != {v2}"
                    )


# ---------------------------------------------------------------------------
# Test #28 — Normalized XLSX determinism
# ---------------------------------------------------------------------------

class TestNormalizedXlsxDeterminism:
    """Two exports of the same data must produce equivalent XLSX (normalized comparison, #28)."""

    def test_disabled_legacy_xlsx_deterministic(self) -> None:
        """Disabled path: two exports must have identical sheet structure and values."""
        df = _make_synthetic_ohlc()
        r1 = _run_legacy(df, tf_cfg=None)
        r2 = _run_legacy(df, tf_cfg=None)
        s1 = _run_signals(df, r1)
        s2 = _run_signals(df, r2)

        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            from supertrend_optimizer.io.excel_tester import export_tester_results
            p1 = export_tester_results(r1, str(Path(d) / "r1.xlsx"), signals_df=s1)
            p2 = export_tester_results(r2, str(Path(d) / "r2.xlsx"), signals_df=s2)
            wb1 = _load_wb_normalized(p1)
            wb2 = _load_wb_normalized(p2)

        diffs = _compare_workbooks(wb1, wb2)
        assert not diffs, (
            "Disabled XLSX not deterministic:\n" + "\n".join(diffs)
        )

    def test_enabled_legacy_xlsx_deterministic(self) -> None:
        """Enabled path: two exports must have identical sheet structure and values."""
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        r1 = _run_legacy(df, tf_cfg=tf_cfg)
        r2 = _run_legacy(df, tf_cfg=tf_cfg)
        s1 = _run_signals(df, r1, tf_cfg=tf_cfg)
        s2 = _run_signals(df, r2, tf_cfg=tf_cfg)

        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            from supertrend_optimizer.io.excel_tester import export_tester_results
            p1 = export_tester_results(r1, str(Path(d) / "r1.xlsx"), signals_df=s1,
                                       trade_filter_config=tf_cfg, df=df)
            p2 = export_tester_results(r2, str(Path(d) / "r2.xlsx"), signals_df=s2,
                                       trade_filter_config=tf_cfg, df=df)
            wb1 = _load_wb_normalized(p1)
            wb2 = _load_wb_normalized(p2)

        diffs = _compare_workbooks(wb1, wb2)
        assert not diffs, (
            "Enabled XLSX not deterministic:\n" + "\n".join(diffs)
        )

    def test_sheet_names_identical_between_runs(self) -> None:
        """Sheet names list must be identical between two exports of same data."""
        df = _make_synthetic_ohlc()
        tf_cfg = _make_enabled_cfg()
        r1 = _run_legacy(df, tf_cfg=tf_cfg)
        r2 = _run_legacy(df, tf_cfg=tf_cfg)
        s1 = _run_signals(df, r1, tf_cfg=tf_cfg)
        s2 = _run_signals(df, r2, tf_cfg=tf_cfg)

        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            from supertrend_optimizer.io.excel_tester import export_tester_results
            p1 = export_tester_results(r1, str(Path(d) / "r1.xlsx"), signals_df=s1,
                                       trade_filter_config=tf_cfg, df=df)
            p2 = export_tester_results(r2, str(Path(d) / "r2.xlsx"), signals_df=s2,
                                       trade_filter_config=tf_cfg, df=df)
            sh1 = openpyxl.load_workbook(p1).sheetnames
            sh2 = openpyxl.load_workbook(p2).sheetnames

        assert sh1 == sh2, f"Sheet names differ: {sh1} vs {sh2}"

    def test_column_orders_identical_between_runs(self) -> None:
        """Column orders in each sheet must be identical between runs."""
        df = _make_synthetic_ohlc()
        r1 = _run_legacy(df, tf_cfg=None)
        r2 = _run_legacy(df, tf_cfg=None)
        s1 = _run_signals(df, r1)
        s2 = _run_signals(df, r2)

        import openpyxl

        def _get_headers(path):
            wb = openpyxl.load_workbook(path)
            return {
                sh: [c.value for c in next(wb[sh].iter_rows(min_row=1, max_row=1))]
                for sh in wb.sheetnames
            }

        with tempfile.TemporaryDirectory() as d:
            from supertrend_optimizer.io.excel_tester import export_tester_results
            p1 = export_tester_results(r1, str(Path(d) / "r1.xlsx"), signals_df=s1)
            p2 = export_tester_results(r2, str(Path(d) / "r2.xlsx"), signals_df=s2)
            h1 = _get_headers(p1)
            h2 = _get_headers(p2)

        for sh in h1:
            assert h1[sh] == h2.get(sh), (
                f"Column order differs in sheet {sh!r}: {h1[sh]} vs {h2.get(sh)}"
            )
