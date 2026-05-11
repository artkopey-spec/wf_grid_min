from __future__ import annotations

from dataclasses import dataclass

import numpy as np
import openpyxl
import pandas as pd
import pytest

from supertrend_optimizer.core.trade_filter_config import (
    build_trade_filter_config_from_raw,
    collect_raw_user_keys,
    resolve_exit_b_immediate_off_in_place,
    resolve_exit_off_mode_in_place,
    resolve_time_filter_in_place,
    resolve_trade_filter_mode_in_place,
    resolve_volume_defaults_in_place,
    resolve_volume_enabled_in_place,
    resolve_zigzag_enabled_in_place,
    validate_trade_filter,
)
from supertrend_optimizer.core.volume_metrics import build_volume_global_metrics
from supertrend_optimizer.engine.result import BacktestResult
from supertrend_optimizer.utils.enums import ExecutionModel
from supertrend_optimizer.utils.time_utils import WFWindowSlice

from wf_grid.baseline.fingerprint import compute_pipeline_fingerprint
from wf_grid.config.schema import (
    BacktestConfig,
    DataConfig,
    ExecutionConfig,
    ExportConfig,
    GridConfig,
    OptimizationConfig,
    ValidationConfig,
    WalkForwardConfig,
)
from wf_grid.config.loader import ConfigError, load_grid_config
from wf_grid.export.xlsx_writer import export_workbook
from wf_grid.grid.enumeration import GridPoint
from wf_grid.pipeline.orchestrator import run_grid_pipeline
from wf_grid.wf import step_executor
from wf_grid.wf._mp_helpers import _strip_filter_diagnostics_arrays
from wf_grid.wf.step_executor import StepResult, execute_oos_step


def _df(n: int = 80) -> pd.DataFrame:
    close = 100.0 + np.linspace(0.0, 12.0, n)
    volume = np.full(n, 100.0)
    return pd.DataFrame(
        {
            "open": close,
            "high": close + 1.0,
            "low": close - 1.0,
            "close": close,
            "volume": volume,
        },
        index=pd.date_range("2025-01-01", periods=n, freq="min"),
    )


def _volume_raw(enabled: bool = True) -> dict:
    if not enabled:
        return {"enabled": False}
    return {
        "enabled": True,
        "mode": "volume_A",
        "short_window": 2,
        "baseline_window": 3,
        "threshold_ratio": 1.0,
        "regime_low_ratio": 0.8,
        "regime_high_ratio": 1.2,
        "direction_lookback_bars": 1,
    }


def _tf_cfg():
    raw = {
        "enabled": True,
        "volume": _volume_raw(True),
    }
    raw_user_keys = collect_raw_user_keys({"trade_filter": raw})
    tf_cfg = build_trade_filter_config_from_raw(raw)
    resolve_zigzag_enabled_in_place(tf_cfg, raw_user_keys)
    resolve_volume_enabled_in_place(tf_cfg, raw_user_keys)
    errors: list[str] = []
    validate_trade_filter(tf_cfg, errors, raw_user_keys, caller_pipeline="wf_grid")
    assert errors == []
    resolve_trade_filter_mode_in_place(tf_cfg, raw_user_keys)
    resolve_exit_off_mode_in_place(tf_cfg, raw_user_keys)
    resolve_exit_b_immediate_off_in_place(tf_cfg, raw_user_keys)
    resolve_time_filter_in_place(tf_cfg, raw_user_keys)
    resolve_volume_defaults_in_place(tf_cfg, raw_user_keys)
    return tf_cfg


def _cfg(*, retain: bool = False) -> GridConfig:
    return GridConfig(
        data=DataConfig(file_path=""),
        optimization=OptimizationConfig(
            atr_period_range=[5, 5],
            multiplier_range=[2.0, 2.0],
            multiplier_step=1.0,
            trade_mode="revers",
        ),
        backtest=BacktestConfig(commission=0.0, min_trades_required=0),
        validation=ValidationConfig(
            warmup_period=0,
            walk_forward=WalkForwardConfig(
                train_size="30m",
                test_size="10m",
                step_size="10m",
                min_train_bars=5,
                min_test_bars=5,
            ),
        ),
        execution=ExecutionConfig(parallel_enabled=False),
        export=ExportConfig(retain_per_bar_filter_diagnostics=retain),
        trade_filter=_tf_cfg(),
        resolved_periods_per_year=365.0,
    )


def _grid_point() -> GridPoint:
    return GridPoint(
        atr_period=5,
        multiplier=2.0,
        trade_mode="revers",
        grid_point_id="atr5_m2.00_revers",
    )


def _step_result(*, retained: bool = True) -> StepResult:
    diag = {
        "trade_filter_state": np.array(["OFF", "ACTIVE_LONG"], dtype=object),
        "filter_block_reason": np.array(["none", "none"], dtype=object),
        "volume_regime": np.array(["volume_warmup", "normal_volume"], dtype=object),
        "volume_initial_direction": np.array(["unknown", "long"], dtype=object),
        "median_relative_volume": np.array([np.nan, 1.0], dtype=np.float64),
    }
    return StepResult(
        grid_point_id="atr5_m2.00_revers",
        wf_step=1,
        test_start_idx=0,
        test_end_idx=2,
        metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
        oos_trades_df=pd.DataFrame(),
        prepend_bars_requested=0,
        prepend_bars_applied=0,
        used_prepend=False,
        used_legacy_oos_path=False,
        used_defensive_fallback=False,
        oos_boundary_index=0,
        warmup_used=0,
        warmup_effective=0,
        effective_oos_bars=2,
        filter_diagnostics_oos=diag if retained else None,
        filter_diagnostics_summary={"diagnostics_available": True},
        filter_config_snapshot={
            "volume_filter_enabled": True,
            "volume_filter_mode": "volume_A",
        },
    )


def test_execute_oos_step_slices_volume_runtime(monkeypatch):
    df = _df(30)
    cfg = _cfg()
    runtime = build_volume_global_metrics(
        df["volume"].to_numpy(),
        df["close"].to_numpy(),
        cfg.trade_filter.volume,
    )
    wf = WFWindowSlice(
        step_index=0,
        train_start_idx=0,
        train_end_idx=10,
        test_start_idx=10,
        test_end_idx=20,
    )
    captured = {}

    def fake_run_single_backtest(**kwargs):
        vr = kwargs["volume_runtime"]
        captured["absolute_offset"] = vr.absolute_offset
        captured["reference_length"] = vr.reference_length
        n = len(kwargs["close"])
        return BacktestResult(
            atr_period=kwargs["atr_period"],
            multiplier=kwargs["multiplier"],
            trade_mode=kwargs["trade_mode"],
            commission=kwargs["commission"],
            warmup=0,
            returns=np.zeros(n - 1, dtype=np.float64),
            equity_curve=np.ones(n, dtype=np.float64),
            positions=np.zeros(n, dtype=np.int8),
            trend=np.ones(n, dtype=np.int8),
            metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
            early_exit=False,
            exit_bar=None,
            exit_drawdown=None,
            trades_df=pd.DataFrame(),
            n_bars_original=n,
            filter_config_snapshot=vr.filter_config_snapshot,
        )

    monkeypatch.setattr(step_executor, "run_single_backtest", fake_run_single_backtest)

    execute_oos_step(
        grid_point=_grid_point(),
        wf_slice=wf,
        full_open=df["open"].to_numpy(),
        full_high=df["high"].to_numpy(),
        full_low=df["low"].to_numpy(),
        full_close=df["close"].to_numpy(),
        full_index=df.index,
        config=cfg,
        prepend_bars_requested=3,
        volume_runtime=runtime,
    )

    assert captured == {"absolute_offset": 7, "reference_length": 13}


def test_strip_preserves_summary_and_snapshot():
    sr = _step_result(retained=True)

    _strip_filter_diagnostics_arrays(
        [sr],
        retain_per_bar_filter_diagnostics=False,
    )

    assert sr.filter_diagnostics_oos is None
    assert sr.filter_diagnostics_summary == {"diagnostics_available": True}
    assert sr.filter_config_snapshot["volume_filter_mode"] == "volume_A"


def test_strip_retain_true_is_noop():
    sr = _step_result(retained=True)

    _strip_filter_diagnostics_arrays(
        [sr],
        retain_per_bar_filter_diagnostics=True,
    )

    assert sr.filter_diagnostics_oos is not None


def test_wf_filter_diagnostics_sheet_respects_retain_flag(tmp_path):
    summary = pd.DataFrame({"grid_point_id": ["atr5_m2.00_revers"]})
    step_long = pd.DataFrame(
        {
            "grid_point_id": ["atr5_m2.00_revers"],
            "wf_step": [1],
            "step_status": ["ok"],
            "sum_pnl_pct": [0.0],
            "sharpe": [0.0],
            "sortino": [0.0],
            "max_drawdown": [0.0],
            "num_trades": [0],
            "profit_factor": [0.0],
            "effective_oos_bars": [2],
        }
    )
    common = dict(
        summary_wide=summary,
        step_oos_long=step_long,
        trades_oos=pd.DataFrame(),
        trades_train=pd.DataFrame(),
        bucket_matrix_median=None,
        step_results_oos={"atr5_m2.00_revers": [_step_result(retained=True)]},
    )

    retained_path = export_workbook(
        **common,
        config=_cfg(retain=True),
        output_path=tmp_path / "retained.xlsx",
    )
    stripped_path = export_workbook(
        **common,
        config=_cfg(retain=False),
        output_path=tmp_path / "stripped.xlsx",
    )

    wb_retained = openpyxl.load_workbook(retained_path, read_only=True)
    wb_stripped = openpyxl.load_workbook(stripped_path, read_only=True)
    try:
        assert "WF_FilterDiagnostics" in wb_retained.sheetnames
        assert "WF_FilterDiagnostics" not in wb_stripped.sheetnames
        rows = list(wb_retained["WF_Config"].iter_rows(values_only=True))
        assert ("filter_config_snapshot", "volume_filter_mode", "volume_A") in rows
        assert ("filter_config_snapshot", "volume_aggregation", "median") in rows
        assert ("filter_config_snapshot", "volume_baseline_session_enabled", "False") in rows
        assert ("filter_config_snapshot", "volume_baseline_session_window", "null") in rows
    finally:
        wb_retained.close()
        wb_stripped.close()


def test_export_retain_config_validation_and_warning(tmp_path):
    base = f"""
schema_version: 1
data:
  file_path: "{(tmp_path / 'data.csv').as_posix()}"
validation:
  walk_forward:
    train_size: 30m
    test_size: 10m
export:
  retain_per_bar_filter_diagnostics: true
"""
    cfg_path = tmp_path / "retain.yaml"
    cfg_path.write_text(base, encoding="utf-8")

    with pytest.warns(RuntimeWarning, match="memory and IPC usage"):
        cfg = load_grid_config(str(cfg_path))
    assert cfg.export.retain_per_bar_filter_diagnostics is True

    bad_path = tmp_path / "bad_retain.yaml"
    bad_path.write_text(
        base.replace("retain_per_bar_filter_diagnostics: true", "retain_per_bar_filter_diagnostics: not_bool"),
        encoding="utf-8",
    )
    with pytest.raises(ConfigError, match="export.retain_per_bar_filter_diagnostics"):
        load_grid_config(str(bad_path))


def test_run_grid_pipeline_volume_sequential_and_parallel_summaries_match(tmp_path):
    df = _df(70)
    data_path = tmp_path / "data.csv"
    df.to_csv(data_path)

    def write_cfg(path, *, parallel: bool) -> None:
        path.write_text(
            f"""
schema_version: 1
data:
  file_path: "{data_path.as_posix()}"
  periods_per_year: 365
optimization:
  atr_period_range: [5, 5]
  multiplier_range: [2.0, 2.0]
  multiplier_step: 1.0
  trade_mode: revers
backtest:
  commission: 0.0
  min_trades_required: 0
validation:
  warmup_period: 0
  walk_forward:
    train_size: 30m
    test_size: 10m
    step_size: 10m
    min_train_bars: 5
    min_test_bars: 5
execution:
  parallel_enabled: {str(parallel).lower()}
  max_workers: 2
export:
  retain_per_bar_filter_diagnostics: false
trade_filter:
  enabled: true
  volume:
    enabled: true
    mode: volume_A
    short_window: 2
    baseline_window: 3
    threshold_ratio: 1.0
    regime_low_ratio: 0.8
    regime_high_ratio: 1.2
    direction_lookback_bars: 1
""",
            encoding="utf-8",
        )

    seq_cfg = tmp_path / "seq.yaml"
    par_cfg = tmp_path / "par.yaml"
    write_cfg(seq_cfg, parallel=False)
    write_cfg(par_cfg, parallel=True)

    seq = run_grid_pipeline(str(seq_cfg), output_path=str(tmp_path / "seq.xlsx"))
    par = run_grid_pipeline(str(par_cfg), output_path=str(tmp_path / "par.xlsx"))

    assert seq.error is None
    assert par.error is None
    volume_cols = [
        "n_volume_started_cycles",
        "n_volume_blocked_start_attempts",
        "avg_median_relative_volume",
    ]
    pd.testing.assert_frame_equal(
        seq.step_oos_long[volume_cols].reset_index(drop=True),
        par.step_oos_long[volume_cols].reset_index(drop=True),
    )


@dataclass
class _BareResult:
    config: GridConfig | None
    step_oos_long: pd.DataFrame | None = None
    step_train_long: pd.DataFrame | None = None
    trades_oos: pd.DataFrame | None = None
    trades_train: pd.DataFrame | None = None
    aggregated: pd.DataFrame | None = None
    ranked: pd.DataFrame | None = None
    summary_wide: pd.DataFrame | None = None
    bucket_matrix_median: pd.DataFrame | None = None
    diagnostics: object | None = None
    bucket_matrix_status: str = "disabled"
    bucket_matrix_error: str | None = None


def test_fingerprint_volume_fields_only_when_enabled():
    baseline = compute_pipeline_fingerprint(_BareResult(config=None))
    zigzag_only = compute_pipeline_fingerprint(
        _BareResult(config=GridConfig(data=DataConfig(file_path="")))
    )
    volume = compute_pipeline_fingerprint(_BareResult(config=_cfg()))

    volume_keys = {
        "volume_filter_enabled",
        "volume_filter_mode",
        "volume_short_window",
        "volume_baseline_window",
        "volume_threshold_ratio",
        "volume_regime_low_ratio",
        "volume_regime_high_ratio",
        "volume_direction_lookback_bars",
    }
    assert volume_keys.isdisjoint(baseline.metadata)
    assert volume_keys.isdisjoint(zigzag_only.metadata)
    assert volume_keys <= set(volume.metadata)
    assert "grid_point_id" not in volume.metadata
