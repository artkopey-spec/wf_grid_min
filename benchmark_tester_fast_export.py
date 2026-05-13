"""
Benchmark legacy Tester XLSX export modes with per-sheet Excel write profiling.

This script runs the legacy tester path in-process for three modes:

  A. baseline: current config defaults
  B. export-only-fast: period splits kept, heavy export sheets disabled
  C. full-fast: only 100% period, heavy export sheets disabled

Outputs are written under --output-dir:
  - one XLSX per mode
  - fast_export_benchmark_<timestamp>.csv
  - fast_export_sheet_profile_<timestamp>.csv
  - fast_export_benchmark_<timestamp>.json

The per-sheet profile is collected by temporarily wrapping
``pandas.DataFrame.to_excel`` around the exporter call only. Production export
APIs are not changed.
"""

from __future__ import annotations

import argparse
import copy
import csv
import json
import sys
import time
import zipfile
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

import pandas as pd
import yaml
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent
DONOR_ROOT = PROJECT_ROOT / "donor"
if str(DONOR_ROOT) not in sys.path:
    sys.path.insert(0, str(DONOR_ROOT))

from supertrend_optimizer.cli.tester import (  # noqa: E402
    load_tester_config,
    merge_cli_and_config,
)
from supertrend_optimizer.core.trade_filter_config import (  # noqa: E402
    is_volume_enabled,
    is_zigzag_enabled,
)
from supertrend_optimizer.core.volume_metrics import (  # noqa: E402
    _warn_if_volume_baseline_window_large,
    build_volume_global_metrics,
)
from supertrend_optimizer.core.zigzag_st_filter import build_zigzag_global_stats  # noqa: E402
from supertrend_optimizer.data.loader import load_ohlc_csv  # noqa: E402
from supertrend_optimizer.data.timeframe import resolve_periods_per_year_from_config  # noqa: E402
from supertrend_optimizer.data.validator import (  # noqa: E402
    validate_ohlc_data,
    validate_volume_filter_data,
)
from supertrend_optimizer.io.excel_tester import export_tester_results  # noqa: E402
from supertrend_optimizer.testing.runner import run_all_periods  # noqa: E402
from supertrend_optimizer.testing.signal_events import build_signal_events  # noqa: E402
from supertrend_optimizer.utils.config import load_config  # noqa: E402
from supertrend_optimizer.utils.enums import ExecutionModel, MarketType  # noqa: E402
from supertrend_optimizer.utils.warmup import calculate_warmup_tester  # noqa: E402


TS_FMT = "%Y%m%d_%H%M%S"


@dataclass(frozen=True)
class Mode:
    name: str
    period: bool
    diagnostics: bool
    signals: bool
    false_start: bool
    cycle: bool
    trades: bool


MODES = (
    Mode("A_baseline", True, True, True, True, True, True),
    Mode("B_export_only_fast", True, False, False, False, False, False),
    Mode("C_full_fast", False, False, False, False, False, False),
)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Benchmark Tester legacy XLSX fast-export modes."
    )
    parser.add_argument("--csv", type=Path, default=PROJECT_ROOT / "data.csv")
    parser.add_argument(
        "--config", type=Path, default=PROJECT_ROOT / "config_tester.yaml"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "results tester" / "fast_export_benchmark",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Also print the JSON summary to stdout.",
    )
    return parser


def _variant_config(raw: dict[str, Any], mode: Mode) -> dict[str, Any]:
    cfg = copy.deepcopy(raw)
    cfg["period"] = mode.period
    export = dict(cfg.get("export") or {})
    export.update(
        {
            "diagnostics": mode.diagnostics,
            "signals": mode.signals,
            "false_start": mode.false_start,
            "cycle": mode.cycle,
            "trades": mode.trades,
        }
    )
    cfg["export"] = export
    segmentation = dict(cfg.get("segmentation") or {})
    segmentation["mode"] = "legacy"
    cfg["segmentation"] = segmentation
    return cfg


@contextmanager
def _profile_to_excel(profile_rows: list[dict[str, Any]]) -> Iterator[None]:
    original = pd.DataFrame.to_excel

    def profiled(self: pd.DataFrame, *args: Any, **kwargs: Any) -> Any:
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        t0 = time.perf_counter()
        try:
            return original(self, *args, **kwargs)
        finally:
            profile_rows.append(
                {
                    "sheet_name": sheet_name,
                    "rows": int(len(self)),
                    "cols": int(len(self.columns)),
                    "to_excel_sec": round(time.perf_counter() - t0, 6),
                }
            )

    pd.DataFrame.to_excel = profiled  # type: ignore[method-assign]
    try:
        yield
    finally:
        pd.DataFrame.to_excel = original  # type: ignore[method-assign]


def _workbook_info(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "sheets": list(wb.sheetnames),
            "sheet_count": len(wb.sheetnames),
            "size_bytes": path.stat().st_size,
        }
    finally:
        wb.close()


def _xlsx_xml_sizes(path: Path) -> dict[str, int]:
    sizes: dict[str, int] = {}
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if info.filename.startswith("xl/worksheets/sheet"):
                sizes[info.filename] = int(info.file_size)
    return sizes


def _summary_100(path: Path) -> tuple[tuple[Any, ...], tuple[Any, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))
        header = tuple(v for v in rows[0] if v is not None)
        period_idx = header.index("Period")
        for row in rows[1:]:
            if len(row) > period_idx and row[period_idx] == "100%":
                return header, tuple(row[: len(header)])
    finally:
        wb.close()
    raise RuntimeError(f"Summary 100% row not found: {path}")


def _sheet_rows(path: Path, sheet_name: str) -> list[tuple[Any, ...]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return [tuple(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _run_mode(
    mode: Mode,
    raw_config: dict[str, Any],
    config_path: Path,
    csv_path: Path,
    output_dir: Path,
    timestamp: str,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    t_total = time.perf_counter()
    raw_variant = _variant_config(raw_config, mode)
    config = load_tester_config(str(config_path), loaded_raw=raw_variant)
    args_stub = argparse.Namespace(
        atr=None,
        mult=None,
        mode=None,
        periods_per_year=None,
        annualization_basis=None,
        market=None,
        execution_model=None,
    )
    params = merge_cli_and_config(args_stub, config)

    t0 = time.perf_counter()
    df = validate_volume_filter_data(
        validate_ohlc_data(load_ohlc_csv(str(csv_path))),
        params.get("trade_filter"),
    )
    data_sec = time.perf_counter() - t0

    tf_cfg = params.get("trade_filter")
    market_enum = MarketType(params["market"]) if params["market"] else None
    periods_per_year = resolve_periods_per_year_from_config(
        config_value=params["annualization_factor"],
        index=df.index,
        explicit_basis=params["annualization_basis"],
        market=market_enum,
    )
    execution_model = (
        ExecutionModel(params["execution_model"])
        if params["execution_model"]
        else ExecutionModel.OPEN_TO_OPEN
    )
    if params["warmup_period_auto"]:
        warmup_period = calculate_warmup_tester(
            n=len(df),
            atr_period=params["atr_period"],
            warmup_period_auto=True,
        )
    else:
        warmup_period = max(params["warmup_period"], params["atr_period"])

    t0 = time.perf_counter()
    full_volume_runtime = None
    if is_volume_enabled(tf_cfg):
        full_volume_runtime = build_volume_global_metrics(
            df["volume"].to_numpy(),
            df["close"].to_numpy(),
            tf_cfg.volume,
            index=df.index,
        )
        _warn_if_volume_baseline_window_large(tf_cfg.volume, len(df))
    volume_sec = time.perf_counter() - t0

    t0 = time.perf_counter()
    zigzag_global_stats = None
    if is_zigzag_enabled(tf_cfg):
        zigzag_global_stats = build_zigzag_global_stats(
            close=df["close"].values,
            trade_filter_config=tf_cfg,
        )
    zigzag_sec = time.perf_counter() - t0

    t0 = time.perf_counter()
    results = run_all_periods(
        df=df,
        atr_period=params["atr_period"],
        multiplier=params["multiplier"],
        trade_mode=params["trade_mode"],
        commission=params["commission"],
        warmup_period=warmup_period,
        periods_per_year=periods_per_year,
        execution_model=execution_model,
        auto_warmup=params["warmup_period_auto"],
        min_trades_required=params["min_trades_required"],
        trade_filter_config=tf_cfg,
        zigzag_global_stats=zigzag_global_stats,
        volume_runtime=full_volume_runtime,
        include_period_splits=params["period"],
    )
    backtest_sec = time.perf_counter() - t0

    t0 = time.perf_counter()
    signals_df = None
    if params["export"]["signals"]:
        signals_df = build_signal_events(
            df=df,
            trend=results[0].result.trend,
            atr_period=params["atr_period"],
            trade_mode=params["trade_mode"],
            execution_model=execution_model,
            filter_diagnostics=results[0].filter_diagnostics,
        )
    signals_sec = time.perf_counter() - t0

    profile_rows: list[dict[str, Any]] = []
    requested_output = output_dir / f"{mode.name}_{timestamp}.xlsx"
    run_metadata = {
        "config_path": str(config_path.resolve()),
        "csv_path": str(csv_path.resolve()),
        "output_path_requested": str(requested_output),
        "segmentation": {"mode": "legacy", "n_parts": params["segmentation"]["n_parts"]},
        "resolved_periods_per_year": periods_per_year,
        "annualization_factor_config": params["annualization_factor"],
        "warmup_period_resolved": warmup_period,
        "warmup_period_auto": params["warmup_period_auto"],
        "execution_model": execution_model.value,
        "market": params["market"],
        "annualization_basis": params["annualization_basis"],
        "warmup_period_effective": results[0].effective_warmup,
        "benchmark_mode": mode.name,
    }
    t0 = time.perf_counter()
    with _profile_to_excel(profile_rows):
        actual_output = Path(
            export_tester_results(
                results,
                str(requested_output),
                signals_df=signals_df,
                false_start_max_bars=params["false_start_max_bars"],
                trade_filter_config=tf_cfg,
                df=df,
                config_yaml_snapshot=raw_variant,
                run_metadata=run_metadata,
                export_diagnostics=params["export"]["diagnostics"],
                export_signals=params["export"]["signals"],
                export_false_start=params["export"]["false_start"],
                export_cycle=params["export"]["cycle"],
                export_trades=params["export"]["trades"],
            )
        )
    export_sec = time.perf_counter() - t0

    wb_info = _workbook_info(actual_output)
    xml_sizes = _xlsx_xml_sizes(actual_output)
    total_sec = time.perf_counter() - t_total
    summary = {
        "mode": mode.name,
        "output_path": str(actual_output),
        "period_count": len(results),
        "sheet_count": wb_info["sheet_count"],
        "sheets": wb_info["sheets"],
        "xlsx_size_bytes": wb_info["size_bytes"],
        "worksheet_xml_bytes": sum(xml_sizes.values()),
        "data_sec": round(data_sec, 6),
        "volume_sec": round(volume_sec, 6),
        "zigzag_sec": round(zigzag_sec, 6),
        "backtest_sec": round(backtest_sec, 6),
        "signals_sec": round(signals_sec, 6),
        "excel_export_sec": round(export_sec, 6),
        "total_sec": round(total_sec, 6),
    }
    for row in profile_rows:
        row["mode"] = mode.name
        row["output_path"] = str(actual_output)
    return summary, profile_rows


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    raw_config = load_config(str(args.config))
    if not isinstance(raw_config, dict):
        raise SystemExit(f"Config root must be a mapping: {args.config}")

    timestamp = datetime.now().strftime(TS_FMT)
    summaries: list[dict[str, Any]] = []
    sheet_profiles: list[dict[str, Any]] = []

    for mode in MODES:
        print(f"[{mode.name}] running...")
        summary, profile = _run_mode(
            mode,
            raw_config,
            args.config,
            args.csv,
            args.output_dir,
            timestamp,
        )
        summaries.append(summary)
        sheet_profiles.extend(profile)
        print(
            f"  total={summary['total_sec']:.3f}s "
            f"export={summary['excel_export_sec']:.3f}s "
            f"size={summary['xlsx_size_bytes']} bytes "
            f"sheets={summary['sheet_count']}"
        )

    baseline_path = Path(str(summaries[0]["output_path"]))
    baseline_summary = _summary_100(baseline_path)
    baseline_metrics = _sheet_rows(baseline_path, "Metrics_100")
    for summary in summaries[1:]:
        path = Path(str(summary["output_path"]))
        summary["summary_100_matches_baseline"] = _summary_100(path) == baseline_summary
        summary["metrics_100_matches_baseline"] = (
            _sheet_rows(path, "Metrics_100") == baseline_metrics
        )

    bench_csv = args.output_dir / f"fast_export_benchmark_{timestamp}.csv"
    sheet_csv = args.output_dir / f"fast_export_sheet_profile_{timestamp}.csv"
    bench_json = args.output_dir / f"fast_export_benchmark_{timestamp}.json"
    _write_csv(bench_csv, summaries)
    _write_csv(sheet_csv, sheet_profiles)
    with bench_json.open("w", encoding="utf-8") as f:
        json.dump({"runs": summaries, "sheet_profile": sheet_profiles}, f, indent=2)

    print(f"\nSummary CSV: {bench_csv}")
    print(f"Sheet profile CSV: {sheet_csv}")
    print(f"JSON: {bench_json}")
    if args.json:
        print(json.dumps({"runs": summaries}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
