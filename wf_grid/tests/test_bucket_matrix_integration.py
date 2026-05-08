"""
Integration tests: step_oos_long → build_median_bucket_matrix → write_bucket_matrix_median_sheet.

Covers шаги 4.1–4.4:
    4.1  Shared fixtures: _make_integration_oos, _make_integration_config
    4.2  Column contract: 24+2N columns, interleaved layout, Block E
    4.3  Round-trip: builder → writer → Excel assertions
    4.4  Data integrity: DD values, sign, NaN-independence, sentinel cleaning
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from wf_grid.bucket.median_matrix_builder import build_median_bucket_matrix
from wf_grid.config.schema import (
    BucketConfig,
    DataConfig,
    GridConfig,
    INVALID_METRIC_VALUE,
    OptimizationConfig,
)
from wf_grid.export.bucket_sheet_writer import write_bucket_matrix_median_sheet
from wf_grid.grid.enumeration import _atr_values


# ===========================================================================
# 4.1 Shared fixtures / helpers
# ===========================================================================

def _make_integration_config(
    atr_range=(10, 14),
    mult_range=(2.0, 2.4),
    mult_step=0.2,
    trade_mode="long",
    atr_bucket_step=2,
    mult_bucket_step=0.2,
    min_buckets_for_median=3,
) -> GridConfig:
    """Minimal GridConfig for integration tests (4 ATR values × 3 mult values)."""
    return GridConfig(
        data=DataConfig(file_path="dummy.csv"),
        optimization=OptimizationConfig(
            atr_period_range=list(atr_range),
            multiplier_range=list(mult_range),
            multiplier_step=mult_step,
            trade_mode=trade_mode,
        ),
        bucket=BucketConfig(
            atr_bucket_step=atr_bucket_step,
            mult_bucket_step=mult_bucket_step,
            min_buckets_for_median=min_buckets_for_median,
        ),
    )


def _make_integration_oos(
    config: GridConfig,
    n_steps: int = 3,
    pnl_base: float = 10.0,
    dd_base: float = -0.10,
    status: str = "ok",
    include_max_drawdown: bool = True,
) -> pd.DataFrame:
    """Create a realistic step_oos_long for integration testing.

    Covers all ATR × mult combinations from the config's ranges.
    """
    opt = config.optimization
    atr_min, atr_max = opt.atr_period_range
    mult_min, mult_max = opt.multiplier_range
    mult_step = opt.multiplier_step
    trade_mode = opt.trade_mode

    atr_vals = _atr_values(atr_min, atr_max, int(opt.atr_period_step))
    # build mult list with rounding to avoid float drift
    mult_vals = []
    m = mult_min
    while round(m, 10) <= round(mult_max, 10):
        mult_vals.append(round(m, 10))
        m = round(m + mult_step, 10)

    rows = []
    for atr in atr_vals:
        for mult in mult_vals:
            gp_id = f"atr{atr}_m{mult}_{trade_mode}"
            for s in range(1, n_steps + 1):
                row: dict = {
                    "atr_period": atr,
                    "multiplier": mult,
                    "trade_mode": trade_mode,
                    "wf_step": s,
                    "step_status": status,
                    "sum_pnl_pct": pnl_base + atr * 0.1 + mult + s * 0.5,
                    "grid_point_id": gp_id,
                }
                if include_max_drawdown:
                    row["max_drawdown"] = dd_base - atr * 0.01 - s * 0.005
                rows.append(row)

    df = pd.DataFrame(rows)
    if not include_max_drawdown:
        # add column as all-NaN to satisfy required columns check
        df["max_drawdown"] = np.nan
    return df


def _round_trip(
    bucket_df: pd.DataFrame,
    atr_step: float = 2,
    mult_step: float = 0.2,
) -> openpyxl.Workbook:
    """Write bucket_df to temp Excel, return loaded workbook."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_bucket_matrix_median_sheet(bucket_df, writer, atr_step, mult_step)
    wb = openpyxl.load_workbook(path)
    Path(path).unlink(missing_ok=True)
    return wb


def _find_col_index(ws, col_name: str) -> int | None:
    """Find 1-based column index by header name in row 1."""
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == col_name:
            return ci
    return None


# ===========================================================================
# 4.2 Column contract
# ===========================================================================

class TestColumnContract:
    def test_column_count_24_plus_2n(self):
        """build_median_bucket_matrix → 24 + 2N columns for N steps."""
        config = _make_integration_config()
        n_steps = 3
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        assert result.shape[1] == 24 + 2 * n_steps, (
            f"Expected {24 + 2 * n_steps} cols, got {result.shape[1]}: {list(result.columns)}"
        )

    def test_column_count_varies_with_steps(self):
        """Column count grows by 2 for each additional step."""
        config = _make_integration_config()
        for n_steps in (1, 2, 4):
            oos = _make_integration_oos(config, n_steps=n_steps)
            result = build_median_bucket_matrix(oos, config)
            assert result.shape[1] == 24 + 2 * n_steps, (
                f"n_steps={n_steps}: expected {24 + 2 * n_steps}, got {result.shape[1]}"
            )

    def test_interleaved_layout(self):
        """Columns are interleaved: Step1, DD_Step1, Step2, DD_Step2, ..."""
        config = _make_integration_config()
        n_steps = 3
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        cols = list(result.columns)
        # Block A ends at index 4 (0-based)
        block_b_start = 5
        for s in range(1, n_steps + 1):
            pnl_idx = block_b_start + (s - 1) * 2
            dd_idx = block_b_start + (s - 1) * 2 + 1
            assert cols[pnl_idx] == f"Step{s}", (
                f"Col[{pnl_idx}] should be Step{s}, got {cols[pnl_idx]}"
            )
            assert cols[dd_idx] == f"DD_Step{s}", (
                f"Col[{dd_idx}] should be DD_Step{s}, got {cols[dd_idx]}"
            )

    def test_block_e_risk_summaries_present(self):
        """Block E: max_drawdown_Median and max_drawdown_Min are in result."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        assert "max_drawdown_Median" in result.columns
        assert "max_drawdown_Min" in result.columns

    def test_block_e_position(self):
        """max_drawdown_Median and max_drawdown_Min follow last DD_Step column."""
        config = _make_integration_config()
        n_steps = 2
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        cols = list(result.columns)
        last_dd_idx = cols.index(f"DD_Step{n_steps}")
        median_idx = cols.index("max_drawdown_Median")
        min_idx = cols.index("max_drawdown_Min")
        assert median_idx == last_dd_idx + 1, (
            f"max_drawdown_Median should be at {last_dd_idx + 1}, got {median_idx}"
        )
        assert min_idx == median_idx + 1

    def test_block_c_metrics_present(self):
        """All Block C metrics are present."""
        block_c = [
            "bucket_presence_steps", "mean_oos_pnl", "std_bucket",
            "pct_params_positive_pnl", "wins_count", "win_steps", "top3_count",
            "above_median_count", "above_median_ratio", "presence_count",
            "above_median_ratio_present", "eligible_median_steps_count",
            "above_median_ratio_eligible", "bucket_stability_score",
            "zone_dominance_score",
        ]
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        for col in block_c:
            assert col in result.columns, f"Block C column missing: {col}"

    def test_block_c_follows_block_e(self):
        """bucket_stability_score comes after max_drawdown_Min."""
        config = _make_integration_config()
        n_steps = 2
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        cols = list(result.columns)
        min_idx = cols.index("max_drawdown_Min")
        stability_idx = cols.index("bucket_stability_score")
        zone_idx = cols.index("zone_dominance_score")
        assert stability_idx > min_idx
        balanced_idx = cols.index("bucket_balanced_score")
        assert zone_idx == balanced_idx - 1, "zone_dominance_score should be just before bucket_balanced_score"
        assert balanced_idx == len(cols) - 1, "bucket_balanced_score should be last column"


# ===========================================================================
# 4.3 Round-trip: builder → writer → Excel
# ===========================================================================

class TestRoundTrip:
    def test_round_trip_sheet_exists(self):
        """Round-trip: sheet 'BucketMatrix_Median' exists after write."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        assert "BucketMatrix_Median" in wb.sheetnames

    def test_round_trip_header_count(self):
        """Round-trip: header row has 24+2N columns."""
        config = _make_integration_config()
        n_steps = 2
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h is not None]
        assert len(headers) == 24 + 2 * n_steps, (
            f"Expected {24 + 2 * n_steps} headers, got {len(headers)}: {headers}"
        )

    def test_round_trip_dd_step_columns_in_header(self):
        """Round-trip: DD_Step* columns appear in Excel header."""
        config = _make_integration_config()
        n_steps = 3
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 2)]
        for s in range(1, n_steps + 1):
            assert f"DD_Step{s}" in headers, f"DD_Step{s} missing from Excel header"

    def test_round_trip_risk_summary_in_header(self):
        """Round-trip: max_drawdown_Median and max_drawdown_Min in Excel header."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 2)]
        assert "max_drawdown_Median" in headers
        assert "max_drawdown_Min" in headers

    def test_round_trip_dd_values_are_negative(self):
        """Round-trip: DD_Step* values in Excel are ≤ 0 (no sign inversion)."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2, dd_base=-0.10)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "DD_Step1")
        assert ci is not None, "DD_Step1 column not found"
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=ci).value
            if val is not None and not (isinstance(val, float) and val != val):
                assert val <= 0, f"DD_Step1 row {row_idx} should be ≤ 0, got {val}"

    def test_round_trip_risk_summary_format_percent(self):
        """Round-trip: max_drawdown_Median uses '0.00%' number format."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "max_drawdown_Median")
        assert ci is not None
        assert ws.cell(row=2, column=ci).number_format == "0.00%"

    def test_round_trip_two_heatmaps(self):
        """Round-trip: exactly 2 conditional formatting rules (2 heatmaps)."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        assert len(ws.conditional_formatting._cf_rules) == 2

    def test_round_trip_risk_heatmap_fixed_scale(self):
        """Round-trip: risk heatmap has fixed scale starting at -0.50."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        risk_cf = cf_keys[1]
        cs = risk_cf.rules[0].colorScale
        assert cs.cfvo[0].type == "num"
        assert float(cs.cfvo[0].val) == pytest.approx(-0.50)

    def test_round_trip_has_data_rows(self):
        """Round-trip: data rows (row 2+) are present in Excel."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result)
        ws = wb["BucketMatrix_Median"]
        assert ws.cell(row=2, column=1).value is not None, "No data in row 2"


# ===========================================================================
# 4.4 Data integrity
# ===========================================================================

class TestDataIntegrity:
    def test_dd_values_are_medians_of_negatives(self):
        """DD_Step* are medians of negative max_drawdown values."""
        config = _make_integration_config()
        # All DD values = -0.10 → medians should be -0.10
        oos = _make_integration_oos(config, n_steps=2, dd_base=-0.10)
        # Adjust so all dd values are exactly -0.10 (override formula)
        oos["max_drawdown"] = -0.10
        result = build_median_bucket_matrix(oos, config)
        dd1_vals = result["DD_Step1"].dropna()
        assert len(dd1_vals) > 0
        assert (dd1_vals <= 0).all(), "All DD_Step values should be ≤ 0"
        assert dd1_vals.abs().max() <= 1.0, "DD values should be fractions (≤ 1.0)"

    def test_dd_step_median_derivation(self):
        """DD_Step1 equals median of max_drawdown for bucket/step, not mean.

        ATR 11, 12, 13 all map to atr_bucket=12 (with atr_bucket_step=2).
        3 values [-0.10, -0.20, -0.30] → median = -0.20.
        """
        config = _make_integration_config(
            atr_range=(11, 13),
            mult_range=(2.0, 2.0),
            mult_step=0.1,
            atr_bucket_step=2,
            mult_bucket_step=0.2,
        )
        # ATR 11, 12, 13 → all map to atr_bucket=12; mult 2.0 → mult_bucket=10
        # DD values: [-0.10, -0.20, -0.30] → median = -0.20
        rows = []
        for atr, dd in [(11, -0.10), (12, -0.20), (13, -0.30)]:
            rows.append({
                "atr_period": atr, "multiplier": 2.0, "trade_mode": "long",
                "wf_step": 1, "step_status": "ok",
                "sum_pnl_pct": 10.0, "max_drawdown": dd,
                "grid_point_id": f"atr{atr}_m2.0_long",
            })
        df = pd.DataFrame(rows)
        result = build_median_bucket_matrix(df, config)
        assert "DD_Step1" in result.columns
        dd1 = result["DD_Step1"].dropna()
        assert len(dd1) > 0
        assert float(dd1.iloc[0]) == pytest.approx(-0.20, abs=1e-6), (
            f"Median of [-0.10, -0.20, -0.30] should be -0.20, got {dd1.iloc[0]}"
        )

    def test_max_drawdown_median_is_step_derived(self):
        """max_drawdown_Median is median of DD_Step* values (step-derived)."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=3)
        result = build_median_bucket_matrix(oos, config)
        if len(result) > 0:
            row = result.iloc[0]
            dd_steps = [row[f"DD_Step{s}"] for s in range(1, 4)
                        if f"DD_Step{s}" in result.columns and not pd.isna(row[f"DD_Step{s}"])]
            if len(dd_steps) > 0:
                expected = np.median(dd_steps)
                assert float(row["max_drawdown_Median"]) == pytest.approx(expected, abs=1e-9)

    def test_max_drawdown_min_is_step_derived(self):
        """max_drawdown_Min is min (worst) of DD_Step* values (step-derived)."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=3)
        result = build_median_bucket_matrix(oos, config)
        if len(result) > 0:
            row = result.iloc[0]
            dd_steps = [row[f"DD_Step{s}"] for s in range(1, 4)
                        if f"DD_Step{s}" in result.columns and not pd.isna(row[f"DD_Step{s}"])]
            if len(dd_steps) > 0:
                expected = min(dd_steps)
                assert float(row["max_drawdown_Min"]) == pytest.approx(expected, abs=1e-9)

    def test_sentinel_is_treated_as_nan(self):
        """INVALID_METRIC_VALUE in max_drawdown → NaN, not propagated."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        # Replace all DD values with sentinel
        oos["max_drawdown"] = INVALID_METRIC_VALUE
        result = build_median_bucket_matrix(oos, config)
        # All DD columns should be NaN (sentinel cleaned, no valid data)
        for col in result.columns:
            if col.startswith("DD_Step"):
                assert result[col].isna().all(), (
                    f"{col} should be all-NaN after sentinel cleaning"
                )

    def test_nan_independence_pnl_and_dd(self):
        """NaN in max_drawdown does not force NaN in Step* columns."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        oos["max_drawdown"] = np.nan
        result = build_median_bucket_matrix(oos, config)
        # Step columns should still have values (PnL data is fine)
        pnl_not_all_nan = result["Step1"].notna().any()
        assert pnl_not_all_nan, "Step1 should have values even when max_drawdown is NaN"
        # But DD columns should be all-NaN
        assert result["DD_Step1"].isna().all()

    def test_nan_independence_dd_and_pnl(self):
        """NaN in sum_pnl_pct does not force NaN in DD_Step* columns."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        oos["sum_pnl_pct"] = np.nan
        result = build_median_bucket_matrix(oos, config)
        # Step columns should be all-NaN
        assert result["Step1"].isna().all()
        # DD columns should still have values
        dd_not_all_nan = result["DD_Step1"].notna().any()
        assert dd_not_all_nan, (
            "DD_Step1 should have values even when sum_pnl_pct is NaN"
        )

    def test_dd_sign_not_inverted(self):
        """max_drawdown values are preserved as-is (no sign inversion)."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2, dd_base=-0.25)
        oos["max_drawdown"] = -0.25  # force consistent value
        result = build_median_bucket_matrix(oos, config)
        dd1_vals = result["DD_Step1"].dropna()
        assert len(dd1_vals) > 0
        assert (dd1_vals < 0).all(), (
            f"DD values should remain negative (no sign inversion), got: {dd1_vals.tolist()}"
        )

    def test_empty_oos_returns_empty_df(self):
        """Empty step_oos_long → empty DataFrame."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2).head(0)
        result = build_median_bucket_matrix(oos, config)
        assert len(result) == 0

    def test_pnl_values_are_positive(self):
        """Step* columns reflect positive pnl_base values."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2, pnl_base=100.0)
        result = build_median_bucket_matrix(oos, config)
        step1_vals = result["Step1"].dropna()
        assert len(step1_vals) > 0
        assert (step1_vals > 0).all(), "Step1 values should be positive with pnl_base=100"


# ===========================================================================
# 4.5 Contract tests (strict enforcement of plan invariants)
# ===========================================================================

class TestContractStepDerived:
    """max_drawdown_Median/Min MUST be computed from DD_Step* values,
    NOT from raw max_drawdown values.  We construct data where
    step-derived ≠ raw-derived to prove independence.
    """

    def _setup_asymmetric(self):
        """Build data where step-derived ≠ raw-derived aggregation.

        2 steps × 3 params (all in single bucket atr_bucket=12):
        - Step 1: DD = [-0.10, -0.20, -0.30] → median = -0.20
        - Step 2: DD = [-0.01, -0.01, -0.01] → median = -0.01

        step-derived Median([-0.20, -0.01]) = -0.105
        step-derived Min([-0.20, -0.01]) = -0.20
        raw Median(all 6 sorted) = median([-0.30,-0.20,-0.10,-0.01,-0.01,-0.01]) = -0.055
        raw Min(all 6) = -0.30
        """
        config = _make_integration_config(
            atr_range=(11, 13),
            mult_range=(2.0, 2.0),
            mult_step=0.1,
            atr_bucket_step=2,
            mult_bucket_step=0.2,
        )
        dd_by_step = {
            1: [-0.10, -0.20, -0.30],
            2: [-0.01, -0.01, -0.01],
        }
        atrs = [11, 12, 13]
        rows = []
        for s in (1, 2):
            for i, atr in enumerate(atrs):
                rows.append({
                    "atr_period": atr, "multiplier": 2.0, "trade_mode": "long",
                    "wf_step": s, "step_status": "ok",
                    "sum_pnl_pct": 10.0,
                    "max_drawdown": dd_by_step[s][i],
                    "grid_point_id": f"atr{atr}_m2.0_long",
                })
        return config, pd.DataFrame(rows)

    def test_step_derived_median_differs_from_raw(self):
        """max_drawdown_Median ≠ raw median(all max_drawdown values)."""
        config, df = self._setup_asymmetric()
        result = build_median_bucket_matrix(df, config)
        assert len(result) > 0
        row = result.iloc[0]
        step_derived = float(row["max_drawdown_Median"])
        raw_all = df["max_drawdown"].dropna().values
        raw_median = float(np.median(raw_all))
        assert step_derived == pytest.approx(-0.105, abs=1e-6), (
            f"step-derived median should be -0.105, got {step_derived}"
        )
        assert raw_median == pytest.approx(-0.055, abs=1e-6), (
            f"raw median should be -0.055, got {raw_median}"
        )
        assert step_derived != pytest.approx(raw_median, abs=1e-6), (
            f"step-derived ({step_derived}) must differ from raw median ({raw_median})"
        )

    def test_step_derived_min_differs_from_raw(self):
        """max_drawdown_Min ≠ raw min(all max_drawdown values)."""
        config, df = self._setup_asymmetric()
        result = build_median_bucket_matrix(df, config)
        row = result.iloc[0]
        step_derived_min = float(row["max_drawdown_Min"])
        raw_min = float(df["max_drawdown"].min())
        assert step_derived_min == pytest.approx(-0.20, abs=1e-6)
        assert raw_min == pytest.approx(-0.30, abs=1e-6)
        assert step_derived_min != pytest.approx(raw_min, abs=1e-6), (
            f"step-derived min ({step_derived_min}) must differ "
            f"from raw min ({raw_min})"
        )

    def test_dd_step_matches_expected_per_step_median(self):
        """Each DD_Step{s} exactly matches np.median of its DD values."""
        config, df = self._setup_asymmetric()
        result = build_median_bucket_matrix(df, config)
        row = result.iloc[0]
        assert float(row["DD_Step1"]) == pytest.approx(-0.20, abs=1e-6)
        assert float(row["DD_Step2"]) == pytest.approx(-0.01, abs=1e-6)


class TestContractNanGuards:
    """NaN-guard contracts: PnL and DD accumulation are fully independent."""

    def test_all_dd_nan_leaves_pnl_intact(self):
        """All max_drawdown=NaN → DD all NaN, PnL & Block C still computed."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        oos["max_drawdown"] = np.nan
        result = build_median_bucket_matrix(oos, config)
        assert len(result) > 0
        assert result["DD_Step1"].isna().all()
        assert result["DD_Step2"].isna().all()
        assert result["max_drawdown_Median"].isna().all()
        assert result["max_drawdown_Min"].isna().all()
        assert result["Step1"].notna().any(), "PnL must survive DD-NaN"
        assert result["bucket_stability_score"].notna().any(), "Block C must survive DD-NaN"

    def test_all_pnl_nan_leaves_dd_intact(self):
        """All sum_pnl_pct=NaN → Step all NaN, DD still computed."""
        config = _make_integration_config()
        oos = _make_integration_oos(config, n_steps=2)
        oos["sum_pnl_pct"] = np.nan
        result = build_median_bucket_matrix(oos, config)
        assert result["Step1"].isna().all()
        assert result["Step2"].isna().all()
        assert result["DD_Step1"].notna().any(), "DD must survive PnL-NaN"
        assert result["DD_Step2"].notna().any()
        assert result["max_drawdown_Median"].notna().any()

    def test_mixed_nan_per_row(self):
        """Row with valid PnL but NaN DD → contributes to Step, not DD_Step."""
        config = _make_integration_config(
            atr_range=(11, 13), mult_range=(2.0, 2.0),
            mult_step=0.1, atr_bucket_step=2, mult_bucket_step=0.2,
        )
        rows = [
            {"atr_period": 11, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 10.0, "max_drawdown": np.nan,
             "grid_point_id": "atr11_m2.0_long"},
            {"atr_period": 12, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": np.nan, "max_drawdown": -0.15,
             "grid_point_id": "atr12_m2.0_long"},
            {"atr_period": 13, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 20.0, "max_drawdown": -0.25,
             "grid_point_id": "atr13_m2.0_long"},
        ]
        df = pd.DataFrame(rows)
        result = build_median_bucket_matrix(df, config)
        row = result.iloc[0]
        # PnL: median([10.0, 20.0]) = 15.0 (NaN row excluded)
        assert float(row["Step1"]) == pytest.approx(15.0, abs=1e-6)
        # DD: median([-0.15, -0.25]) = -0.20 (NaN row excluded)
        assert float(row["DD_Step1"]) == pytest.approx(-0.20, abs=1e-6)


class TestContractSentinel:
    """INVALID_METRIC_VALUE (-999.0) must be cleaned to NaN before accumulation."""

    def test_sentinel_dd_not_propagated_to_median(self):
        """Sentinel in max_drawdown → cleaned, does not distort DD_Step median."""
        config = _make_integration_config(
            atr_range=(11, 13), mult_range=(2.0, 2.0),
            mult_step=0.1, atr_bucket_step=2, mult_bucket_step=0.2,
        )
        rows = [
            {"atr_period": 11, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 10.0, "max_drawdown": INVALID_METRIC_VALUE,
             "grid_point_id": "atr11_m2.0_long"},
            {"atr_period": 12, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 20.0, "max_drawdown": -0.15,
             "grid_point_id": "atr12_m2.0_long"},
            {"atr_period": 13, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 30.0, "max_drawdown": -0.25,
             "grid_point_id": "atr13_m2.0_long"},
        ]
        df = pd.DataFrame(rows)
        result = build_median_bucket_matrix(df, config)
        row = result.iloc[0]
        # DD: sentinel cleaned → [-0.15, -0.25] → median = -0.20
        assert float(row["DD_Step1"]) == pytest.approx(-0.20, abs=1e-6), (
            f"Sentinel should not distort median, got {row['DD_Step1']}"
        )

    def test_sentinel_pnl_not_propagated_to_step(self):
        """Sentinel in sum_pnl_pct → cleaned, does not distort Step median."""
        config = _make_integration_config(
            atr_range=(11, 13), mult_range=(2.0, 2.0),
            mult_step=0.1, atr_bucket_step=2, mult_bucket_step=0.2,
        )
        rows = [
            {"atr_period": 11, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": INVALID_METRIC_VALUE, "max_drawdown": -0.10,
             "grid_point_id": "atr11_m2.0_long"},
            {"atr_period": 12, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 10.0, "max_drawdown": -0.20,
             "grid_point_id": "atr12_m2.0_long"},
            {"atr_period": 13, "multiplier": 2.0, "trade_mode": "long",
             "wf_step": 1, "step_status": "ok",
             "sum_pnl_pct": 20.0, "max_drawdown": -0.30,
             "grid_point_id": "atr13_m2.0_long"},
        ]
        df = pd.DataFrame(rows)
        result = build_median_bucket_matrix(df, config)
        row = result.iloc[0]
        # PnL: sentinel cleaned → [10, 20] → median = 15
        assert float(row["Step1"]) == pytest.approx(15.0, abs=1e-6)
        # DD: all valid → [-0.10, -0.20, -0.30] → median = -0.20
        assert float(row["DD_Step1"]) == pytest.approx(-0.20, abs=1e-6)


class TestContractColumnOrder:
    """Strict column ordering contract across full integration pipeline."""

    def test_full_column_sequence(self):
        """24+2N columns in exact order: Block A → B+D interleaved → E → C."""
        config = _make_integration_config()
        n_steps = 3
        oos = _make_integration_oos(config, n_steps=n_steps)
        result = build_median_bucket_matrix(oos, config)
        cols = list(result.columns)

        block_a = ["bucket_param", "bucket_key", "atr_bucket",
                    "mult_bucket_ticks", "bucket_size"]
        interleaved = []
        for s in range(1, n_steps + 1):
            interleaved.extend([f"Step{s}", f"DD_Step{s}"])
        block_e = ["max_drawdown_Median", "max_drawdown_Min"]
        block_c = [
            "bucket_presence_steps", "mean_oos_pnl", "median_oos_pnl",
            "std_bucket", "pct_params_positive_pnl",
            "wins_count", "win_steps", "top3_count",
            "above_median_count", "above_median_ratio", "presence_count",
            "above_median_ratio_present", "eligible_median_steps_count",
            "above_median_ratio_eligible", "bucket_stability_score",
            "zone_dominance_score", "bucket_balanced_score",
        ]
        expected = block_a + interleaved + block_e + block_c
        assert cols == expected, (
            f"Column mismatch.\n"
            f"Expected: {expected}\n"
            f"Got:      {cols}"
        )


# ===========================================================================
# 4.6 End-to-End: full pipeline → Excel → read back all assertions
# ===========================================================================

class TestE2E:
    """Full pipeline: step_oos_long → build → write → Excel read-back."""

    @pytest.fixture()
    def e2e_wb(self):
        config = _make_integration_config()
        n_steps = 3
        oos = _make_integration_oos(config, n_steps=n_steps, dd_base=-0.12)
        result = build_median_bucket_matrix(oos, config)
        wb = _round_trip(result, atr_step=2, mult_step=0.2)
        return wb, n_steps, result

    def test_e2e_sheet_present(self, e2e_wb):
        wb, _, _ = e2e_wb
        assert "BucketMatrix_Median" in wb.sheetnames

    def test_e2e_header_complete(self, e2e_wb):
        wb, n_steps, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h is not None]
        assert len(headers) == 24 + 2 * n_steps

    def test_e2e_dd_columns_present_in_excel(self, e2e_wb):
        wb, n_steps, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]
        for s in range(1, n_steps + 1):
            assert f"DD_Step{s}" in headers

    def test_e2e_dd_values_negative_in_excel(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "DD_Step1")
        assert ci is not None
        for ri in range(2, ws.max_row + 1):
            val = ws.cell(row=ri, column=ci).value
            if val is not None and isinstance(val, (int, float)):
                assert val <= 0, f"Row {ri}: DD_Step1={val} must be ≤ 0"

    def test_e2e_dd_format_is_percent(self, e2e_wb):
        wb, n_steps, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        for s in range(1, n_steps + 1):
            ci = _find_col_index(ws, f"DD_Step{s}")
            assert ci is not None
            assert ws.cell(row=2, column=ci).number_format == "0.00%", (
                f"DD_Step{s} format must be '0.00%'"
            )

    def test_e2e_pnl_format_unchanged(self, e2e_wb):
        wb, n_steps, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        for s in range(1, n_steps + 1):
            ci = _find_col_index(ws, f"Step{s}")
            assert ci is not None
            assert ws.cell(row=2, column=ci).number_format == "0.000000", (
                f"Step{s} format must be '0.000000'"
            )

    def test_e2e_risk_summary_format(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        for col in ("max_drawdown_Median", "max_drawdown_Min"):
            ci = _find_col_index(ws, col)
            assert ci is not None
            assert ws.cell(row=2, column=ci).number_format == "0.00%"

    def test_e2e_two_heatmaps_present(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        cf_rules = ws.conditional_formatting._cf_rules
        assert len(cf_rules) == 2

    def test_e2e_stability_heatmap_auto_scale(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        stability_cs = cf_keys[0].rules[0].colorScale
        assert stability_cs.cfvo[0].type == "min"
        assert stability_cs.cfvo[2].type == "max"

    def test_e2e_risk_heatmap_fixed_scale(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        risk_cs = cf_keys[1].rules[0].colorScale
        assert risk_cs.cfvo[0].type == "num"
        assert float(risk_cs.cfvo[0].val) == pytest.approx(-0.50)
        assert float(risk_cs.cfvo[1].val) == pytest.approx(-0.15)
        assert float(risk_cs.cfvo[2].val) == pytest.approx(0.0)
        assert risk_cs.color[0].rgb == "FFFF0000"
        assert risk_cs.color[1].rgb == "FFFFFF99"
        assert risk_cs.color[2].rgb == "FF00AA00"

    def test_e2e_risk_heatmap_title_present(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        found = False
        for ri in range(1, ws.max_row + 1):
            val = ws.cell(row=ri, column=1).value
            if val and "RISK HEATMAP" in str(val).upper():
                found = True
                break
        assert found, "Risk heatmap title not found in Excel sheet"

    def test_e2e_data_matches_builder_output(self, e2e_wb):
        """Excel data matches builder DataFrame (spot check first row, first DD col)."""
        wb, _, result = e2e_wb
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "DD_Step1")
        assert ci is not None
        excel_val = ws.cell(row=2, column=ci).value
        builder_val = float(result.iloc[0]["DD_Step1"])
        if not pd.isna(builder_val):
            assert excel_val == pytest.approx(builder_val, abs=1e-6)

    def test_e2e_freeze_panes(self, e2e_wb):
        wb, _, _ = e2e_wb
        ws = wb["BucketMatrix_Median"]
        assert ws.freeze_panes == "E2"
