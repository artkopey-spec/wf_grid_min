"""
FIX-5.2a / FIX-5.2b — Integration tests for wf_step 1-based indexing.

FIX-5.2a: Data-layer path
    make_walk_forward_slices -> run_wf_for_grid_point -> collect_oos_steps

FIX-5.2b: Export-layer path
    synthetic step_oos_long (wf_step=[1,2,3]) -> build_summary_wide
    -> export_workbook -> openpyxl workbook sheet assertions

Contract under test (FIX-5.1):
    StepResult.wf_step is a 1-based export label.
    wf_slice.step_index is 0-based (donor internal) and is normalised to
    wf_step = step_index + 1 at StepResult construction time.

Expected properties (FIX-5.2a):
    - wf_step values form the exact set {1, 2, ..., N}
    - wf_step.min() == 1  (no 0-indexed leakage)
    - wf_step.max() == N  (no step is missing)
    - no wf_step == 0 anywhere in the DataFrame
    - every grid_point_id has exactly N rows (completeness)

Expected properties (FIX-5.2b):
    - workbook sheets include DISCLAIMER, WF_Config, WF_01, WF_02, WF_03,
      WF_Trades, WF_Train_Trades, summary
    - workbook does NOT contain WF_00
    - summary_wide["n_segments"] == 3
    - each WF_0X sheet contains only rows for that step

These assertions would FAIL against the pre-FIX-5.1 code where
step_index was used directly (0-based) and wf_step=0 appeared in the
first step of every grid point.
"""

from __future__ import annotations

import tempfile
import textwrap
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from supertrend_optimizer.utils.time_utils import make_walk_forward_slices
from supertrend_optimizer.utils.warmup import calculate_warmup

from wf_grid.collect.step_collector import collect_oos_steps
from wf_grid.config.loader import load_grid_config
from wf_grid.config.schema import DataConfig, GridConfig, RankingConfig
from wf_grid.export.summary_builder import build_summary_wide
from wf_grid.export.xlsx_writer import export_workbook
from wf_grid.grid.enumeration import GridPoint
from wf_grid.status.status_model import StepStatus
from wf_grid.wf.runner import run_wf_for_grid_point


# ---------------------------------------------------------------------------
# Shared helpers (mirror patterns from test_step_executor.py)
# ---------------------------------------------------------------------------

def _write_yaml(tmp_path: Path, content: str) -> str:
    p = tmp_path / "config.yaml"
    p.write_text(textwrap.dedent(content), encoding="utf-8")
    return str(p)


def _make_config(tmp_path: Path) -> GridConfig:
    """Minimal config: 50-bar train / 40-bar test → yields N >= 3 slices on 300-bar data."""
    yaml_text = """\
data:
  file_path: data.csv
  periods_per_year: 252
optimization:
  atr_period_range: [5, 10]
  multiplier_range: [2.0, 3.0]
  multiplier_step: 0.5
  trade_mode: both
backtest:
  commission: 0.0002
  min_trades_required: 1
  early_exit_enabled: false
  early_exit_max_drawdown: 0.50
  early_exit_check_bars: 50
validation:
  warmup_period: 0
  warmup_period_auto: false
  walk_forward:
    train_size: "50bars"
    test_size: "40bars"
status:
  min_meaningful_bars: 5
"""
    path = _write_yaml(tmp_path, yaml_text)
    cfg = load_grid_config(path)
    cfg.resolved_periods_per_year = 252.0
    return cfg


def _make_trending_data(n: int = 300, seed: int = 7) -> pd.DataFrame:
    """Synthetic OHLC with a drift so the engine generates some trades."""
    rng = np.random.RandomState(seed)
    close = 100.0 + np.cumsum(rng.randn(n) * 0.5 + 0.05)
    close = np.maximum(close, 10.0)
    high = close + rng.uniform(0.1, 1.0, n)
    low = close - rng.uniform(0.1, 1.0, n)
    open_ = np.maximum(close + rng.randn(n) * 0.3, low + 0.01)
    idx = pd.date_range("2020-01-01", periods=n, freq="1D")
    return pd.DataFrame(
        {"open": open_, "high": high, "low": low, "close": close},
        index=idx,
    )


def _make_grid_point(atr: int = 5, mult: float = 2.0, mode: str = "both") -> GridPoint:
    gid = f"atr{atr}_m{mult:.2f}_{mode}"
    return GridPoint(atr_period=atr, multiplier=mult, trade_mode=mode, grid_point_id=gid)


# ---------------------------------------------------------------------------
# Test: wf_step 1-based indexing through the data layer
# ---------------------------------------------------------------------------

class TestWFStepDataLayerIndexing:
    """Integration test: slices -> runner -> collector wf_step contract."""

    def test_wf_step_is_one_based_min(self, tmp_path):
        """wf_step.min() == 1 — no 0-indexed leakage from step_index."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()
        gp = _make_grid_point()

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )
        assert len(wf_slices) >= 3, (
            f"Fixture produced only {len(wf_slices)} slices; need >= 3 for a meaningful test. "
            "Increase data length or reduce train/test sizes."
        )

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })

        oos_results = {
            gp.grid_point_id: run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )
        }

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=len(wf_slices),
            grid_points=[gp],
        )

        assert df["wf_step"].min() == 1, (
            f"wf_step.min() is {df['wf_step'].min()}, expected 1. "
            "FIX-5.1 (+1 normalisation) may not be applied."
        )

    def test_wf_step_is_one_based_max(self, tmp_path):
        """wf_step.max() == N — all steps present, no off-by-one at the top."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()
        gp = _make_grid_point()

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })
        n = len(wf_slices)

        oos_results = {
            gp.grid_point_id: run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )
        }

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=n,
            grid_points=[gp],
        )

        assert df["wf_step"].max() == n, (
            f"wf_step.max() is {df['wf_step'].max()}, expected {n}."
        )

    def test_wf_step_no_zero(self, tmp_path):
        """No wf_step == 0 anywhere — pre-FIX-5.1 code would fail this."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()
        gp = _make_grid_point()

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })

        oos_results = {
            gp.grid_point_id: run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )
        }

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=len(wf_slices),
            grid_points=[gp],
        )

        zero_rows = df[df["wf_step"] == 0]
        assert zero_rows.empty, (
            f"Found {len(zero_rows)} row(s) with wf_step == 0. "
            "These correspond to the first WF step with 0-based step_index leakage."
        )

    def test_wf_step_complete_set(self, tmp_path):
        """wf_step values form exactly {1, 2, ..., N} — no gaps, no duplicates per grid point."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()
        gp = _make_grid_point()

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )
        n = len(wf_slices)

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })

        oos_results = {
            gp.grid_point_id: run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )
        }

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=n,
            grid_points=[gp],
        )

        gp_df = df[df["grid_point_id"] == gp.grid_point_id]
        actual_steps = set(gp_df["wf_step"].tolist())
        expected_steps = set(range(1, n + 1))

        assert actual_steps == expected_steps, (
            f"wf_step set is {sorted(actual_steps)}, expected {sorted(expected_steps)}. "
            f"Missing: {sorted(expected_steps - actual_steps)}, "
            f"extra: {sorted(actual_steps - expected_steps)}."
        )

    def test_wf_step_completeness_row_count(self, tmp_path):
        """Exactly N rows per grid_point_id — no step dropped, no step duplicated."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()
        gp = _make_grid_point()

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )
        n = len(wf_slices)

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })

        oos_results = {
            gp.grid_point_id: run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )
        }

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=n,
            grid_points=[gp],
        )

        gp_df = df[df["grid_point_id"] == gp.grid_point_id]
        assert len(gp_df) == n, (
            f"Expected {n} rows for grid_point_id={gp.grid_point_id}, got {len(gp_df)}."
        )

    def test_wf_step_multiple_grid_points(self, tmp_path):
        """Multiple grid points: each has wf_step in {1..N}, no cross-contamination."""
        cfg = _make_config(tmp_path)
        data = _make_trending_data()

        gp1 = _make_grid_point(atr=5, mult=2.0)
        gp2 = _make_grid_point(atr=10, mult=3.0)
        grid_points = [gp1, gp2]

        wf_slices = make_walk_forward_slices(
            data.index, "50bars", "40bars", scheme="rolling",
            min_train_bars=50, min_test_bars=10,
        )
        n = len(wf_slices)

        prepend = calculate_warmup(len(data), {
            "optimization": {"atr_period_range": [5, 10]},
            "validation": {"warmup_period": 0},
        })

        oos_results = {}
        for gp in grid_points:
            oos_results[gp.grid_point_id] = run_wf_for_grid_point(
                grid_point=gp,
                wf_slices=wf_slices,
                full_data=data,
                config=cfg,
                prepend_bars_requested=prepend,
            )

        df = collect_oos_steps(
            oos_results, cfg,
            expected_n_steps=n,
            grid_points=grid_points,
        )

        expected_steps = set(range(1, n + 1))

        for gp in grid_points:
            gp_df = df[df["grid_point_id"] == gp.grid_point_id]
            actual = set(gp_df["wf_step"].tolist())
            assert actual == expected_steps, (
                f"grid_point_id={gp.grid_point_id}: "
                f"wf_step set {sorted(actual)} != expected {sorted(expected_steps)}"
            )

        # Global: no zeros
        assert (df["wf_step"] == 0).sum() == 0, "Found wf_step == 0 in multi-gp run."


# ===========================================================================
# FIX-5.2b — Export-layer test: sheet naming and n_segments
# ===========================================================================
#
# Purpose: verify that 1-based wf_step values correctly drive:
#   1. Sheet names WF_01..WF_N (not WF_00..WF_{N-1})
#   2. summary_wide["n_segments"] == N
#   3. Each WF_0X sheet contains only rows for that step
#   4. All expected sheets are present; WF_00 is absent
#
# This test does NOT run the real pipeline runner. It builds synthetic
# DataFrames with wf_step=[1,2,3] directly and calls export_workbook,
# isolating the export layer from the execution layer.
# ---------------------------------------------------------------------------

_N_STEPS = 3
_GP_IDS = ("atr5_m2.00_both", "atr10_m3.00_long")


def _export_cfg() -> GridConfig:
    return GridConfig(
        data=DataConfig(file_path="dummy.csv"),
        ranking=RankingConfig(mode="gates_score"),
    )


def _make_export_step_row(gp_id: str, wf_step: int) -> dict:
    """One row for step_oos_long with minimal required fields."""
    return {
        "grid_point_id": gp_id,
        "atr_period": int(gp_id.split("_")[0].replace("atr", "")),
        "multiplier": float(gp_id.split("_")[1].replace("m", "")),
        "trade_mode": gp_id.split("_")[2],
        "wf_step": wf_step,
        "test_start_idx": (wf_step - 1) * 40,
        "test_end_idx": wf_step * 40,
        "step_status": StepStatus.OK.value,
        "sum_pnl_pct": 3.0 + wf_step,
        "sharpe": 1.0,
        "sortino": 1.2,
        "max_drawdown": -0.10,
        "cagr": 0.12,
        "win_rate": 0.55,
        "num_trades": 8,
        "profit_factor": 1.5,
        "avg_trade": 0.4,
        "prepend_bars_applied": 20,
        "effective_oos_bars": 39,
        "used_prepend": True,
        "used_legacy_oos_path": False,
        "used_defensive_fallback": False,
        "oos_boundary_index": 20,
        "warmup_used": 0,
        "warmup_effective": 0,
        "prepend_bars_requested": 20,
        "error_message": None,
        "error_type": None,
    }


def _make_export_agg_row(gp_id: str) -> dict:
    """Minimal aggregated row for one grid point (N=3 steps all ok)."""
    return {
        "grid_point_id": gp_id,
        "n_ok_steps": _N_STEPS,
        "n_total_steps": _N_STEPS,
        "ok_ratio": 1.0,
        "sum_pnl_pct_Mean": 5.0,
        "sum_pnl_pct_Median": 5.0,
        "sum_pnl_pct_Std": 1.0,
        "sum_pnl_pct_Min": 4.0,
        "sum_pnl_pct_Max": 6.0,
        "num_trades_Median": 8.0,
        "max_drawdown_Min": -0.10,
        "profit_factor_Median": 1.5,
        "sharpe_Median": 1.0,
        "sortino_Median": 1.2,
        "cagr_Median": 0.12,
        "win_rate_Median": 0.55,
        "avg_trade_Median": 0.4,
        "profitable_segments_count": _N_STEPS,
        "total_oos_trades": _N_STEPS * 8,
        "has_defensive_fallback_steps": False,
        "abs_max_drawdown_Min": 0.10,
    }


def _make_export_ranked_row(gp_id: str, grid_rank: int) -> dict:
    """Minimal ranked row (Tier 1, all gates passed)."""
    return {
        "grid_rank": grid_rank,
        "grid_point_id": gp_id,
        "tier": 1,
        "n_ok_steps": _N_STEPS,
        "n_total_steps": _N_STEPS,
        "ok_ratio": 1.0,
        "seed_gate_passed": True,
        "tester_seed_score": 0.9 - grid_rank * 0.1,
        "score_contract_status": "ok",
        "score_discrimination_status": "ok",
        "gate_ok_positive_median": True,
        "gate_ok_min_trades": True,
        "gate_ok_worst_segment": True,
        "gate_ok_drawdown": True,
        "gate_ok_min_total_trades": True,
        "seed_gate_fail_reason": "",
        "sum_pnl_pct_Median": 5.0,
        "sum_pnl_pct_Min": 4.0,
        "sum_pnl_pct_Std": 1.0,
        "max_drawdown_Min": -0.10,
        "num_trades_Median": 8.0,
        "profit_factor_Median": 1.5,
        "sharpe_Median": 1.0,
        "sortino_Median": 1.2,
        "cagr_Median": 0.12,
        "win_rate_Median": 0.55,
        "avg_trade_Median": 0.4,
        "profitable_segments_count": _N_STEPS,
        "total_oos_trades": _N_STEPS * 8,
        "has_defensive_fallback_steps": False,
        "abs_max_drawdown_Min": 0.10,
    }


def _build_export_inputs():
    """Build (step_oos_long, aggregated, ranked, summary_wide, cfg) for export tests."""
    cfg = _export_cfg()

    step_rows = []
    for gp_id in _GP_IDS:
        for s in range(1, _N_STEPS + 1):
            step_rows.append(_make_export_step_row(gp_id, s))
    step_oos_long = pd.DataFrame(step_rows)

    agg_rows = [_make_export_agg_row(gp_id) for gp_id in _GP_IDS]
    aggregated = pd.DataFrame(agg_rows)

    ranked_rows = [
        _make_export_ranked_row(gp_id, i + 1) for i, gp_id in enumerate(_GP_IDS)
    ]
    ranked = pd.DataFrame(ranked_rows)

    summary_wide = build_summary_wide(step_oos_long, aggregated, ranked, cfg)

    return step_oos_long, aggregated, ranked, summary_wide, cfg


class TestWFStepExportLayerIndexing:
    """Export-layer tests: 1-based wf_step → correct XLSX sheet names and n_segments."""

    def test_expected_sheets_present(self, tmp_path):
        """All mandatory sheets are present in the exported workbook."""
        import openpyxl

        step_oos_long, _, _, summary_wide, cfg = _build_export_inputs()

        out = tmp_path / "export_test.xlsx"
        export_workbook(
            summary_wide=summary_wide,
            step_oos_long=step_oos_long,
            trades_oos=pd.DataFrame(),
            trades_train=pd.DataFrame(),
            config=cfg,
            output_path=out,
        )

        wb = openpyxl.load_workbook(out, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        expected = {"DISCLAIMER", "WF_Config", "WF_01", "WF_02", "WF_03",
                    "WF_Trades", "WF_Train_Trades", "summary"}
        missing = expected - set(sheets)
        assert not missing, f"Sheets missing from workbook: {sorted(missing)}"

    def test_no_wf_00_sheet(self, tmp_path):
        """WF_00 must NOT exist — would indicate 0-indexed leakage (pre-FIX-5.1)."""
        import openpyxl

        step_oos_long, _, _, summary_wide, cfg = _build_export_inputs()

        out = tmp_path / "export_test.xlsx"
        export_workbook(
            summary_wide=summary_wide,
            step_oos_long=step_oos_long,
            trades_oos=pd.DataFrame(),
            trades_train=pd.DataFrame(),
            config=cfg,
            output_path=out,
        )

        wb = openpyxl.load_workbook(out, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        assert "WF_00" not in sheets, (
            "Sheet WF_00 found in workbook. This indicates wf_step=0 in "
            "step_oos_long, which means FIX-5.1 (+1 normalisation) was not applied."
        )

    def test_disclaimer_is_first_sheet(self, tmp_path):
        """DISCLAIMER must be the first sheet (§2.7 sheet order)."""
        import openpyxl

        step_oos_long, _, _, summary_wide, cfg = _build_export_inputs()

        out = tmp_path / "export_test.xlsx"
        export_workbook(
            summary_wide=summary_wide,
            step_oos_long=step_oos_long,
            trades_oos=pd.DataFrame(),
            trades_train=pd.DataFrame(),
            config=cfg,
            output_path=out,
        )

        wb = openpyxl.load_workbook(out, read_only=True)
        first = wb.sheetnames[0]
        wb.close()

        assert first == "DISCLAIMER", (
            f"First sheet is '{first}', expected 'DISCLAIMER'."
        )

    def test_n_segments_equals_n_steps(self, tmp_path):
        """summary_wide['n_segments'] must equal N (3), not N-1 (pre-FIX-5.1 bug)."""
        _, _, _, summary_wide, _ = _build_export_inputs()

        assert "n_segments" in summary_wide.columns, (
            "n_segments column missing from summary_wide."
        )

        actual = summary_wide["n_segments"].unique()
        assert len(actual) == 1, f"n_segments has multiple values: {actual}"
        assert actual[0] == _N_STEPS, (
            f"n_segments == {actual[0]}, expected {_N_STEPS}. "
            "Pre-FIX-5.1 code produced N-1 because wf_step started at 0 "
            "and max(wf_step) was N-1."
        )

    def test_wf_step_sheets_contain_correct_rows(self, tmp_path):
        """Each WF_0X sheet contains exactly the rows for that step number."""
        step_oos_long, _, _, summary_wide, cfg = _build_export_inputs()

        out = tmp_path / "export_test.xlsx"
        export_workbook(
            summary_wide=summary_wide,
            step_oos_long=step_oos_long,
            trades_oos=pd.DataFrame(),
            trades_train=pd.DataFrame(),
            config=cfg,
            output_path=out,
        )

        for step_num in range(1, _N_STEPS + 1):
            sheet_name = f"WF_{step_num:02d}"
            df = pd.read_excel(out, sheet_name=sheet_name)

            assert len(df) > 0, (
                f"Sheet {sheet_name} is empty — step {step_num} data lost."
            )
            assert len(df) == len(_GP_IDS), (
                f"Sheet {sheet_name}: expected {len(_GP_IDS)} rows (one per grid point), "
                f"got {len(df)}."
            )

    def test_wf_step_sheets_count(self, tmp_path):
        """Exactly N WF step sheets exist (WF_01..WF_03), not N-1 or N+1."""
        import openpyxl

        step_oos_long, _, _, summary_wide, cfg = _build_export_inputs()

        out = tmp_path / "export_test.xlsx"
        export_workbook(
            summary_wide=summary_wide,
            step_oos_long=step_oos_long,
            trades_oos=pd.DataFrame(),
            trades_train=pd.DataFrame(),
            config=cfg,
            output_path=out,
        )

        wb = openpyxl.load_workbook(out, read_only=True)
        wf_sheets = [s for s in wb.sheetnames if s.startswith("WF_") and s[3:].isdigit()]
        wb.close()

        assert len(wf_sheets) == _N_STEPS, (
            f"Expected {_N_STEPS} WF step sheets, found {len(wf_sheets)}: {wf_sheets}. "
            "Pre-FIX-5.1 code would produce N-1 sheets because WF_00 was generated "
            "and WF_N was missing (off-by-one in range(1, max+1))."
        )
