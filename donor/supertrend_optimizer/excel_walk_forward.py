"""
Excel export for Walk-Forward optimization results.

This module exports WalkForwardResult to Excel format with multiple sheets.
"""

import logging
import re
from typing import Any, Optional
import pandas as pd
import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

from supertrend_optimizer.validation.walk_forward import (
    WalkForwardResult,
    WFStepResult,
    MiniGridState,
    GridEvaluationCounters
)
from supertrend_optimizer.scoring.aggregation import (
    format_atr_range as _format_atr_range,
    format_mult_range as _format_mult_range,
    format_bucket_label as _format_bucket_label,
)
from supertrend_optimizer.scoring.diversification import (
    _build_full_bucket_step_matrix,
    _build_median_bucket_step_matrix,
    _build_surface_bucket_step_matrix,
)
from supertrend_optimizer.utils.constants import (
    DEFAULT_ATR_BUCKET_STEP,
    DEFAULT_MULT_BUCKET_STEP,
)
from supertrend_optimizer.io.excel_format_helpers import (
    format_excel_export_df,
    apply_openpyxl_column_formats,
    apply_gate_status_fills,
)


def _raw_dd_to_pct(value, is_fraction: bool = True) -> float | None:
    """
    Convert a raw max_drawdown value to a positive percentage.

    calculate_max_drawdown() returns a negative fraction (e.g. -0.01394 for
    a 1.394% drawdown).  This helper normalises it to a positive percent for
    display purposes *without* modifying the original column.

    F-25: explicit is_fraction parameter replaces the > 1.0 heuristic.

    Rules
    -----
    - None / NaN / INVALID_METRIC_VALUE (-999.0) -> None
    - is_fraction=True  (default): abs(value) * 100.0
    - is_fraction=False: abs(value)  (already a percentage)
    """
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(f) or f == -999.0:
        return None
    abs_f = abs(f)
    return abs_f * 100.0 if is_fraction else abs_f


def _flatten_config_to_rows(cfg: dict | None, prefix: str = "") -> list[dict]:
    """Flatten nested config dict to flat rows (Section, Parameter, Value)."""
    rows = []
    for k, v in (cfg or {}).items():
        key = f"{prefix}.{k}" if prefix else str(k)
        if isinstance(v, dict):
            rows.extend(_flatten_config_to_rows(v, key))
        else:
            if v is None:
                val = "null"
            elif isinstance(v, (list, tuple)):
                val = str(list(v))
            else:
                val = str(v)
            if "." in key:
                section, param = key.split(".", 1)
            else:
                section, param = "root", key
            rows.append({"Section": section, "Parameter": param, "Value": val})
    return rows


def _write_config_sheet(writer: pd.ExcelWriter, config: dict | None) -> None:
    """Write WF_Config sheet with full flattened config."""
    rows = _flatten_config_to_rows(config)
    df = pd.DataFrame(rows, columns=["Section", "Parameter", "Value"])
    if not df.empty:
        df = df.sort_values(["Section", "Parameter"], kind="mergesort")
    df.to_excel(writer, sheet_name="WF_Config", index=False)


def _compute_worst_step_summary(result: WalkForwardResult) -> dict:
    """
    Compute worst step metrics across all WF steps.
    
    Returns dictionary with:
    - worst_step_index_sortino, worst_step_sortino
    - worst_step_index_sum_pnl, worst_step_sum_pnl_pct
    - worst_step_index_max_dd, worst_step_max_dd_pct
    
    Rules:
    - Uses step.test_metrics values
    - Ignores missing/INVALID_METRIC_VALUE/NaN
    - max_dd: compares by absolute DD in percent (abs(dd * 100))
    - Returns NaN if no valid data
    """
    worst_sortino = None
    worst_sortino_idx = np.nan
    
    worst_sum_pnl = None
    worst_sum_pnl_idx = np.nan
    
    worst_max_dd_pct = None
    worst_max_dd_idx = np.nan
    
    for step in result.steps:
        if step.test_metrics is None:
            continue
        
        # Check sortino (find minimum)
        sortino = step.test_metrics.get("sortino")
        if sortino is not None and sortino != -999.0 and not np.isnan(sortino):
            if worst_sortino is None or sortino < worst_sortino:
                worst_sortino = sortino
                worst_sortino_idx = step.step_index
        
        # Check sum_pnl_pct (find minimum)
        sum_pnl = step.test_metrics.get("sum_pnl_pct")
        if sum_pnl is not None and sum_pnl != -999.0 and not np.isnan(sum_pnl):
            if worst_sum_pnl is None or sum_pnl < worst_sum_pnl:
                worst_sum_pnl = sum_pnl
                worst_sum_pnl_idx = step.step_index
        
        # Check max_drawdown (find maximum absolute %)
        max_dd = step.test_metrics.get("max_drawdown")
        if max_dd is not None and max_dd != -999.0 and not np.isnan(max_dd):
            max_dd_pct = abs(max_dd * 100)
            if worst_max_dd_pct is None or max_dd_pct > worst_max_dd_pct:
                worst_max_dd_pct = max_dd_pct
                worst_max_dd_idx = step.step_index
    
    return {
        "worst_step_index_sortino": worst_sortino_idx,
        "worst_step_sortino": worst_sortino if worst_sortino is not None else np.nan,
        "worst_step_index_sum_pnl": worst_sum_pnl_idx,
        "worst_step_sum_pnl_pct": worst_sum_pnl if worst_sum_pnl is not None else np.nan,
        "worst_step_index_max_dd": worst_max_dd_idx,
        "worst_step_max_dd_pct": worst_max_dd_pct if worst_max_dd_pct is not None else np.nan
    }


def _write_step_sheet(step: WFStepResult, writer: pd.ExcelWriter, sheet_name: str) -> None:
    """
    Write TOP-K entries for a single WF step into a dedicated sheet.
    
    Creates a sheet (e.g., WF_01, WF_02) with TOP-K parameters and their
    train + OOS metrics.
    
    Parameters
    ----------
    step : WFStepResult
        Walk-Forward step result containing top_entries
    writer : pd.ExcelWriter
        Excel writer object
    sheet_name : str
        Name of the sheet (e.g., "WF_01")
    """
    # Define columns in strict order
    # entry_train_objective = per-entry train objective (from trial for that atr_period, multiplier)
    # step_best_train_objective = step-level best (same for all rows in step)
    columns = [
        "rank",
        "atr_period",
        "multiplier",
        "entry_train_objective",
        "step_best_train_objective",
        "train_sum_pnl_pct",
        "robust_score",
        "oos_sortino",
        "oos_sum_pnl_pct",
        "worst_oos_max_dd_pct",
        "oos_num_trades",
        "in_consensus"
    ]
    
    # Convert top_entries to DataFrame
    if len(step.top_entries) == 0:
        # Empty step - create DataFrame with headers only
        df = pd.DataFrame(columns=columns)
    else:
        step_best = _to_excel_safe(step.train_objective_value)
        rows = []
        for entry in step.top_entries:
            rows.append({
                "rank": entry.rank,
                "atr_period": entry.atr_period,
                "multiplier": entry.multiplier,
                "entry_train_objective": _to_excel_safe(entry.train_objective),
                "step_best_train_objective": step_best,
                "train_sum_pnl_pct": entry.train_sum_pnl_pct,
                "robust_score": entry.robust_score,
                "oos_sortino": entry.oos_sortino,
                "oos_sum_pnl_pct": entry.oos_sum_pnl_pct,
                "worst_oos_max_dd_pct": entry.oos_max_dd_pct,
                "oos_num_trades": entry.oos_num_trades,
                "in_consensus": entry.in_consensus
            })
        df = pd.DataFrame(rows, columns=columns)
    
    # Write to Excel
    format_excel_export_df(df).to_excel(writer, sheet_name=sheet_name, index=False)


def _build_aggregated_topk_table(
    result: WalkForwardResult,
    config: dict | None = None,
    _return_step_df: bool = False,
) -> "pd.DataFrame | tuple[pd.DataFrame, pd.DataFrame]":
    """Build aggregated TOP-K table (delegates to scoring.aggregation)."""
    from supertrend_optimizer.scoring.aggregation import build_aggregated_topk_table
    return build_aggregated_topk_table(result, config=config, _return_step_df=_return_step_df)


def _write_consensus_sheet(consensus_df: Optional[pd.DataFrame], writer: pd.ExcelWriter) -> None:
    """
    Write consensus DataFrame into 'WF_Consensus' sheet.
    
    Only writes the ready DataFrame - no computation here.
    
    Parameters
    ----------
    consensus_df : Optional[pd.DataFrame]
        Consensus DataFrame with columns: atr_period, multiplier, coverage_count, coverage_ratio
    writer : pd.ExcelWriter
        Excel writer object
    """
    # Define expected columns
    columns = ["atr_period", "multiplier", "coverage_count", "coverage_ratio"]
    
    # Handle None or empty DataFrame
    if consensus_df is None or len(consensus_df) == 0:
        df = pd.DataFrame(columns=columns)
    else:
        df = consensus_df.copy()
    
    # Write to Excel
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Consensus", index=False)


def _get_consensus_param_no1(result: WalkForwardResult) -> tuple[Optional[int], Optional[float]]:
    """
    Get consensus parameter #1 (atr_period, multiplier).
    
    Logic: use selected_params if available, else first row of consensus_df.
    Same logic as _compute_selected_params for final consensus choice.
    """
    sel_atr = getattr(result, "selected_atr_period", None)
    sel_mult = getattr(result, "selected_multiplier", None)
    if sel_atr is not None and sel_mult is not None:
        return int(sel_atr), float(sel_mult)
    if result.consensus_df is not None and len(result.consensus_df) > 0:
        row = result.consensus_df.iloc[0]
        return int(row["atr_period"]), float(row["multiplier"])
    return None, None


def _build_consensus_details_df(result: WalkForwardResult) -> pd.DataFrame:
    """
    Build WF_Consensus_Details DataFrame.
    
    One row per WF step: OOS and TRAIN metrics for consensus parameter #1 when
    it appears in that step's top_entries. If consensus param not in step → row with NaN.
    
    Columns: wf_step, atr_period, multiplier, coverage_count, coverage_ratio,
    oos_sum_pnl_pct, oos_sortino, oos_max_dd_pct, oos_num_trades,
    train_sum_pnl_pct, train_sortino, train_num_trades, train_sharpe, train_cagr,
    robust_score, in_consensus
    """
    from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE
    
    columns = [
        "wf_step", "atr_period", "multiplier", "coverage_count", "coverage_ratio",
        "oos_sum_pnl_pct", "oos_sortino", "worst_oos_max_dd_pct", "oos_num_trades",
        "train_sum_pnl_pct", "train_sortino", "train_num_trades", "train_sharpe", "train_cagr",
        "robust_score", "in_consensus"
    ]
    
    consensus_atr, consensus_mult = _get_consensus_param_no1(result)
    if consensus_atr is None or consensus_mult is None:
        return pd.DataFrame(columns=columns)
    
    coverage_count = np.nan
    coverage_ratio = np.nan
    if result.consensus_df is not None and len(result.consensus_df) > 0:
        match = result.consensus_df[
            (result.consensus_df["atr_period"] == consensus_atr) &
            (np.abs(result.consensus_df["multiplier"] - consensus_mult) < 1e-6)
        ]
        if not match.empty:
            coverage_count = int(match.iloc[0]["coverage_count"])
            coverage_ratio = float(match.iloc[0]["coverage_ratio"])
    
    rows = []
    for step_index, step in enumerate(result.steps, start=1):
        wf_step = f"WF_{step_index:02d}"
        
        entry = None
        for e in step.top_entries:
            if e.atr_period == consensus_atr and _mult_eq(e.multiplier, consensus_mult):
                entry = e
                break
        
        if entry is not None:
            oos_sum_pnl_pct = _to_excel_safe(entry.oos_sum_pnl_pct)
            if oos_sum_pnl_pct is not None and oos_sum_pnl_pct == INVALID_METRIC_VALUE:
                oos_sum_pnl_pct = None
            oos_sortino = _to_excel_safe(entry.oos_sortino)
            if oos_sortino is not None and oos_sortino == INVALID_METRIC_VALUE:
                oos_sortino = None
            worst_oos_max_dd_pct = _to_excel_safe(entry.oos_max_dd_pct)
            if worst_oos_max_dd_pct is not None and worst_oos_max_dd_pct == INVALID_METRIC_VALUE:
                worst_oos_max_dd_pct = None
            oos_num_trades = _to_excel_safe(entry.oos_num_trades)
            if oos_num_trades is not None and oos_num_trades == INVALID_METRIC_VALUE:
                oos_num_trades = None
            train_sum_pnl_pct = _to_excel_safe(getattr(entry, "train_sum_pnl_pct", None))
            if train_sum_pnl_pct is not None and train_sum_pnl_pct == INVALID_METRIC_VALUE:
                train_sum_pnl_pct = None
            train_sortino = _to_excel_safe(getattr(entry, "train_sortino", None))
            if train_sortino is not None and train_sortino == INVALID_METRIC_VALUE:
                train_sortino = None
            train_num_trades = _to_excel_safe(getattr(entry, "train_num_trades", None))
            if train_num_trades is not None and train_num_trades == INVALID_METRIC_VALUE:
                train_num_trades = None
            train_sharpe = _to_excel_safe(getattr(entry, "train_sharpe", None))
            if train_sharpe is not None and train_sharpe == INVALID_METRIC_VALUE:
                train_sharpe = None
            train_cagr = _to_excel_safe(getattr(entry, "train_cagr", None))
            if train_cagr is not None and train_cagr == INVALID_METRIC_VALUE:
                train_cagr = None
            robust_score = _to_excel_safe(entry.robust_score)
            if robust_score is not None and robust_score == INVALID_METRIC_VALUE:
                robust_score = None
            rows.append({
                "wf_step": wf_step,
                "atr_period": consensus_atr,
                "multiplier": consensus_mult,
                "coverage_count": coverage_count,
                "coverage_ratio": coverage_ratio,
                "oos_sum_pnl_pct": oos_sum_pnl_pct,
                "oos_sortino": oos_sortino,
                "worst_oos_max_dd_pct": worst_oos_max_dd_pct,
                "oos_num_trades": oos_num_trades,
                "train_sum_pnl_pct": train_sum_pnl_pct,
                "train_sortino": train_sortino,
                "train_num_trades": train_num_trades,
                "train_sharpe": train_sharpe,
                "train_cagr": train_cagr,
                "robust_score": robust_score,
                "in_consensus": entry.in_consensus,
            })
        else:
            rows.append({
                "wf_step": wf_step,
                "atr_period": consensus_atr,
                "multiplier": consensus_mult,
                "coverage_count": coverage_count,
                "coverage_ratio": coverage_ratio,
                "oos_sum_pnl_pct": np.nan,
                "oos_sortino": np.nan,
                "worst_oos_max_dd_pct": np.nan,
                "oos_num_trades": np.nan,
                "train_sum_pnl_pct": np.nan,
                "train_sortino": np.nan,
                "train_num_trades": np.nan,
                "train_sharpe": np.nan,
                "train_cagr": np.nan,
                "robust_score": np.nan,
                "in_consensus": False,
            })
    
    df = pd.DataFrame(rows, columns=columns)
    df["wf_step"] = df["wf_step"].astype(str)
    # worst_oos_max_dd_pct is already normalized to positive percent in
    # walk_forward._fill_oos_metrics() via _normalize_dd_to_pct().
    # Do NOT re-normalize here: re-applying abs()<=1.0 -> *100 would corrupt
    # values smaller than 1% (e.g. 0.5% would become 50%).
    return df


def _write_consensus_details_sheet(result: WalkForwardResult, writer: pd.ExcelWriter) -> None:
    """
    Write WF_Consensus_Details sheet.
    
    One row per WF step: OOS metrics for consensus parameter #1 across all steps.
    If consensus is empty → empty sheet with headers only.
    """
    df = _build_consensus_details_df(result)
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Consensus_Details", index=False)


def _write_aggregated_topk_sheet(result: WalkForwardResult, writer: pd.ExcelWriter, config: dict | None = None) -> None:
    """
    Write aggregated TOP-K table into 'WF_TOP_K_Aggregated' sheet.
    
    Builds and writes a deduplicated TOP-K table that combines entries
    from all WF steps, grouped by (atr_period, multiplier).
    
    Parameters
    ----------
    result : WalkForwardResult
        Walk-Forward result with steps containing top_entries
    writer : pd.ExcelWriter
        Excel writer object
    config : dict | None
        Configuration dictionary containing consensus settings
    """
    agg_df = _build_aggregated_topk_table(result, config)
    format_excel_export_df(agg_df).to_excel(writer, sheet_name="WF_TOP_K_Aggregated", index=False)


def _cfg_gate_threshold(cfg: dict, key: str, default: float) -> "float | None":
    """Read a canonical gate threshold from a config dict with null-as-disabled semantics."""
    from supertrend_optimizer.scoring.canonical import cfg_gate_threshold
    return cfg_gate_threshold(cfg, key, default)


def _apply_tradeability_gates(
    df: pd.DataFrame,
    config: dict,
) -> pd.DataFrame:
    """Evaluate tradeability gate conditions (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import apply_tradeability_gates
    return apply_tradeability_gates(df, config)


def _canonical_gates_status(scored_df: pd.DataFrame, config: dict) -> dict:
    """Return canonical gate evaluation summary (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import canonical_gates_status
    return canonical_gates_status(scored_df, config)


def _calculate_trade_score(
    df: pd.DataFrame,
    config: dict,
) -> pd.DataFrame:
    """Compute trade_score from OOS metrics (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import calculate_trade_score
    return calculate_trade_score(df, config)


def _calculate_final_score(
    df: pd.DataFrame,
    config: dict,
) -> pd.DataFrame:
    """Compute final_score from structure + trade (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import calculate_final_score
    return calculate_final_score(df, config)


def _compute_coverage_mix(
    df: pd.DataFrame,
    canonical_cfg: dict,
) -> "pd.Series":
    """Compute coverage_mix blend (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import compute_coverage_mix
    return compute_coverage_mix(df, canonical_cfg)


def _calculate_canonical_rank_score(
    agg_df: pd.DataFrame,
    atr_bucket_step: int,
    mult_bucket_step: float,
    config: dict | None = None,
    total_steps: int | None = None,
) -> pd.DataFrame:
    """Calculate canonical rank score (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import calculate_canonical_rank_score
    return calculate_canonical_rank_score(
        agg_df, atr_bucket_step, mult_bucket_step,
        config=config, total_steps=total_steps,
    )


def _compute_regime_penalties(
    *,
    mu: float,
    nu: float,
    total_steps: int,
    K: "pd.Series",
    present: "pd.Series",
    valid_sort: "pd.Series",
) -> "tuple[pd.Series, pd.Series]":
    """Compute regime penalties (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import compute_regime_penalties
    return compute_regime_penalties(
        mu=mu, nu=nu, total_steps=total_steps,
        K=K, present=present, valid_sort=valid_sort,
    )


def _compute_regime_score(
    df: pd.DataFrame,
    config: dict,
    total_steps: int | None,
) -> "pd.Series":
    """Compute regime_robust score (delegates to scoring.canonical)."""
    from supertrend_optimizer.scoring.canonical import compute_regime_score
    return compute_regime_score(df, config=config, total_steps=total_steps)


def _enrich_soft_gate_columns(
    agg_df: pd.DataFrame,
    step_df: pd.DataFrame,
    config: dict | None,
) -> pd.DataFrame:
    """
    Compute per-parameter soft PnL gate statistics from step-level data and
    merge them onto agg_df as new columns.

    New columns added to agg_df:
        oos_sum_pnl_pct_violation_count  — number of WF steps where
            oos_sum_pnl_pct < threshold (int, NaN when config absent)
        oos_sum_pnl_pct_p25             — 25th-percentile of per-step
            oos_sum_pnl_pct values (float, NaN when config absent)

    When soft_pnl_gate_enabled is absent/false, both columns are filled
    with NaN so that apply_tradeability_gates falls back to the hard gate.

    The function never modifies agg_df in-place; it returns a new DataFrame.
    """
    import numpy as np

    agg_df = agg_df.copy()
    agg_df["oos_sum_pnl_pct_violation_count"] = pd.array([np.nan] * len(agg_df), dtype="Float64")
    agg_df["oos_sum_pnl_pct_p25"] = pd.array([np.nan] * len(agg_df), dtype="Float64")

    cfg = config or {}
    gates_cfg = cfg.get("walk_forward", {}).get("canonical", {}).get("gates", {})
    if not bool(gates_cfg.get("soft_pnl_gate_enabled", False)):
        return agg_df

    threshold = gates_cfg.get("min_oos_sum_pnl_pct_min", None)
    if threshold is None:
        return agg_df

    if step_df is None or step_df.empty:
        return agg_df

    if "oos_sum_pnl_pct" not in step_df.columns:
        return agg_df

    group_cols = [c for c in ["atr_period", "multiplier"] if c in step_df.columns]
    if not group_cols:
        return agg_df

    pnl_col = pd.to_numeric(step_df["oos_sum_pnl_pct"], errors="coerce")
    step_work = step_df[group_cols].copy()
    step_work["_pnl"] = pnl_col

    def _violation_count(s):
        return int((s < threshold).sum())

    def _p25(s):
        return float(np.nanpercentile(s.dropna().values, 25)) if s.notna().any() else np.nan

    grp = step_work.groupby(group_cols)["_pnl"]
    violation_counts = grp.apply(_violation_count).reset_index()
    violation_counts.columns = group_cols + ["oos_sum_pnl_pct_violation_count"]

    p25_vals = grp.apply(_p25).reset_index()
    p25_vals.columns = group_cols + ["oos_sum_pnl_pct_p25"]

    stats = pd.merge(violation_counts, p25_vals, on=group_cols, how="outer")

    agg_df = agg_df.drop(columns=["oos_sum_pnl_pct_violation_count", "oos_sum_pnl_pct_p25"])
    agg_df = pd.merge(agg_df, stats, on=group_cols, how="left")

    # Ensure Float64 dtype (nullable) for the new columns to avoid object dtype on empty df
    for col in ("oos_sum_pnl_pct_violation_count", "oos_sum_pnl_pct_p25"):
        if col in agg_df.columns:
            agg_df[col] = pd.to_numeric(agg_df[col], errors="coerce").astype("Float64")

    return agg_df


def _write_wf_topk_canonical_sheet(
    agg_df: pd.DataFrame,
    writer: pd.ExcelWriter,
    config: dict | None = None,
    total_steps: int | None = None,
) -> pd.DataFrame:
    """
    Build and write the WF_TOP_K_Canonical sheet.

    Takes the same aggregated TOP-K dataframe used for WF_TOP_K_Aggregated
    (does NOT modify the original), calculates canonical rank scores, sorts,
    and writes to a new sheet.

    Sorting (CRITICAL FIX B):
        Primary  : final_score DESC  — ensures -inf rows (gate-blocked) go to bottom
        Secondary: canonical_rank_score DESC  — tie-breaker on structure
        Tertiary : robust_score DESC, coverage_count DESC, atr_period ASC, multiplier ASC

    Backward compat: in legacy mode (no balance/trade/gates config) final_score ==
    canonical_rank_score, so sort order is identical to the old behaviour.

    canonical_rank:
        Rows where tradeability_passed=True  → 1..N (ascending from best)
        Rows where tradeability_passed=False → NaN  (shown at bottom of sheet)

    Column order:
        canonical_rank, final_score, regime_score, canonical_rank_score,
        structure_score, trade_score, gate_exclusion_reason, atr_period,
        multiplier, coverage_count, rank_component_robust, atr_bucket,
        mult_bucket, coverage_count_bucket, train_sum_pnl_pct, oos_sum_pnl_pct,
        oos_sum_pnl_pct_min, worst_oos_max_dd_pct, pnl_norm, dd_norm,
        [remaining agg_df columns in original order]

    Parameters
    ----------
    agg_df : pd.DataFrame
        Aggregated TOP-K dataframe (not modified).
    writer : pd.ExcelWriter
        Active Excel writer.
    config : dict | None
        Full run config forwarded to _calculate_canonical_rank_score.

    Returns
    -------
    pd.DataFrame
        The fully scored and sorted dataframe (with canonical_rank assigned,
        metadata rows prepended).  Returned so callers can reuse it without
        recomputing canonical scoring a second time.
    """
    if config is None:
        config = {}
    consensus_cfg = config.get("walk_forward", {}).get("consensus", {})
    atr_bucket_step = int(consensus_cfg.get("atr_bucket_step", DEFAULT_ATR_BUCKET_STEP))
    mult_bucket_step = float(consensus_cfg.get("mult_bucket_step", DEFAULT_MULT_BUCKET_STEP))

    df = _calculate_canonical_rank_score(
        agg_df.copy(), atr_bucket_step, mult_bucket_step, config=config,
        total_steps=total_steps,
    )

    # Sort order depends on ranking mode.
    # regime_robust: regime_score leads; gate-excluded rows (-inf) sink naturally.
    # classic: final_score leads (CRITICAL FIX B preserved — gate-blocked rows go
    #          to bottom because their final_score == -inf).
    ranking_mode = (
        config.get("walk_forward", {})
              .get("canonical", {})
              .get("ranking", {})
              .get("mode", "classic")
    )
    if ranking_mode == "regime_robust" and "regime_score" in df.columns:
        sort_cols = [
            "regime_score",
            "coverage_count_bucket",
            "coverage_count",
            "final_score",
            "canonical_rank_score",
            "atr_period",
            "multiplier",
        ]
        sort_asc = [False, False, False, False, False, True, True]
    else:
        sort_cols = [
            "final_score",
            "canonical_rank_score",
            "robust_score",
            "coverage_count",
            "atr_period",
            "multiplier",
        ]
        sort_asc = [False, False, False, False, True, True]
    existing_sort = [(c, a) for c, a in zip(sort_cols, sort_asc) if c in df.columns]
    if existing_sort:
        df = df.sort_values(
            by=[c for c, _ in existing_sort],
            ascending=[a for _, a in existing_sort],
            kind="mergesort",
        ).reset_index(drop=True)

    # canonical_rank: 1..N only for tradeability_passed rows; NaN for blocked.
    if "tradeability_passed" in df.columns:
        passed_mask = df["tradeability_passed"].astype(bool)
    else:
        passed_mask = pd.Series(True, index=df.index)

    canonical_rank = pd.array([float("nan")] * len(df), dtype="Float64")
    rank_counter = 1
    for i in df.index:
        if passed_mask[i]:
            canonical_rank[i] = rank_counter
            rank_counter += 1
    df.insert(0, "canonical_rank", canonical_rank)

    # Add bucket_param (same format as BucketMatrix_Full: "ATR 41–43 | Mult 2.8–2.9")
    def _row_bucket_param(row):
        ab = row.get("atr_bucket")
        mb = row.get("mult_bucket")
        mt = row.get("mult_bucket_ticks")
        if pd.isna(ab) or (pd.isna(mb) and (mt is None or pd.isna(mt))):
            return ""
        if mt is None or pd.isna(mt):
            mt = int(round(float(mb) / mult_bucket_step))
        return _format_bucket_label(int(ab), int(mt), atr_bucket_step, mult_bucket_step)

    df["bucket_param"] = df.apply(_row_bucket_param, axis=1)

    # bucket_key: stable string surrogate for (atr_bucket, mult_bucket_ticks) — I-5
    def _row_bucket_key(row):
        ab = row.get("atr_bucket")
        mt = row.get("mult_bucket_ticks")
        mb = row.get("mult_bucket")
        if pd.isna(ab):
            return ""
        if mt is None or (isinstance(mt, float) and pd.isna(mt)):
            if mb is None or (isinstance(mb, float) and pd.isna(mb)):
                return ""
            mt = int(round(float(mb) / mult_bucket_step))
        return f"{int(ab)}_{int(mt)}"

    df["bucket_key"] = df.apply(_row_bucket_key, axis=1)

    # Build ordered column list (no duplicates)
    priority_cols = [
        "canonical_rank",
        "final_score",
        "regime_score",
        "canonical_rank_score",
        "structure_score",
        "trade_score",
        "gate_exclusion_reason",
        "gate_status",
        "atr_period",
        "multiplier",
        "coverage_count",
        "rank_component_robust",
        "atr_bucket",
        "mult_bucket",
        "mult_bucket_ticks",
        "bucket_key",
        "bucket_param",
        "coverage_count_bucket",
        "train_sum_pnl_pct",
        "oos_sum_pnl_pct",
        "oos_sum_pnl_pct_min",
        "worst_oos_max_dd_pct",
        "pnl_norm",
        "dd_norm",
    ]
    existing_priority = [c for c in priority_cols if c in df.columns]
    remaining = [c for c in df.columns if c not in set(existing_priority)]
    final_cols = existing_priority + remaining

    # Determine gates status for metadata header
    gates_status_dict = _canonical_gates_status(df, config)
    gates_status = gates_status_dict["status"]

    # Build gate-threshold metadata rows to prepend above the main table.
    canonical_cfg = config.get("walk_forward", {}).get("canonical", {})
    oos_gates_cfg = canonical_cfg.get("gates", None)
    tg_cfg = canonical_cfg.get("tradeability_gates", {})

    meta_rows: list[dict] = []
    # Always emit status line when any gate is active
    if oos_gates_cfg is not None or bool(tg_cfg.get("enabled", False)):
        meta_rows.append({
            "canonical_rank": "# gates_status",
            "final_score": gates_status,
        })
    if oos_gates_cfg is not None:
        meta_rows.append({
            "canonical_rank": "# gate: max_worst_oos_dd_pct",
            "final_score": _cfg_gate_threshold(oos_gates_cfg, "max_worst_oos_dd_pct", 35.0),
        })
        meta_rows.append({
            "canonical_rank": "# gate: min_oos_sortino_min",
            "final_score": _cfg_gate_threshold(oos_gates_cfg, "min_oos_sortino_min", 0.0),
        })
        meta_rows.append({
            "canonical_rank": "# gate: min_oos_sum_pnl_pct_min",
            "final_score": _cfg_gate_threshold(oos_gates_cfg, "min_oos_sum_pnl_pct_min", 0.0),
        })
        meta_rows.append({
            "canonical_rank": "# gate: min_oos_trades_median",
            "final_score": _cfg_gate_threshold(oos_gates_cfg, "min_oos_trades_median", 0.0),
        })
    if bool(tg_cfg.get("enabled", False)):
        for key in ("min_sortino", "max_dd", "min_trades"):
            val = tg_cfg.get(key)
            if val is not None:
                meta_rows.append({
                    "canonical_rank": f"# tradeability_gate: {key}",
                    "final_score": float(val),
                })

    # scored_df: clean data rows only (no metadata header rows).
    # Used by _canonical_gates_status and run_diversification_filter — both
    # need a purely numeric dataframe without the "# gates_status" header rows.
    scored_df = df[final_cols].copy()

    # Ensure mult_bucket_ticks is present — required by run_diversification_filter
    # for Table D (bucket-aggregated step performance). If missing (e.g. legacy
    # agg_df built before mult_bucket_ticks was added to aggregation.py), recompute.
    if "mult_bucket_ticks" not in scored_df.columns and "atr_period" in scored_df.columns and "multiplier" in scored_df.columns:
        scored_df = _apply_param_buckets(scored_df, atr_step=atr_bucket_step, mult_step=mult_bucket_step)

    if meta_rows:
        meta_df = pd.DataFrame(meta_rows, columns=list(scored_df.columns))
        meta_df = meta_df.dropna(axis=1, how="all")
        scored_df = scored_df.dropna(axis=1, how="all")
        out_df = pd.concat([meta_df, scored_df], ignore_index=True)
    else:
        out_df = scored_df

    formatted_out = format_excel_export_df(out_df)
    formatted_out.to_excel(writer, sheet_name="WF_TOP_K_Canonical", index=False)

    # Apply gate_status color fills (green=PASS, amber=INSUFFICIENT, red=FAIL).
    # data_start_row=2 because row 1 is the header (no banner row on this sheet).
    if "WF_TOP_K_Canonical" in writer.sheets and "gate_status" in scored_df.columns:
        n_meta = len(meta_rows)
        ws_can = writer.sheets["WF_TOP_K_Canonical"]
        # Metadata rows are prepended above the data; data starts at row n_meta+2.
        apply_gate_status_fills(
            ws_can, formatted_out, data_start_row=n_meta + 2
        )

    # Return the clean scored df (without metadata rows) so callers can reuse
    # it for _canonical_gates_status and run_diversification_filter without
    # a second call to _calculate_canonical_rank_score.
    return scored_df


def _mult_eq(a: float, b: float, tol: float = 1e-6) -> bool:
    """Compare multipliers with tolerance for float equality."""
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) < tol


def _apply_param_buckets(
    df: pd.DataFrame,
    atr_step: int,
    mult_step: float,
) -> pd.DataFrame:
    """Add atr_bucket, mult_bucket_ticks, mult_bucket columns (delegates to scoring.aggregation)."""
    from supertrend_optimizer.scoring.aggregation import _apply_param_buckets
    return _apply_param_buckets(df, atr_step=atr_step, mult_step=mult_step)


def _build_aggregated_details_df(result: WalkForwardResult, config: dict | None = None) -> pd.DataFrame:
    """
    Build WF_Aggregated_Details DataFrame.
    
    One row per WF step: OOS and TRAIN metrics for rank=1 param from WF_TOP_K_Aggregated
    when it appears in that step's top_entries. If param absent in step → row with
    atr_period, multiplier filled, metrics NaN, in_consensus=False.
    
    Columns: wf_step, atr_period, multiplier, coverage_count, coverage_ratio,
    oos_sum_pnl_pct, oos_sortino, oos_max_dd_pct, oos_num_trades,
    train_sum_pnl_pct, train_sortino, train_num_trades, train_sharpe, train_cagr,
    robust_score, in_consensus
    """
    from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE
    
    columns = [
        "wf_step", "atr_period", "multiplier", "coverage_count", "coverage_ratio",
        "oos_sum_pnl_pct", "oos_sortino", "worst_oos_max_dd_pct", "oos_num_trades",
        "train_sum_pnl_pct", "train_sortino", "train_num_trades", "train_sharpe", "train_cagr",
        "robust_score", "in_consensus"
    ]
    
    agg_df = _build_aggregated_topk_table(result, config)
    if agg_df.empty or "atr_period" not in agg_df.columns:
        return pd.DataFrame(columns=columns)
    if "rank" in agg_df.columns:
        agg_df = agg_df.sort_values("rank", ascending=True).reset_index(drop=True)
    row1 = agg_df.iloc[0]
    agg_atr = int(row1["atr_period"])
    agg_mult = float(row1["multiplier"])
    coverage_count = int(row1["coverage_count"]) if "coverage_count" in row1 else np.nan
    coverage_ratio = float(row1["coverage_ratio"]) if "coverage_ratio" in row1 else np.nan
    
    rows = []
    for step_index, step in enumerate(result.steps, start=1):
        wf_step = f"WF_{step_index:02d}"
        
        entry = None
        for e in step.top_entries:
            if e.atr_period == agg_atr and _mult_eq(e.multiplier, agg_mult):
                entry = e
                break
        
        if entry is not None:
            oos_sum_pnl_pct = _to_excel_safe(entry.oos_sum_pnl_pct)
            if oos_sum_pnl_pct is not None and oos_sum_pnl_pct == INVALID_METRIC_VALUE:
                oos_sum_pnl_pct = None
            oos_sortino = _to_excel_safe(entry.oos_sortino)
            if oos_sortino is not None and oos_sortino == INVALID_METRIC_VALUE:
                oos_sortino = None
            worst_oos_max_dd_pct = _to_excel_safe(entry.oos_max_dd_pct)
            if worst_oos_max_dd_pct is not None and worst_oos_max_dd_pct == INVALID_METRIC_VALUE:
                worst_oos_max_dd_pct = None
            oos_num_trades = _to_excel_safe(entry.oos_num_trades)
            if oos_num_trades is not None and oos_num_trades == INVALID_METRIC_VALUE:
                oos_num_trades = None
            train_sum_pnl_pct = _to_excel_safe(getattr(entry, "train_sum_pnl_pct", None))
            if train_sum_pnl_pct is not None and train_sum_pnl_pct == INVALID_METRIC_VALUE:
                train_sum_pnl_pct = None
            train_sortino = _to_excel_safe(getattr(entry, "train_sortino", None))
            if train_sortino is not None and train_sortino == INVALID_METRIC_VALUE:
                train_sortino = None
            train_num_trades = _to_excel_safe(getattr(entry, "train_num_trades", None))
            if train_num_trades is not None and train_num_trades == INVALID_METRIC_VALUE:
                train_num_trades = None
            train_sharpe = _to_excel_safe(getattr(entry, "train_sharpe", None))
            if train_sharpe is not None and train_sharpe == INVALID_METRIC_VALUE:
                train_sharpe = None
            train_cagr = _to_excel_safe(getattr(entry, "train_cagr", None))
            if train_cagr is not None and train_cagr == INVALID_METRIC_VALUE:
                train_cagr = None
            robust_score = _to_excel_safe(entry.robust_score)
            if robust_score is not None and robust_score == INVALID_METRIC_VALUE:
                robust_score = None
            rows.append({
                "wf_step": wf_step,
                "atr_period": agg_atr,
                "multiplier": agg_mult,
                "coverage_count": coverage_count,
                "coverage_ratio": coverage_ratio,
                "oos_sum_pnl_pct": oos_sum_pnl_pct,
                "oos_sortino": oos_sortino,
                "worst_oos_max_dd_pct": worst_oos_max_dd_pct,
                "oos_num_trades": oos_num_trades,
                "train_sum_pnl_pct": train_sum_pnl_pct,
                "train_sortino": train_sortino,
                "train_num_trades": train_num_trades,
                "train_sharpe": train_sharpe,
                "train_cagr": train_cagr,
                "robust_score": robust_score,
                "in_consensus": entry.in_consensus,
            })
        else:
            rows.append({
                "wf_step": wf_step,
                "atr_period": agg_atr,
                "multiplier": agg_mult,
                "coverage_count": coverage_count,
                "coverage_ratio": coverage_ratio,
                "oos_sum_pnl_pct": np.nan,
                "oos_sortino": np.nan,
                "worst_oos_max_dd_pct": np.nan,
                "oos_num_trades": np.nan,
                "train_sum_pnl_pct": np.nan,
                "train_sortino": np.nan,
                "train_num_trades": np.nan,
                "train_sharpe": np.nan,
                "train_cagr": np.nan,
                "robust_score": np.nan,
                "in_consensus": False,
            })
    
    df = pd.DataFrame(rows, columns=columns)
    df["wf_step"] = df["wf_step"].astype(str)
    # worst_oos_max_dd_pct is already normalized to positive percent in
    # walk_forward._fill_oos_metrics() via _normalize_dd_to_pct().
    # Do NOT re-normalize here: re-applying abs()<=1.0 -> *100 would corrupt
    # values smaller than 1% (e.g. 0.5% would become 50%).
    return df


def _write_aggregated_details_sheet(
    result: WalkForwardResult, writer: pd.ExcelWriter, config: dict | None = None
) -> None:
    """
    Write WF_Aggregated_Details sheet.
    
    One row per WF step for rank=1 param from WF_TOP_K_Aggregated.
    If aggregated table empty → empty sheet with headers only.
    """
    df = _build_aggregated_details_df(result, config)
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Aggregated_Details", index=False)


def _build_minigrid_results_table(result: WalkForwardResult, config: dict | None = None) -> pd.DataFrame:
    """
    Build aggregated results table for MiniGrid points across all WF steps.
    
    Uses mini_grid_trials for ALL metrics including robust_score.
    robust_score is now calculated for ALL mini-grid points during evaluation.
    Ensures exactly grid_size rows matching the MiniGrid grid points.
    
    Parameters
    ----------
    result : WalkForwardResult
        Walk-Forward result with mini_grid_state and mini_grid_trials
        
    Returns
    -------
    pd.DataFrame
        Aggregated MiniGrid results table with exactly grid_size rows (one per grid point):
        grid_index, atr_period, multiplier, rank, robust_score, robust_score_std,
        oos_sortino, oos_sortino_min, oos_sum_pnl_pct, oos_sum_pnl_pct_min,
        worst_oos_max_dd_pct, oos_num_trades_median, coverage_count, coverage_ratio
        Sorted by robust_score DESC, coverage_count DESC, oos_sortino DESC,
        atr_period ASC, multiplier ASC
    """
    from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE
    from supertrend_optimizer.imputation.fill_missing import fill_missing_step_metrics
    from supertrend_optimizer.imputation.full_matrix import ensure_full_step_candidates

    # Handle disabled or empty mini-grid
    if result.mini_grid_state is None or len(result.mini_grid_state.params) == 0:
        return pd.DataFrame(columns=[
            "grid_index", "atr_period", "multiplier", "rank", "robust_score", "robust_score_std",
            "oos_sortino", "oos_sortino_min", "oos_sum_pnl_pct", "oos_sum_pnl_pct_min",
            "worst_oos_max_dd_pct", "oos_num_trades_median", "coverage_count", "coverage_ratio"
        ])
    
    # Build canonical grid_points DataFrame
    grid_rows = []
    for idx, (atr, mult) in enumerate(result.mini_grid_state.params, start=1):
        grid_rows.append({
            "grid_index": idx,
            "atr_period": atr,
            "multiplier": round(mult, 6)  # Quantize for matching
        })
    grid_points_df = pd.DataFrame(grid_rows)
    
    # Prepare metrics from mini_grid_trials
    if result.mini_grid_trials is None or len(result.mini_grid_trials) == 0:
        # No trial data - return grid with NaN metrics
        grid_points_df["rank"] = np.arange(1, len(grid_points_df) + 1)
        grid_points_df["robust_score"] = np.nan
        grid_points_df["robust_score_std"] = np.nan
        grid_points_df["oos_sortino"] = np.nan
        grid_points_df["oos_sortino_min"] = np.nan
        grid_points_df["oos_sum_pnl_pct"] = np.nan
        grid_points_df["oos_sum_pnl_pct_min"] = np.nan
        grid_points_df["worst_oos_max_dd_pct"] = np.nan
        grid_points_df["oos_num_trades_median"] = pd.NA
        grid_points_df["coverage_count"] = 0
        grid_points_df["coverage_ratio"] = 0.0
        return grid_points_df
    
    eval_df = result.mini_grid_trials.copy()
    eval_df["multiplier"] = eval_df["multiplier"].round(6)

    # Normalize max_drawdown to positive percent BEFORE imputation so that
    # the imputation layer works with consistent positive-percent values.
    if "max_drawdown" in eval_df.columns:
        dd_values = eval_df["max_drawdown"].copy()
        valid_mask = dd_values.notna() & (dd_values != INVALID_METRIC_VALUE)
        if valid_mask.any():
            abs_values = dd_values.abs()
            eval_df.loc[valid_mask, "max_drawdown"] = np.where(
                abs_values[valid_mask] <= 1.0,
                abs_values[valid_mask] * 100.0,
                abs_values[valid_mask]
            )

    # Inner join: only keep evaluations for grid points
    eval_df = eval_df.merge(
        grid_points_df[["atr_period", "multiplier", "grid_index"]],
        on=["atr_period", "multiplier"],
        how="inner"
    )

    total_steps = len(result.steps)

    # ── Full matrix: ensure every grid candidate appears in every window ─────
    # Universe = grid_points_df (fixed MiniGrid lattice).
    # Missing rows are added with INVALID_METRIC_VALUE so that
    # fill_missing_step_metrics can impute them uniformly.
    # Rename short MiniGrid columns to canonical oos_* names first so that
    # ensure_full_step_candidates can fill the correct metric columns.
    _mg_rename_in = {
        "sortino":     "oos_sortino",
        "sum_pnl_pct": "oos_sum_pnl_pct",
        "max_drawdown": "oos_max_dd_pct",
        "num_trades":  "oos_num_trades",
    }
    _mg_rename_out = {v: k for k, v in _mg_rename_in.items()}
    eval_df = eval_df.rename(columns={k: v for k, v in _mg_rename_in.items() if k in eval_df.columns})

    eval_df = ensure_full_step_candidates(eval_df, config, universe_df=grid_points_df)

    # ── Imputation layer ─────────────────────────────────────────────────────
    # fill_missing_step_metrics operates on the full matrix (including newly
    # added missing rows) using canonical oos_* column names.
    eval_df = fill_missing_step_metrics(eval_df, config)
    eval_df = eval_df.rename(columns={k: v for k, v in _mg_rename_out.items() if k in eval_df.columns})

    # Replace INVALID_METRIC_VALUE in non-imputed column (robust_score)
    if "robust_score" in eval_df.columns:
        eval_df["robust_score"] = eval_df["robust_score"].replace(INVALID_METRIC_VALUE, np.nan)

    # Group by grid point and aggregate
    group_cols = ["grid_index", "atr_period", "multiplier"]

    # Build aggregation dictionary
    # CRITICAL: For robust_score_std, we need population std (ddof=0)
    agg_dict = {}

    if "robust_score" in eval_df.columns:
        agg_dict["robust_score"] = ["mean", lambda x: x.std(ddof=0)]

    if "sortino" in eval_df.columns:
        agg_dict["sortino"] = ["mean", "min"]
    if "sum_pnl_pct" in eval_df.columns:
        agg_dict["sum_pnl_pct"] = ["mean", "min"]
    if "max_drawdown" in eval_df.columns:
        agg_dict["max_drawdown"] = "max"
    if "num_trades" in eval_df.columns:
        agg_dict["num_trades"] = "median"

    if len(agg_dict) == 0:
        grid_points_df["rank"] = np.arange(1, len(grid_points_df) + 1)
        grid_points_df["robust_score"] = np.nan
        grid_points_df["robust_score_std"] = np.nan
        grid_points_df["oos_sortino"] = np.nan
        grid_points_df["oos_sortino_min"] = np.nan
        grid_points_df["oos_sum_pnl_pct"] = np.nan
        grid_points_df["oos_sum_pnl_pct_min"] = np.nan
        grid_points_df["worst_oos_max_dd_pct"] = np.nan
        grid_points_df["oos_num_trades_median"] = pd.NA
        grid_points_df["coverage_count"] = 0
        grid_points_df["coverage_ratio"] = 0.0
        grid_points_df["coverage_count_effective"] = 0
        grid_points_df["coverage_ratio_effective"] = 0.0
        grid_points_df["fallback_share"] = 0.0
        return grid_points_df

    # Aggregate
    agg_df = eval_df.groupby(group_cols, as_index=False).agg(agg_dict)

    # Flatten multi-level columns
    new_columns = []
    for col in agg_df.columns.values:
        if isinstance(col, tuple):
            if col[1]:
                agg_func_name = col[1] if not callable(col[1]) else "std"
                if "<lambda" in str(agg_func_name):
                    agg_func_name = "std"
                new_columns.append(f"{col[0]}_{agg_func_name}")
            else:
                new_columns.append(col[0])
        else:
            new_columns.append(col)
    agg_df.columns = new_columns

    # Rename columns to match expected output
    rename_map = {
        "robust_score_mean": "robust_score",
        "robust_score_std": "robust_score_std",
        "sortino_mean": "oos_sortino",
        "sortino_min": "oos_sortino_min",
        "sum_pnl_pct_mean": "oos_sum_pnl_pct",
        "sum_pnl_pct_min": "oos_sum_pnl_pct_min",
        "max_drawdown_max": "worst_oos_max_dd_pct",
        "num_trades_median": "oos_num_trades_median",
    }
    agg_df = agg_df.rename(columns=rename_map)

    # Add coverage_count (number of unique steps where this point was evaluated).
    # Exclude missing-rows added by ensure_full_step_candidates.
    _mg_real_mask = eval_df["is_missing_row"].eq(False) if "is_missing_row" in eval_df.columns else pd.Series(True, index=eval_df.index)
    coverage = eval_df[_mg_real_mask].groupby(group_cols)["step_id"].nunique().reset_index()
    coverage.columns = list(group_cols) + ["coverage_count"]
    agg_df = agg_df.merge(coverage, on=group_cols, how="left")
    agg_df["coverage_count"] = agg_df["coverage_count"].fillna(0).astype(int)

    # Add coverage_ratio
    agg_df["coverage_ratio"] = agg_df["coverage_count"] / total_steps

    # ── Effective coverage ───────────────────────────────────────────────────
    _src_col = "sum_pnl_pct_is_imputed"  # short name after rename-back
    _imp_src_col = "sum_pnl_pct_imputation_source"
    if _src_col in eval_df.columns:
        _eff_mask = (
            eval_df[_src_col].eq(False)
            | eval_df[_imp_src_col].eq("neighbors")
        )
    else:
        _eff_mask = pd.Series(True, index=eval_df.index)

    _eff_series = (
        eval_df[_eff_mask]
        .groupby(group_cols)["step_id"]
        .nunique()
        .reset_index()
        .rename(columns={"step_id": "coverage_count_effective"})
    )
    agg_df = agg_df.merge(_eff_series, on=group_cols, how="left")
    agg_df["coverage_count_effective"] = agg_df["coverage_count_effective"].fillna(0).astype(int)
    agg_df["coverage_ratio_effective"] = agg_df["coverage_count_effective"] / total_steps

    _mg_src_cols = [c for c in eval_df.columns if c.endswith("_imputation_source")]
    if _mg_src_cols:
        _mg_any_fallback = eval_df[_mg_src_cols].eq("fallback_worst").any(axis=1)
        _fb_series = (
            eval_df[_mg_any_fallback]
            .groupby(group_cols)["step_id"]
            .nunique()
            .reset_index()
            .rename(columns={"step_id": "_fb_count"})
        )
        agg_df = agg_df.merge(_fb_series, on=group_cols, how="left")
        agg_df["fallback_share"] = agg_df["_fb_count"].fillna(0) / total_steps
        agg_df = agg_df.drop(columns=["_fb_count"])
    else:
        agg_df["fallback_share"] = 0.0
    
    # Left join with grid_points to ensure all grid_size rows are present
    result_df = grid_points_df.merge(
        agg_df.drop(columns=["atr_period", "multiplier"], errors="ignore"),
        on="grid_index",
        how="left"
    )
    
    # Fill missing coverage values with 0
    result_df["coverage_count"] = result_df["coverage_count"].fillna(0).astype(int)
    result_df["coverage_ratio"] = result_df["coverage_ratio"].fillna(0.0)
    result_df["coverage_count_effective"] = result_df["coverage_count_effective"].fillna(0).astype(int)
    result_df["coverage_ratio_effective"] = result_df["coverage_ratio_effective"].fillna(0.0)
    result_df["fallback_share"] = result_df["fallback_share"].fillna(0.0)

    # Deterministic sorting
    sort_cols = ["robust_score", "coverage_count", "oos_sortino", "atr_period", "multiplier"]
    sort_ascending = [False, False, False, True, True]

    existing_sort_cols = [col for col in sort_cols if col in result_df.columns]
    existing_sort_ascending = [sort_ascending[i] for i, col in enumerate(sort_cols) if col in result_df.columns]

    result_df = result_df.sort_values(
        by=existing_sort_cols,
        ascending=existing_sort_ascending,
        na_position='last'
    ).reset_index(drop=True)

    # Add rank column (deterministic)
    result_df.insert(3, "rank", np.arange(1, len(result_df) + 1))

    # Stabilize dtypes for Excel export
    int_cols = ["rank", "grid_index", "atr_period", "coverage_count", "coverage_count_effective"]
    for col in int_cols:
        if col in result_df.columns:
            numeric_series = pd.to_numeric(result_df[col], errors='coerce')
            result_df[col] = numeric_series.round().astype("Int64")

    # oos_num_trades_median as nullable Int64
    if "oos_num_trades_median" in result_df.columns:
        result_df["oos_num_trades_median"] = pd.to_numeric(
            result_df["oos_num_trades_median"], errors='coerce'
        ).round().astype("Int64")

    # Float columns
    float_cols = [
        "multiplier", "robust_score", "robust_score_std",
        "oos_sortino", "oos_sortino_min", "oos_sum_pnl_pct", "oos_sum_pnl_pct_min",
        "worst_oos_max_dd_pct", "coverage_ratio",
        "coverage_ratio_effective", "fallback_share",
    ]
    for col in float_cols:
        if col in result_df.columns:
            result_df[col] = result_df[col].astype("float64")
    
    # Sanity check: worst_oos_max_dd_pct must be >= 0 and <= 1000
    if "worst_oos_max_dd_pct" in result_df.columns:
        valid_dd = result_df["worst_oos_max_dd_pct"].notna()
        if valid_dd.any():
            dd_values = result_df.loc[valid_dd, "worst_oos_max_dd_pct"]
            if (dd_values < 0).any() or (dd_values > 1000).any():
                raise ValueError(
                    f"worst_oos_max_dd_pct sanity check failed: values must be in [0, 1000]. "
                    f"Got min={dd_values.min():.2f}, max={dd_values.max():.2f}"
                )
    
    # Sanity check: robust_score_std should be reasonable (< 10) for points with coverage >= 2
    # This catches scaling errors (e.g., multiplying by 100 incorrectly)
    if "robust_score_std" in result_df.columns and "coverage_count" in result_df.columns:
        multi_step_mask = (result_df["coverage_count"] >= 2) & result_df["robust_score_std"].notna()
        if multi_step_mask.any():
            std_values = result_df.loc[multi_step_mask, "robust_score_std"]
            if (std_values >= 10).any():
                import warnings
                max_std = std_values.max()
                problematic_rows = result_df.loc[multi_step_mask & (result_df["robust_score_std"] >= 10)]
                warnings.warn(
                    f"robust_score_std sanity check: found unusually high std values (>= 10). "
                    f"Max std={max_std:.2f}. This may indicate a scaling error. "
                    f"Problematic points: {problematic_rows[['atr_period', 'multiplier', 'robust_score_std']].to_dict('records')}",
                    UserWarning
                )
    
    # Final check: must have exactly grid_size rows
    expected_size = len(result.mini_grid_state.params)
    if len(result_df) != expected_size:
        raise AssertionError(
            f"MiniGrid results table size mismatch: expected {expected_size} rows, got {len(result_df)}"
        )
    
    return result_df


def _resolve_plateau_metric_col(config: dict | None) -> str:
    """
    Resolve which column in mini_grid_trials to use for per-step plateau analysis.

    Returns 'robust_score' when ranking_metric=='robust_score', else objective_metric
    (defaulting to 'sortino').
    """
    cfg = config or {}
    ranking_metric = cfg.get("optimization", {}).get("ranking_metric", "value")
    objective_metric = cfg.get("optimization", {}).get("objective_metric", "sortino")
    return "robust_score" if ranking_metric == "robust_score" else objective_metric


def _calculate_plateau_stats(scores: np.ndarray, threshold: float) -> dict:
    """
    Calculate plateau statistics for an array of scores.

    Plateau boundary: best_score - abs(best_score) * (1 - threshold)

    This formula correctly handles negative best values: for negative best,
    best*threshold would move *towards zero* (less negative), the opposite of
    allowing a % degradation. Using abs(best)*(1-threshold) always subtracts a
    non-negative band regardless of sign.
    """
    from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE

    x = np.asarray(scores, dtype=np.float64)
    x = x[~np.isnan(x)]
    x = x[~np.isclose(x, INVALID_METRIC_VALUE)]

    total = int(len(x))
    if total == 0:
        return dict(
            best_score=np.nan,
            threshold_score=np.nan,
            plateau_count=0,
            total_points=0,
            plateau_share=0.0,
        )

    best = float(np.max(x))
    thr_score = best - abs(best) * (1.0 - float(threshold))
    count = int(np.sum(x >= thr_score))
    share = float(count / total)

    return dict(
        best_score=best,
        threshold_score=thr_score,
        plateau_count=count,
        total_points=total,
        plateau_share=share,
    )


def _build_minigrid_plateau_per_step_table(result, config: dict | None = None) -> pd.DataFrame:
    """
    Build per-step plateau statistics table from mini_grid_trials (TRAIN data).

    When plateau is disabled, returns a table with the correct schema but NaN/NA values
    so downstream code never breaks on missing columns.
    """
    cfg = config or {}
    plateau_cfg = (
        cfg.get("walk_forward", {}).get("mini_grid", {}).get("plateau", {}) or {}
    )
    enabled = bool(plateau_cfg.get("enabled", False))
    threshold = float(plateau_cfg.get("threshold", 0.95))

    cols = [
        "step_id",
        "metric_used",
        "plateau_threshold",
        "best_score",
        "threshold_score",
        "plateau_count",
        "total_points",
        "plateau_share",
    ]

    trials = getattr(result, "mini_grid_trials", None)
    if trials is None or len(trials) == 0 or "step_id" not in trials.columns:
        return pd.DataFrame(columns=cols)

    metric_col = _resolve_plateau_metric_col(cfg)
    step_ids = sorted(pd.unique(trials["step_id"]))
    rows: list[dict] = []

    if not enabled:
        for sid in step_ids:
            rows.append(
                {
                    "step_id": int(sid),
                    "metric_used": metric_col,
                    "plateau_threshold": np.nan,
                    "best_score": np.nan,
                    "threshold_score": np.nan,
                    "plateau_count": pd.NA,
                    "total_points": pd.NA,
                    "plateau_share": np.nan,
                }
            )
        return pd.DataFrame(rows, columns=cols)

    if metric_col not in trials.columns:
        for sid in step_ids:
            rows.append(
                {
                    "step_id": int(sid),
                    "metric_used": metric_col,
                    "plateau_threshold": threshold,
                    "best_score": np.nan,
                    "threshold_score": np.nan,
                    "plateau_count": pd.NA,
                    "total_points": pd.NA,
                    "plateau_share": np.nan,
                }
            )
        return pd.DataFrame(rows, columns=cols)

    for sid, g in trials.groupby("step_id"):
        scores = g[metric_col].to_numpy(dtype="float64", na_value=np.nan)
        stats = _calculate_plateau_stats(scores, threshold)
        rows.append(
            {
                "step_id": int(sid),
                "metric_used": metric_col,
                "plateau_threshold": threshold,
                "best_score": stats["best_score"],
                "threshold_score": stats["threshold_score"],
                "plateau_count": stats["plateau_count"],
                "total_points": stats["total_points"],
                "plateau_share": float(round(stats["plateau_share"], 6)),
            }
        )
        logger.info(
            "MiniGrid plateau (per-step) | step=%d | metric=%s | best=%.6f"
            " | thr=%.2f | share=%.3f | count=%d/%d",
            int(sid),
            metric_col,
            stats["best_score"] if np.isfinite(stats["best_score"]) else float("nan"),
            threshold,
            stats["plateau_share"],
            stats["plateau_count"],
            stats["total_points"],
        )

    df = pd.DataFrame(rows, columns=cols)
    df["plateau_count"] = df["plateau_count"].astype("Int64", errors="ignore")
    df["total_points"] = df["total_points"].astype("Int64", errors="ignore")
    return df


def _append_aggregated_plateau_columns(df: pd.DataFrame, config: dict | None) -> pd.DataFrame:
    """
    Append plateau_threshold, plateau_count, plateau_share columns to the aggregated
    MiniGrid results table.

    Column selection maps directly to what the aggregated table actually contains:
    - ranking_metric=='robust_score' and column present → use 'robust_score'
    - otherwise → 'oos_sortino' (sortino) or 'oos_sum_pnl_pct' (sum_pnl_pct)

    When disabled, columns are added with stable NaN/NA schema so existing Excel
    readers never break on missing columns.
    """
    cfg = config or {}
    plateau_cfg = (
        cfg.get("walk_forward", {}).get("mini_grid", {}).get("plateau", {}) or {}
    )
    enabled = bool(plateau_cfg.get("enabled", False))
    threshold = float(plateau_cfg.get("threshold", 0.95))

    out = df.copy()
    out["plateau_threshold"] = np.nan
    out["plateau_count"] = pd.NA
    out["plateau_share"] = np.nan

    if not enabled or len(out) == 0:
        return out

    ranking_metric = cfg.get("optimization", {}).get("ranking_metric", "value")
    objective_metric = cfg.get("optimization", {}).get("objective_metric", "sortino")

    if ranking_metric == "robust_score" and "robust_score" in out.columns:
        col = "robust_score"
    else:
        col = "oos_sortino" if objective_metric == "sortino" else "oos_sum_pnl_pct"

    if col not in out.columns:
        return out

    stats = _calculate_plateau_stats(
        out[col].to_numpy(dtype="float64", na_value=np.nan), threshold
    )

    out["plateau_threshold"] = threshold
    out["plateau_count"] = stats["plateau_count"]
    out["plateau_share"] = float(round(stats["plateau_share"], 6))
    out["plateau_count"] = out["plateau_count"].astype("Int64", errors="ignore")

    logger.info(
        "MiniGrid plateau (aggregated) | col=%s | best=%.6f | thr=%.2f"
        " | share=%.3f | count=%d/%d",
        col,
        stats["best_score"] if np.isfinite(stats["best_score"]) else float("nan"),
        threshold,
        stats["plateau_share"],
        stats["plateau_count"],
        stats["total_points"],
    )
    return out


def _write_mini_grid_sheet(
    mini_grid_state: Optional[MiniGridState],
    eval_counters: Optional[GridEvaluationCounters],
    result: WalkForwardResult,
    writer: pd.ExcelWriter,
    config: Optional[dict] = None,
) -> None:
    """
    Write WF_MiniGrid sheet (Phase 6).
    
    If mini_grid_state is None (disabled mode):
        Creates sheet with single message row
    Else:
        Creates sheet with 4 sections:
        - SECTION 1: Metadata
        - SECTION 2: Seed Parameters
        - SECTION 3: Grid Parameters
        - SECTION 4: MiniGrid Results (Aggregated) - NEW
    
    Parameters
    ----------
    mini_grid_state : Optional[MiniGridState]
        Mini-grid state from WF result
    eval_counters : Optional[GridEvaluationCounters]
        Evaluation counters from WF result
    result : WalkForwardResult
        Full Walk-Forward result (for aggregating metrics)
    writer : pd.ExcelWriter
        Excel writer object
    """
    sheet_name = "WF_MiniGrid"
    
    if mini_grid_state is None:
        # Disabled mode: single message
        df = pd.DataFrame({"Message": ["Mini-grid mode not enabled"]})
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        return
    
    # ═══ SECTION 1: Metadata ═══
    metadata_rows = []

    # Basic grid info
    metadata_rows.append(("source_step", mini_grid_state.source_step))
    metadata_rows.append(("seed_source", mini_grid_state.seed_source))
    metadata_rows.append(("grid_size", mini_grid_state.grid_size_after_truncate))
    metadata_rows.append(("grid_size_before_truncate", mini_grid_state.grid_size_before_truncate))
    metadata_rows.append(("truncated", "Yes" if mini_grid_state.truncated else "No"))
    metadata_rows.append(("truncate_policy_used", mini_grid_state.truncate_policy_used or "N/A"))

    # Phase 2: discovery metadata (kept close to top for existing test compat)
    _merged = getattr(mini_grid_state, "merged_seeds", None) or []
    _disc_steps_count = len({s for ms in _merged for s in ms.source_steps}) if _merged else 1
    _is_normalized = any(getattr(ms, "normalized", False) for ms in _merged)
    _single_step_mode = _disc_steps_count <= 1
    metadata_rows.append(("discovery_steps_count", _disc_steps_count))
    metadata_rows.append(("score_normalized", "Yes" if _is_normalized else "No"))
    metadata_rows.append(("single_step_mode", "Yes" if _single_step_mode else "No"))

    # Config snapshot
    config_snap = mini_grid_state.config_snapshot
    metadata_rows.append(("top_n_seeds", config_snap.get("top_n_seeds", "N/A")))
    metadata_rows.append(("atr_delta", config_snap.get("atr_delta", "N/A")))
    metadata_rows.append(("mult_delta", config_snap.get("mult_delta", "N/A")))
    metadata_rows.append(("atr_step", config_snap.get("atr_step", "N/A")))
    metadata_rows.append(("mult_step", config_snap.get("mult_step", "N/A")))

    max_grid_size = config_snap.get("max_grid_size")
    metadata_rows.append(("max_grid_size", max_grid_size if max_grid_size is not None else "unlimited"))

    # Evaluation counters
    if eval_counters is not None:
        metadata_rows.append(("eval_attempted", eval_counters.attempted))
        metadata_rows.append(("eval_kept", eval_counters.kept))
        metadata_rows.append(("eval_early_exit", eval_counters.early_exit_count))
        metadata_rows.append(("eval_invalid_objective", eval_counters.invalid_objective_count))
    else:
        metadata_rows.append(("eval_attempted", "N/A"))
        metadata_rows.append(("eval_kept", "N/A"))
        metadata_rows.append(("eval_early_exit", "N/A"))
        metadata_rows.append(("eval_invalid_objective", "N/A"))

    # Phase 2.1 summary: selected seed
    _sel_atr = getattr(mini_grid_state, "selected_seed_atr", None)
    _sel_mult = getattr(mini_grid_state, "selected_seed_multiplier", None)
    _sel_mt = getattr(mini_grid_state, "selected_seed_mult_ticks", None)
    # Fall back to first seed_param when selected fields not set
    if _sel_atr is None and mini_grid_state.seed_params:
        _sp0 = mini_grid_state.seed_params[0]
        _sel_atr = _sp0[0]
        _sel_mult = _sp0[1]
        _cfg_snap2 = mini_grid_state.config_snapshot
        _ms_val = _cfg_snap2.get("mult_step", 0.1)
        _sel_mt = int(round(_sel_mult / _ms_val)) if _sel_mult is not None else None
    metadata_rows.append(("selected_seed_atr", _sel_atr if _sel_atr is not None else "N/A"))
    metadata_rows.append(("selected_seed_multiplier", _sel_mult if _sel_mult is not None else "N/A"))
    metadata_rows.append(("selected_seed_mult_ticks", _sel_mt if _sel_mt is not None else "N/A"))

    # Phase 2.1 summary: merged seed count + appear_count distribution
    metadata_rows.append(("merged_seed_count", len(_merged)))
    if _merged:
        import statistics as _stats
        _appear_counts = [getattr(ms, "appear_count", 1) for ms in _merged]
        metadata_rows.append(("appear_count_min", min(_appear_counts)))
        metadata_rows.append(("appear_count_median", _stats.median(_appear_counts)))
        metadata_rows.append(("appear_count_max", max(_appear_counts)))
    else:
        metadata_rows.append(("appear_count_min", "N/A"))
        metadata_rows.append(("appear_count_median", "N/A"))
        metadata_rows.append(("appear_count_max", "N/A"))

    # Phase 2.1 reproducibility fields
    _resolved_disc = getattr(mini_grid_state, "resolved_discovery_steps", None)
    _resolved_top_m = getattr(mini_grid_state, "resolved_top_m_per_step", None)
    _resolved_mac = getattr(mini_grid_state, "resolved_min_appear_count", None)
    metadata_rows.append(("resolved_discovery_steps",
                          str(_resolved_disc) if _resolved_disc is not None else "N/A"))
    metadata_rows.append(("resolved_top_m_per_step",
                          _resolved_top_m if _resolved_top_m is not None else "N/A"))
    metadata_rows.append(("resolved_min_appear_count",
                          _resolved_mac if _resolved_mac is not None else "N/A"))

    # Reproducibility: optuna_seed from config (if provided)
    _optuna_seed: Any = "N/A"
    if config is not None:
        _optuna_seed = config.get("optimization", {}).get("seed", "N/A")
    metadata_rows.append(("optuna_seed", _optuna_seed))
    
    metadata_df = pd.DataFrame(metadata_rows, columns=["Property", "Value"])
    
    # ═══ SECTION 2: Seed Parameters ═══
    # Phase 1: use merged_seeds when available for richer metadata
    merged_seeds = getattr(mini_grid_state, "merged_seeds", None)
    config_snap = mini_grid_state.config_snapshot
    mult_step_val = config_snap.get("mult_step", 0.1)

    seed_rows = []
    if merged_seeds:
        _any_normalized = any(getattr(ms, "normalized", False) for ms in merged_seeds)
        for rank, ms in enumerate(merged_seeds, start=1):
            row: dict = {
                "seed_rank": rank,
                "atr_period": ms.atr_period,
                "multiplier": ms.multiplier,
                "mult_ticks": ms.mult_ticks,
                "score_mean": ms.score_mean,
                "score_min": ms.score_min,
                "appear_count": ms.appear_count,
                "source_steps": str(ms.source_steps),
            }
            if _any_normalized:
                row["raw_score_mean"] = getattr(ms, "raw_score_mean", None)
                row["raw_score_min"] = getattr(ms, "raw_score_min", None)
            seed_rows.append(row)
        _base_cols = ["seed_rank", "atr_period", "multiplier", "mult_ticks",
                      "score_mean", "score_min", "appear_count", "source_steps"]
        _norm_cols = ["raw_score_mean", "raw_score_min"] if _any_normalized else []
        seeds_df = pd.DataFrame(seed_rows) if seed_rows else pd.DataFrame(
            columns=_base_cols + _norm_cols
        )
    else:
        # Legacy fallback: seed_params as (atr, mult) tuples
        for rank, (atr, mult) in enumerate(mini_grid_state.seed_params, start=1):
            mt = int(round(mult / mult_step_val))
            seed_rows.append({
                "seed_rank": rank,
                "atr_period": atr,
                "multiplier": mult,
                "mult_ticks": mt,
                "score_mean": None,
                "score_min": None,
                "appear_count": 1,
                "source_steps": "[0]",
            })
        seeds_df = pd.DataFrame(seed_rows) if seed_rows else pd.DataFrame(
            columns=["seed_rank", "atr_period", "multiplier", "mult_ticks",
                     "score_mean", "score_min", "appear_count", "source_steps"]
        )
    
    # ═══ SECTION 3: Grid Parameters ═══
    grid_rows = []
    for idx, (atr, mult) in enumerate(mini_grid_state.params, start=1):
        grid_rows.append({
            "grid_index": idx,
            "atr_period": atr,
            "multiplier": mult
        })
    
    grid_df = pd.DataFrame(grid_rows) if grid_rows else pd.DataFrame(
        columns=["grid_index", "atr_period", "multiplier"]
    )
    
    # ═══ SECTION 4: MiniGrid Results (Aggregated) ═══
    minigrid_results_df = _build_minigrid_results_table(result, config)
    minigrid_results_df = _append_aggregated_plateau_columns(minigrid_results_df, config)

    # ═══ Write all sections with separators ═══
    startrow = 0

    # Section 1: Metadata
    metadata_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(metadata_df) + 2  # +2 for header and blank row

    # Section 2: Seeds
    seeds_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(seeds_df) + 2

    # Section 3: Grid
    grid_df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(grid_df) + 2

    # Section 4: MiniGrid Results (Aggregated)
    wb = writer.book
    ws = wb[sheet_name]
    ws.cell(row=startrow + 1, column=1, value="MiniGrid Results (Aggregated)")
    startrow += 1
    format_excel_export_df(minigrid_results_df).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(minigrid_results_df) + 2

    # Section 5: MiniGrid Plateau (Per Step)
    plateau_step_df = _build_minigrid_plateau_per_step_table(result, config=config)
    ws.cell(row=startrow + 1, column=1, value="MiniGrid Plateau (Per Step)")
    startrow += 1
    format_excel_export_df(plateau_step_df).to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
    startrow += len(plateau_step_df) + 2


def _write_refine_details_sheet(result: WalkForwardResult, writer: pd.ExcelWriter, config: dict | None = None) -> None:
    """
    Write WF_Refine_Details sheet — per-step trials from MiniGrid Refine.

    Only created when at least one step has refine_enabled=True.
    Columns: step_idx, atr_period, multiplier, <objective_metric>,
             ranking_metric (robust_score or objective_metric), source, early_exit.
    """
    refine_steps = [s for s in result.steps if getattr(s, "refine_enabled", False)]
    if not refine_steps:
        return  # Sheet not created when refine is globally disabled

    obj_metric = (config or {}).get("optimization", {}).get("objective_metric", "sortino")

    rows = []
    for step in refine_steps:
        meta = getattr(step, "refine_meta", None) or {}
        # trials_df is stored on refine_meta["trials_df"] if we choose to carry it,
        # but the Blueprint says "data from opt_result.trials_df on each step".
        # We store a snapshot in refine_meta["trials_df"] (added in this stage).
        trials_df = meta.get("trials_df")
        if trials_df is None or len(trials_df) == 0:
            # Fallback: emit a single summary row from the meta
            rows.append({
                "step_idx": step.step_index,
                "atr_period": meta.get("refined_atr"),
                "multiplier": meta.get("refined_mult"),
                obj_metric: None,
                "robust_score": None,
                "source": "summary_only",
                "early_exit": None,
            })
            continue
        for _, tr in trials_df.iterrows():
            rows.append({
                "step_idx": step.step_index,
                "atr_period": _to_excel_safe(tr.get("atr_period")),
                "multiplier": _to_excel_safe(tr.get("multiplier")),
                obj_metric: _to_excel_safe(tr.get(obj_metric)),
                "robust_score": _to_excel_safe(tr.get("robust_score")),
                "source": tr.get("source", "refine"),
                "early_exit": _to_excel_safe(tr.get("early_exit")),
            })

    if not rows:
        return

    df = pd.DataFrame(rows)
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Refine_Details", index=False)


def export_walk_forward_to_excel(
    result: WalkForwardResult,
    output_path: str,
    config: dict | None = None,
    plateau_result: "Optional[Any]" = None,
) -> str:
    """
    Export Walk-Forward results to Excel.

    Creates an Excel file with multiple sheets:
    - WF_TOP_K_Aggregated: Aggregated TOP-K table across all steps (deduplicated)
    - WF_Consensus: Coverage consensus across steps (PR7)
    - WF_Consensus_Details: OOS and TRAIN metrics for consensus param #1 per WF step
    - WF_Aggregated_Details: OOS and TRAIN metrics for rank=1 from aggregated TOP-K per WF step
    - WF_01, WF_02, ...: TOP-K parameters for each step (PR7)
    - WF_Summary: Aggregated OOS metrics
    - WF_Windows: Per-step details
    - WF_Refine_Details: MiniGrid Refine trials per step (only when refine enabled)
    - WF_Trades: OOS/Test trades (per-step best params)
    - WF_Train_Trades: TRAIN trades (per-step best params)
    - WF_Equity: OOS equity curve
    - WF_Gates: Gates evaluation (if enabled)

    plateau_result : PlateauAnalysisResult, optional
        Pre-computed plateau analysis result.  When provided, writes
        PlateauAnalysis, PlateauSurface (deep mode only), and PlateauPivot_N
        sheets.  This function does NOT run backtests — it is write-only.

    Args:
        result: WalkForwardResult object
        output_path: Path to output Excel file
        config: Configuration dictionary containing consensus settings
        plateau_result: Pre-computed PlateauAnalysisResult (optional)

    Returns:
        output_path (same as input)
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # === WF_Config first (ensures at least one sheet before any build) ===
        _write_config_sheet(writer, config)

        # === WF_Gates second (if enabled) ===
        if result.gates_result is not None and result.gates_result.enabled:
            _write_gates_sheet(result, writer)

        # === Build data (needed for priority sheets) ===
        _agg_df, _step_df_filled = _build_aggregated_topk_table(
            result, config, _return_step_df=True
        )
        _agg_df = _enrich_soft_gate_columns(_agg_df, _step_df_filled, config)
        _canonical_scored_df = _write_wf_topk_canonical_sheet(
            _agg_df, writer, config=config, total_steps=len(result.steps)
        )
        _append_cag_diagnostics_to_gates(writer, _canonical_scored_df, config or {})
        _full_bucket_df = _build_full_bucket_step_matrix(
            result, _canonical_scored_df, config or {}
        )
        _median_bucket_df = _build_median_bucket_step_matrix(
            result, _canonical_scored_df, config or {}
        )
        _surface_bucket_df = None
        _surface_trials = getattr(result, "all_step_trials", None)
        if (
            _surface_trials is not None
            and isinstance(_surface_trials, pd.DataFrame)
            and not _surface_trials.empty
        ):
            _surface_bucket_df = _build_surface_bucket_step_matrix(
                all_step_trials=_surface_trials,
                config=config or {},
                wf_total_steps=len(result.steps),
            )

        # === PRIORITY ORDER: WF_Config, WF_Gates, WF_TOP_K_Canonical already written ===

        # 3. WF_TOP_K_Canonical (written during build above; sheet order fixed in _reorder_sheets)

        # 4. BucketMatrix_Full
        _write_bucket_matrix_full_sheet(_full_bucket_df, writer)

        # 5. BucketMatrix_Median
        _write_bucket_matrix_median_sheet(_median_bucket_df, writer)

        # 6. BucketTrainSurface (if exists)
        if _surface_bucket_df is not None:
            _write_bucket_train_surface_sheet(_surface_bucket_df, writer, config or {})

        # 7. PlateauSummary — full candidate pool ranking (deep mode only, when surface present)
        # Written before PlateauAnalysis/PlateauSurface; final sheet position is fixed
        # by _reorder_sheets (PlateauSummary = index 1, after WF_Config).
        #
        # Two-step pattern: compute df first (retained for downstream use by
        # UnifiedProductionRank), then write.
        _plateau_summary_df = None
        if plateau_result is not None and getattr(plateau_result, "surface", None) is not None:
            from supertrend_optimizer.analysis.bucket_plateau import (
                _PLATEAU_SUMMARY_COLUMNS,
                _compute_plateau_summary_scores,
            )
            try:
                _plateau_summary_df = _compute_plateau_summary_scores(
                    plateau_result=plateau_result,
                    median_bucket_df=_median_bucket_df,
                    surface_bucket_df=_surface_bucket_df,
                    config=config or {},
                )
            except Exception:
                logger.warning(
                    "export: _compute_plateau_summary_scores failed",
                    exc_info=True,
                )
                _plateau_summary_df = None

            _write_plateau_summary_sheet_from_df(
                summary_df=(
                    _plateau_summary_df
                    if _plateau_summary_df is not None
                    else pd.DataFrame(columns=_PLATEAU_SUMMARY_COLUMNS)
                ),
                writer=writer,
            )

        # 7b. UnifiedProductionRank (ensemble layer over canonical + plateau + topology)
        # _component_map and _surface_raw_df are extracted unconditionally here so that
        # FullInformationRank (block 7c below) can always reference them regardless of
        # whether unified_ranking is enabled.  Both default to None when plateau surface
        # is absent — FIR handles None inputs gracefully via degraded-mode channels.
        _component_map = (
            getattr(plateau_result.surface, "component_map", None)
            if plateau_result is not None
            and getattr(plateau_result, "surface", None) is not None
            else None
        )
        _surface_raw_df = (
            getattr(plateau_result.surface, "raw_df", None)
            if plateau_result is not None
            and getattr(plateau_result, "surface", None) is not None
            else None
        )

        _unified_cfg = (config or {}).get("walk_forward", {}).get(
            "unified_ranking", {}
        )
        if _unified_cfg.get("enabled", False):
            from supertrend_optimizer.analysis.unified_ranking import (
                compute_unified_ranking,
            )

            try:
                _unified_df, _plateau_only_df = compute_unified_ranking(
                    canonical_df=_canonical_scored_df,
                    plateau_summary_df=_plateau_summary_df,
                    component_map=_component_map,
                    raw_df=_surface_raw_df,
                    config=config or {},
                )
                _write_unified_production_rank_sheet(
                    _unified_df, _plateau_only_df, writer, config or {},
                )
            except Exception:
                logger.exception(
                    "export: compute_unified_ranking failed — writing error placeholder sheet"
                )
                try:
                    _write_unified_error_placeholder(writer)
                except Exception:
                    logger.exception(
                        "export: _write_unified_error_placeholder also failed"
                    )

        # 7c. FullInformationRank (FIR) — 14-channel orthogonal scoring
        # Runs unconditionally when enabled; does not depend on unified_ranking.
        # _component_map and _surface_raw_df are always defined above (may be None).
        # Trade-metric availability is checked inside FIR via trade_quality_min_observations.
        _fir_cfg = (config or {}).get("walk_forward", {}).get(
            "full_information_rank", {}
        )
        if _fir_cfg.get("enabled", False):
            try:
                from supertrend_optimizer.analysis.full_information_ranking import (
                    compute_full_information_rank,
                    write_fir_sheets,
                )
                _fir_df, _fir_meta = compute_full_information_rank(
                    canonical_df=_canonical_scored_df,
                    step_df=_step_df_filled,
                    plateau_summary_df=_plateau_summary_df,
                    surface_raw_df=_surface_raw_df,
                    component_map=_component_map,
                    result=result,
                    config=config or {},
                )
                write_fir_sheets(_fir_df, _fir_meta, writer, config or {})
            except Exception:
                logger.exception(
                    "export: compute_full_information_rank failed — skipping FIR sheets"
                )

        # 8–10. PlateauAnalysis, PlateauSurface, PlateauPivot_1..N (if plateau_result)
        if plateau_result is not None:
            _write_plateau_sheets(
                plateau_result, writer, config or {},
                bucket_matrix_df=_median_bucket_df,
                bucket_train_surface_df=_surface_bucket_df,
            )

        # 11. WF_01, WF_02, ...
        for i, step in enumerate(result.steps, start=1):
            sheet_name = f"WF_{i:02d}"
            _write_step_sheet(step, writer, sheet_name)

        # 11. WF_Trades
        _write_trades_sheet(result, writer)

        # 12. WF_Train_Trades
        _write_train_trades_sheet(result, writer)

        # === Remaining sheets (unchanged content, different order) ===
        _write_mini_grid_sheet(
            result.mini_grid_state, result.mini_grid_eval_counters, result, writer, config=config
        )
        _write_consensus_details_sheet(result, writer)
        _write_aggregated_details_sheet(result, writer, config)
        format_excel_export_df(_agg_df).to_excel(writer, sheet_name="WF_TOP_K_Aggregated", index=False)
        _write_consensus_sheet(result.consensus_df, writer)

        _cfg = config or {}
        _canonical_status = {"status": "DISABLED", "fallback_used": False, "top1_row": None}
        if len(_agg_df) > 0:
            _canonical_status = _canonical_gates_status(_canonical_scored_df, _cfg)

        _write_summary_sheet(
            result, writer,
            canonical_top1=_canonical_status["top1_row"],
            canonical_gates_status=_canonical_status["status"],
            canonical_fallback_used=_canonical_status["fallback_used"],
            config=config,
        )
        _write_windows_sheet(result, writer, config=config)
        _write_windows_bucket_sheet(result, writer, config)
        _write_refine_details_sheet(result, writer, config=config)
        _write_equity_sheet(result, writer)

        _imp_cfg = (config or {}).get("walk_forward", {}).get("imputation", {})
        if bool(_imp_cfg.get("enabled", True)):
            _write_imputation_sheet(_step_df_filled, writer, total_steps=len(result.steps))

        from supertrend_optimizer.scoring.diversification import run_diversification_filter
        _div_result = run_diversification_filter(_canonical_scored_df, result, config or {})
        if _div_result.enabled or _div_result.placeholder:
            _write_wf_ensemble_sheet(_div_result, writer)
            _write_wf_diversification_matrix_sheet(_div_result, writer)

    # Apply freeze panes
    _apply_freeze_panes(output_path)
    
    return output_path


def _write_summary_sheet(
    result: WalkForwardResult,
    writer: pd.ExcelWriter,
    canonical_top1: "pd.Series | None" = None,
    canonical_gates_status: str = "DISABLED",
    canonical_fallback_used: bool = False,
    config: "dict | None" = None,
) -> None:
    """Write WF_Summary sheet with aggregated OOS metrics.

    Parameters
    ----------
    result : WalkForwardResult
    writer : pd.ExcelWriter
    canonical_top1 : pd.Series | None
        Row from canonical-ranked agg_df corresponding to rank #1.
        None when status==ALL_FAILED or agg_df is empty.
        Display-only — does not affect any parameter-selection logic.
    canonical_gates_status : str
        "PASS" | "ALL_FAILED" | "DISABLED"
    canonical_fallback_used : bool
        True when ALL_FAILED (fallback was *not* applied — kept for traceability).
    config : dict | None
        Full run config; used to read canonical gate thresholds for display.
    """
    metrics = result.oos_metrics
    
    # Create summary data as single row
    # n_steps = total WF steps (including skipped); valid_step_count = steps with top_entries
    valid_step_count = getattr(result, "valid_step_count", None)
    if valid_step_count is None:
        valid_step_count = sum(1 for s in result.steps if len(getattr(s, "top_entries", [])) > 0)
    total_steps = len(result.steps)
    skipped_step_count = getattr(result, "skipped_step_count", total_steps - valid_step_count)
    skipped_step_ratio = getattr(result, "skipped_step_ratio", (skipped_step_count / total_steps if total_steps > 0 else 0.0))
    summary_data = {
        "n_steps": total_steps,
        "valid_step_count": valid_step_count,
        "skipped_step_count": int(skipped_step_count),
        "skipped_step_ratio": round(float(skipped_step_ratio), 2),
        "oos_num_trades": metrics.get("num_trades", np.nan),
        "oos_win_rate": metrics.get("win_rate", np.nan),
        "oos_sum_pnl_pct": metrics.get("sum_pnl_pct", np.nan),
        "oos_avg_trade": metrics.get("avg_trade", np.nan),
        "oos_profit_factor": metrics.get("profit_factor", np.nan),
        "oos_sharpe": metrics.get("sharpe", np.nan),
        "oos_sortino": metrics.get("sortino", np.nan),
        "oos_max_drawdown": metrics.get("max_drawdown", np.nan),
        "oos_cagr": metrics.get("cagr", np.nan)
    }
    
    # Add gates columns if gates are enabled
    if result.gates_result is not None and result.gates_result.enabled:
        summary_data["gates_enabled"] = "Yes"
        summary_data["gates_passed"] = "PASS" if result.gates_result.passed else "FAIL"
        summary_data["gates_fail_count"] = result.gates_result.fail_count
    else:
        summary_data["gates_enabled"] = "No"
        summary_data["gates_passed"] = "N/A"
        summary_data["gates_fail_count"] = np.nan
    
    # Add worst step metrics
    worst_step = _compute_worst_step_summary(result)
    summary_data["worst_step_index_sortino"] = worst_step["worst_step_index_sortino"]
    summary_data["worst_step_sortino"] = worst_step["worst_step_sortino"]
    summary_data["worst_step_index_sum_pnl"] = worst_step["worst_step_index_sum_pnl"]
    summary_data["worst_step_sum_pnl_pct"] = worst_step["worst_step_sum_pnl_pct"]
    summary_data["worst_step_index_max_dd"] = worst_step["worst_step_index_max_dd"]
    summary_data["worst_step_max_dd_pct"] = worst_step["worst_step_max_dd_pct"]
    
    # === PHASE 6: Add Mini-Grid summary fields ===
    if result.mini_grid_state is not None:
        summary_data["mini_grid_mode"] = "Yes"
        summary_data["mini_grid_size"] = int(result.mini_grid_state.grid_size_after_truncate)
        summary_data["mini_grid_truncated"] = "Yes" if result.mini_grid_state.truncated else "No"
        summary_data["mini_grid_source_step"] = int(result.mini_grid_state.source_step)
        summary_data["mini_grid_seed_source"] = str(result.mini_grid_state.seed_source)
        
        # Calculate kept percentage
        if result.mini_grid_eval_counters is not None and result.mini_grid_eval_counters.attempted > 0:
            kept_pct = (result.mini_grid_eval_counters.kept / result.mini_grid_eval_counters.attempted) * 100
            summary_data["mini_grid_eval_kept_pct"] = f"{kept_pct:.1f}%"
        else:
            summary_data["mini_grid_eval_kept_pct"] = "N/A"
    else:
        summary_data["mini_grid_mode"] = "No"
        summary_data["mini_grid_size"] = np.nan  # Use NaN instead of "N/A" for numeric columns
        summary_data["mini_grid_truncated"] = "N/A"
        summary_data["mini_grid_source_step"] = np.nan  # Use NaN instead of "N/A" for numeric columns
        summary_data["mini_grid_seed_source"] = "N/A"
        summary_data["mini_grid_eval_kept_pct"] = "N/A"
    
    # === Selected params (consensus / mini-grid winner) ===
    # Backward compat: if selected_* absent (old result), use NaN/unknown
    sel_atr = getattr(result, "selected_atr_period", None)
    sel_mult = getattr(result, "selected_multiplier", None)
    sel_rank = getattr(result, "selected_rank", None)
    sel_cov_ct = getattr(result, "selected_coverage_count", None)
    sel_cov_ratio = getattr(result, "selected_coverage_ratio", None)
    sel_src = getattr(result, "selected_source", "unknown")
    summary_data["selected_atr_period"] = int(sel_atr) if sel_atr is not None else np.nan
    summary_data["selected_multiplier"] = float(sel_mult) if sel_mult is not None else np.nan
    summary_data["selected_rank"] = int(sel_rank) if sel_rank is not None else np.nan
    summary_data["selected_coverage_count"] = int(sel_cov_ct) if sel_cov_ct is not None else np.nan
    summary_data["selected_coverage_ratio"] = float(sel_cov_ratio) if sel_cov_ratio is not None else np.nan
    summary_data["selected_source"] = str(sel_src) if sel_src else "unknown"

    # === Stage 3: MiniGrid Refine aggregates ===
    _refine_steps = [s for s in result.steps if getattr(s, "refine_enabled", False)]
    _any_refine = len(_refine_steps) > 0
    summary_data["refine_enabled"] = "Yes" if _any_refine else "No"
    summary_data["refine_plateau_width_mean"] = (
        getattr(result, "refine_plateau_width_mean", None)
        if _any_refine else np.nan
    )
    summary_data["refine_plateau_width_min"] = (
        getattr(result, "refine_plateau_width_min", None)
        if _any_refine else np.nan
    )
    summary_data["refine_grid_size_avg"] = (
        getattr(result, "refine_grid_size_avg", None)
        if _any_refine else np.nan
    )
    summary_data["refine_best_changed_count"] = (
        getattr(result, "refine_best_changed_count", None)
        if _any_refine else np.nan
    )
    # Normalise None → NaN for numeric refine fields
    for _k in ("refine_plateau_width_mean", "refine_plateau_width_min",
               "refine_grid_size_avg", "refine_best_changed_count"):
        if summary_data[_k] is None:
            summary_data[_k] = np.nan

    # === Canonical Top-1 (display-only; does not affect parameter selection) ===
    # canonical_top1 is None when ALL_FAILED or agg_df empty.
    if canonical_top1 is not None:
        summary_data["canonical_top1_atr"] = int(canonical_top1["atr_period"])
        summary_data["canonical_top1_mult"] = float(canonical_top1["multiplier"])
        summary_data["canonical_top1_score"] = round(float(canonical_top1["canonical_rank_score"]), 6)
        summary_data["canonical_top1_regime_score"] = (
            round(float(canonical_top1["regime_score"]), 6)
            if "regime_score" in canonical_top1.index
            else np.nan
        )
    else:
        summary_data["canonical_top1_atr"] = np.nan
        summary_data["canonical_top1_mult"] = np.nan
        summary_data["canonical_top1_score"] = np.nan
        summary_data["canonical_top1_regime_score"] = np.nan

    # === Canonical gates status / fallback ===
    summary_data["canonical_gates_status"] = canonical_gates_status  # PASS/ALL_FAILED/DISABLED
    summary_data["canonical_fallback_used"] = canonical_fallback_used

    # === Canonical OOS risk gates (thresholds used; NaN when section absent) ===
    _canonical_gates_cfg = (
        (config or {}).get("walk_forward", {}).get("canonical", {}).get("gates", None)
    )
    if _canonical_gates_cfg is not None:
        summary_data["canonical_gate_max_worst_oos_dd_pct"] = _cfg_gate_threshold(
            _canonical_gates_cfg, "max_worst_oos_dd_pct", 35.0
        )
        summary_data["canonical_gate_min_oos_sortino_min"] = _cfg_gate_threshold(
            _canonical_gates_cfg, "min_oos_sortino_min", 0.0
        )
        summary_data["canonical_gate_min_oos_sum_pnl_pct_min"] = _cfg_gate_threshold(
            _canonical_gates_cfg, "min_oos_sum_pnl_pct_min", 0.0
        )
        summary_data["canonical_gate_min_oos_trades_median"] = _cfg_gate_threshold(
            _canonical_gates_cfg, "min_oos_trades_median", 0.0
        )
    else:
        summary_data["canonical_gate_max_worst_oos_dd_pct"] = np.nan
        summary_data["canonical_gate_min_oos_sortino_min"] = np.nan
        summary_data["canonical_gate_min_oos_sum_pnl_pct_min"] = np.nan
        summary_data["canonical_gate_min_oos_trades_median"] = np.nan

    # Replace INVALID_METRIC_VALUE (-999.0) with NaN, and None gate thresholds with NaN
    # (None means "gate disabled"; NaN renders as blank in Excel).
    _gate_threshold_keys = (
        "canonical_gate_max_worst_oos_dd_pct",
        "canonical_gate_min_oos_sortino_min",
        "canonical_gate_min_oos_sum_pnl_pct_min",
        "canonical_gate_min_oos_trades_median",
    )
    for key, value in summary_data.items():
        if isinstance(value, (int, float)) and value == -999.0:
            summary_data[key] = np.nan
        elif key in _gate_threshold_keys and value is None:
            summary_data[key] = np.nan
    
    df = pd.DataFrame([summary_data])
    
    # Rename columns for display
    column_names = [
        "N Steps",
        "Valid Step Count",
        "Skipped Step Count",
        "Skipped Step Ratio",
        "OOS Num Trades",
        "OOS Win Rate",
        "OOS Sum PnL %",
        "OOS Avg Trade",
        "OOS Profit Factor",
        "OOS Sharpe",
        "OOS Sortino",
        "OOS Max DD",
        "OOS CAGR",
        "Gates Enabled",
        "Gates Passed",
        "Gates Fail Count",
        "Worst Step Index (Sortino)",
        "Worst Step Sortino",
        "Worst Step Index (Sum PnL)",
        "Worst Step Sum PnL %",
        "Worst Step Index (Max DD)",
        "Worst Step Max DD %",
        "Mini-Grid Mode",
        "Mini-Grid Size",
        "Mini-Grid Truncated",
        "Mini-Grid Source Step",
        "Mini-Grid Seed Source",
        "Mini-Grid Eval Kept %",
        "Selected ATR Period",
        "Selected Multiplier",
        "Selected Rank",
        "Selected Coverage Count",
        "Selected Coverage Ratio",
        "Selected Source",
        "Refine Enabled",
        "Refine Plateau Width (Mean)",
        "Refine Plateau Width (Min)",
        "Refine Grid Size (Avg)",
        "Refine Best Changed Count",
        "Canonical Top-1 ATR",
        "Canonical Top-1 Mult",
        "Canonical Top-1 Score",
        "Canonical Top-1 Regime Score",
        "Canonical Gates Status",
        "Canonical Fallback Used",
        "Canonical Gate Max Worst OOS DD %",
        "Canonical Gate Min OOS Sortino Min",
        "Canonical Gate Min OOS Sum PnL % (Min Step)",
        "Canonical Gate Min OOS Trades Median",
    ]
    df.columns = column_names

    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Summary", index=False)


def _to_excel_safe(val):
    """Convert value to Excel-compatible type (naive datetime, native int/float)."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    if isinstance(val, pd.Timestamp):
        return val.tz_localize(None) if val.tz is not None else val
    if isinstance(val, (np.integer, np.int64, np.int32)):
        return int(val)
    if isinstance(val, (np.floating, np.float64, np.float32)):
        return float(val) if not np.isnan(val) else None
    return val


def _resolve_train_objective_col_name(config: dict | None) -> str:
    """
    Choose human-readable column name for step.train_objective_value
    based on optimization.objective_metric.
    """
    obj = None
    if config is not None:
        obj = (config.get("optimization", {}) or {}).get("objective_metric")

    if obj == "sortino":
        return "Train Sortino"
    if obj == "sharpe":
        return "Train Sharpe"
    if obj == "cagr":
        return "Train CAGR"
    if obj == "sum_pnl_pct":
        # Avoid duplicate column name with "Train Sum PnL %"
        return "Train Objective: Sum PnL %"

    return "Train Objective"


def _write_windows_sheet(
    result: WalkForwardResult,
    writer: pd.ExcelWriter,
    config: dict | None = None,
) -> None:
    """Write WF_Windows sheet with per-step details."""
    train_obj_col = _resolve_train_objective_col_name(config)
    rows = []

    for step in result.steps:
        # Determine what to show for start/end (timestamps if available, else indices)
        if step.train_start_time is not None:
            train_start = step.train_start_time
            train_end = step.train_end_time
            test_start = step.test_start_time
            test_end = step.test_end_time
        else:
            train_start = step.train_start_idx
            train_end = step.train_end_idx
            test_start = step.test_start_idx
            test_end = step.test_end_idx

        # Extract test metrics
        test_metrics = step.test_metrics or {}
        train_metrics = getattr(step, "train_metrics", None) or {}

        step_status = getattr(step, "step_status", "OK" if (step.best_atr_period is not None and step.best_multiplier is not None) else "SKIPPED")
        ref_metrics = getattr(step, "ref_test_metrics", None) or {}
        ref_train_metrics = getattr(step, "ref_train_metrics", None) or {}
        row = {
            "Step": _to_excel_safe(step.step_index),
            "Step Status": step_status,
            "Train Start": _to_excel_safe(train_start),
            "Train End": _to_excel_safe(train_end),
            "Test Start": _to_excel_safe(test_start),
            "Test End": _to_excel_safe(test_end),
            "Best ATR": _to_excel_safe(step.best_atr_period),
            "Best Mult": _to_excel_safe(step.best_multiplier),
            train_obj_col: _to_excel_safe(step.train_objective_value),
            "Train Sum PnL %": _to_excel_safe(train_metrics.get("sum_pnl_pct", np.nan)),
            "Test Obj": _to_excel_safe(step.test_objective_value),
            "Test Trades": _to_excel_safe(test_metrics.get("num_trades", np.nan)),
            "Test Sum PnL %": _to_excel_safe(test_metrics.get("sum_pnl_pct", np.nan)),
            "Test Max DD": _to_excel_safe(test_metrics.get("max_drawdown", np.nan)),
            "Test Max DD %": _raw_dd_to_pct(test_metrics.get("max_drawdown")),
            "Test Sortino": _to_excel_safe(test_metrics.get("sortino", np.nan)),
            "Ref ATR": _to_excel_safe(getattr(step, "ref_atr_period", None)),
            "Ref Mult": _to_excel_safe(getattr(step, "ref_multiplier", None)),
            "Ref Test Trades": _to_excel_safe(ref_metrics.get("num_trades", np.nan)),
            "Ref Test Sum PnL %": _to_excel_safe(ref_metrics.get("sum_pnl_pct", np.nan)),
            "Ref Test Max DD": _to_excel_safe(ref_metrics.get("max_drawdown", np.nan)),
            "Ref Test Sortino": _to_excel_safe(ref_metrics.get("sortino", np.nan)),
            "Ref Train Trades": _to_excel_safe(ref_train_metrics.get("num_trades", np.nan)),
            "Ref Train Sum PnL %": _to_excel_safe(ref_train_metrics.get("sum_pnl_pct", np.nan)),
            "Ref Train Max DD": _to_excel_safe(ref_train_metrics.get("max_drawdown", np.nan)),
            "Ref Train Sortino": _to_excel_safe(ref_train_metrics.get("sortino", np.nan)),
            # MiniGrid Refine columns (Stage 3)
            "Refine Enabled": _to_excel_safe(getattr(step, "refine_enabled", False)),
            "Refine Grid Total": _to_excel_safe(getattr(step, "refine_grid_total", None)),
            "Refine Grid Valid": _to_excel_safe(getattr(step, "refine_grid_valid", None)),
            "Refine Grid Early Exit": _to_excel_safe(getattr(step, "refine_grid_early_exit", None)),
            "Refine Plateau Width": _to_excel_safe(getattr(step, "refine_plateau_width", None)),
            "Refine Plateau Depth": _to_excel_safe(getattr(step, "refine_plateau_depth", None)),
            "Refine Plateau Monotonicity": _to_excel_safe(getattr(step, "refine_plateau_monotonicity", None)),
            "Refine Best Changed": _to_excel_safe(getattr(step, "refine_best_changed", None)),
        }

        # Replace INVALID_METRIC_VALUE with None (will show as empty in Excel)
        for key, value in row.items():
            if isinstance(value, (int, float)) and value == -999.0:
                row[key] = None

        rows.append(row)

    # Table 1: Walk-Forward окно + Optuna Best (Discovery)
    cols1 = [
        "Step", "Step Status", "Train Start", "Train End", "Test Start", "Test End",
        "Best ATR", "Best Mult", train_obj_col, "Train Sum PnL %",
        "Test Obj", "Test Trades", "Test Sum PnL %", "Test Max DD", "Test Max DD %", "Test Sortino",
    ]
    df1 = format_excel_export_df(pd.DataFrame(rows, columns=cols1))
    df1.to_excel(writer, sheet_name="WF_Windows", index=False, startrow=0)

    # Table 2: Refined (MiniGrid Refine Stage 2)
    cols2 = [
        "Step", "Ref ATR", "Ref Mult",
        "Ref Test Trades", "Ref Test Sum PnL %", "Ref Test Max DD", "Ref Test Sortino",
        "Ref Train Trades", "Ref Train Sum PnL %", "Ref Train Max DD", "Ref Train Sortino",
    ]
    df2 = format_excel_export_df(pd.DataFrame(rows, columns=cols2))
    startrow2 = len(df1) + 2
    df2.to_excel(writer, sheet_name="WF_Windows", index=False, startrow=startrow2)

    # Table 3: Refine Diagnostics
    cols3 = [
        "Step", "Refine Enabled", "Refine Grid Total", "Refine Grid Valid",
        "Refine Grid Early Exit", "Refine Plateau Width", "Refine Plateau Depth",
        "Refine Plateau Monotonicity", "Refine Best Changed",
    ]
    df3 = pd.DataFrame(rows, columns=cols3)
    startrow3 = startrow2 + len(df2) + 2
    df3.to_excel(writer, sheet_name="WF_Windows", index=False, startrow=startrow3)


def _build_windows_bucket_df(result: WalkForwardResult, config: dict | None = None) -> pd.DataFrame:
    """
    Build per-step bucket-aggregated DataFrame from step.top_entries.

    One group of rows per WF step. Within each step, entries are grouped by
    (atr_bucket, mult_bucket_ticks, mult_bucket) and aggregated.

    Skipped steps (empty top_entries) produce a single sentinel row with
    step_status="SKIPPED", count=0, and NaN for all bucket/metric columns.

    Returns
    -------
    pd.DataFrame sorted by: step ASC, contains_best DESC, avg_robust_score DESC,
    atr_bucket ASC, mult_bucket_ticks ASC.
    """
    from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE

    if config is None:
        config = {}
    consensus_config = config.get("walk_forward", {}).get("consensus", {})
    atr_bucket_step = int(consensus_config.get("atr_bucket_step", DEFAULT_ATR_BUCKET_STEP))
    mult_bucket_step = float(consensus_config.get("mult_bucket_step", DEFAULT_MULT_BUCKET_STEP))

    output_rows: list[dict] = []

    for step in result.steps:
        # Resolve date/index columns
        if step.train_start_time is not None:
            train_start = _to_excel_safe(step.train_start_time)
            train_end = _to_excel_safe(step.train_end_time)
            test_start = _to_excel_safe(step.test_start_time)
            test_end = _to_excel_safe(step.test_end_time)
        else:
            train_start = step.train_start_idx
            train_end = step.train_end_idx
            test_start = step.test_start_idx
            test_end = step.test_end_idx

        step_status = getattr(step, "step_status", "OK" if step.best_atr_period is not None else "SKIPPED")

        if len(step.top_entries) == 0:
            output_rows.append({
                "step": step.step_index,
                "step_status": step_status,
                "train_start": train_start,
                "train_end": train_end,
                "test_start": test_start,
                "test_end": test_end,
                "atr_bucket": pd.NA,
                "mult_bucket_ticks": pd.NA,
                "mult_bucket": np.nan,
                "count": 0,
                "avg_train_objective": np.nan,
                "avg_robust_score": np.nan,
                "robust_score_std": np.nan,
                "avg_train_sum_pnl_pct": np.nan,
                "avg_oos_sortino": np.nan,
                "avg_oos_sum_pnl_pct": np.nan,
                "worst_oos_max_dd_pct": np.nan,
                "median_oos_num_trades": pd.NA,
                "in_consensus": False,
                "contains_best": False,
            })
            continue

        # Build DataFrame from entries
        entry_rows = []
        for e in step.top_entries:
            entry_rows.append({
                "atr_period": e.atr_period,
                "multiplier": e.multiplier,
                "train_objective": e.train_objective,
                "robust_score": e.robust_score,
                "train_sum_pnl_pct": e.train_sum_pnl_pct,
                "oos_sortino": e.oos_sortino,
                "oos_sum_pnl_pct": e.oos_sum_pnl_pct,
                "oos_max_dd_pct": e.oos_max_dd_pct,
                "oos_num_trades": e.oos_num_trades,
                "in_consensus": e.in_consensus,
            })
        edf = pd.DataFrame(entry_rows)

        # Replace INVALID_METRIC_VALUE with NaN
        metric_cols = [
            "train_objective", "robust_score", "train_sum_pnl_pct",
            "oos_sortino", "oos_sum_pnl_pct", "oos_max_dd_pct", "oos_num_trades",
        ]
        for col in metric_cols:
            if col in edf.columns:
                edf[col] = pd.to_numeric(edf[col], errors="coerce").replace(INVALID_METRIC_VALUE, np.nan)

        # Normalize oos_max_dd_pct: entries created directly (e.g. in tests) may
        # carry a raw fraction (e.g. -0.30) rather than a positive percent (30.0).
        # Apply _normalize_dd_to_pct so both formats are handled uniformly.
        if "oos_max_dd_pct" in edf.columns:
            from supertrend_optimizer.validation.walk_forward import _normalize_dd_to_pct
            edf["oos_max_dd_pct"] = edf["oos_max_dd_pct"].apply(
                lambda v: _normalize_dd_to_pct(v) if pd.notna(v) else np.nan
            )

        # Ensure in_consensus is bool
        edf["in_consensus"] = edf["in_consensus"].astype(bool)

        # Apply unified bucket helper
        edf = _apply_param_buckets(edf, atr_step=atr_bucket_step, mult_step=mult_bucket_step)

        # Determine contains_best using tick-based bucket comparison
        best_atr = step.best_atr_period
        best_mult = step.best_multiplier
        best_atr_bucket: int | None = None
        best_mult_ticks: int | None = None
        if best_atr is not None and best_mult is not None:
            best_atr_bucket = int(round(best_atr / atr_bucket_step) * atr_bucket_step)
            best_mult_ticks = int(round(best_mult / mult_bucket_step))

        # Group by bucket
        group_cols = ["atr_bucket", "mult_bucket_ticks", "mult_bucket"]

        agg = edf.groupby(group_cols, as_index=False).agg(
            count=("train_objective", "size"),
            avg_train_objective=("train_objective", "mean"),
            avg_robust_score=("robust_score", "mean"),
            robust_score_std=("robust_score", lambda x: x.std(ddof=0)),
            avg_train_sum_pnl_pct=("train_sum_pnl_pct", "mean"),
            avg_oos_sortino=("oos_sortino", "mean"),
            avg_oos_sum_pnl_pct=("oos_sum_pnl_pct", "mean"),
            worst_oos_max_dd_pct=("oos_max_dd_pct", "max"),  # Worst drawdown (already positive %), so max() = worst
            median_oos_num_trades=("oos_num_trades", "median"),
            in_consensus=("in_consensus", "max"),
        )

        for _, brow in agg.iterrows():
            b_atr = int(brow["atr_bucket"])
            b_ticks = int(brow["mult_bucket_ticks"])

            if best_atr_bucket is not None and best_mult_ticks is not None:
                contains_best = (b_atr == best_atr_bucket) and (b_ticks == best_mult_ticks)
            else:
                contains_best = False

            output_rows.append({
                "step": step.step_index,
                "step_status": step_status,
                "train_start": train_start,
                "train_end": train_end,
                "test_start": test_start,
                "test_end": test_end,
                "atr_bucket": b_atr,
                "mult_bucket_ticks": b_ticks,
                "mult_bucket": float(brow["mult_bucket"]),
                "count": int(brow["count"]),
                "avg_train_objective": brow["avg_train_objective"],
                "avg_robust_score": brow["avg_robust_score"],
                "robust_score_std": brow["robust_score_std"],
                "avg_train_sum_pnl_pct": brow["avg_train_sum_pnl_pct"],
                "avg_oos_sortino": brow["avg_oos_sortino"],
                "avg_oos_sum_pnl_pct": brow["avg_oos_sum_pnl_pct"],
                "worst_oos_max_dd_pct": brow["worst_oos_max_dd_pct"],
                "median_oos_num_trades": brow["median_oos_num_trades"],
                "in_consensus": bool(brow["in_consensus"]),
                "contains_best": contains_best,
            })

    if len(output_rows) == 0:
        return pd.DataFrame(columns=[
            "step", "step_status", "train_start", "train_end", "test_start", "test_end",
            "atr_bucket", "mult_bucket_ticks", "mult_bucket", "count",
            "avg_train_objective", "avg_robust_score", "robust_score_std",
            "avg_train_sum_pnl_pct", "avg_oos_sortino", "avg_oos_sum_pnl_pct",
            "worst_oos_max_dd_pct", "median_oos_num_trades", "in_consensus", "contains_best",
        ])

    out = pd.DataFrame(output_rows)

    # Stabilise dtypes
    out["count"] = pd.to_numeric(out["count"], errors="coerce").astype("Int64")
    out["mult_bucket_ticks"] = pd.to_numeric(out["mult_bucket_ticks"], errors="coerce").astype("Int64")
    out["median_oos_num_trades"] = pd.to_numeric(out["median_oos_num_trades"], errors="coerce").round().astype("Int64")
    out["in_consensus"] = out["in_consensus"].astype(bool)
    out["contains_best"] = out["contains_best"].astype(bool)

    float_cols = [
        "mult_bucket", "avg_train_objective", "avg_robust_score", "robust_score_std",
        "avg_train_sum_pnl_pct", "avg_oos_sortino", "avg_oos_sum_pnl_pct",
        "worst_oos_max_dd_pct",
    ]
    for col in float_cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").astype("float64")

    # Deterministic sort: step ASC, contains_best DESC, avg_robust_score DESC,
    # atr_bucket ASC, mult_bucket_ticks ASC
    out = out.sort_values(
        by=["step", "contains_best", "avg_robust_score", "atr_bucket", "mult_bucket_ticks"],
        ascending=[True, False, False, True, True],
        na_position="last",
        kind="mergesort",
    ).reset_index(drop=True)

    return out


def _write_windows_bucket_sheet(
    result: WalkForwardResult, writer: pd.ExcelWriter, config: dict | None = None
) -> None:
    """Write WF_Windows_Bucket sheet aggregating top_entries by bucket per WF step."""
    df = _build_windows_bucket_df(result, config)
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Windows_Bucket", index=False)


def _write_trades_sheet(result: WalkForwardResult, writer: pd.ExcelWriter) -> None:
    """Write WF_Trades sheet with all OOS/Test trades (per-step best params)."""
    if result.oos_trades_df is not None and len(result.oos_trades_df) > 0:
        df = result.oos_trades_df.copy()
        
        # Remove timezone from datetime columns if present
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                if hasattr(df[col].dtype, 'tz') and df[col].dtype.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)
        
        df.to_excel(writer, sheet_name="WF_Trades", index=False)
    else:
        # Create empty DataFrame
        df = pd.DataFrame()
        df.to_excel(writer, sheet_name="WF_Trades", index=False)


def _write_train_trades_sheet(result: WalkForwardResult, writer: pd.ExcelWriter) -> None:
    """Write WF_Train_Trades sheet with all TRAIN trades (per-step best params)."""
    if result.train_trades_df is not None and len(result.train_trades_df) > 0:
        df = result.train_trades_df.copy()

        # Remove timezone from datetime columns if present
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                if hasattr(df[col].dtype, 'tz') and df[col].dtype.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)

        df.to_excel(writer, sheet_name="WF_Train_Trades", index=False)
    else:
        df = pd.DataFrame()
        df.to_excel(writer, sheet_name="WF_Train_Trades", index=False)


def _write_equity_sheet(result: WalkForwardResult, writer: pd.ExcelWriter) -> None:
    """Write WF_Equity sheet with OOS equity curve."""
    if result.oos_equity_curve is not None and len(result.oos_equity_curve) > 0:
        equity = result.oos_equity_curve
        
        # Create DataFrame with bar index and equity
        df = pd.DataFrame({
            "Bar": range(len(equity)),
            "Equity": equity
        })
        
        format_excel_export_df(df).to_excel(writer, sheet_name="WF_Equity", index=False)
    else:
        # Create empty DataFrame
        df = pd.DataFrame({"Bar": [], "Equity": []})
        df.to_excel(writer, sheet_name="WF_Equity", index=False)


def _write_gates_sheet(result: WalkForwardResult, writer: pd.ExcelWriter) -> None:
    """Write WF_Gates sheet with gate check details."""
    if result.gates_result is None or not result.gates_result.enabled:
        return
    
    rows = []
    for check in result.gates_result.checks:
        row = {
            "Name": check.name,
            "Passed": "PASS" if check.passed else "FAIL",
            "Value": check.value if check.value is not None else np.nan,
            "Threshold": check.threshold if check.threshold is not None else np.nan,
            "Message": check.message if check.message else ""
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    format_excel_export_df(df).to_excel(writer, sheet_name="WF_Gates", index=False)


def _append_cag_diagnostics_to_gates(
    writer: pd.ExcelWriter,
    canonical_df: pd.DataFrame,
    config: dict,
) -> None:
    """Append coverage-aware gating diagnostics rows to the WF_Gates sheet.

    Called after the canonical sheet is built so that gate_status counts are
    available.  If WF_Gates sheet does not exist (gates not enabled), no-ops.
    """
    if "WF_Gates" not in writer.sheets:
        return

    gates_cfg = (
        config.get("walk_forward", {})
        .get("canonical", {})
        .get("gates", {})
    )
    cag_cfg = gates_cfg.get("coverage_aware_gating", {})
    cag_enabled = bool(cag_cfg.get("enabled", False))

    from supertrend_optimizer.scoring.constants import GATE_FAIL, GATE_INSUFFICIENT, GATE_PASS

    gs = canonical_df.get("gate_status", pd.Series(dtype=str)) if canonical_df is not None else pd.Series(dtype=str)

    rows = [
        {
            "Name": "coverage_aware_gating",
            "Passed": "ENABLED" if cag_enabled else "DISABLED",
            "Value": np.nan,
            "Threshold": np.nan,
            "Message": f"min_coverage_for_hard_gate={cag_cfg.get('min_coverage_for_hard_gate', 3)}, "
                       f"confidence_beta={cag_cfg.get('confidence_beta', 0.5)}, "
                       f"confidence_min={cag_cfg.get('confidence_min', 0.3)}",
        },
        {
            "Name": "gate_status_PASS_count",
            "Passed": np.nan,
            "Value": int((gs == GATE_PASS).sum()),
            "Threshold": np.nan,
            "Message": "candidates with gate_status == PASS",
        },
        {
            "Name": "gate_status_INSUFFICIENT_count",
            "Passed": np.nan,
            "Value": int((gs == GATE_INSUFFICIENT).sum()),
            "Threshold": np.nan,
            "Message": "candidates with gate_status == INSUFFICIENT_OOS_DATA",
        },
        {
            "Name": "gate_status_FAIL_count",
            "Passed": np.nan,
            "Value": int((gs == GATE_FAIL).sum()),
            "Threshold": np.nan,
            "Message": "candidates with gate_status == FAIL",
        },
        {
            "Name": "soft_pnl_gate",
            "Passed": "ENABLED" if bool(gates_cfg.get("soft_pnl_gate_enabled", False)) else "DISABLED",
            "Value": gates_cfg.get("soft_pnl_gate_max_violations", np.nan),
            "Threshold": np.nan,
            "Message": f"mode={gates_cfg.get('soft_pnl_gate_mode', 'n/a')}",
        },
    ]

    ws = writer.sheets["WF_Gates"]
    # openpyxl: append below existing rows
    try:
        existing_rows = ws.max_row
        cag_df = format_excel_export_df(pd.DataFrame(rows))
        for r_idx, row_data in enumerate(cag_df.itertuples(index=False), start=existing_rows + 2):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val if not (isinstance(val, float) and np.isnan(val)) else None)
    except Exception:
        logger.debug("_append_cag_diagnostics_to_gates: failed to write rows", exc_info=True)


def _write_imputation_sheet(
    step_df: pd.DataFrame,
    writer: pd.ExcelWriter,
    total_steps: int,
) -> None:
    """
    Write WF_Imputation sheet with per-candidate and run-level imputation stats.

    Sheet layout:
        Section 1 (rows 1+): Per-candidate summary
            atr_period, multiplier, n_steps_total, n_real, n_neighbors,
            n_fallback_worst, n_missing_row, n_invalid_original,
            fallback_share, neighbors_share, coverage_ratio
        Blank row separator
        Section 2: Run-level totals per OOS metric
            metric, n_real, n_neighbors, n_fallback_worst, n_total_cells,
            pct_real, pct_neighbors, pct_fallback

    Parameters
    ----------
    step_df : pd.DataFrame
        Filled step-level table (output of fill_missing_step_metrics).
    writer : pd.ExcelWriter
        Excel writer object.
    total_steps : int
        Total number of walk-forward windows.
    """
    from supertrend_optimizer.imputation.provenance import (
        build_imputation_summary,
        build_imputation_run_totals,
    )

    if len(step_df) == 0:
        pd.DataFrame(columns=["atr_period", "multiplier", "n_steps_total",
                               "n_real", "n_neighbors", "n_fallback_worst",
                               "fallback_share"]).to_excel(
            writer, sheet_name="WF_Imputation", index=False
        )
        return

    per_candidate = build_imputation_summary(step_df, total_steps)
    run_totals = build_imputation_run_totals(step_df, total_steps)

    # Write both sections to the same sheet using startrow
    format_excel_export_df(per_candidate).to_excel(writer, sheet_name="WF_Imputation", index=False, startrow=0)
    # Section 2 starts after per_candidate rows + 2 blank rows for separation
    section2_start = len(per_candidate) + 3  # header + data rows + 2 blank
    format_excel_export_df(run_totals).to_excel(writer, sheet_name="WF_Imputation", index=False, startrow=section2_start)


def _build_stability_heatmap(
    bucket_step_df: pd.DataFrame,
    atr_bucket_step: int = DEFAULT_ATR_BUCKET_STEP,
    mult_bucket_step: float = DEFAULT_MULT_BUCKET_STEP,
) -> pd.DataFrame:
    """Build ATR × Mult stability heatmap DataFrame.

    Rows = atr_bucket (sorted ascending), columns = mult_bucket_ticks (sorted ascending).
    Cell values = bucket_stability_score.  Missing combinations are NaN.
    Top-left index label is "ATR \\ MULT".
    Index and column labels are human-readable ranges (e.g. "23–25", "2.6–2.7").
    """
    required = {"atr_bucket", "mult_bucket_ticks", "bucket_stability_score"}
    if bucket_step_df.empty or not required.issubset(bucket_step_df.columns):
        return pd.DataFrame()

    pivot = bucket_step_df.pivot_table(
        index="atr_bucket",
        columns="mult_bucket_ticks",
        values="bucket_stability_score",
        aggfunc="first",
    )
    pivot = pivot.sort_index(axis=0).sort_index(axis=1)

    # Replace raw bucket integers with human-readable range labels
    pivot.index = [
        _format_atr_range(int(v), atr_bucket_step) for v in pivot.index
    ]
    pivot.columns = [
        _format_mult_range(int(v), mult_bucket_step) for v in pivot.columns
    ]

    pivot.index.name = "ATR \\ MULT"
    pivot.columns.name = None
    return pivot


def _build_stability_zone(
    bucket_step_df: pd.DataFrame,
    atr_bucket_step: int = DEFAULT_ATR_BUCKET_STEP,
    mult_bucket_step: float = DEFAULT_MULT_BUCKET_STEP,
) -> pd.DataFrame:
    """Identify the most stable parameter region (top-30% buckets).

    Returns a two-column DataFrame with Metric / Value rows.
    """
    required = {"atr_bucket", "mult_bucket_ticks", "bucket_stability_score"}
    if bucket_step_df.empty or not required.issubset(bucket_step_df.columns):
        return pd.DataFrame()

    bdf = bucket_step_df.dropna(subset=["bucket_stability_score"]).copy()
    if bdf.empty:
        return pd.DataFrame()

    threshold_idx = max(1, int(np.ceil(len(bdf) * 0.30)))
    top_buckets = bdf.nlargest(threshold_idx, "bucket_stability_score")

    atr_min = int(top_buckets["atr_bucket"].min())
    atr_max = int(top_buckets["atr_bucket"].max())
    mult_min = int(top_buckets["mult_bucket_ticks"].min())
    mult_max = int(top_buckets["mult_bucket_ticks"].max())
    atr_center_bucket = int(np.median(top_buckets["atr_bucket"]))
    mult_center_bucket = int(np.median(top_buckets["mult_bucket_ticks"]))
    zone_bucket_count = len(top_buckets)
    zone_mean_stability = round(float(top_buckets["bucket_stability_score"].mean()), 6)

    # Human-readable ranges using canonical formatters
    atr_range_display = (
        f"{_format_atr_range(atr_min, atr_bucket_step)}–{_format_atr_range(atr_max, atr_bucket_step)}"
    )
    mult_range_display = (
        f"{_format_mult_range(mult_min, mult_bucket_step)}–{_format_mult_range(mult_max, mult_bucket_step)}"
    )
    center_display = _format_bucket_label(atr_center_bucket, mult_center_bucket, atr_bucket_step, mult_bucket_step)

    rows = [
        {"Metric": "ATR range",           "Value": atr_range_display},
        {"Metric": "Multiplier range",     "Value": mult_range_display},
        {"Metric": "Zone center",          "Value": center_display},
        {"Metric": "Buckets in zone",      "Value": zone_bucket_count},
        {"Metric": "Mean stability score", "Value": zone_mean_stability},
    ]
    return pd.DataFrame(rows)


def _write_wf_ensemble_sheet(
    div_result: Any,
    writer: pd.ExcelWriter,
) -> None:
    """Write WF_ENSEMBLE sheet.

    Sheet layout:
        Table A — Selected Models
        [blank row]
        Table B — Pair Diagnostics
        [blank row]
        Table C — WF Step Winners matrix
        [blank row]
        Title: BUCKET AGGREGATED STEP PERFORMANCE
        Table D — Bucket-aggregated step performance
        [blank row]
        Title: PARAMETER STABILITY HEATMAP (ATR × MULT BUCKET)
        Heatmap — stability score pivot
        [blank row]
        Title: PARAMETER STABILITY ZONE
        Zone table — top-30% bucket statistics

    Placeholder mode writes metadata + column headers only.
    """
    from openpyxl.styles import Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "WF_ENSEMBLE"

    if div_result.placeholder or not div_result.enabled:
        # Stable schema: write metadata row + headers for each table
        meta_a = pd.DataFrame([{"# diversification": "disabled"}])
        meta_a.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

        header_a = pd.DataFrame(columns=[
            "atr_period", "multiplier", "canonical_rank", "final_score",
            "robust_score", "regime_score", "coverage_ratio_bucket",
            "atr_bucket", "mult_bucket_ticks",
            "in_ensemble", "ensemble_rank", "portfolio_score",
        ])
        header_b = pd.DataFrame(columns=[
            "model_A", "model_B", "corr_returns",
            "trade_overlap_rate", "phase_similarity", "pair_passed",
        ])
        header_c = pd.DataFrame(columns=["param", "wins_count", "win_steps", "top3_count"])

        start_b = 4   # meta row + header + 1 blank
        start_c = start_b + 3

        header_a.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        header_b.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_b)
        header_c.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_c)
        return

    # --- Full mode ---
    candidates_df       = div_result.candidates_df
    pair_diagnostics_df = div_result.pair_diagnostics_df
    step_winners_df     = div_result.step_winners_df
    bucket_step_df      = getattr(div_result, "bucket_step_df", pd.DataFrame())

    # ------------------------------------------------------------------ #
    # Optional feature metadata summary block                             #
    # ------------------------------------------------------------------ #
    repr_mode            = getattr(div_result, "representative_mode", "off")
    repr_pool_size       = getattr(div_result, "representative_pool_size", None)
    ss_enabled           = getattr(div_result, "subset_search_enabled", False)
    ss_subsets_evaluated = getattr(div_result, "subsets_evaluated", None)
    ss_best_score        = getattr(div_result, "best_subset_score", None)

    meta_rows = []
    if repr_mode != "off":
        meta_rows.append({
            "# feature": "representative_mode",
            "value": repr_mode,
        })
        if repr_pool_size is not None:
            meta_rows.append({
                "# feature": "representative_pool_size",
                "value": repr_pool_size,
            })
    if ss_enabled:
        meta_rows.append({
            "# feature": "subset_search_enabled",
            "value": True,
        })
        if ss_subsets_evaluated is not None:
            meta_rows.append({
                "# feature": "subsets_evaluated",
                "value": ss_subsets_evaluated,
            })
        if ss_best_score is not None:
            meta_rows.append({
                "# feature": "best_subset_score",
                "value": round(ss_best_score, 6),
            })

    meta_row_offset = 0
    if meta_rows:
        meta_df = pd.DataFrame(meta_rows)
        meta_df.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=0
        )
        meta_row_offset = len(meta_rows) + 2  # data rows + header + blank

    # ------------------------------------------------------------------ #
    # Write tables via pandas (data + headers)                            #
    # ------------------------------------------------------------------ #

    # Table A
    format_excel_export_df(candidates_df).to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=meta_row_offset
    )

    # Table B — starts after Table A + 1 blank row
    start_b = meta_row_offset + len(candidates_df) + 2  # meta + header row + data rows + 1 blank
    format_excel_export_df(pair_diagnostics_df).to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_b
    )

    # Table C — starts after Table B + 1 blank row
    start_c = start_b + len(pair_diagnostics_df) + 2
    format_excel_export_df(step_winners_df).to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_c
    )

    # Determine where Table D starts (always computed for row tracking)
    start_title_d = start_c + len(step_winners_df) + 2
    start_d = start_title_d + 2  # title row + header row offset handled by to_excel

    # Table D — bucket-aggregated step performance
    # atr_range_label / mult_range_label / bucket_range are internal display
    # helpers; they are excluded from Excel to avoid redundancy with bucket_param.
    # atr_bucket / mult_bucket_ticks / bucket_key are identity columns — kept visible (I-5).
    _INTERNAL_COLS_D = {"atr_range_label", "mult_range_label", "bucket_range"}
    _display_cols_d: list = []
    if not bucket_step_df.empty:
        step_cols_d = [c for c in bucket_step_df.columns if c.startswith("Step")]
        _preferred_order = (
            ["bucket_param", "bucket_key", "atr_bucket", "mult_bucket_ticks",
             "bucket_size", "bucket_stability_score"]
            + step_cols_d
            + ["wins_count", "win_steps", "top3_count",
               "above_median_count", "above_median_ratio", "dominance_score"]
        )
        _display_cols_d = [
            c for c in _preferred_order
            if c in bucket_step_df.columns and c not in _INTERNAL_COLS_D
        ]
        # Append any remaining columns not covered by preferred order
        _remaining = [
            c for c in bucket_step_df.columns
            if c not in _display_cols_d and c not in _INTERNAL_COLS_D
        ]
        _display_cols_d.extend(_remaining)
        title_d_df = pd.DataFrame([{"bucket_param": "BUCKET AGGREGATED STEP PERFORMANCE"}])
        title_d_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_d,
        )
        bucket_step_df[_display_cols_d].to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=start_d,
        )

    # Heatmap section
    heatmap_df = _build_stability_heatmap(bucket_step_df)
    start_title_hm = start_d + (len(bucket_step_df) + 2 if not bucket_step_df.empty else 2)
    start_hm = start_title_hm + 2  # title + blank before data

    if not heatmap_df.empty:
        title_hm_df = pd.DataFrame(
            [{"ATR \\ MULT": "PARAMETER STABILITY HEATMAP (ATR × MULT BUCKET)"}]
        )
        title_hm_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_hm,
        )
        heatmap_df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=start_hm)

    # Stability zone section
    zone_df = _build_stability_zone(bucket_step_df)
    hm_rows = (len(heatmap_df) + 1) if not heatmap_df.empty else 0  # +1 for index header row
    start_title_zone = start_hm + hm_rows + 2
    start_zone = start_title_zone + 2

    if not zone_df.empty:
        title_zone_df = pd.DataFrame([{"Metric": "PARAMETER STABILITY ZONE"}])
        title_zone_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_zone,
        )
        zone_df.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=start_zone,
        )

    # ------------------------------------------------------------------ #
    # Post-write openpyxl formatting                                      #
    # ------------------------------------------------------------------ #
    ws = writer.sheets[sheet_name]

    # --- Freeze panes at Table C header row ---
    # Table C header is at start_c (0-based) → Excel row = start_c + 1
    # Freeze below that row so headers stay visible
    freeze_row = start_c + 2  # +1 for 1-based, +1 to freeze below header
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # --- Autofilter on Table C header row ---
    if not step_winners_df.empty:
        n_cols_c = len(step_winners_df.columns)
        filter_ref = (
            f"A{start_c + 1}:"
            f"{get_column_letter(n_cols_c)}{start_c + 1 + len(step_winners_df)}"
        )
        ws.auto_filter.ref = filter_ref

    # --- Autofilter on Table D header row (same as Table C) ---
    if not bucket_step_df.empty and _display_cols_d:
        n_cols_d = len(_display_cols_d)
        filter_ref_d = (
            f"A{start_d + 1}:"
            f"{get_column_letter(n_cols_d)}{start_d + 1 + len(bucket_step_df)}"
        )
        ws.auto_filter.ref = filter_ref_d

    # --- Auto-adjust column widths ---
    col_widths: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col_letter = get_column_letter(cell.column)
            cell_len = len(str(cell.value)) + 2
            if col_widths.get(col_letter, 0) < cell_len:
                col_widths[col_letter] = cell_len
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = min(width, 60)

    # --- Numeric formats ---
    # Detect Step column indices in Table C
    step_col_indices_c: list = []
    if not step_winners_df.empty:
        for ci, col_name in enumerate(step_winners_df.columns, start=1):
            if col_name.startswith("Step"):
                step_col_indices_c.append(ci)

    # Format Step cells in Table C
    for ci in step_col_indices_c:
        for ri in range(start_c + 2, start_c + 2 + len(step_winners_df)):
            cell = ws.cell(row=ri, column=ci)
            cell.number_format = "0.00"

    # Format Step cells and bucket_stability_score in Table D
    if not bucket_step_df.empty and _display_cols_d:
        for ci, col_name in enumerate(_display_cols_d, start=1):
            if col_name.startswith("Step"):
                for ri in range(start_d + 2, start_d + 2 + len(bucket_step_df)):
                    ws.cell(row=ri, column=ci).number_format = "0.00"
            elif col_name == "bucket_stability_score":
                for ri in range(start_d + 2, start_d + 2 + len(bucket_step_df)):
                    ws.cell(row=ri, column=ci).number_format = "0.000000"

    # --- Heatmap conditional formatting + alignment ---
    if not heatmap_df.empty:
        n_hm_rows = len(heatmap_df)
        n_hm_cols = len(heatmap_df.columns)
        # Data starts at column 2 (col 1 = index label), row start_hm+2 (1-based header+1)
        hm_data_row_start = start_hm + 2
        hm_data_row_end   = start_hm + 1 + n_hm_rows
        hm_col_start = 2
        hm_col_end   = 1 + n_hm_cols

        hm_range = (
            f"{get_column_letter(hm_col_start)}{hm_data_row_start}:"
            f"{get_column_letter(hm_col_end)}{hm_data_row_end}"
        )

        ws.conditional_formatting.add(
            hm_range,
            ColorScaleRule(
                start_type="min",  start_color="FFFFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF99",
                end_type="max",    end_color="FF00AA00",
            ),
        )

        # Center-align heatmap cells and set column width ≈ 11
        for ri in range(start_hm + 1, start_hm + 2 + n_hm_rows):
            for ci in range(1, hm_col_end + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.alignment = Alignment(horizontal="center")
        for ci in range(hm_col_start, hm_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11


def _write_bucket_matrix_full_sheet(
    full_bucket_df: pd.DataFrame,
    writer: pd.ExcelWriter,
) -> None:
    """Write BucketMatrix_Full sheet.

    Sheet layout:
        Row 1  — column headers (autofilter on full table range)
        Rows 2+ — one row per (atr_bucket × mult_bucket_ticks), sorted by
                  bucket_stability_score DESC before writing
        [blank row]
        Title row: PARAMETER STABILITY HEATMAP (ATR × MULT BUCKET) — bold, size 12
        [blank row]
        Heatmap — stability score pivot (ATR rows × Mult columns)

    Freeze panes: E2 (header row + first 4 columns frozen).
    Column widths: fixed spec + auto-widen for bucket_param if needed.
    Numeric formats: Step* → "0.000000", ratio/score cols → "0.000000".
    NaN cells stay empty (no zero substitution).
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "BucketMatrix_Full"

    # ------------------------------------------------------------------ #
    # Placeholder for empty input                                         #
    # ------------------------------------------------------------------ #
    _PLACEHOLDER_COLS = [
        "bucket_param", "atr_bucket", "mult_bucket_ticks", "bucket_size",
        "bucket_presence_steps", "mean_oos_pnl",
        "wins_count", "win_steps", "top3_count",
        "above_median_count", "above_median_ratio",
        "bucket_stability_score", "dominance_score",
    ]
    if full_bucket_df.empty:
        pd.DataFrame(columns=_PLACEHOLDER_COLS).to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=0,
        )
        ws = writer.sheets[sheet_name]
        # Freeze and autofilter still work on an empty table
        ws.freeze_panes = ws.cell(row=2, column=5)
        n_ph = len(_PLACEHOLDER_COLS)
        ws.auto_filter.ref = f"A1:{get_column_letter(n_ph)}1"
        return

    # ------------------------------------------------------------------ #
    # Sort by bucket_stability_score DESC before writing                  #
    # ------------------------------------------------------------------ #
    if "bucket_stability_score" in full_bucket_df.columns:
        display_df = full_bucket_df.sort_values(
            "bucket_stability_score", ascending=False
        ).reset_index(drop=True)
    else:
        display_df = full_bucket_df.copy()

    # ------------------------------------------------------------------ #
    # Main table                                                          #
    # ------------------------------------------------------------------ #
    start_main = 0  # 0-based row for pandas startrow
    display_df = format_excel_export_df(display_df)
    display_df.to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_main,
    )

    # ------------------------------------------------------------------ #
    # Heatmap section                                                     #
    # ------------------------------------------------------------------ #
    heatmap_df = _build_stability_heatmap(full_bucket_df)
    # +1 for header row written by to_excel, +1 blank separator
    start_title_hm = start_main + len(display_df) + 2
    start_hm = start_title_hm + 2  # title row + blank before heatmap header

    if not heatmap_df.empty:
        title_hm_df = pd.DataFrame(
            [{"ATR \\ MULT": "PARAMETER STABILITY HEATMAP (ATR × MULT BUCKET)"}]
        )
        title_hm_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_hm,
        )
        heatmap_df.to_excel(
            writer, sheet_name=sheet_name, index=True, startrow=start_hm,
        )

    # ------------------------------------------------------------------ #
    # Post-write openpyxl formatting                                      #
    # ------------------------------------------------------------------ #
    ws = writer.sheets[sheet_name]

    # ── 1. Freeze panes: header row + first 4 columns → E2 ────────────
    ws.freeze_panes = ws.cell(row=2, column=5)

    # ── 2. Autofilter: full table range (header + data rows) ──────────
    n_cols = len(display_df.columns)
    n_data_rows = len(display_df)
    if n_cols > 0:
        # Range covers header row (row 1) through last data row
        filter_last_row = start_main + 1 + n_data_rows  # 1-based
        ws.auto_filter.ref = (
            f"A{start_main + 1}:{get_column_letter(n_cols)}{filter_last_row}"
        )

    # ── 3. Column widths ───────────────────────────────────────────────
    # Fixed spec widths keyed by column name
    _FIXED_WIDTHS = {
        "bucket_param":           28,
        "atr_bucket":             12,
        "mult_bucket_ticks":      16,
        "bucket_size":            12,
        "bucket_presence_steps":  20,
        "mean_oos_pnl":           14,
        "wins_count":             10,
        "win_steps":              16,
        "top3_count":             10,
        "above_median_count":     16,
        "above_median_ratio":     18,
        "bucket_stability_score": 22,
        "dominance_score":        18,
    }
    _WF_COL_WIDTH = 12  # default for Step* / WF* columns

    for ci, col_name in enumerate(display_df.columns, start=1):
        col_letter = get_column_letter(ci)
        if col_name in _FIXED_WIDTHS:
            width = _FIXED_WIDTHS[col_name]
        elif col_name.startswith("Step"):
            width = _WF_COL_WIDTH
        else:
            # Fallback: auto-size from cell content (capped at 40)
            max_len = len(str(col_name)) + 2
            for ri in range(start_main + 2, start_main + 2 + n_data_rows):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)) + 2)
            width = min(max_len, 40)
        ws.column_dimensions[col_letter].width = width

    # ── 4. Numeric formats ─────────────────────────────────────────────
    _step_cols_set  = {c for c in display_df.columns if c.startswith("Step")}
    _score_cols_set = {"above_median_ratio", "bucket_stability_score", "dominance_score"}
    _pnl_cols_set   = {"mean_oos_pnl"}
    _int_cols_set   = {"atr_bucket", "mult_bucket_ticks", "bucket_size",
                       "wins_count", "top3_count", "above_median_count"}

    data_row_start = start_main + 2               # 1-based first data row
    data_row_end   = start_main + 1 + n_data_rows  # 1-based last data row

    for ci, col_name in enumerate(display_df.columns, start=1):
        if col_name in _step_cols_set:
            fmt = "0.0"          # compact — step values are OOS pnl %
        elif col_name in _score_cols_set:
            fmt = "0.000000"
        elif col_name in _pnl_cols_set:
            fmt = "0.00"
        elif col_name in _int_cols_set:
            fmt = "0"
        else:
            continue
        for ri in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = fmt

    # ── 5. Heatmap title: bold, font size 12, left-aligned ────────────
    if not heatmap_df.empty:
        title_excel_row = start_title_hm + 1  # 1-based
        title_cell = ws.cell(row=title_excel_row, column=1)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left")

    # ── 6. Heatmap conditional formatting + alignment ──────────────────
    if not heatmap_df.empty:
        n_hm_rows = len(heatmap_df)
        n_hm_cols = len(heatmap_df.columns)
        # Heatmap data starts one row below the header row written by to_excel
        hm_data_row_start = start_hm + 2   # 1-based
        hm_data_row_end   = start_hm + 1 + n_hm_rows
        hm_col_start = 2                   # column 1 = ATR index label
        hm_col_end   = 1 + n_hm_cols

        hm_range = (
            f"{get_column_letter(hm_col_start)}{hm_data_row_start}:"
            f"{get_column_letter(hm_col_end)}{hm_data_row_end}"
        )
        ws.conditional_formatting.add(
            hm_range,
            ColorScaleRule(
                start_type="min",      start_color="FFFFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF99",
                end_type="max",        end_color="FF00AA00",
            ),
        )
        # Center-align all heatmap cells (header + data)
        for ri in range(start_hm + 1, start_hm + 2 + n_hm_rows):
            for ci in range(1, hm_col_end + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
        # Fixed column width for heatmap mult columns
        for ci in range(hm_col_start, hm_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11


def _write_bucket_matrix_median_sheet(
    median_bucket_df: pd.DataFrame,
    writer: pd.ExcelWriter,
) -> None:
    """Write BucketMatrix_Median sheet.

    Median-based companion to BucketMatrix_Full.  The winner-based sheet
    (BucketMatrix_Full) shows max(oos_sum_pnl_pct) per bucket per step;
    this sheet shows median(oos_sum_pnl_pct) — reflecting zone strength
    rather than the best local winner.

    Extra columns vs BucketMatrix_Full:
        std_bucket            — population std of all raw step pnl values
                                across all params in the bucket.
        pct_params_positive_pnl — fraction of params with positive mean
                                  observed OOS PnL (conditional on top-K
                                  presence; no zero-fill for absent steps).
        zone_dominance_score  — same formula as dominance_score on the
                                winner-based sheet but computed from median
                                step values (renamed to avoid confusion).

    Sheet layout mirrors BucketMatrix_Full:
        Row 1  — column headers (autofilter)
        Rows 2+ — one row per bucket, sorted by bucket_stability_score DESC
        [blank row]
        Title row: ZONE STRENGTH HEATMAP (MEDIAN-BASED, ATR × MULT BUCKET)
        [blank row]
        Heatmap — stability score pivot (ATR rows × Mult columns)

    Freeze panes: E2.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "BucketMatrix_Median"

    _PLACEHOLDER_COLS = [
        "bucket_param", "atr_bucket", "mult_bucket_ticks", "bucket_size",
        "bucket_presence_steps", "mean_oos_pnl", "std_bucket",
        "pct_params_positive_pnl",
        "wins_count", "win_steps", "top3_count",
        "above_median_count", "above_median_ratio",
        "bucket_stability_score", "zone_dominance_score",
    ]
    if median_bucket_df.empty:
        pd.DataFrame(columns=_PLACEHOLDER_COLS).to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=0,
        )
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = ws.cell(row=2, column=5)
        n_ph = len(_PLACEHOLDER_COLS)
        ws.auto_filter.ref = f"A1:{get_column_letter(n_ph)}1"
        return

    # Sort by bucket_stability_score DESC
    if "bucket_stability_score" in median_bucket_df.columns:
        display_df = median_bucket_df.sort_values(
            "bucket_stability_score", ascending=False
        ).reset_index(drop=True)
    else:
        display_df = median_bucket_df.copy()

    start_main = 0
    display_df = format_excel_export_df(display_df)
    display_df.to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_main,
    )

    # Heatmap section
    heatmap_df = _build_stability_heatmap(median_bucket_df)
    start_title_hm = start_main + len(display_df) + 2
    start_hm = start_title_hm + 2

    if not heatmap_df.empty:
        title_hm_df = pd.DataFrame(
            [{"ATR \\ MULT": "ZONE STRENGTH HEATMAP (MEDIAN-BASED, ATR \u00d7 MULT BUCKET)"}]
        )
        title_hm_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_hm,
        )
        heatmap_df.to_excel(
            writer, sheet_name=sheet_name, index=True, startrow=start_hm,
        )

    ws = writer.sheets[sheet_name]

    # Freeze panes: header row + first 4 columns → E2
    ws.freeze_panes = ws.cell(row=2, column=5)

    # Autofilter
    n_cols = len(display_df.columns)
    n_data_rows = len(display_df)
    if n_cols > 0:
        filter_last_row = start_main + 1 + n_data_rows
        ws.auto_filter.ref = (
            f"A{start_main + 1}:{get_column_letter(n_cols)}{filter_last_row}"
        )

    # Column widths
    _FIXED_WIDTHS = {
        "bucket_param":              28,
        "atr_bucket":                12,
        "mult_bucket_ticks":         16,
        "bucket_size":               12,
        "bucket_presence_steps":     20,
        "mean_oos_pnl":              14,
        "std_bucket":                14,
        "pct_params_positive_pnl":   22,
        "wins_count":                10,
        "win_steps":                 16,
        "top3_count":                10,
        "above_median_count":        16,
        "above_median_ratio":        18,
        "bucket_stability_score":    22,
        "zone_dominance_score":      22,
    }
    _WF_COL_WIDTH = 12

    for ci, col_name in enumerate(display_df.columns, start=1):
        col_letter = get_column_letter(ci)
        if col_name in _FIXED_WIDTHS:
            width = _FIXED_WIDTHS[col_name]
        elif col_name.startswith("Step"):
            width = _WF_COL_WIDTH
        else:
            max_len = len(str(col_name)) + 2
            for ri in range(start_main + 2, start_main + 2 + n_data_rows):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)) + 2)
            width = min(max_len, 40)
        ws.column_dimensions[col_letter].width = width

    # Numeric formats
    _step_cols_set  = {c for c in display_df.columns if c.startswith("Step")}
    _score_cols_set = {
        "above_median_ratio", "bucket_stability_score",
        "zone_dominance_score", "pct_params_positive_pnl",
    }
    _pnl_cols_set   = {"mean_oos_pnl", "std_bucket"}
    _int_cols_set   = {"atr_bucket", "mult_bucket_ticks", "bucket_size",
                       "wins_count", "top3_count", "above_median_count"}

    data_row_start = start_main + 2
    data_row_end   = start_main + 1 + n_data_rows

    for ci, col_name in enumerate(display_df.columns, start=1):
        if col_name in _step_cols_set:
            fmt = "0.0"
        elif col_name in _score_cols_set:
            fmt = "0.000000"
        elif col_name in _pnl_cols_set:
            fmt = "0.00"
        elif col_name in _int_cols_set:
            fmt = "0"
        else:
            continue
        for ri in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = fmt

    # Heatmap title: bold, font size 12, left-aligned
    if not heatmap_df.empty:
        title_excel_row = start_title_hm + 1
        title_cell = ws.cell(row=title_excel_row, column=1)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left")

    # Heatmap conditional formatting + alignment
    if not heatmap_df.empty:
        n_hm_rows = len(heatmap_df)
        n_hm_cols = len(heatmap_df.columns)
        hm_data_row_start = start_hm + 2
        hm_data_row_end   = start_hm + 1 + n_hm_rows
        hm_col_start = 2
        hm_col_end   = 1 + n_hm_cols

        hm_range = (
            f"{get_column_letter(hm_col_start)}{hm_data_row_start}:"
            f"{get_column_letter(hm_col_end)}{hm_data_row_end}"
        )
        ws.conditional_formatting.add(
            hm_range,
            ColorScaleRule(
                start_type="min",      start_color="FFFFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF99",
                end_type="max",        end_color="FF00AA00",
            ),
        )
        for ri in range(start_hm + 1, start_hm + 2 + n_hm_rows):
            for ci in range(1, hm_col_end + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
        for ci in range(hm_col_start, hm_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11


def _write_bucket_train_surface_sheet(
    surface_bucket_df: pd.DataFrame,
    writer: pd.ExcelWriter,
    config: dict,
) -> None:
    """Write BucketTrainSurface sheet.

    Surface-layer companion to BucketMatrix_Median.  Built from all_step_trials
    (the full search-space surface across all WF steps) rather than top-k winners.

    Layout
    ------
    Row 1  — metadata header (source semantics notice)
    Row 2  — blank
    Row 3  — column headers (autofilter on data table)
    Rows 4+ — one row per bucket, sorted by surface_stability_score DESC
    [blank row]
    Title: SURFACE STABILITY HEATMAP (ATR x MULT BUCKET)
    [blank row]
    Heatmap — surface_stability_score pivot (ATR rows x Mult columns)

    Summary columns written (before Step* columns)
    -----------------------------------------------
    bucket_param, atr_bucket, mult_bucket_ticks,
    surface_presence_count, surface_presence_ratio,
    surface_point_count_mean, surface_fill_ratio_mean,
    reliable_steps_count, reliable_steps_ratio,
    surface_above_median_count, surface_above_median_ratio,
    surface_stability_score

    Freeze panes: E4 (3 header rows + first 4 data columns).
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "BucketTrainSurface"

    _PLACEHOLDER_COLS = [
        "bucket_param", "atr_bucket", "mult_bucket_ticks",
        "surface_presence_count", "surface_presence_ratio",
        "surface_point_count_mean", "surface_fill_ratio_mean",
        "reliable_steps_count", "reliable_steps_ratio",
        "surface_above_median_count", "surface_above_median_ratio",
        "surface_stability_score",
    ]

    if surface_bucket_df.empty:
        pd.DataFrame(columns=_PLACEHOLDER_COLS).to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=0,
        )
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = ws.cell(row=2, column=5)
        n_ph = len(_PLACEHOLDER_COLS)
        ws.auto_filter.ref = f"A1:{get_column_letter(n_ph)}1"
        return

    # Sort by surface_stability_score DESC
    if "surface_stability_score" in surface_bucket_df.columns:
        display_df = surface_bucket_df.sort_values(
            "surface_stability_score", ascending=False
        ).reset_index(drop=True)
    else:
        display_df = surface_bucket_df.copy()

    # --- metadata header (2 rows above data) ---
    # Row 1: source semantics notice
    # Row 2: blank
    # Row 3: column headers  (startrow=2 → 0-based → Excel row 3)
    _wf_cfg = config.get("walk_forward", {})
    _total_steps = display_df["surface_presence_count"].max() if "surface_presence_count" in display_df.columns else "?"
    _reliable_thr = _wf_cfg.get("surface", {}).get("reliable_fill_threshold", 0.5)
    _min_rel = _wf_cfg.get("surface", {}).get("min_reliable_ratio", 0.2)

    meta_rows = [
        {
            "A": (
                f"SOURCE: all_step_trials (TRAIN metrics, full search surface). "
                f"reliable_fill_threshold={_reliable_thr}, "
                f"min_reliable_ratio={_min_rel}. "
                f"Step cells = median(sum_pnl_pct) for structurally present buckets; "
                f"scoring metrics use reliable steps only."
            )
        },
        {"A": ""},
    ]

    start_meta = 0
    start_data = 2   # 0-based → Excel row 3

    # Write metadata rows manually (single-column label)
    pd.DataFrame([r.get("A", "") for r in meta_rows], columns=["note"]).to_excel(
        writer, sheet_name=sheet_name, index=False, header=False, startrow=start_meta,
    )

    # Write main data table
    display_df = format_excel_export_df(display_df)
    display_df.to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=start_data,
    )

    # Resolve canonical step values for label formatting in this sheet
    _cons_cfg = config.get("walk_forward", {}).get("consensus", {})
    _atr_bucket_step = int(_cons_cfg.get("atr_bucket_step", DEFAULT_ATR_BUCKET_STEP))
    _mult_bucket_step = float(_cons_cfg.get("mult_bucket_step", DEFAULT_MULT_BUCKET_STEP))

    # Heatmap section
    _hm_stability_col = "surface_stability_score"
    heatmap_df: pd.DataFrame = pd.DataFrame()
    if (
        _hm_stability_col in surface_bucket_df.columns
        and {"atr_bucket", "mult_bucket_ticks"}.issubset(surface_bucket_df.columns)
    ):
        _pivot_src = surface_bucket_df.copy()
        _pivot = _pivot_src.pivot_table(
            index="atr_bucket",
            columns="mult_bucket_ticks",
            values=_hm_stability_col,
            aggfunc="first",
        )
        if not _pivot.empty:
            _pivot.index = [_format_atr_range(int(v), _atr_bucket_step) for v in _pivot.index]
            _pivot.columns = [_format_mult_range(int(v), _mult_bucket_step) for v in _pivot.columns]
            _pivot.index.name = "ATR \\ MULT"
            heatmap_df = _pivot

    start_title_hm = start_data + len(display_df) + 2
    start_hm = start_title_hm + 2

    if not heatmap_df.empty:
        title_df = pd.DataFrame(
            [{"ATR \\ MULT": "SURFACE STABILITY HEATMAP (ATR \u00d7 MULT BUCKET)"}]
        )
        title_df.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False,
            startrow=start_title_hm,
        )
        heatmap_df.to_excel(
            writer, sheet_name=sheet_name, index=True, startrow=start_hm,
        )

    ws = writer.sheets[sheet_name]

    # Freeze panes: 2 meta rows + 1 header row + first 4 cols → E4
    ws.freeze_panes = ws.cell(row=start_data + 2, column=5)

    # Autofilter on data table header row
    n_cols = len(display_df.columns)
    n_data_rows = len(display_df)
    if n_cols > 0:
        header_excel_row = start_data + 1
        last_data_excel_row = start_data + 1 + n_data_rows
        ws.auto_filter.ref = (
            f"A{header_excel_row}:{get_column_letter(n_cols)}{last_data_excel_row}"
        )

    # Style metadata row 1 — italic, light grey fill
    _meta_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid")
    meta_cell = ws.cell(row=1, column=1)
    meta_cell.font = Font(italic=True, size=9)
    meta_cell.fill = _meta_fill
    meta_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30
    # Merge across all data columns for the meta label
    if n_cols > 1:
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=min(n_cols, 20)
        )

    # Column widths
    _FIXED_WIDTHS = {
        "bucket_param":               28,
        "atr_bucket":                 12,
        "mult_bucket_ticks":          16,
        "surface_presence_count":     22,
        "surface_presence_ratio":     22,
        "surface_point_count_mean":   24,
        "surface_fill_ratio_mean":    22,
        "reliable_steps_count":       22,
        "reliable_steps_ratio":       20,
        "surface_above_median_count": 26,
        "surface_above_median_ratio": 26,
        "surface_stability_score":    24,
    }
    _WF_COL_WIDTH = 12
    data_row_start_xl = start_data + 2
    data_row_end_xl   = start_data + 1 + n_data_rows

    for ci, col_name in enumerate(display_df.columns, start=1):
        col_letter = get_column_letter(ci)
        if col_name in _FIXED_WIDTHS:
            width = _FIXED_WIDTHS[col_name]
        elif col_name.startswith("Step"):
            width = _WF_COL_WIDTH
        else:
            max_len = len(str(col_name)) + 2
            for ri in range(data_row_start_xl, data_row_end_xl + 1):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)) + 2)
            width = min(max_len, 40)
        ws.column_dimensions[col_letter].width = width

    # Numeric formats
    _step_cols_set  = {c for c in display_df.columns if c.startswith("Step")}
    _ratio_cols_set = {
        "surface_presence_ratio", "surface_fill_ratio_mean",
        "reliable_steps_ratio", "surface_above_median_ratio",
        "surface_stability_score",
    }
    _mean_cols_set  = {"surface_point_count_mean"}
    _int_cols_set   = {
        "atr_bucket", "mult_bucket_ticks",
        "surface_presence_count", "reliable_steps_count",
        "surface_above_median_count",
    }

    for ci, col_name in enumerate(display_df.columns, start=1):
        if col_name in _step_cols_set:
            fmt = "0.0"
        elif col_name in _ratio_cols_set:
            fmt = "0.000000"
        elif col_name in _mean_cols_set:
            fmt = "0.0"
        elif col_name in _int_cols_set:
            fmt = "0"
        else:
            continue
        for ri in range(data_row_start_xl, data_row_end_xl + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = fmt

    # Heatmap title styling
    if not heatmap_df.empty:
        title_xl_row = start_title_hm + 1
        title_cell = ws.cell(row=title_xl_row, column=1)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left")

        n_hm_rows = len(heatmap_df)
        n_hm_cols = len(heatmap_df.columns)
        hm_data_row_start = start_hm + 2
        hm_data_row_end   = start_hm + 1 + n_hm_rows
        hm_col_start = 2
        hm_col_end   = 1 + n_hm_cols

        hm_range = (
            f"{get_column_letter(hm_col_start)}{hm_data_row_start}:"
            f"{get_column_letter(hm_col_end)}{hm_data_row_end}"
        )
        ws.conditional_formatting.add(
            hm_range,
            ColorScaleRule(
                start_type="min",      start_color="FFFFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF99",
                end_type="max",        end_color="FF00AA00",
            ),
        )
        for ri in range(start_hm + 1, start_hm + 2 + n_hm_rows):
            for ci in range(1, hm_col_end + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
        for ci in range(hm_col_start, hm_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11


def _write_wf_diversification_matrix_sheet(
    div_result: Any,
    writer: pd.ExcelWriter,
) -> None:
    """Write WF_DIVERSIFICATION_MATRIX sheet.

    Sheet layout (three blocks stacked vertically):
        Block 1 — Correlation Matrix (N×N)
        [blank row + label row]
        Block 2 — Trade Overlap Matrix (M×M)
        [blank row + label row]
        Block 3 — Signal Phase Similarity Matrix (M×M)

    When div_result.placeholder is True the sheet is written with a single
    metadata row and no matrix data.
    """
    sheet_name = "WF_DIVERSIFICATION_MATRIX"

    if div_result.placeholder or not div_result.enabled:
        meta = pd.DataFrame([{"# diversification": "disabled"}])
        meta.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        return

    current_row = 0

    def _write_matrix_block(
        mat: pd.DataFrame,
        label: str,
        startrow: int,
    ) -> int:
        """Write a label row then the matrix. Returns next available row."""
        if mat.empty:
            label_df = pd.DataFrame([{label: "(not computed)"}])
            label_df.to_excel(
                writer, sheet_name=sheet_name, index=True, startrow=startrow
            )
            return startrow + 3

        # Label row
        label_df = pd.DataFrame([{label: ""}])
        label_df.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=startrow
        )
        # Matrix (index = row labels, columns = col labels)
        mat.to_excel(
            writer, sheet_name=sheet_name, index=True,
            startrow=startrow + 2
        )
        return startrow + 2 + len(mat) + 2  # label + blank + matrix + blank

    current_row = _write_matrix_block(
        div_result.corr_matrix, "# Correlation Matrix (N x N)", current_row
    )
    current_row = _write_matrix_block(
        div_result.overlap_matrix, "# Trade Overlap Matrix (M x M)", current_row
    )
    current_row = _write_matrix_block(
        div_result.phase_matrix, "# Signal Phase Similarity Matrix (M x M)", current_row
    )

    # Block 4 — Correlation Metadata (long format, one row per pair)
    meta_df = getattr(div_result, "corr_meta_df", None)
    if meta_df is not None and not meta_df.empty:
        label_df = pd.DataFrame([{"# Correlation Metadata (pair detail)": ""}])
        label_df.to_excel(
            writer, sheet_name=sheet_name, index=False, startrow=current_row
        )
        meta_df.to_excel(
            writer, sheet_name=sheet_name, index=False,
            startrow=current_row + 2
        )


# =============================================================================
# Public helper — build canonical-scored candidates DataFrame
# =============================================================================

def build_candidates_df(
    result: WalkForwardResult,
    config: "dict | None" = None,
) -> pd.DataFrame:
    """
    Build the canonical-scored candidates DataFrame used for plateau bucket selection.

    Replicates the scoring logic from export_walk_forward_to_excel without
    writing any Excel output.  This is the single source of truth for
    canonical scoring; both the Excel export and the plateau orchestration
    call this path.

    Parameters
    ----------
    result : WalkForwardResult
        Completed Walk-Forward result.
    config : dict, optional
        Full run config (same object passed to export_walk_forward_to_excel).

    Returns
    -------
    pd.DataFrame
        Canonical-scored candidates table including ``atr_bucket`` and
        ``mult_bucket_ticks`` columns required by select_target_buckets.

    Raises
    ------
    RuntimeError
        If the scored DataFrame is missing required bucket columns after scoring.
    """
    if config is None:
        config = {}

    agg_df, _step_df = _build_aggregated_topk_table(result, config, _return_step_df=True)
    agg_df = _enrich_soft_gate_columns(agg_df, _step_df, config)

    consensus_cfg = config.get("walk_forward", {}).get("consensus", {})
    atr_bucket_step  = int(consensus_cfg.get("atr_bucket_step",  DEFAULT_ATR_BUCKET_STEP))
    mult_bucket_step = float(consensus_cfg.get("mult_bucket_step", DEFAULT_MULT_BUCKET_STEP))
    total_steps      = len(result.steps) if result is not None else None

    candidates_df = _calculate_canonical_rank_score(
        agg_df,
        atr_bucket_step=atr_bucket_step,
        mult_bucket_step=mult_bucket_step,
        config=config,
        total_steps=total_steps,
    )

    # Fallback: recompute bucket columns if missing (mirrors the same guard in
    # _write_wf_topk_canonical_sheet for legacy / empty-result edge cases).
    if (
        "mult_bucket_ticks" not in candidates_df.columns
        and "atr_period" in candidates_df.columns
        and "multiplier" in candidates_df.columns
    ):
        candidates_df = _apply_param_buckets(
            candidates_df, atr_step=atr_bucket_step, mult_step=mult_bucket_step
        )

    required_cols = {"atr_bucket", "mult_bucket_ticks"}
    missing = required_cols - set(candidates_df.columns)
    if missing:
        raise RuntimeError(
            f"build_candidates_df: missing required columns {missing}. "
            f"Available: {sorted(candidates_df.columns.tolist())}"
        )

    return candidates_df


# =============================================================================
# Plateau Analysis Excel export (Phase 5)
# =============================================================================

def _write_plateau_sheets(
    plateau_result: Any,
    writer: pd.ExcelWriter,
    config: dict,
    bucket_matrix_df: "Optional[pd.DataFrame]" = None,
    bucket_train_surface_df: "Optional[pd.DataFrame]" = None,
) -> None:
    """
    Write all Plateau Analysis sheets to the workbook.

    Dispatcher: calls individual sheet writers based on plateau_result content
    and config flags.  Does NOT run backtests — write-only function.

    Sheets written (in order, after BucketMatrix_Median):
        PlateauAnalysis        — summary table (always, if plateau_result has data)
        PlateauSurface         — parameter surface heatmaps (deep mode only)
        ProductionCandidates   — top-K candidates + funnel table (deep mode only)
        PlateauPivot_1..N      — per-bucket step pivots (controlled by export_pivots)

    Parameters
    ----------
    bucket_matrix_df : pd.DataFrame or None
        BucketMatrix_Median table, pre-built in export_walk_forward_to_excel.
        Passed through to ProductionCandidates sheet for funnel enrichment.
    bucket_train_surface_df : pd.DataFrame or None
        BucketTrainSurface table (available only when all_step_trials is present).
        Passed through to ProductionCandidates sheet for funnel enrichment.
    """
    pa_cfg = config.get("plateau_analysis", {})

    if not getattr(plateau_result, "buckets", None):
        logger.debug("_write_plateau_sheets: no buckets in plateau_result, skipping.")
        return

    # PlateauAnalysis — always
    _write_plateau_analysis_sheet(plateau_result, writer)

    # PlateauSurface — deep mode only, when surface data is present
    surface = getattr(plateau_result, "surface", None)
    if surface is not None and not getattr(surface, "raw_df", pd.DataFrame()).empty:
        _write_plateau_surface_sheet(surface, writer, config=config)

    # ProductionCandidates — deep mode only, when surface is present
    # Written regardless of whether candidates were actually found (shows
    # placeholder message if selection produced no results).
    if surface is not None:
        _write_production_candidates_sheet(
            surface=surface,
            plateau_summary_df=getattr(plateau_result, "summary_df", pd.DataFrame()),
            bucket_matrix_df=bucket_matrix_df,
            bucket_train_surface_df=bucket_train_surface_df,
            writer=writer,
        )

    # PlateauPivot_N — controlled by export_pivots flag
    if pa_cfg.get("export_pivots", True):
        max_sheets = int(pa_cfg.get("max_pivot_sheets", 5))
        _write_plateau_pivot_sheets(plateau_result, writer, max_sheets)


def _write_unified_error_placeholder(writer: pd.ExcelWriter) -> None:
    """Write a minimal UnifiedProductionRank sheet indicating a compute failure.

    Writes a human-readable error notice so the analyst knows the sheet
    was attempted but failed.  Full traceback is in the application log
    (logged via logger.exception at the call site).
    """
    SHEET_NAME = "UnifiedProductionRank"
    pd.DataFrame().to_excel(writer, sheet_name=SHEET_NAME, index=False)
    ws = writer.sheets[SHEET_NAME]
    ws["A1"] = "UnifiedProductionRank — COMPUTE ERROR"
    ws["A2"] = "Unified ranking computation failed. Check logs for details."


def _write_unified_production_rank_sheet(
    unified_df: pd.DataFrame,
    plateau_only_df: pd.DataFrame | None,
    writer: pd.ExcelWriter,
    config: dict,
) -> None:
    """Write UnifiedProductionRank sheet.

    Layout
    ------
    Row 1       : section header "UnifiedProductionRank"
    Row 2       : column headers for unified_df
    Rows 3..N+2 : data rows for unified_df
    Row N+4     : section header "PlateauOnly" (if plateau_only_df non-empty)
    Row N+5     : column headers for plateau_only_df
    Rows N+6..  : data rows for plateau_only_df

    Formatting
    ----------
    - Dark-teal header (#1F4E79 bg, white bold text)
    - Section banners in amber (#C65911 bg, white bold)
    - Color scale on unified_score (red→yellow→green, 0–1)
    - consensus_class cell fill: STRONG=green, MODERATE=yellow,
      WEAK=orange, BLOCKED=red
    - Freeze panes at row 3 (below section banner + header)
    - Auto-filter on header row
    - Column widths tuned to content
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    SHEET_NAME = "UnifiedProductionRank"
    HDR_BG = "1F4E79"   # dark blue — matches other ranking sheets
    HDR_FG = "FFFFFF"
    BANNER_BG = "C65911"  # amber — section separator
    BANNER_FG = "FFFFFF"

    _CONSENSUS_FILLS = {
        "GOLD":        PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid"),
        "SILVER":      PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid"),
        "BRONZE":      PatternFill(start_color="FFAA44", end_color="FFAA44", fill_type="solid"),
        "UNCONFIRMED": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }

    def _write_section_banner(ws, row: int, text: str, n_cols: int) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
        cell.font = Font(bold=True, color=BANNER_FG, size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx in range(2, n_cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.fill = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")

    def _write_header_row(ws, row: int, columns) -> None:
        header_fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")
        header_font = Font(bold=True, color=HDR_FG, size=10)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_data_rows(ws, df: pd.DataFrame, start_row: int) -> None:
        import math
        for r_offset, (_, row_data) in enumerate(df.iterrows()):
            for col_idx, value in enumerate(row_data, start=1):
                # openpyxl cannot serialize pd.NA / pd.NaT — convert to None
                if value is pd.NA or value is pd.NaT:
                    value = None
                elif isinstance(value, float) and math.isnan(value):
                    value = None
                ws.cell(row=start_row + r_offset, column=col_idx, value=value)

    def _apply_float_format(ws, df: pd.DataFrame, start_row: int) -> None:
        apply_openpyxl_column_formats(ws, df, data_start_row=start_row)

    def _apply_color_scale(ws, df: pd.DataFrame, col_name: str,
                           data_start_row: int, n_rows: int) -> None:
        if col_name not in df.columns or n_rows == 0:
            return
        col_idx = df.columns.get_loc(col_name) + 1
        letter = get_column_letter(col_idx)
        cell_range = f"{letter}{data_start_row}:{letter}{data_start_row + n_rows - 1}"
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=0.5, mid_color="FFEB84",
                end_type="num", end_value=1, end_color="63BE7B",
            ),
        )

    def _apply_consensus_fills(ws, df: pd.DataFrame, data_start_row: int) -> None:
        if "consensus_class" not in df.columns:
            return
        col_idx = df.columns.get_loc("consensus_class") + 1
        for r_offset, (_, row_data) in enumerate(df.iterrows()):
            val = row_data.get("consensus_class", None)
            fill = _CONSENSUS_FILLS.get(str(val).upper() if val else "", None)
            if fill is not None:
                ws.cell(row=data_start_row + r_offset, column=col_idx).fill = fill

    def _set_column_widths(ws, df: pd.DataFrame) -> None:
        _NARROW = {"unified_rank", "canonical_rank", "prod_rank",
                   "agreement_count", "max_possible", "atr_period",
                   "component_id", "unified_component_size"}
        _MEDIUM = {"multiplier", "atr_bucket", "mult_bucket_ticks",
                   "bucket_param", "consensus_class", "fusion_case",
                   "source_presence", "hard_gate_passed", "in_canonical",
                   "in_plateau", "has_topology_data", "canonical_top_quartile",
                   "plateau_confirmed", "topology_confirmed",
                   "tradeability_passed", "gate_passed"}
        for col_idx, col_name in enumerate(df.columns, start=1):
            letter = get_column_letter(col_idx)
            if col_name in _NARROW:
                ws.column_dimensions[letter].width = 10
            elif col_name in _MEDIUM:
                ws.column_dimensions[letter].width = 16
            elif col_name in ("disqualification_reason", "gate_exclusion_reason"):
                ws.column_dimensions[letter].width = 30
            else:
                ws.column_dimensions[letter].width = 20

    # ── Prepare data ─────────────────────────────────────────────────────────
    main_df = format_excel_export_df(unified_df) if not unified_df.empty else unified_df
    n_main_cols = len(main_df.columns)

    # ── Placeholder branch ────────────────────────────────────────────────────
    if main_df.empty:
        placeholder_df = pd.DataFrame(columns=unified_df.columns)
        placeholder_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        ws = writer.sheets[SHEET_NAME]
        _write_header_row(ws, row=1, columns=unified_df.columns)
        ws["A2"] = "No rows in unified ranking result"
        ws["A2"].font = Font(italic=True, color="9C0006")
        return

    # ── Write to Excel via to_excel (handles dtypes cleanly) ─────────────────
    # We write manually row-by-row to support multi-section layout.
    # Use a blank sheet first.
    pd.DataFrame(columns=main_df.columns).to_excel(
        writer, sheet_name=SHEET_NAME, index=False
    )
    ws = writer.sheets[SHEET_NAME]

    # ── Section A: UnifiedProductionRank ─────────────────────────────────────
    _write_section_banner(ws, row=1, text="UnifiedProductionRank", n_cols=n_main_cols)
    _write_header_row(ws, row=2, columns=main_df.columns)
    _write_data_rows(ws, main_df, start_row=3)
    _apply_float_format(ws, main_df, start_row=3)
    _apply_color_scale(ws, main_df, "unified_score", data_start_row=3, n_rows=len(main_df))
    _apply_consensus_fills(ws, main_df, data_start_row=3)
    apply_gate_status_fills(ws, main_df, data_start_row=3)
    _set_column_widths(ws, main_df)

    # Freeze below banner + header; auto-filter on header row
    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.auto_filter.ref = f"A2:{get_column_letter(n_main_cols)}2"

    # ── Section B: PlateauOnly (if present) ──────────────────────────────────
    if plateau_only_df is not None and not plateau_only_df.empty:
        plat_df = format_excel_export_df(plateau_only_df)
        n_plat_cols = len(plat_df.columns)
        banner_row = len(main_df) + 4   # gap of 1 blank row after section A
        hdr_row = banner_row + 1
        data_start = hdr_row + 1

        _write_section_banner(
            ws, row=banner_row,
            text="PlateauOnly — in plateau but not in canonical universe",
            n_cols=max(n_main_cols, n_plat_cols),
        )
        _write_header_row(ws, row=hdr_row, columns=plat_df.columns)
        _write_data_rows(ws, plat_df, start_row=data_start)
        _apply_float_format(ws, plat_df, start_row=data_start)
        _apply_color_scale(
            ws, plat_df, "final_prod_score",
            data_start_row=data_start, n_rows=len(plat_df),
        )


def _write_plateau_summary_sheet_from_df(
    summary_df: pd.DataFrame,
    writer: pd.ExcelWriter,
) -> None:
    """Write PlateauSummary sheet from a pre-computed DataFrame.

    Handles all formatting: header styling, numeric format, color scales,
    freeze panes, auto-filter, and column widths.

    If ``summary_df`` is empty, writes a placeholder sheet with the correct
    column schema and a message in A2.

    This is the canonical write path used by ``export_walk_forward_to_excel``.
    The original ``_write_plateau_summary_sheet`` is kept as a backward-
    compatible wrapper for external callers and existing tests.
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from supertrend_optimizer.analysis.bucket_plateau import _PLATEAU_SUMMARY_COLUMNS

    SHEET_NAME = "PlateauSummary"
    HDR_BG = "1F4E79"   # dark blue — matches PlateauAnalysis header
    HDR_FG = "FFFFFF"

    # ── Placeholder branch (empty pool) ──────────────────────────────────────
    if summary_df.empty:
        # Write header row with correct schema, then message in A2
        placeholder_df = pd.DataFrame(columns=_PLATEAU_SUMMARY_COLUMNS)
        placeholder_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        ws = writer.sheets[SHEET_NAME]

        # Style header row
        header_fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")
        header_font = Font(bold=True, color=HDR_FG, size=10)
        for col_idx in range(1, len(_PLATEAU_SUMMARY_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws["A2"] = "No candidates with gate_passed == True"
        ws["A2"].font = Font(italic=True, color="9C0006")
        return

    # ── Write data ────────────────────────────────────────────────────────────
    summary_df = format_excel_export_df(summary_df)
    summary_df.to_excel(writer, sheet_name=SHEET_NAME, index=False, startrow=0)
    ws = writer.sheets[SHEET_NAME]
    n_cols = len(summary_df.columns)
    n_rows = len(summary_df)

    # ── Header styling ────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG, fill_type="solid")
    header_font = Font(bold=True, color=HDR_FG, size=10)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply per-column number formats (int → "0", metric → "0.0000", etc.)
    apply_openpyxl_column_formats(ws, summary_df, data_start_row=2)

    # ── Color scale on final_prod_score ──────────────────────────────────────
    if "final_prod_score" in summary_df.columns:
        fps_col_idx = summary_df.columns.get_loc("final_prod_score") + 1
        fps_letter = get_column_letter(fps_col_idx)
        fps_range = f"{fps_letter}2:{fps_letter}{n_rows + 1}"
        ws.conditional_formatting.add(
            fps_range,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=0.5, mid_color="FFEB84",
                end_type="num", end_value=1, end_color="63BE7B",
            ),
        )

    # ── Color scale on final_rank_score ──────────────────────────────────────
    if "final_rank_score" in summary_df.columns:
        frs_col_idx = summary_df.columns.get_loc("final_rank_score") + 1
        frs_letter = get_column_letter(frs_col_idx)
        frs_range = f"{frs_letter}2:{frs_letter}{n_rows + 1}"
        ws.conditional_formatting.add(
            frs_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # ── Freeze panes + auto-filter ────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # prod_rank
    ws.column_dimensions["B"].width = 10   # atr_period
    ws.column_dimensions["C"].width = 12   # multiplier
    for col_idx in range(4, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def _write_plateau_summary_sheet(
    plateau_result: Any,
    median_bucket_df: "Optional[pd.DataFrame]",
    surface_bucket_df: "Optional[pd.DataFrame]",
    writer: pd.ExcelWriter,
    config: dict,
) -> None:
    """Backward-compatible wrapper: compute PlateauSummary scores then write.

    .. deprecated::
        For new internal code use the two-step pattern instead::

            summary_df = _compute_plateau_summary_scores(...)
            _write_plateau_summary_sheet_from_df(summary_df, writer)

        This wrapper is kept for external callers and existing tests that
        call ``_write_plateau_summary_sheet`` directly.
    """
    from supertrend_optimizer.analysis.bucket_plateau import (
        _PLATEAU_SUMMARY_COLUMNS,
        _compute_plateau_summary_scores,
    )

    try:
        summary_df = _compute_plateau_summary_scores(
            plateau_result=plateau_result,
            median_bucket_df=median_bucket_df,
            surface_bucket_df=surface_bucket_df,
            config=config,
        )
    except Exception:
        logger.warning(
            "_write_plateau_summary_sheet: _compute_plateau_summary_scores failed — "
            "writing placeholder sheet.",
            exc_info=True,
        )
        summary_df = pd.DataFrame(columns=_PLATEAU_SUMMARY_COLUMNS)

    _write_plateau_summary_sheet_from_df(summary_df, writer)


def _write_plateau_analysis_sheet(
    plateau_result: Any,
    writer: pd.ExcelWriter,
) -> None:
    """Write PlateauAnalysis sheet — one row per bucket, sorted by quality score DESC."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "PlateauAnalysis"
    summary_df = getattr(plateau_result, "summary_df", pd.DataFrame())

    if summary_df.empty:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        return

    summary_df = format_excel_export_df(summary_df)
    summary_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

    ws = writer.sheets[sheet_name]
    n_cols = len(summary_df.columns)
    n_rows = len(summary_df)

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply per-column number formats (int → "0", metric → "0.0000", etc.)
    apply_openpyxl_column_formats(ws, summary_df, data_start_row=2)

    # Color scale on plateau_quality_score
    if "plateau_quality_score" in summary_df.columns:
        qs_col_idx = summary_df.columns.get_loc("plateau_quality_score") + 1
        qs_letter = get_column_letter(qs_col_idx)
        data_range = f"{qs_letter}2:{qs_letter}{n_rows + 1}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # Color scale on cross_step_robustness: fixed 0→red, 0.5→yellow, 1→green
    if "cross_step_robustness" in summary_df.columns:
        csr_col_idx = summary_df.columns.get_loc("cross_step_robustness") + 1
        csr_letter = get_column_letter(csr_col_idx)
        csr_range = f"{csr_letter}2:{csr_letter}{n_rows + 1}"
        ws.conditional_formatting.add(
            csr_range,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8696B",
                mid_type="num", mid_value=0.5, mid_color="FFEB84",
                end_type="num", end_value=1, end_color="63BE7B",
            ),
        )

    # Color scale on worst_step_penalty: 0→green, 0.5→yellow, 1→red
    if "worst_step_penalty" in summary_df.columns:
        wsp_col_idx = summary_df.columns.get_loc("worst_step_penalty") + 1
        wsp_letter = get_column_letter(wsp_col_idx)
        wsp_range = f"{wsp_letter}2:{wsp_letter}{n_rows + 1}"
        ws.conditional_formatting.add(
            wsp_range,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="63BE7B",
                mid_type="num", mid_value=0.5, mid_color="FFEB84",
                end_type="num", end_value=1, end_color="F8696B",
            ),
        )

    # Autofilter + freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Column widths
    ws.column_dimensions["A"].width = 28  # bucket_param
    for col_idx in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _write_plateau_surface_sheet(
    surface: Any,
    writer: pd.ExcelWriter,
    config: "dict | None" = None,
) -> None:
    """
    Write PlateauSurface sheet: raw parameter table + 4 heatmap pivots.

    Layout (openpyxl startrow):
        Title row
        (blank)
        raw_df table
        (2 blank rows)
        Title: MEDIAN PnL SURFACE
        (blank)
        median_pnl_pivot
        (2 blank rows)
        Title: PLATEAU WIDTH MAP — Positive Ratio
        (blank)
        positive_ratio_pivot
        (2 blank rows)
        Title: VARIANCE MAP — Std PnL
        (blank)
        std_pnl_pivot
        (2 blank rows)
        Title: PLATEAU MASK
        (blank)
        plateau_mask_pivot
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    sheet_name = "PlateauSurface"
    raw_df = getattr(surface, "raw_df", pd.DataFrame())

    if raw_df.empty:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # ── Add bucket_param column (human-readable ATR × Mult bucket label) ────
    # Uses the same format_bucket_label function as BucketMatrix / PlateauAnalysis.
    # Requires apply_param_buckets columns (atr_bucket, mult_bucket_ticks) to be
    # present in raw_df; they are added by _build_surface_data.
    raw_df = raw_df.copy()
    _cons_cfg = (config or {}).get("walk_forward", {}).get("consensus", {})
    from supertrend_optimizer.utils.constants import (
        DEFAULT_ATR_BUCKET_STEP as _SURF_ATR_STEP,
        DEFAULT_MULT_BUCKET_STEP as _SURF_MULT_STEP,
    )
    _surf_atr_step = int(_cons_cfg.get("atr_bucket_step", _SURF_ATR_STEP))
    _surf_mult_step = float(_cons_cfg.get("mult_bucket_step", _SURF_MULT_STEP))

    if "atr_bucket" not in raw_df.columns or "mult_bucket_ticks" not in raw_df.columns:
        logger.error(
            "PlateauSurface export requires atr_bucket and mult_bucket_ticks in raw_df; "
            "columns present: %s. "
            "These columns must be added by _build_surface_data via apply_param_buckets.",
            list(raw_df.columns),
        )
        raise ValueError(
            "Missing bucket columns (atr_bucket, mult_bucket_ticks) in PlateauSurface raw_df. "
            "Ensure _build_surface_data calls apply_param_buckets before returning."
        )

    def _surf_bucket_param(row: "pd.Series") -> str:
        try:
            return _format_bucket_label(
                int(row["atr_bucket"]),
                int(row["mult_bucket_ticks"]),
                _surf_atr_step,
                _surf_mult_step,
            )
        except (KeyError, ValueError, TypeError):
            logger.warning(
                "bucket_param fallback to UNKNOWN: atr_bucket=%s, mult_bucket_ticks=%s",
                row.get("atr_bucket"), row.get("mult_bucket_ticks"),
            )
            return "UNKNOWN"
    raw_df["bucket_param"] = raw_df.apply(_surf_bucket_param, axis=1)

    # ── Reorder: final_rank, atr_period, multiplier, bucket_param first ──────
    # mult_bucket (float64, added by apply_param_buckets alongside atr_bucket /
    # mult_bucket_ticks) is a derived convenience column redundant with
    # bucket_param; hide it from Excel to avoid clutter.
    _priority_cols = ["final_rank", "atr_period", "multiplier", "bucket_param"]
    _hidden_cols = {"mult_bucket"}
    _remaining_cols = [
        c for c in raw_df.columns
        if c not in _priority_cols and c not in _hidden_cols
    ]
    raw_df = raw_df[
        [c for c in _priority_cols if c in raw_df.columns] + _remaining_cols
    ]

    # Track current row (0-indexed for pd.DataFrame.to_excel startrow)
    current_row = 0

    # Title
    title_df = pd.DataFrame([["PARAMETER SURFACE — DEEP PLATEAU ANALYSIS"]])
    title_df.to_excel(writer, sheet_name=sheet_name, index=False,
                      header=False, startrow=current_row)
    current_row += 2  # title + blank

    # Raw table — rename columns for display clarity (S5).
    # Internal column names stay canonical in raw_df; Excel shows friendly tags.
    _display_rename = {
        # Coverage vs WF length (P3-2 style — avoid confusing n_steps with total WF steps)
        "n_steps": "n_observations",
        # S0 DD provenance
        "worst_max_dd_pct": "worst_max_dd_pct [filled, for ranking]",
        "worst_max_dd_pct_raw": "worst_max_dd_pct_raw [diagnostic]",
        "dd_imputed": "dd_imputed [diagnostic]",
        # DD quality — continuous penalty (stage 1 / penalty etap)
        "dd_observed_count": "dd_observed_count [diagnostic]",
        "dd_observed_ratio": "dd_observed_ratio [diagnostic]",
        "dd_quality_penalty": "dd_quality_penalty [confidence multiplier, gate-passed only]",
        # final_rank_score pipeline
        "gate_passed": "gate_passed [diagnostic]",
        "final_rank": "final_rank",
        "profit_norm": "profit_norm [factor]",
        "stability_norm": "stability_norm [factor]",
        "risk_norm": "risk_norm [factor]",
        "plateau_norm": "plateau_norm [factor]",
        "confidence_norm": "confidence_norm [factor]",
        "plateau_density": "plateau_density [diagnostic]",
        "local_pnl": "local_pnl [diagnostic]",
        "local_risk": "local_risk [diagnostic]",
        "local_risk_norm": "local_risk_norm [diagnostic]",
        "consistency_score": "consistency_score [diagnostic]",
        "support_score": "support_score [= consistency_score] [diagnostic]",
    }
    _raw_df_display = raw_df.rename(
        columns={k: v for k, v in _display_rename.items() if k in raw_df.columns}
    )
    format_excel_export_df(_raw_df_display).to_excel(writer, sheet_name=sheet_name, index=False, startrow=current_row)
    current_row += len(_raw_df_display) + 3  # data rows + header + 2 blank

    # Helper: write one pivot heatmap block
    def _write_pivot_block(pivot_df: pd.DataFrame, title: str, color_rule) -> None:
        nonlocal current_row
        title_df_ = pd.DataFrame([[title]])
        title_df_.to_excel(writer, sheet_name=sheet_name, index=False,
                            header=False, startrow=current_row)
        current_row += 2  # title + blank
        pivot_df.to_excel(writer, sheet_name=sheet_name, index=True,
                          startrow=current_row)
        ws = writer.sheets[sheet_name]
        n_pr, n_pc = pivot_df.shape
        data_start_row = current_row + 2  # +1 header +1 1-based
        data_end_row   = data_start_row + n_pr - 1
        data_start_col = 2  # col A is index
        data_end_col   = data_start_col + n_pc - 1
        if color_rule is not None and n_pr > 0 and n_pc > 0:
            r1 = get_column_letter(data_start_col) + str(data_start_row)
            r2 = get_column_letter(data_end_col)   + str(data_end_row)
            ws.conditional_formatting.add(f"{r1}:{r2}", color_rule)
        current_row += n_pr + 3  # data rows + header + 2 blank

    # 1. Median PnL Surface — 3-color centered at 0
    med_pivot = getattr(surface, "median_pnl_pivot", pd.DataFrame())
    all_vals = pd.to_numeric(med_pivot.values.flatten(), errors="coerce")
    all_vals = all_vals[~np.isnan(all_vals)]
    min_val = float(np.min(all_vals)) if len(all_vals) > 0 else -1.0
    max_val = float(np.max(all_vals)) if len(all_vals) > 0 else 1.0
    if not med_pivot.empty:
        _write_pivot_block(
            med_pivot,
            "MEDIAN PnL SURFACE (ATR × Multiplier)",
            ColorScaleRule(
                start_type="num", start_value=min_val, start_color="F8696B",
                mid_type="num",   mid_value=0,         mid_color="FFFFFF",
                end_type="num",   end_value=max_val,   end_color="63BE7B",
            ),
        )

    # 2. Positive Ratio — white→green
    pos_pivot = getattr(surface, "positive_ratio_pivot", pd.DataFrame())
    if not pos_pivot.empty:
        _write_pivot_block(
            pos_pivot,
            "PLATEAU WIDTH MAP — Positive Ratio (ATR × Multiplier)",
            ColorScaleRule(
                start_type="num", start_value=0.0, start_color="FFFFFF",
                end_type="num",   end_value=1.0,   end_color="63BE7B",
            ),
        )

    # 3. Variance — green→yellow→red
    std_pivot = getattr(surface, "std_pnl_pivot", pd.DataFrame())
    if not std_pivot.empty:
        _write_pivot_block(
            std_pivot,
            "VARIANCE MAP — Std PnL (ATR × Multiplier)",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max",   end_color="F8696B",
            ),
        )

    # 4. Plateau Mask — 0=white, 1=dark green
    mask_pivot = getattr(surface, "plateau_mask_pivot", pd.DataFrame())
    std_thr = getattr(surface, "std_threshold", 0.0)
    pos_thr = getattr(surface, "positive_ratio_threshold", 0.6)
    max_dd_thr = getattr(surface, "max_allowed_dd", None)
    std_ceiling_applied = getattr(surface, "std_ceiling_applied", False)
    std_q = float(getattr(surface, "std_quantile", 0.65))
    std_q_raw = float(getattr(surface, "std_pnl_quantile_raw", std_thr))
    std_max_c = float(getattr(surface, "std_max_ceiling", 50.0))
    gate_min_st = int(getattr(surface, "gate_min_steps", 3))
    if not mask_pivot.empty:
        _std_hint = " [std_max ceiling applied]" if std_ceiling_applied else ""
        # Formula matches _build_surface_data: std_threshold = min(quantile, std_max).
        _std_formula = (
            f"std_pnl <= {std_thr:.4f} = min(quantile_{std_q:.2f}({std_q_raw:.4f}), "
            f"std_max={std_max_c:.2f}){_std_hint}"
        )
        _cov_clause = f" AND n_observations >= {gate_min_st}"
        if max_dd_thr is not None:
            _dd_clause = (
                f" AND dd_observed_count >= 1 AND worst_max_dd_pct <= {max_dd_thr:.1f}"
                f" [DD quality penalizes confidence; mixed DD allowed]"
            )
        else:
            _dd_clause = " [DD gate disabled — no max_allowed_dd_pct set]"
        _mask_title = (
            f"PLATEAU MASK (stable_flag=1 when: positive_ratio >= {pos_thr} "
            f"AND {_std_formula}{_cov_clause}{_dd_clause})"
        )
        _write_pivot_block(
            mask_pivot,
            _mask_title,
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FFFFFF",
                end_type="num",   end_value=1,   end_color="375623",
            ),
        )

    # Freeze panes on the surface sheet
    ws = writer.sheets[sheet_name]
    ws.freeze_panes = ws.cell(row=2, column=1)


def _write_production_candidates_sheet(
    surface: Any,
    plateau_summary_df: pd.DataFrame,
    bucket_matrix_df: "Optional[pd.DataFrame]",
    bucket_train_surface_df: "Optional[pd.DataFrame]",
    writer: pd.ExcelWriter,
) -> None:
    """Write ProductionCandidates sheet.

    Layout (all sections on one sheet, separated by blank rows):

        Row 1   — section header "TOP-3 PRODUCTION CANDIDATES"
        Row 2   — blank
        Rows 3+ — production_candidates_df (compact view)
        [2 blank rows]
        Section header "FUNNEL TABLE (all pipeline stages)"
        [1 blank row]
        funnel_df
        [2 blank rows]
        Section header "COMPONENT MAP"  (omitted if component_map is None/empty)
        [1 blank row]
        component_map pivot

    If no candidates were found, the sheet is still created and shows a
    "No production candidates available" placeholder message.
    """
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    SHEET_NAME = "ProductionCandidates"  # ≤31 chars

    # ── Colours ──────────────────────────────────────────────────────────────
    HDR_BG   = "1F4E79"   # dark blue — section headers
    HDR_FG   = "FFFFFF"
    TBL_HDR  = "2E75B6"   # mid blue — table column headers
    TBL_FG   = "FFFFFF"

    # ── Retrieve data ─────────────────────────────────────────────────────────
    production_candidates_df = getattr(surface, "production_candidates", None)
    component_map            = getattr(surface, "component_map", None)

    # Compact columns for Section 1 (human-readable top candidates)
    _CAND_COLS = [
        "rank", "atr_period", "multiplier", "bucket_param",
        "component_id", "component_size", "distance_to_center", "distance_to_medoid",
        "plateau_quality_score", "point_selection_score", "positive_pnl_ratio",
        "median_pnl", "std_pnl", "worst_max_dd_pct",
        # DD quality: risk → data quality → score (causal chain)
        "dd_observed_count", "dd_observed_ratio", "dd_quality_penalty",
        "final_selection_score",
        "selection_mode", "selection_reason",
    ]

    has_candidates = (
        production_candidates_df is not None
        and isinstance(production_candidates_df, pd.DataFrame)
        and not production_candidates_df.empty
    )

    # ── Helper: write a styled section header ────────────────────────────────
    def _write_section_header(ws, row: int, text: str, n_cols: int) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font      = Font(bold=True, color=HDR_FG, size=11)
        cell.fill      = PatternFill(start_color=HDR_BG, end_color=HDR_BG,
                                     fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        # Extend fill across columns
        for col in range(2, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill(start_color=HDR_BG, end_color=HDR_BG,
                                 fill_type="solid")

    # ── Helper: write a DataFrame with styled column headers ─────────────────
    def _write_df_block(ws, df: pd.DataFrame, start_row: int) -> int:
        """Write df to ws starting at start_row. Returns next free row."""
        if df is None or df.empty:
            return start_row
        cols = list(df.columns)
        # Column header row
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=start_row, column=ci, value=str(col))
            c.font      = Font(bold=True, color=TBL_FG, size=10)
            c.fill      = PatternFill(start_color=TBL_HDR, end_color=TBL_HDR,
                                      fill_type="solid")
            c.alignment = Alignment(horizontal="center")
        # Data rows
        for ri, (_, row_data) in enumerate(df.iterrows(), 1):
            for ci, col in enumerate(cols, 1):
                val = row_data[col]
                # Convert numpy/pandas scalar types to native Python
                if hasattr(val, "item"):
                    try:
                        val = val.item()
                    except Exception:
                        val = str(val)
                elif pd.isna(val) if not isinstance(val, str) else False:
                    val = None
                ws.cell(row=start_row + ri, column=ci, value=val)
        return start_row + len(df) + 1  # +1 for header row

    # ── No-candidates placeholder ─────────────────────────────────────────────
    if not has_candidates:
        placeholder = pd.DataFrame(
            [["No production candidates available — "
              "run plateau analysis in deep mode to generate candidates."]],
            columns=["message"],
        )
        placeholder.to_excel(writer, sheet_name=SHEET_NAME, index=False, startrow=0)
        ws = writer.sheets[SHEET_NAME]
        ws.cell(row=1, column=1).font = Font(bold=True, color="9C0006")
        return

    # ── Build funnel table (enriched with upstream bucket data) ───────────────
    from supertrend_optimizer.analysis.bucket_plateau import (
        build_production_candidates_funnel_table,
    )
    try:
        funnel_df = build_production_candidates_funnel_table(
            production_candidates_df=production_candidates_df,
            bucket_matrix_df=bucket_matrix_df,
            bucket_train_surface_df=bucket_train_surface_df,
            plateau_summary_df=plateau_summary_df,
        )
    except Exception:
        logger.warning(
            "_write_production_candidates_sheet: funnel table build failed.",
            exc_info=True,
        )
        funnel_df = pd.DataFrame()

    # ── Prepare compact candidates view ───────────────────────────────────────
    cand_cols_avail = [c for c in _CAND_COLS if c in production_candidates_df.columns]
    cand_view = format_excel_export_df(production_candidates_df[cand_cols_avail].copy())

    # ── Determine column widths (use widest table) ────────────────────────────
    max_cols = max(len(cand_cols_avail), len(funnel_df.columns) if not funnel_df.empty else 0)

    # ── Write to a temp DataFrame first so we get a sheet handle ─────────────
    # We seed the sheet with an empty frame, then use openpyxl directly.
    pd.DataFrame().to_excel(writer, sheet_name=SHEET_NAME, index=False)
    ws = writer.sheets[SHEET_NAME]

    current_row = 1  # openpyxl is 1-based

    # ── Section 1: Top candidates ─────────────────────────────────────────────
    n_cand = len(cand_view)
    _write_section_header(
        ws, current_row,
        f"TOP-{n_cand} PRODUCTION CANDIDATES",
        max(len(cand_cols_avail), 4),
    )
    current_row += 2  # header + blank

    current_row = _write_df_block(ws, cand_view, current_row)
    current_row += 2  # trailing blank rows

    # ── Section 2: Funnel table ───────────────────────────────────────────────
    _write_section_header(
        ws, current_row,
        "FUNNEL TABLE — ALL PIPELINE STAGES",
        max(len(funnel_df.columns) if not funnel_df.empty else 4, 4),
    )
    current_row += 2

    if funnel_df.empty:
        ws.cell(row=current_row, column=1,
                value="Funnel table unavailable (upstream data missing).")
        current_row += 2
    else:
        current_row = _write_df_block(ws, format_excel_export_df(funnel_df), current_row)
        current_row += 2

    # ── Section 3: Component map ──────────────────────────────────────────────
    has_comp_map = (
        component_map is not None
        and isinstance(component_map, pd.DataFrame)
        and not component_map.empty
    )
    if has_comp_map:
        _write_section_header(
            ws, current_row,
            "COMPONENT MAP (plateau connected components, 8-neighborhood BFS)",
            max(len(component_map.columns) + 1, 4),
        )
        current_row += 2

        # component_map is a pivot: index=atr_period, columns=multiplier
        # Write index as first column
        comp_reset = component_map.reset_index()
        comp_reset.columns = [
            "atr_period" if str(c) == "index" else c
            for c in comp_reset.columns
        ]
        current_row = _write_df_block(ws, comp_reset, current_row)

    # ── Column widths ─────────────────────────────────────────────────────────
    # Auto-width: cap at 30 chars, min 8
    for col_cells in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                try:
                    max_len = max(max_len, min(len(str(cell.value)), 30))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max_len + 2

    # ── Freeze top row ────────────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=2, column=1)


def _write_plateau_pivot_sheets(
    plateau_result: Any,
    writer: pd.ExcelWriter,
    max_sheets: int,
) -> None:
    """
    Write PlateauPivot_1..N sheets — one per analyzed bucket.

    Each sheet: pivot_df from plateau_result.pivot_tables with conditional
    formatting on Step columns.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    pivot_tables = getattr(plateau_result, "pivot_tables", {})
    buckets      = getattr(plateau_result, "buckets", [])

    # Build label map for sheet titles
    label_map = {
        (br.atr_bucket, br.mult_bucket_ticks): br.bucket_label
        for br in buckets
    }

    written = 0
    for (ab, mt), pivot_df in pivot_tables.items():
        if written >= max_sheets:
            break
        if pivot_df is None or (hasattr(pivot_df, "empty") and pivot_df.empty):
            continue

        sheet_name = f"PlateauPivot_{written + 1}"
        label = label_map.get((ab, mt), f"ATR {ab} | Ticks {mt}")

        # Write title row then pivot
        title_df = pd.DataFrame([[f"PLATEAU PIVOT: {label}"]])
        title_df.to_excel(writer, sheet_name=sheet_name, index=False,
                          header=False, startrow=0)
        pivot_df.to_excel(writer, sheet_name=sheet_name, index=True, startrow=2)

        ws = writer.sheets[sheet_name]

        # Header styling
        n_cols = len(pivot_df.columns) + 1  # +1 for index col
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center")

        # Conditional formatting on Step columns
        step_col_indices = [
            pivot_df.columns.get_loc(c) + 2  # +1 for 1-base, +1 for index col
            for c in pivot_df.columns
            if str(c).startswith("Step")
        ]
        n_data_rows = len(pivot_df)
        if step_col_indices and n_data_rows > 0:
            for col_idx in step_col_indices:
                col_letter = get_column_letter(col_idx)
                data_range = f"{col_letter}4:{col_letter}{3 + n_data_rows}"
                ws.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B",
                    ),
                )

        ws.freeze_panes = ws.cell(row=4, column=2)
        ws.column_dimensions["A"].width = 16
        written += 1


def _reorder_sheets(wb) -> None:
    """Reorder sheets: user-priority sheets first, then the rest in original order.

    Priority order:
        WF_Config, PlateauSummary, UnifiedProductionRank, ProductionCandidates,
        WF_Gates, WF_TOP_K_Canonical, BucketMatrix_Full, BucketMatrix_Median,
        BucketTrainSurface, PlateauAnalysis, PlateauSurface,
        PlateauPivot_1..N, WF_01.., WF_Trades, WF_Train_Trades

    Sheets absent from the workbook are silently skipped.
    PlateauSummary is placed at index 1 (after WF_Config).
    UnifiedProductionRank is placed at index 2 (after PlateauSummary) when present.
    """
    priority_static = [
        "WF_Config",
        "PlateauSummary",
        "UnifiedProductionRank",
        "FullInformationRank",
        "FIR_Meta",
        "ProductionCandidates",
        "WF_Gates",
        "WF_TOP_K_Canonical",
        "BucketMatrix_Full",
        "BucketMatrix_Median",
        "BucketTrainSurface",
        "PlateauAnalysis",
        "PlateauSurface",
    ]
    plateau_pivots = sorted(
        [s for s in wb.sheetnames if re.match(r"PlateauPivot_\d+$", s)],
        key=lambda s: int(re.search(r"\d+", s).group()),
    )
    wf_steps = sorted(
        [s for s in wb.sheetnames if re.match(r"WF_\d{2}$", s)],
        key=lambda s: int(s.split("_")[1]),
    )
    priority_tail = ["WF_Trades", "WF_Train_Trades"]

    priority_all = (
        [s for s in priority_static if s in wb.sheetnames]
        + plateau_pivots
        + wf_steps
        + [s for s in priority_tail if s in wb.sheetnames]
    )
    other = [s for s in wb.sheetnames if s not in priority_all]
    new_order = priority_all + other

    by_name = {ws.title: ws for ws in wb._sheets}
    wb._sheets = [by_name[n] for n in new_order]


def _apply_freeze_panes(output_path: str) -> None:
    """Apply freeze panes to all sheets (freeze first row).

    WF_ENSEMBLE is skipped here because _write_wf_ensemble_sheet sets a
    custom freeze pane positioned at the Table C header row.
    """
    wb = load_workbook(output_path)
    _reorder_sheets(wb)

    _CUSTOM_FREEZE_SHEETS = {"WF_ENSEMBLE", "BucketMatrix_Full", "BucketMatrix_Median"}
    for sheet_name in wb.sheetnames:
        if sheet_name in _CUSTOM_FREEZE_SHEETS:
            continue  # custom freeze already applied during sheet write
        ws = wb[sheet_name]
        # Freeze first row (A2 means freeze everything above row 2)
        ws.freeze_panes = "A2"

    wb.save(output_path)
