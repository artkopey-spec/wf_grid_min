"""
Excel export functionality for SuperTrend Tester.

This module exports tester results to Excel format with metrics and trades.
"""

from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path
import dataclasses
import warnings
import numpy as np
import pandas as pd

from supertrend_optimizer.testing.runner import PeriodResult, SegmentResult
from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE
from supertrend_optimizer.io.excel_format_helpers import format_excel_export_df


# Metrics column display names
METRICS_DISPLAY_NAMES = {
    "sum_pnl_pct": "Sum PnL %",
    "sharpe": "Sharpe",
    "sortino": "Sortino",
    "max_drawdown": "Max Drawdown",
    "cagr": "CAGR",
    "win_rate": "Win Rate",
    "num_trades": "Num Trades",
    "profit_factor": "Profit Factor",
    "avg_trade": "Avg Trade"
}

# Signals column display names
SIGNALS_DISPLAY_NAMES = {
    "signal_time": "Signal Time",
    "signal_bar_index": "Signal Bar Index",
    "event_type": "Event Type",
    "direction": "Direction",
    "st_color_before": "ST Color Before",
    "st_color_after": "ST Color After",
    "is_reversal": "Is Reversal",
    "exec_price": "Exec Price",
    "signal_open_price": "Signal Open Price",
    "signal_close_price": "Signal Close Price",
    "signal_body_pct": "Signal Body %",
    "signal_range_pct": "Signal Range %",
    "signal_body_atr": "Signal Body ATR",
    "signal_range_atr": "Signal Range ATR",
    "t1_return_pct": "T+1 Return %",
    "t2_return_pct": "T+2 Return %",
    "t3_return_pct": "T+3 Return %",
    # filter columns (plan §7.9 / PR 9)
    "atr_pct":          "ATR %",
    "volume_ratio":     "Volume Ratio",
    "volatility_pass":  "Vol ATR Pass",   # in amp modes = amplitude-side pass (patch §J)
    "volume_pass":      "Vol Flow Pass",
    "allow_entry":      "Allow Entry",
    "filtered_reason":  "Filtered Reason",
    "entry_bar_index":  "Entry Bar Index",
    # amplitude diagnostic columns (v1.3, patch §J, schema v2)
    # _NA in legacy / none modes and on close rows.
    "amp_n":         "Amp N",
    "amp_threshold": "Amp Threshold",
    "separation":    "Separation",
    "amp_valid":     "Amp Valid",
    "amp_ok":        "Amp OK",
    "not_dead":      "Not Dead",
    # median ratio columns
    "signal_body_pct_median_ratio":  "Signal Body % / Median",
    "signal_range_pct_median_ratio": "Signal Range % / Median",
    # zigzag diagnostic columns (v2.0, plan §3.8.1) — _NA in non-zz modes
    "zz_leg_direction":   "ZZ Leg Dir",
    "zz_cand_height_pct": "ZZ Cand Height %",
    "zz_global_median":   "ZZ Global Median %",
    "zz_global_p80":      "ZZ Global P80 %",
    "zz_local_median":    "ZZ Local Median %",
    "zz_n_legs":          "ZZ N Legs",
    "zz_regime_state":    "ZZ Regime",
    "zz_armed":           "ZZ Armed",
    "zz_armed_side":      "ZZ Armed Side",
    # RFC v3.1 §7.4 (fix N-10) — Phase 5 Signals sheet additions.
    "zz_ready_a":         "ZZ Ready A",
    "zz_ready_b":         "ZZ Ready B",
    "zz_arm_source":      "ZZ Arm Source",
}

# ZigZag int8-enum → display string mappings (plan §3.8.1a).
_ZZ_ENUM_LABELS: Dict[str, Dict[int, str]] = {
    "leg_direction": {0: "",    1: "up",        -1: "down"},
    "regime_state":  {0: "closed", 1: "grace",   2: "active"},
    "armed_side":    {0: "",    1: "long",      -1: "short"},
    # RFC v3.1 §4.5 / fix N-11: +6 "no_regime_off" for FIRED_NO_REGIME_OFF.
    "fired":         {0: "",    1: "yes_shot",  2: "no_new_pivot",
                      3: "no_timeout_soft", 4: "no_timeout_hard",
                      5: "session_reset",  6: "no_regime_off"},
    # RFC v3.1 §4.1 / §5.3: arm_source enum (ARM_SRC_*).
    "arm_source":    {0: "NONE", 1: "A", 2: "B", 3: "BOTH"},
}

# Trades column display names
TRADES_DISPLAY_NAMES = {
    "trade_id": "Trade ID",
    "direction": "Direction",
    "entry_time": "Entry Time",
    "entry_index": "Entry Index",
    "entry_price": "Entry Price",
    "exit_time": "Exit Time",
    "exit_index": "Exit Index",
    "exit_price": "Exit Price",
    "bars_held": "Bars Held",
    "gross_pnl_pct": "Gross PnL %",
    "commission_pct": "Commission %",
    "net_pnl_pct": "Net PnL %",
    "supertrend_color": "SuperTrend Color",
    # ZigZag triggering leg columns (plan §3.8.3) — N/A for non-zz modes
    "triggering_leg_id":                    "Triggering Leg ID",
    "triggering_leg_height_pct":            "Triggering Leg Height %",
    "triggering_leg_height_ratio":          "Triggering Leg Height Ratio",
    "triggering_leg_direction":             "Triggering Leg Direction",
    "bars_from_leg_confirm_to_entry":       "Bars From Leg Confirm To Entry",
    "correction_height_pct_before_entry":   "Correction Height % Before Entry",
    # Phase 5 (RFC v3.1 §7.7 / §7.8) — sign-ful + A-vs-B analytics + orphan columns.
    "entry_vs_leg_phase":                   "Entry vs Leg Phase",
    "triggering_arm_source":                "Triggering Arm Source",
    "cand_side_at_entry":                   "Cand Side At Entry",
    "cand_height_pct_at_entry":             "Cand Height % At Entry",
    "cand_leg_id_at_entry":                 "Cand Leg ID At Entry",
}

# Legacy export: diagnostic sheet for very short trades (100% slice only)
FALSE_START_SHEET_NAME = "false start"

FALSE_START_COLUMNS: Tuple[str, ...] = (
    TRADES_DISPLAY_NAMES["trade_id"],
    TRADES_DISPLAY_NAMES["direction"],
    TRADES_DISPLAY_NAMES["entry_time"],
    TRADES_DISPLAY_NAMES["entry_price"],
    TRADES_DISPLAY_NAMES["exit_time"],
    TRADES_DISPLAY_NAMES["exit_price"],
    TRADES_DISPLAY_NAMES["bars_held"],
    TRADES_DISPLAY_NAMES["net_pnl_pct"],
    TRADES_DISPLAY_NAMES["commission_pct"],
    SIGNALS_DISPLAY_NAMES["signal_time"],
    SIGNALS_DISPLAY_NAMES["event_type"],
    SIGNALS_DISPLAY_NAMES["is_reversal"],
    SIGNALS_DISPLAY_NAMES["exec_price"],
    SIGNALS_DISPLAY_NAMES["signal_body_pct_median_ratio"],
    SIGNALS_DISPLAY_NAMES["signal_range_pct_median_ratio"],
    SIGNALS_DISPLAY_NAMES["t1_return_pct"],
    SIGNALS_DISPLAY_NAMES["t2_return_pct"],
    SIGNALS_DISPLAY_NAMES["t3_return_pct"],
    "False Start Type",
)


def _normalize_timestamp_naive(ts: Any) -> pd.Timestamp:
    """Strip timezone for comparisons / Excel compatibility."""
    if pd.isna(ts):
        return pd.NaT
    t = pd.Timestamp(ts)
    if t.tzinfo is not None:
        return t.tz_localize(None)
    return t


def _price_matches_exec_entry(exec_price: Any, entry_price: Any, rtol: float = 1e-5, atol: float = 1e-8) -> bool:
    if pd.isna(exec_price) or pd.isna(entry_price):
        return False
    try:
        return bool(np.isclose(float(exec_price), float(entry_price), rtol=rtol, atol=atol))
    except (TypeError, ValueError):
        return False


def _match_trade_to_signal_row(
    trade_row: pd.Series,
    signals_prep: Optional[pd.DataFrame],
) -> Optional[pd.Series]:
    """
    Deterministic trade → signal match for legacy false-start diagnostics.

    1) Same Direction, signal_time <= entry_time, Exec Price matches Entry Price
       → choose the signal with the maximum signal_time (closest preceding among price hits).
    2) Else: same Direction, signal_time <= entry_time → nearest preceding (max signal_time).
    3) Else: None (caller fills NaNs).
    """
    if signals_prep is None or len(signals_prep) == 0:
        return None

    t_dir = trade_row.get("direction")
    if pd.isna(t_dir):
        return None

    entry_t = _normalize_timestamp_naive(trade_row.get("entry_time"))
    if pd.isna(entry_t):
        return None

    entry_price = trade_row.get("entry_price")

    same_dir = signals_prep[signals_prep["direction"] == t_dir]
    if same_dir.empty:
        return None

    before = same_dir[same_dir["signal_time"] <= entry_t]
    if before.empty:
        return None

    # Step 1 — exec price matches entry price
    price_hits = before[
        before.apply(
            lambda r: _price_matches_exec_entry(r.get("exec_price"), entry_price),
            axis=1,
        )
    ]
    if len(price_hits) > 0:
        best_idx = price_hits["signal_time"].idxmax()
        return price_hits.loc[best_idx]

    # Step 2 — nearest preceding (max signal_time)
    best_idx = before["signal_time"].idxmax()
    return before.loc[best_idx]


def _prepare_signals_lookup(signals_df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    """Copy signals with naive signal_time for matching (internal column names)."""
    if signals_df is None or len(signals_df) == 0:
        return None
    df = signals_df.copy()
    if "signal_time" not in df.columns or "direction" not in df.columns:
        return None
    st = pd.to_datetime(df["signal_time"])
    if pd.api.types.is_datetime64_any_dtype(st):
        if st.dt.tz is not None:
            st = st.dt.tz_localize(None)
    df["signal_time"] = st
    return df


def _classify_false_start_type(
    gross: Any,
    net: Any,
    bars_held: Any,
    t1: Any,
    t2: Any,
    t3: Any,
) -> str:
    """Rule-based False Start Type (informational only; priority order)."""
    if pd.notna(gross) and pd.notna(net) and gross > 0 and net <= 0:
        return "commission_drag"

    if pd.notna(net) and pd.notna(bars_held):
        try:
            b = float(bars_held)
        except (TypeError, ValueError):
            b = float("nan")
        if not np.isnan(b) and b == 1.0 and net < 0:
            return "instant_loss"

    if pd.notna(net) and net <= 0:
        t2_pos = pd.notna(t2) and t2 > 0
        t3_pos = pd.notna(t3) and t3 > 0
        if t2_pos or t3_pos:
            return "missed_followthrough"

    if pd.notna(net) and net <= 0:
        available = [x for x in (t1, t2, t3) if pd.notna(x)]
        if len(available) > 0 and all(x <= 0 for x in available):
            return "correct_early_exit"

    return "flat_noise"


def _build_false_start_sheet_df(
    trades_df_raw: Optional[pd.DataFrame],
    signals_df: Optional[pd.DataFrame],
    false_start_max_bars: int = 4,
) -> pd.DataFrame:
    """
    Build the false start DataFrame (19 columns); may be empty with headers only.

    Rows are trades from the 100% slice with ``bars_held`` strictly below
    ``false_start_max_bars`` (same threshold as the legacy Excel export setting
    ``export.false_start_max_bars``; default 4).
    """
    empty = pd.DataFrame(columns=list(FALSE_START_COLUMNS))
    if trades_df_raw is None or len(trades_df_raw) == 0:
        return empty
    if "bars_held" not in trades_df_raw.columns:
        return empty

    short_tr = trades_df_raw[trades_df_raw["bars_held"] < false_start_max_bars].copy()
    if len(short_tr) == 0:
        return empty

    need_gross = "gross_pnl_pct" in short_tr.columns

    signals_lookup = _prepare_signals_lookup(signals_df)

    rows_out: List[dict] = []
    for _, trow in short_tr.iterrows():
        gross = trow["gross_pnl_pct"] if need_gross else float("nan")
        net = trow.get("net_pnl_pct")
        bars = trow.get("bars_held")

        matched = _match_trade_to_signal_row(trow, signals_lookup)

        sig_time = matched.get("signal_time") if matched is not None else np.nan
        evt = matched.get("event_type") if matched is not None else np.nan
        rev = matched.get("is_reversal") if matched is not None else np.nan
        ex_px = matched.get("exec_price") if matched is not None else np.nan
        body_m = matched.get("signal_body_pct_median_ratio") if matched is not None else np.nan
        rng_m = matched.get("signal_range_pct_median_ratio") if matched is not None else np.nan
        t1 = matched.get("t1_return_pct") if matched is not None else np.nan
        t2 = matched.get("t2_return_pct") if matched is not None else np.nan
        t3 = matched.get("t3_return_pct") if matched is not None else np.nan

        ftype = _classify_false_start_type(gross, net, bars, t1, t2, t3)

        # Display names for trade columns (same as Trades_*)
        row_dict = {
            TRADES_DISPLAY_NAMES["trade_id"]: trow.get("trade_id"),
            TRADES_DISPLAY_NAMES["direction"]: trow.get("direction"),
            TRADES_DISPLAY_NAMES["entry_time"]: trow.get("entry_time"),
            TRADES_DISPLAY_NAMES["entry_price"]: trow.get("entry_price"),
            TRADES_DISPLAY_NAMES["exit_time"]: trow.get("exit_time"),
            TRADES_DISPLAY_NAMES["exit_price"]: trow.get("exit_price"),
            TRADES_DISPLAY_NAMES["bars_held"]: trow.get("bars_held"),
            TRADES_DISPLAY_NAMES["net_pnl_pct"]: trow.get("net_pnl_pct"),
            TRADES_DISPLAY_NAMES["commission_pct"]: trow.get("commission_pct"),
            SIGNALS_DISPLAY_NAMES["signal_time"]: sig_time,
            SIGNALS_DISPLAY_NAMES["event_type"]: evt,
            SIGNALS_DISPLAY_NAMES["is_reversal"]: rev,
            SIGNALS_DISPLAY_NAMES["exec_price"]: ex_px,
            SIGNALS_DISPLAY_NAMES["signal_body_pct_median_ratio"]: body_m,
            SIGNALS_DISPLAY_NAMES["signal_range_pct_median_ratio"]: rng_m,
            SIGNALS_DISPLAY_NAMES["t1_return_pct"]: t1,
            SIGNALS_DISPLAY_NAMES["t2_return_pct"]: t2,
            SIGNALS_DISPLAY_NAMES["t3_return_pct"]: t3,
            "False Start Type": ftype,
        }
        rows_out.append(row_dict)

    out = pd.DataFrame(rows_out, columns=list(FALSE_START_COLUMNS))
    return out


def _strip_tz_from_false_start_datetimes(df: pd.DataFrame) -> None:
    """In-place: remove tz from Entry/Exit/Signal Time for Excel."""
    for col in (
        TRADES_DISPLAY_NAMES["entry_time"],
        TRADES_DISPLAY_NAMES["exit_time"],
        SIGNALS_DISPLAY_NAMES["signal_time"],
    ):
        if col not in df.columns:
            continue
        s = df[col]
        if pd.api.types.is_datetime64_any_dtype(s) and s.dt.tz is not None:
            df[col] = s.dt.tz_localize(None)


def _build_false_start_summary_df(
    trades_df_raw: Optional[pd.DataFrame],
    false_start_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Summary metrics for the ``false start`` sheet (Metric / Value, 7 rows).

    All counts and ratios are derived from ``false_start_df``, i.e. the same
    trade subset as the main table (``bars_held`` below the effective export
    threshold). Reversal share uses ``Is Reversal`` on ``false_start_df`` only
    where values are non-null (no 0 when signal data is missing).
    """
    n_trades = len(trades_df_raw) if trades_df_raw is not None else 0
    n_fs = len(false_start_df)

    col_bars = TRADES_DISPLAY_NAMES["bars_held"]
    col_net = TRADES_DISPLAY_NAMES["net_pnl_pct"]
    col_rev = SIGNALS_DISPLAY_NAMES["is_reversal"]

    # Total false starts
    total = float(n_fs)

    # False starts / all Trades_100
    if n_trades == 0:
        ratio_all = np.nan
    else:
        ratio_all = n_fs / n_trades

    if n_fs == 0:
        med_bars = np.nan
        med_net = np.nan
        losing_share = np.nan
        bars1_share = np.nan
    else:
        med_bars = false_start_df[col_bars].median() if col_bars in false_start_df.columns else np.nan
        med_net = false_start_df[col_net].median() if col_net in false_start_df.columns else np.nan
        net_s = false_start_df[col_net] if col_net in false_start_df.columns else pd.Series(dtype=float)
        losing_share = float((net_s < 0).sum()) / n_fs
        bh = false_start_df[col_bars] if col_bars in false_start_df.columns else pd.Series(dtype=float)
        bars1_share = float((bh == 1).sum()) / n_fs

    if n_fs == 0 or col_rev not in false_start_df.columns:
        rev_share = np.nan
    else:
        rev_s = false_start_df[col_rev]
        valid_rev = rev_s.notna()
        n_valid = int(valid_rev.sum())
        if n_valid == 0:
            rev_share = np.nan
        else:
            rv = rev_s[valid_rev]
            is_rev = (rv == True) | (rv == 1)
            rev_share = float(is_rev.sum()) / n_valid

    metrics = [
        "Total false starts",
        "False starts / all Trades_100",
        "Median Bars Held",
        "Median Net PnL %",
        "Losing false starts share",
        "Bars Held = 1 share",
        "Reversal-signal share among false starts",
    ]
    values = [
        total,
        ratio_all,
        med_bars,
        med_net,
        losing_share,
        bars1_share,
        rev_share,
    ]
    return pd.DataFrame({"Metric": metrics, "Value": values})


def _write_false_start_sheet(
    writer: pd.ExcelWriter,
    trades_df_raw: Optional[pd.DataFrame],
    false_start_df: pd.DataFrame,
) -> None:
    """Write ``false start`` sheet: summary block, main table, autofilter on main header, datetimes."""
    from openpyxl.utils import get_column_letter

    _strip_tz_from_false_start_datetimes(false_start_df)

    summary_df = _build_false_start_summary_df(trades_df_raw, false_start_df)
    # Header + len(summary) rows, then one blank row, then main table (pandas 0-based startrow).
    main_startrow = len(summary_df) + 2

    summary_df.to_excel(writer, sheet_name=FALSE_START_SHEET_NAME, index=False)
    false_start_df.to_excel(
        writer, sheet_name=FALSE_START_SHEET_NAME, index=False, startrow=main_startrow
    )
    ws = writer.sheets[FALSE_START_SHEET_NAME]
    n_cols = len(false_start_df.columns)
    # Autofilter on main table header row (below summary + blank row).
    if n_cols > 0:
        header_excel_row = main_startrow + 1
        ws.auto_filter.ref = (
            f"A{header_excel_row}:{get_column_letter(n_cols)}{header_excel_row}"
        )

    if len(false_start_df) == 0:
        return

    for col_name in (
        TRADES_DISPLAY_NAMES["entry_time"],
        TRADES_DISPLAY_NAMES["exit_time"],
        SIGNALS_DISPLAY_NAMES["signal_time"],
    ):
        if col_name not in false_start_df.columns:
            continue
        series = false_start_df[col_name]
        if not pd.api.types.is_datetime64_any_dtype(series):
            continue
        drop = series.dropna()
        if len(drop) == 0:
            continue
        first_val = drop.iloc[0]
        if isinstance(first_val, pd.Timestamp) and (
            first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
        ):
            col_idx = list(false_start_df.columns).index(col_name) + 1
            first_data_excel_row = main_startrow + 2
            last_data_excel_row = main_startrow + 1 + len(false_start_df)
            for row_idx in range(first_data_excel_row, last_data_excel_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "YYYY-MM-DD HH:MM:SS"


def add_test_timestamp_to_filename(filepath: str, timestamp_format: str = "%Y%m%d_%H%M%S") -> str:
    """
    Add timestamp with "TEST" prefix to filename before extension.
    
    Args:
        filepath: Original file path (e.g., "result.xlsx")
        timestamp_format: Format for timestamp (default: YYYYMMDD_HHMMSS)
        
    Returns:
        New filepath with test timestamp (e.g., "test_result_20260211_142104.xlsx")
    """
    path = Path(filepath)
    timestamp = datetime.now().strftime(timestamp_format)
    new_name = f"test_{path.stem}_{timestamp}{path.suffix}"
    return str(path.parent / new_name)


def export_tester_results(
    period_results: List[PeriodResult],
    output_path: str,
    signals_df: Optional[pd.DataFrame] = None,
    false_start_max_bars: int = 4,
) -> str:
    """
    Export tester results to Excel file.

    Creates sheets:
    - Summary: 5 rows with main metrics for each period
    - Metrics_100, Metrics_75, Metrics_50, Metrics_33, Metrics_25: detailed metrics
    - Trades_100, Trades_75, Trades_50, Trades_33, Trades_25: all trades
    - Signals (optional): SuperTrend signal events from the 100% period — written
      only when ``signals_df`` is passed.
    - false start (always): trades from the 100% slice whose holding length
      (``bars_held``) is below the configured false-start threshold
      (``bars_held`` < ``false_start_max_bars``). Default threshold is 4 when the
      setting is omitted from YAML (``export.false_start_max_bars``). Summary
      metrics on that sheet are computed from this same filtered set.
      Signal-related columns use ``signals_df`` when provided; otherwise they are
      empty (NaN). This extends the legacy workbook layout versus older builds that
      had no such sheet.

    All sheet names are <= 31 characters.
    Trades sheets are created even if there are 0 trades (empty table with headers).

    Args:
        period_results: List of PeriodResult for 100%, 75%, 50%, 33%, 25%
        output_path: Path to output Excel file (timestamp with TEST will be added automatically)
        signals_df: Optional DataFrame from ``build_signal_events()``. When
            provided, the ``Signals`` sheet is written and its rows are used to
            populate signal columns on ``false start``. When ``None``, no
            ``Signals`` sheet is written and signal columns on ``false start`` are
            left empty.
        false_start_max_bars: Integer ``N`` >= 1; false-start rows are trades with
            ``bars_held`` < ``N``. Should match ``export.false_start_max_bars`` from
            tester config (default 4).

    Returns:
        Actual output path used (with TEST timestamp)
    """
    # Add test timestamp to filename
    output_path = add_test_timestamp_to_filename(output_path)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_data = []
        for pr in period_results:
            row = {
                "Period": pr.period_label,
                "ATR Period": pr.atr_period,
                "Multiplier": pr.multiplier,
                "Mode": pr.trade_mode,
                "Sum PnL %": pr.metrics.get("sum_pnl_pct", 0),
                "Sharpe": pr.metrics.get("sharpe", 0),
                "Sortino": pr.metrics.get("sortino", 0),
                "Max Drawdown": pr.metrics.get("max_drawdown", 0),
                "Win Rate": pr.metrics.get("win_rate", 0),
                "Num Trades": pr.metrics.get("num_trades", 0),
            }
            summary_data.append(row)
        
        summary_df = pd.DataFrame(summary_data)
        summary_df = format_excel_export_df(summary_df)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
        # Metrics and Trades sheets for each period
        for pr in period_results:
            # Determine sheet suffix (100, 75, 50, 33, 25)
            period_num = pr.period_label.replace("%", "")
            
            # Metrics sheet
            metrics_sheet = f"Metrics_{period_num}"
            if len(metrics_sheet) > 31:
                metrics_sheet = metrics_sheet[:31]
            
            # Create metrics DataFrame
            metrics_row = {
                "Period": pr.period_label,
                "ATR Period": pr.atr_period,
                "Multiplier": pr.multiplier,
                "Mode": pr.trade_mode,
                "Commission": pr.commission,
                "Warmup (requested)": pr.warmup,
                "Warmup (effective)": pr.result.effective_warmup,
            }
            # Add all metrics
            for key, display_name in METRICS_DISPLAY_NAMES.items():
                metrics_row[display_name] = pr.metrics.get(key, 0)
            
            metrics_df = pd.DataFrame([metrics_row])
            metrics_df = format_excel_export_df(metrics_df)
            metrics_df.to_excel(writer, sheet_name=metrics_sheet, index=False)
            
            # Trades sheet
            trades_sheet = f"Trades_{period_num}"
            if len(trades_sheet) > 31:
                trades_sheet = trades_sheet[:31]
            
            # Get trades from result (DD-01: extracted from full history)
            trades_df = pr.trades_df
            
            # Rename columns for display
            if trades_df is not None and len(trades_df) > 0:
                trades_df = trades_df.rename(columns=TRADES_DISPLAY_NAMES)
                # Remove timezone from datetime columns (Excel doesn't support timezone-aware datetimes)
                entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
                exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
                if entry_time_col in trades_df.columns:
                    if pd.api.types.is_datetime64_any_dtype(trades_df[entry_time_col]):
                        # Remove timezone info without changing time values (keep local time)
                        if trades_df[entry_time_col].dt.tz is not None:
                            trades_df[entry_time_col] = trades_df[entry_time_col].dt.tz_localize(None)
                if exit_time_col in trades_df.columns:
                    if pd.api.types.is_datetime64_any_dtype(trades_df[exit_time_col]):
                        # Remove timezone info without changing time values (keep local time)
                        if trades_df[exit_time_col].dt.tz is not None:
                            trades_df[exit_time_col] = trades_df[exit_time_col].dt.tz_localize(None)
            else:
                # Create empty DataFrame with display column names
                trades_df = pd.DataFrame(columns=list(TRADES_DISPLAY_NAMES.values()))
            
            trades_df.to_excel(writer, sheet_name=trades_sheet, index=False)
            
            # Format datetime columns if they contain time (not just date)
            if trades_df is not None and len(trades_df) > 0:
                worksheet = writer.sheets[trades_sheet]
                
                # Check if timestamps include time (non-midnight)
                entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
                exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
                
                if entry_time_col in trades_df.columns:
                    # Get first non-null entry_time to check format
                    first_entry = trades_df[entry_time_col].dropna()
                    if len(first_entry) > 0:
                        first_val = first_entry.iloc[0]
                        # Check if it's a datetime with time component
                        if isinstance(first_val, pd.Timestamp) and (first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0):
                            # Apply HH:MM:SS format to entry_time and exit_time columns
                            entry_col_idx = list(trades_df.columns).index(entry_time_col) + 1  # +1 for Excel 1-based
                            exit_col_idx = list(trades_df.columns).index(exit_time_col) + 1
                            
                            # Format all rows (skip header)
                            for row_idx in range(2, len(trades_df) + 2):  # Excel rows are 1-based, +1 for header
                                entry_cell = worksheet.cell(row=row_idx, column=entry_col_idx)
                                exit_cell = worksheet.cell(row=row_idx, column=exit_col_idx)
                                entry_cell.number_format = "YYYY-MM-DD HH:MM:SS"
                                exit_cell.number_format = "YYYY-MM-DD HH:MM:SS"

        if signals_df is not None:
            _write_signals_sheet(writer, "Signals", signals_df)

        pr_100 = next((pr for pr in period_results if pr.period_label == "100%"), None)
        trades_100_raw = pr_100.trades_df if pr_100 is not None else None

        # ZigZag Legs sheet (plan §3.8.2) — written for zz modes.
        _fd_100 = pr_100.result.filter_diagnostics if pr_100 is not None else None
        if _fd_100 and _fd_100.get("mode") in ("zigzag", "zigzag_and_volume"):
            from supertrend_optimizer.utils.enums import ExecutionModel as _EM
            _zz_legs_raw = _fd_100.get("zz_legs") or ()
            # D1: read execution_model from filter_diagnostics (set in engine/run.py).
            _em_val_100 = _fd_100.get("execution_model", "open_to_open")
            _exec_model_100 = _EM(_em_val_100)
            _zz_legs_linked = _link_trades_to_legs(
                trades_100_raw,
                _zz_legs_raw,
                _exec_model_100,
                zz_cand_leg_id=_fd_100.get("zz_cand_leg_id"),
            )
            # D2: use real DatetimeIndex from filter_diagnostics for Legs timestamps.
            _zz_index_100 = _fd_100.get("zz_index", pd.Index([]))
            _write_legs_sheet(writer, _zz_legs_linked, _zz_index_100, _exec_model_100)

            # D3+D4: enrich Trades_100 with Triggering Leg columns (§3.8.3).
            if trades_100_raw is not None and len(trades_100_raw) > 0:
                _is_o2o_100 = (_exec_model_100.value == "open_to_open")
                _enriched = _enrich_trades_with_leg_columns(
                    trades_100_raw,
                    _zz_legs_linked,
                    _is_o2o_100,
                    fd=_fd_100,
                )
                # Re-write Trades_100 sheet with enriched data.
                _enriched_display = _enriched.rename(columns=TRADES_DISPLAY_NAMES)
                # Strip tz info — openpyxl rejects tz-aware datetimes.  The
                # first-pass write via _write_trades_sheet handled this, but
                # the enriched re-write bypasses that guard because it
                # operates on ``trades_100_raw`` (untouched source frame).
                # Mirror the canonical pattern in _write_trades_sheet.
                for _ts_col in (
                    TRADES_DISPLAY_NAMES.get("entry_time"),
                    TRADES_DISPLAY_NAMES.get("exit_time"),
                ):
                    if _ts_col and _ts_col in _enriched_display.columns:
                        _col = _enriched_display[_ts_col]
                        if pd.api.types.is_datetime64_any_dtype(_col):
                            if _col.dt.tz is not None:
                                _enriched_display[_ts_col] = _col.dt.tz_localize(None)
                _enriched_display.to_excel(writer, sheet_name="Trades_100", index=False)

        false_start_df = _build_false_start_sheet_df(
            trades_100_raw, signals_df, false_start_max_bars=false_start_max_bars
        )
        _write_false_start_sheet(writer, trades_100_raw, false_start_df)

        # filters_summary sheet (plan §7.10)
        fd_list = [pr.result.filter_diagnostics for pr in period_results]
        labels = [pr.period_label for pr in period_results]
        _write_filters_summary_sheet(writer, fd_list, labels, include_total_row=False)

    return output_path


def _add_eqblk_timestamp_to_filename(
    filepath: str, timestamp_format: str = "%Y%m%d_%H%M%S"
) -> str:
    """Add timestamp with 'test_' prefix and '_eqblk' suffix to filename."""
    path = Path(filepath)
    timestamp = datetime.now().strftime(timestamp_format)
    new_name = f"test_{path.stem}_eqblk_{timestamp}{path.suffix}"
    return str(path.parent / new_name)


def _write_trades_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    trades_df: "pd.DataFrame | None",
) -> None:
    """Write a single trades sheet, reusing TRADES_DISPLAY_NAMES formatting."""
    if trades_df is not None and len(trades_df) > 0:
        trades_df = trades_df.rename(columns=TRADES_DISPLAY_NAMES)
        entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
        exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
        for col in (entry_time_col, exit_time_col):
            if col in trades_df.columns:
                if pd.api.types.is_datetime64_any_dtype(trades_df[col]):
                    if trades_df[col].dt.tz is not None:
                        trades_df[col] = trades_df[col].dt.tz_localize(None)
    else:
        trades_df = pd.DataFrame(columns=list(TRADES_DISPLAY_NAMES.values()))

    trades_df.to_excel(writer, sheet_name=sheet_name, index=False)

    if len(trades_df) > 0:
        worksheet = writer.sheets[sheet_name]
        entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
        exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
        if entry_time_col in trades_df.columns:
            first_entry = trades_df[entry_time_col].dropna()
            if len(first_entry) > 0:
                first_val = first_entry.iloc[0]
                if isinstance(first_val, pd.Timestamp) and (
                    first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
                ):
                    entry_col_idx = list(trades_df.columns).index(entry_time_col) + 1
                    exit_col_idx = list(trades_df.columns).index(exit_time_col) + 1
                    for row_idx in range(2, len(trades_df) + 2):
                        worksheet.cell(row=row_idx, column=entry_col_idx).number_format = "YYYY-MM-DD HH:MM:SS"
                        worksheet.cell(row=row_idx, column=exit_col_idx).number_format = "YYYY-MM-DD HH:MM:SS"


def _write_signals_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    signals_df: "pd.DataFrame | None",
) -> None:
    """Write the Signals sheet with autofilter and conditional formatting."""
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles import PatternFill

    has_data = signals_df is not None and len(signals_df) > 0

    if has_data:
        signals_df = signals_df.rename(columns=SIGNALS_DISPLAY_NAMES)
        time_col = SIGNALS_DISPLAY_NAMES["signal_time"]
        if time_col in signals_df.columns:
            if pd.api.types.is_datetime64_any_dtype(signals_df[time_col]):
                if signals_df[time_col].dt.tz is not None:
                    signals_df[time_col] = signals_df[time_col].dt.tz_localize(None)
    else:
        signals_df = pd.DataFrame(columns=list(SIGNALS_DISPLAY_NAMES.values()))

    signals_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    n_cols = len(signals_df.columns)

    # Autofilter on header row (always, even for empty sheet)
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if has_data:
        n_rows = len(signals_df)
        cols = list(signals_df.columns)

        # Datetime formatting for Signal Time
        time_col = SIGNALS_DISPLAY_NAMES["signal_time"]
        if time_col in cols:
            first_vals = signals_df[time_col].dropna()
            if len(first_vals) > 0:
                first_val = first_vals.iloc[0]
                if isinstance(first_val, pd.Timestamp) and (
                    first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
                ):
                    col_idx = cols.index(time_col) + 1
                    for row_idx in range(2, n_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = (
                            "YYYY-MM-DD HH:MM:SS"
                        )

        # Conditional formatting: green→yellow→red (bigger value = worse signal)
        # Applied to Signal Body % and Signal Range % only
        for internal_key in ("signal_body_pct", "signal_range_pct"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                ws.conditional_formatting.add(
                    cf_range,
                    ColorScaleRule(
                        start_type="min", start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="F8696B",
                    ),
                )

        # Two-level threshold rules for Median Ratio columns:
        #   > 2   → bright project-red  (rule added first = highest priority)
        #   > 1.6 → project-yellow      (only reached when the > 2 rule did not fire)
        _fill_ratio_strong = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
        _fill_ratio_soft   = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
        for internal_key in ("signal_body_pct_median_ratio", "signal_range_pct_median_ratio"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                # Add > 2 first (priority 1 — highest). stopIfTrue prevents > 1.6 from also firing.
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="greaterThan", formula=["2"],
                               fill=_fill_ratio_strong, stopIfTrue=True),
                )
                # Add > 1.6 second (priority 2). Only fires for 1.6 < x ≤ 2.
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="greaterThan", formula=["1.6"],
                               fill=_fill_ratio_soft),
                )

        # Highlight negative T+N returns
        _fill_neg = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for internal_key in ("t1_return_pct", "t2_return_pct", "t3_return_pct"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="lessThan", formula=["0"], fill=_fill_neg),
                )


def _entry_bar_for_decision(entry_index: int, is_o2o: bool) -> int:
    """Plan §3.8.3: convert trade.entry_index to decision bar index."""
    return entry_index - 1 if is_o2o else entry_index


def _enrich_trades_with_leg_columns(
    trades_df: pd.DataFrame,
    zz_legs_linked: list,
    is_o2o: bool,
    fd: Optional[dict] = None,
) -> pd.DataFrame:
    """
    Phase 5 (RFC v3.1 §7.7 / §7.8): enrich trades with Triggering-Leg columns
    and A-vs-B + orphan analytics.

    Columns added (in addition to pre-Phase-5 5 legacy):
      Legacy (§3.8.3):
        triggering_leg_id, triggering_leg_height_pct,
        triggering_leg_height_ratio, triggering_leg_direction,
        bars_from_leg_confirm_to_entry,          # SIGN-FUL per §7.7
        correction_height_pct_before_entry       # None when pre_confirm per §7.7
      Phase 5 (§7.7 / §7.8 fix N-07):
        entry_vs_leg_phase        ∈ {pre_confirm, post_confirm, unlinked}
        triggering_arm_source     ∈ {A, B, BOTH, NONE} (from lg.arm_source;
                                    orphan → from zz_arm_source[decision_bar])
        cand_side_at_entry        ∈ {LONG, SHORT, NONE} (from leg_direction)
        cand_height_pct_at_entry  from zz_cand_height_pct[decision_bar]
        cand_leg_id_at_entry      from zz_cand_leg_id[decision_bar]

    Parameters
    ----------
    fd : filter_diagnostics dict, optional
        When supplied, enables Phase 5 per-bar columns and orphan arm_source
        attribution.  When omitted (pre-Phase-5 call-sites / fixture tests),
        the Phase 5 columns are filled with None / legacy defaults and
        semantics reduce to the pre-Phase-5 contract.
    """
    result = trades_df.copy()
    leg_by_trade: dict = {
        int(lg.trade_id_if_fired): lg
        for lg in zz_legs_linked
        if lg.trade_id_if_fired is not None
    }

    # Per-bar arrays from filter_diagnostics (optional).
    zz_cand_leg_id = fd.get("zz_cand_leg_id") if fd else None
    zz_cand_height_pct = fd.get("zz_cand_height_pct") if fd else None
    zz_leg_direction = fd.get("zz_leg_direction") if fd else None
    zz_arm_source_arr = fd.get("zz_arm_source") if fd else None

    def _decision_bar(entry_index: int) -> int:
        return _entry_bar_for_decision(int(entry_index), is_o2o)

    def _cand_side_label(dec: int) -> Optional[str]:
        """cand_side is opposite of leg_direction on the candidate (§5.3)."""
        if zz_leg_direction is None or dec < 0 or dec >= len(zz_leg_direction):
            return None
        ld = int(zz_leg_direction[dec])
        if ld == 1:    # LEG_DIR_UP → candidate short
            return "SHORT"
        if ld == -1:   # LEG_DIR_DOWN → candidate long
            return "LONG"
        return "NONE"

    def _cand_height_pct_at(dec: int):
        if zz_cand_height_pct is None or dec < 0 or dec >= len(zz_cand_height_pct):
            return None
        v = float(zz_cand_height_pct[dec])
        return v if np.isfinite(v) else None

    def _cand_leg_id_at(dec: int):
        if zz_cand_leg_id is None or dec < 0 or dec >= len(zz_cand_leg_id):
            return None
        v = int(zz_cand_leg_id[dec])
        return v if v >= 0 else None

    def _arm_source_at(dec: int) -> Optional[str]:
        if zz_arm_source_arr is None or dec < 0 or dec >= len(zz_arm_source_arr):
            return None
        return _ZZ_ENUM_LABELS["arm_source"].get(int(zz_arm_source_arr[dec]), "NONE")

    def _row_dec(row) -> int:
        return _decision_bar(row.get("entry_index", -1))

    def _tl_id(tid):
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        return int(lg.leg_id) if lg else None

    def _tl_height_pct(tid):
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        return float(lg.height_pct) if lg else None

    def _tl_height_ratio(tid):
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is None:
            return None
        gm = float(lg.global_median_at_confirm)
        return float(lg.height_pct) / gm if gm > 0 else None

    def _tl_direction(tid):
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is None:
            return None
        return _ZZ_ENUM_LABELS["leg_direction"].get(int(lg.direction), "")

    def _bars_to_entry(row):
        tid = row.get("trade_id")
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is None:
            return None
        # §7.7: SIGN-FUL.  May be negative for pre-confirm trades.
        return _row_dec(row) - int(lg.confirm_bar)

    def _entry_vs_leg_phase(row):
        tid = row.get("trade_id")
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is None:
            return "unlinked"
        # Post-confirm is the legacy case: decision_bar >= confirm_bar.
        return "pre_confirm" if _row_dec(row) < int(lg.confirm_bar) else "post_confirm"

    def _triggering_arm_source(row):
        tid = row.get("trade_id")
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is not None:
            return _ZZ_ENUM_LABELS["arm_source"].get(int(getattr(lg, "arm_source", 0)), "NONE")
        # Orphan: attribute arm_source from per-bar diagnostics at decision_bar.
        return _arm_source_at(_row_dec(row))

    def _correction_height(row):
        tid = row.get("trade_id")
        lg = leg_by_trade.get(int(tid)) if tid is not None else None
        if lg is None:
            return None
        dec = _row_dec(row)
        # §7.7 pre_confirm → always None (semantic not defined).
        if dec < int(lg.confirm_bar):
            return None
        trig_dir = int(lg.direction)
        opposite_dir = -trig_dir
        trig_confirm = int(lg.confirm_bar)
        for candidate in zz_legs_linked:
            if (
                int(candidate.direction) == opposite_dir
                and int(candidate.confirm_bar) > trig_confirm
                and int(candidate.confirm_bar) < dec
            ):
                return float(candidate.height_pct)
        return None

    if "trade_id" not in result.columns:
        for col in (
            "triggering_leg_id",
            "triggering_leg_height_pct",
            "triggering_leg_height_ratio",
            "triggering_leg_direction",
            "bars_from_leg_confirm_to_entry",
            "correction_height_pct_before_entry",
            "entry_vs_leg_phase",
            "triggering_arm_source",
            "cand_side_at_entry",
            "cand_height_pct_at_entry",
            "cand_leg_id_at_entry",
        ):
            result[col] = None
        return result

    trade_ids = result["trade_id"]
    result["triggering_leg_id"]          = trade_ids.map(_tl_id)
    result["triggering_leg_height_pct"]  = trade_ids.map(_tl_height_pct)
    result["triggering_leg_height_ratio"] = trade_ids.map(_tl_height_ratio)
    result["triggering_leg_direction"]   = trade_ids.map(_tl_direction)
    result["bars_from_leg_confirm_to_entry"] = result.apply(_bars_to_entry, axis=1)
    result["correction_height_pct_before_entry"] = result.apply(_correction_height, axis=1)
    # Phase 5 additive columns (§7.7 / §7.8 / fix N-07).
    result["entry_vs_leg_phase"]    = result.apply(_entry_vs_leg_phase, axis=1)
    result["triggering_arm_source"] = result.apply(_triggering_arm_source, axis=1)
    result["cand_side_at_entry"]       = result["entry_index"].map(
        lambda ei: _cand_side_label(_decision_bar(ei)) if ei is not None else None
    )
    result["cand_height_pct_at_entry"] = result["entry_index"].map(
        lambda ei: _cand_height_pct_at(_decision_bar(ei)) if ei is not None else None
    )
    result["cand_leg_id_at_entry"]     = result["entry_index"].map(
        lambda ei: _cand_leg_id_at(_decision_bar(ei)) if ei is not None else None
    )
    return result


def _link_trades_to_legs(
    trades_df: Optional[pd.DataFrame],
    zz_legs: tuple,
    execution_model,
    zz_cand_leg_id: Optional[np.ndarray] = None,
) -> list:
    """
    RFC v3.1 §7.6 two-step trade↔leg linkage (audit fix B-02, RP-1 Block 1-D).

    Returns a list of LegRecord (dataclass.replace used for frozen dataclasses).

    Step 1 — post-confirm legacy criterion:
        match = { lg | lg.shot_bar == decision_bar AND lg.fired == YES_SHOT }
        → sets lg.trade_id_if_fired only.  Legacy shot_bar / fired stay
          as written by core; pre_confirm_shot_bar is NOT touched.

    Step 2 — pre-confirm candidate linkage (Phase 5, any arm_source):
        cand_id_at_dec = zz_cand_leg_id[decision_bar]
        match = { lg | lg.start_bar <= decision_bar < lg.confirm_bar
                       AND lg._cand_leg_id_at_confirm == cand_id_at_dec }
        → sets lg.trade_id_if_fired.
        → sets lg.pre_confirm_shot_bar ONLY if it is still -1 (legacy
          backfill path).  If core already wrote a value (RP-1C — core
          is source-of-truth for pre-confirm shot events):
            core value == decision_bar   → keep core value
            core value != decision_bar   → raise RuntimeError (io/core
                                            divergence)

    Invariants enforced:
    - Legacy shot_bar/arm_bar on LegRecord unchanged (pre-confirm data
      lives in pre_confirm_* fields only — inv 8.1.10 / 8.3.7).
    - Mutual-exclusion (LX-08 / RP-1D): a single trade_id never links
      through both Step 1 and Step 2.  The structural `continue` on
      Step 1 guarantees this; a belt-and-suspenders assertion at the
      end of the function also verifies it across the whole pass.
    - Multiple-match → warn, skip linkage.
    - Orphan trade (no match on either step) → leg stays unlinked.

    Backward-compat: zz_cand_leg_id is optional.  Pre-Phase-5 call-sites
    (and fixture-based tests constructing raw LegRecord tuples) may omit it;
    in that case Step 2 is skipped silently and only Step 1 runs, matching
    legacy behaviour.
    """
    from supertrend_optimizer.core.zigzag_filter import FIRED_YES_SHOT

    legs_list = list(zz_legs)
    if trades_df is None or len(trades_df) == 0 or not legs_list:
        return legs_list

    from supertrend_optimizer.utils.enums import ExecutionModel
    is_o2o = (execution_model == ExecutionModel.OPEN_TO_OPEN)

    # Index of already-linked leg positions (prevents one leg being linked
    # to two trades).  Both steps respect it.
    linked_positions: set = set()

    # RP-1D / LX-08 — defensive mutual-exclusion tracking.  Each trade_id
    # can land in AT MOST ONE of the two sets.  Verified at the end of
    # the function.  `None` trade_ids (shouldn't happen in practice) are
    # skipped from the guard.
    step1_linked_trades: set = set()
    step2_linked_trades: set = set()

    # Cache: cand_id → leg_idx for Step 2 (only legs with valid _cand_leg_id_at_confirm).
    cand_to_leg_idx: Dict[int, int] = {}
    for i, lg in enumerate(legs_list):
        cid = int(getattr(lg, "_cand_leg_id_at_confirm", -1))
        if cid >= 0:
            # First-write-wins if multiple legs somehow share a cand_id (shouldn't).
            cand_to_leg_idx.setdefault(cid, i)

    for trade in trades_df.itertuples(index=False):
        entry_idx = int(getattr(trade, "entry_index", -1))
        if entry_idx < 0:
            continue
        dec = _entry_bar_for_decision(entry_idx, is_o2o)
        trade_id = getattr(trade, "trade_id", None)

        # ---------- Step 1 — post-confirm legacy criterion ----------
        matched_positions = [
            i for i, lg in enumerate(legs_list)
            if i not in linked_positions
            and int(lg.shot_bar) == dec
            and int(lg.fired) == int(FIRED_YES_SHOT)
        ]
        if len(matched_positions) == 1:
            pos = matched_positions[0]
            legs_list[pos] = dataclasses.replace(
                legs_list[pos], trade_id_if_fired=trade_id
            )
            linked_positions.add(pos)
            if trade_id is not None:
                step1_linked_trades.add(int(trade_id))
            # Mutual-exclusion: skip Step 2 for this trade entirely.
            continue
        if len(matched_positions) > 1:
            warnings.warn(
                f"trade {trade_id} matches multiple legs with "
                f"fired=YES_SHOT and shot_bar={dec}; linkage dropped",
                stacklevel=2,
            )
            continue

        # ---------- Step 2 — pre-confirm candidate linkage (§7.6) ----------
        if zz_cand_leg_id is None or len(zz_cand_leg_id) == 0:
            continue
        if dec < 0 or dec >= len(zz_cand_leg_id):
            continue
        cand_id_at_dec = int(zz_cand_leg_id[dec])
        if cand_id_at_dec < 0:
            warnings.warn(
                f"trade {trade_id} — candidate id unknown at decision_bar={dec} "
                f"(orphan trade; triggering_leg_id will be None)",
                stacklevel=2,
            )
            continue
        pos = cand_to_leg_idx.get(cand_id_at_dec)
        if pos is None or pos in linked_positions:
            # No leg ever confirmed for this candidate id — orphan.
            warnings.warn(
                f"trade {trade_id} — no post-hoc candidate match for "
                f"cand_leg_id={cand_id_at_dec} (orphan trade)",
                stacklevel=2,
            )
            continue
        lg = legs_list[pos]
        # Window check: pre-confirm decision must fall strictly before confirm.
        if not (int(lg.start_bar) <= dec < int(lg.confirm_bar)):
            warnings.warn(
                f"trade {trade_id} — cand_leg_id={cand_id_at_dec} matched "
                f"leg_id={lg.leg_id} but decision_bar={dec} is outside "
                f"[{lg.start_bar},{lg.confirm_bar}); orphan",
                stacklevel=2,
            )
            continue

        # RP-1D R-3 — pre_confirm_shot_bar reconciliation with core.
        # After RP-1C, core (`_unified_armament_fsm` flush) is the
        # authoritative writer for `pre_confirm_shot_bar`.  The io layer
        # may fill the field for legacy call-sites that never ran core
        # (e.g. fixture-based unit tests, backfill tools), but must
        # NEVER silently overwrite a value already set by core.
        existing_pcsb = int(lg.pre_confirm_shot_bar)
        if existing_pcsb == -1:
            new_pcsb = int(dec)
        elif existing_pcsb == int(dec):
            # Core already set the same value — io agrees, keep it.
            new_pcsb = existing_pcsb
        else:
            raise RuntimeError(
                f"trade {trade_id} Step 2 linkage for leg_id={lg.leg_id}: "
                f"core-set pre_confirm_shot_bar={existing_pcsb} disagrees "
                f"with io decision_bar={int(dec)}.  Core is source-of-truth "
                "for pre-confirm shot events (RFC v3.1 §7.5 / RP-1C); io "
                "must not overwrite with a different value."
            )

        # RP-1D LX-08 — defensive mutual-exclusion assert (belt-and-
        # suspenders; the `continue` above already guarantees this).
        if trade_id is not None and int(trade_id) in step1_linked_trades:
            raise RuntimeError(
                f"trade {trade_id} reached Step 2 after Step 1 already "
                "linked it — two-step linkage mutual-exclusion violated "
                "(LX-08 / RFC §7.6)."
            )

        legs_list[pos] = dataclasses.replace(
            lg,
            trade_id_if_fired=trade_id,
            pre_confirm_shot_bar=new_pcsb,
        )
        linked_positions.add(pos)
        if trade_id is not None:
            step2_linked_trades.add(int(trade_id))

    # RP-1D LX-08 — end-of-pass mutual-exclusion invariant.
    both = step1_linked_trades & step2_linked_trades
    if both:
        raise RuntimeError(
            f"two-step linkage mutual-exclusion violated (LX-08): trade_ids "
            f"linked via BOTH Step 1 and Step 2: {sorted(both)}"
        )

    return legs_list


def _write_legs_sheet(
    writer: "pd.ExcelWriter",
    zz_legs: list,
    index: "pd.DatetimeIndex",
    execution_model,
) -> None:
    """
    Plan §3.8.2: write the Legs sheet (one row per LegRecord).

    Columns as specified in §3.8.2; enum fields rendered via _ZZ_ENUM_LABELS.
    """
    if not zz_legs:
        pd.DataFrame().to_excel(writer, sheet_name="Legs", index=False)
        return

    rows = []
    n_idx = len(index)
    for lg in zz_legs:
        cb = int(lg.confirm_bar)
        eb = int(lg.end_bar)
        sb = int(lg.start_bar)

        def _ts(bar: int):
            if 0 <= bar < n_idx:
                return index[bar]
            return pd.NaT

        confirm_ts = _ts(cb)
        end_ts     = _ts(eb)
        start_ts   = _ts(sb)

        session_date = (
            confirm_ts.date() if not pd.isna(confirm_ts) else None
        )

        gm = getattr(lg, "global_median_at_confirm", None)
        height_ratio = (
            round(float(lg.height_pct) / float(gm), 6)
            if gm and float(gm) > 0
            else float("nan")
        )

        fired_int = int(getattr(lg, "fired", 0))
        arm_bar   = getattr(lg, "arm_bar", None)
        shot_bar  = getattr(lg, "shot_bar", None)
        armed_side_int = int(getattr(lg, "armed_side", 0)) if hasattr(lg, "armed_side") else 0
        dir_int        = int(lg.direction)

        rows.append({
            "Leg ID":              int(lg.leg_id),
            "Session Date":        session_date,
            "Direction":           _ZZ_ENUM_LABELS["leg_direction"].get(dir_int, str(dir_int)),
            "Start Bar":           sb,
            "End Bar":             eb,
            "Confirm Bar":         cb,
            "Start Time":          start_ts,
            "End Time":            end_ts,
            "Confirm Time":        confirm_ts,
            "Start Hour":          start_ts.hour if not pd.isna(start_ts) else None,
            "End Hour":            end_ts.hour if not pd.isna(end_ts) else None,
            "Confirm Hour":        confirm_ts.hour if not pd.isna(confirm_ts) else None,
            "Start Price":         float(lg.start_price),
            "End Price":           float(lg.end_price),
            "Height %":            float(lg.height_pct),
            "Height Ratio":        height_ratio,
            "Length Bars":         int(lg.length_bars),
            "Confirm Lag Bars":    int(lg.confirm_lag_bars),
            "N Legs Before":       int(getattr(lg, "n_legs_before", 0)),
            "Global Median %":     float(gm) if gm is not None else float("nan"),
            "Global P80 %":        float(getattr(lg, "global_p80_at_confirm", 0) or 0),
            "Local Median %":      float(getattr(lg, "local_median_at_confirm", 0) or 0),
            "Regime At Confirm":   _ZZ_ENUM_LABELS["regime_state"].get(
                                       int(getattr(lg, "regime_state_at_confirm", 0)), ""),
            "Opened Regime":       bool(getattr(lg, "opened_regime", False)),
            "Closed Regime":       bool(getattr(lg, "closed_regime", False)),
            "Is Strong":           bool(getattr(lg, "is_strong", False)),
            "Armed Side":          _ZZ_ENUM_LABELS["armed_side"].get(armed_side_int, ""),
            "Arm Bar":             int(arm_bar) if arm_bar is not None else None,
            "Fired":               _ZZ_ENUM_LABELS["fired"].get(fired_int, str(fired_int)),
            "Shot Bar":            int(shot_bar) if shot_bar is not None else None,
            "Trade ID":            lg.trade_id_if_fired,
            # Phase 5 (RFC v3.1 §7.5) — pre-confirm lifecycle + arm-source.
            "Arm Source":          _ZZ_ENUM_LABELS["arm_source"].get(
                                        int(getattr(lg, "arm_source", 0)), "NONE"),
            "Armed By Candidate":  bool(getattr(lg, "armed_by_candidate", False)),
            "Pre-Confirm Arm Bar": int(getattr(lg, "pre_confirm_arm_bar", -1)),
            "Pre-Confirm Shot Bar": int(getattr(lg, "pre_confirm_shot_bar", -1)),
        })

    legs_df = pd.DataFrame(rows)
    # Strip timezone info from datetime columns — openpyxl rejects
    # tz-aware datetimes.  Other sheets (Trades / Signals) already do
    # the same at their write-sites; the Legs sheet was missing this
    # guard and tripped on tz-aware CSVs (e.g. `data.csv` is UTC+03:00).
    for _ts_col in ("Start Time", "End Time", "Confirm Time"):
        if _ts_col in legs_df.columns:
            _s = pd.to_datetime(legs_df[_ts_col], errors="coerce")
            if getattr(_s.dt, "tz", None) is not None:
                legs_df[_ts_col] = _s.dt.tz_localize(None)
    legs_df.to_excel(writer, sheet_name="Legs", index=False)


def _blocked_volatility_label(mode: str) -> str:
    """
    Return the display label for the ``blocked_by_volatility`` bucket.

    In amplitude modes the bucket contains amp-side blockings (patch §D.2).
    In legacy modes it contains ATR-side blockings.
    """
    if mode in ("amplitude", "amplitude_and_volume"):
        return "Blocked: Amplitude (incl. ATR floor)"
    return "Blocked: Volatility"


def _write_filters_summary_sheet(
    writer: "pd.ExcelWriter",
    filter_diagnostics_list: List[Optional[dict]],
    labels: List[str],
    include_total_row: bool = False,
) -> None:
    """
    Write a ``filters_summary`` sheet.

    Plan §7.10 contract:
    - Секция параметров: mode, thresholds (including amp params for amp modes).
    - Секция агрегатов: 7 counter keys.
    - For equal_blocks: include_total_row=True adds a totals row.
    - Does not read ``filters_cfg`` directly — only ``filter_diagnostics``.
    - Does not raise when ``trades_df`` is empty (counters may be 0).

    v1.3 (patch §D.2): ``Blocked: Volatility`` label is mode-dependent:
    - legacy modes  → ``"Blocked: Volatility"``
    - amplitude modes → ``"Blocked: Amplitude (incl. ATR floor)"``
    The underlying counter key (``blocked_by_volatility``) is unchanged.

    Args:
        writer: Open ``pd.ExcelWriter``.
        filter_diagnostics_list: One dict per period/segment (or None for legacy
            results that predate filter support).
        labels: Row labels (e.g. ["100%","75%",...] or ["S1","S2",...]).
        include_total_row: If True, append a "Total" row summing the counters.
    """
    rows = []
    for label, fd in zip(labels, filter_diagnostics_list):
        if fd is None:
            fd = {
                "mode": "none",
                "thresholds": {},
                "counters": {
                    "raw_entry_signals": 0,
                    "passed_entry_signals": 0,
                    "blocked_entry_signals": 0,
                    "blocked_by_volatility": 0,
                    "blocked_by_volume": 0,
                    "blocked_by_both": 0,
                    "blocked_by_vol_ma_invalid": 0,
                },
            }
        thresh = fd.get("thresholds", {}) or {}
        c = fd.get("counters", {}) or {}
        mode = fd.get("mode", "none")
        vol_label = _blocked_volatility_label(mode)

        # Amplitude parameter columns (patch §D.2 / §4.4.2).
        # Echoed from thresholds["amplitude"] if present (engine writes it
        # for amp modes via _thresholds_for_diagnostics).
        amp_thresh = thresh.get("amplitude") or {}

        # Amp ATR Floor in both fraction (config unit) and percent (display unit)
        # so that users can directly compare with the ATR % column in Signals
        # without mental unit conversion (patch §F8).
        _amp_atr_floor_frac = amp_thresh.get("atr_floor")
        _amp_atr_floor_pct = (
            round(_amp_atr_floor_frac * 100.0, 6)
            if _amp_atr_floor_frac is not None
            else None
        )

        # ZigZag parameter columns (plan §3.8.4).
        zz_thresh = thresh.get("zigzag") or {}
        _is_zz_mode_row = mode in ("zigzag", "zigzag_and_volume")

        # ZigZag leg aggregates (from zz_legs in filter_diagnostics).
        _zz_legs = fd.get("zz_legs") or ()
        _n_legs = len(_zz_legs)
        _n_up = sum(1 for lg in _zz_legs if int(lg.direction) == 1)
        _n_dn = _n_legs - _n_up
        _heights = [float(lg.height_pct) for lg in _zz_legs] if _zz_legs else []
        _med_h = float(np.median(_heights)) if _heights else float("nan")
        _p80_h = float(np.percentile(_heights, 80)) if _heights else float("nan")
        _max_h = float(max(_heights)) if _heights else float("nan")
        _lengths = [int(lg.length_bars) for lg in _zz_legs] if _zz_legs else []
        _med_len = float(np.median(_lengths)) if _lengths else float("nan")
        _lags = [int(lg.confirm_lag_bars) for lg in _zz_legs] if _zz_legs else []
        _med_lag = float(np.median(_lags)) if _lags else float("nan")

        # ZigZag armament aggregates.
        from supertrend_optimizer.core.zigzag_filter import (
            FIRED_NONE, FIRED_YES_SHOT, FIRED_NO_NEW_PIVOT, FIRED_NO_TIMEOUT_SOFT,
            FIRED_NO_TIMEOUT_HARD, FIRED_SESSION_RESET, FIRED_NO_REGIME_OFF,
            ARM_SRC_NONE, ARM_SRC_A, ARM_SRC_B, ARM_SRC_BOTH,
        )
        # arm_bar is int with sentinel -1 (not Optional), so `is not None` is
        # always True.  Correct check: arm_bar != -1 (or equivalently fired !=
        # FIRED_NONE, which is the canonical way to detect an armed leg).
        _n_armed = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", FIRED_NONE)) != int(FIRED_NONE))
        _n_yes   = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_YES_SHOT))
        _n_nnp   = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_NO_NEW_PIVOT))
        _n_soft  = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_NO_TIMEOUT_SOFT))
        _n_hard  = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_NO_TIMEOUT_HARD))
        _n_sess  = sum(1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_SESSION_RESET))
        # Phase 5 (RFC v3.1 §7.8 / fix N-12) — FIRED_NO_REGIME_OFF breakdown.
        _n_no_regime_off = sum(
            1 for lg in _zz_legs if int(getattr(lg, "fired", 0)) == int(FIRED_NO_REGIME_OFF)
        )
        _fired_rate = round(_n_yes / _n_armed * 100, 2) if _n_armed > 0 else float("nan")

        # Phase 5 (RFC v3.1 §7.8) — A-vs-B trade attribution aggregates.
        # Uses enriched trades_df if available via outer scope (trades_100_raw).
        # Falls back to leg-level arm_source distribution so the columns are
        # always populated for zigzag runs (non-zz rows stay at None).
        _trades_from_a = _trades_from_b = _trades_from_both = _trades_orphan = 0
        # Prefer trade-level attribution when an enriched Trades sheet is in
        # scope (populated by the caller).  The `trades_df_enriched` symbol
        # is attached by the outer writer before calling this builder.
        _enriched = locals().get("_fd_trades_enriched")
        if _enriched is not None and "triggering_arm_source" in _enriched.columns:
            _tas = _enriched["triggering_arm_source"].fillna("").astype(str)
            _tid = _enriched.get("triggering_leg_id")
            _is_orphan = (_tid.isna()) if _tid is not None else False
            _trades_from_a   = int(((_tas == "A")   & ~_is_orphan).sum())
            _trades_from_b   = int(((_tas == "B")   & ~_is_orphan).sum())
            _trades_from_both = int(((_tas == "BOTH") & ~_is_orphan).sum())
            _trades_orphan   = int(_is_orphan.sum()) if hasattr(_is_orphan, "sum") else 0
        else:
            # Fallback: legs with trade_id_if_fired set, grouped by arm_source.
            for lg in _zz_legs:
                if lg.trade_id_if_fired is None:
                    continue
                _src = int(getattr(lg, "arm_source", 0))
                if _src == int(ARM_SRC_A):
                    _trades_from_a += 1
                elif _src == int(ARM_SRC_B):
                    _trades_from_b += 1
                elif _src == int(ARM_SRC_BOTH):
                    _trades_from_both += 1

        # ZigZag regime aggregates (from per-bar arrays).
        _zz_regime_arr = fd.get("zz_regime_state")
        if _zz_regime_arr is not None and _is_zz_mode_row:
            _n_closed = int(np.sum(_zz_regime_arr == 0))
            _n_grace  = int(np.sum(_zz_regime_arr == 1))
            _n_active = int(np.sum(_zz_regime_arr == 2))
        else:
            _n_closed = _n_grace = _n_active = None

        # Per-reason breakdown for blocked_by_volatility (plan §3.8.4).
        _reason_arr = fd.get("filtered_reason")
        def _count_reason(code: str) -> int:
            if _reason_arr is None:
                return 0
            return int(np.sum(_reason_arr == code))

        row: dict = {
            "Label": label,
            "Filter Mode": mode,
            # Legacy volatility params
            "Min ATR %": thresh.get("min_atr_pct"),
            "Max ATR %": thresh.get("max_atr_pct"),
            # Amplitude params (None for non-amp modes)
            "Amp N": amp_thresh.get("n"),
            "Amp Lookback": amp_thresh.get("lookback"),
            "Amp Q": amp_thresh.get("q"),
            "Amp Min Sep": amp_thresh.get("min_separation"),
            "Amp ATR Floor": _amp_atr_floor_frac,
            "Amp ATR Floor %": _amp_atr_floor_pct,
            # Volume params
            "Volume MA Col": thresh.get("volume_ma_column"),
            "Global Volume MA Mean": thresh.get("global_volume_ma_mean"),
            "Min Ratio": thresh.get("min_ratio"),
            "Max Ratio": thresh.get("max_ratio"),
            # ZigZag params (None for non-zz modes)
            "ZZ Reversal Threshold": zz_thresh.get("reversal_threshold"),
            "ZZ Min Legs Global": zz_thresh.get("min_legs_global"),
            "ZZ Q Strong": zz_thresh.get("q_strong"),
            "ZZ K Local": zz_thresh.get("k_local"),
            "ZZ Entry Side": zz_thresh.get("entry_side"),
            "ZZ Arm Timeout Soft": zz_thresh.get("arm_timeout_bars_since_extreme"),
            "ZZ Arm Timeout Hard": zz_thresh.get("arm_timeout_bars_hard"),
            # Counters
            "Raw Signals": c.get("raw_entry_signals", 0),
            "Passed Signals": c.get("passed_entry_signals", 0),
            "Blocked Signals": c.get("blocked_entry_signals", 0),
            vol_label: c.get("blocked_by_volatility", 0),
            "Blocked: Volume": c.get("blocked_by_volume", 0),
            "Blocked: Both": c.get("blocked_by_both", 0),
            "Blocked: Vol MA Invalid": c.get("blocked_by_vol_ma_invalid", 0),
        }

        if _is_zz_mode_row:
            # ZigZag per-reason breakdown of blocked_by_volatility (plan §3.8.4).
            row.update({
                "ZZ Blocked: Warmup":        _count_reason("zz_warmup"),
                "ZZ Blocked: Regime Off":    _count_reason("zz_regime_off"),
                "ZZ Blocked: Expired Time":  _count_reason("zz_expired_time"),
                "ZZ Blocked: Expired Pivot": _count_reason("zz_expired_new_pivot"),
                "ZZ Blocked: Same Leg":      _count_reason("zz_locked_same_leg"),
                "ZZ Blocked: Not Armed":     _count_reason("zz_not_armed"),
                "ZZ Blocked: Armed Waiting": _count_reason("zz_armed_waiting"),
                "ZZ Blocked: Pathological":  _count_reason("zz_pathological"),
                # ZigZag leg aggregates
                "ZZ Total Legs": _n_legs,
                "ZZ Legs Up": _n_up,
                "ZZ Legs Down": _n_dn,
                "ZZ Median Height %": _med_h,
                "ZZ P80 Height %": _p80_h,
                "ZZ Max Height %": _max_h,
                "ZZ Median Length Bars": _med_len,
                "ZZ Median Confirm Lag": _med_lag,
                # ZigZag armament aggregates
                "ZZ N Armed Legs": _n_armed,
                "ZZ N Fired YES": _n_yes,
                "ZZ N Fired No New Pivot": _n_nnp,
                "ZZ N Fired Timeout Soft": _n_soft,
                "ZZ N Fired Timeout Hard": _n_hard,
                "ZZ N Fired Session Reset": _n_sess,
                # Phase 5 (RFC v3.1 §7.8 / fix N-12).
                "ZZ Legs No Regime Off": _n_no_regime_off,
                "ZZ Fired Rate %": _fired_rate,
                # ZigZag regime aggregates
                "ZZ Bars Closed": _n_closed,
                "ZZ Bars Grace": _n_grace,
                "ZZ Bars Active": _n_active,
                # Phase 5 (RFC v3.1 §7.8) — A-vs-B trade attribution.
                "ZZ Trades From Contour A":    _trades_from_a,
                "ZZ Trades From Contour B":    _trades_from_b,
                "ZZ Trades From Both":         _trades_from_both,
                "ZZ Trades Orphan":            _trades_orphan,
            })

        rows.append(row)

    if include_total_row and rows:
        # Counter columns that should be summed. The volatility label may
        # differ per row in theory (all rows share the same mode in practice),
        # so use the first row's label for the total column name.
        first_vol_label = _blocked_volatility_label(
            rows[0].get("Filter Mode", "none") if rows else "none"
        )
        int_cols = [
            "Raw Signals", "Passed Signals", "Blocked Signals",
            first_vol_label, "Blocked: Volume",
            "Blocked: Both", "Blocked: Vol MA Invalid",
            # ZigZag per-reason counters (present only in zz rows)
            "ZZ Blocked: Warmup", "ZZ Blocked: Regime Off",
            "ZZ Blocked: Expired Time", "ZZ Blocked: Expired Pivot",
            "ZZ Blocked: Same Leg", "ZZ Blocked: Not Armed",
            "ZZ Blocked: Armed Waiting", "ZZ Blocked: Pathological",
            "ZZ Total Legs", "ZZ Legs Up", "ZZ Legs Down",
            "ZZ N Armed Legs", "ZZ N Fired YES", "ZZ N Fired No New Pivot",
            "ZZ N Fired Timeout Soft", "ZZ N Fired Timeout Hard",
            "ZZ N Fired Session Reset",
            "ZZ Legs No Regime Off",
            "ZZ Bars Closed", "ZZ Bars Grace", "ZZ Bars Active",
            "ZZ Trades From Contour A", "ZZ Trades From Contour B",
            "ZZ Trades From Both", "ZZ Trades Orphan",
        ]
        total_row = {col: "" for col in rows[0]}
        total_row["Label"] = "Total"
        total_row["Filter Mode"] = rows[0].get("Filter Mode", "none") if rows else "none"
        for col in int_cols:
            total_row[col] = sum(r.get(col, 0) or 0 for r in rows)
        # Thresholds come from first row (same for all segments).
        for col in ("Min ATR %", "Max ATR %", "Amp N", "Amp Lookback",
                    "Amp Q", "Amp Min Sep", "Amp ATR Floor", "Amp ATR Floor %",
                    "Volume MA Col", "Global Volume MA Mean",
                    "Min Ratio", "Max Ratio",
                    "ZZ Reversal Threshold", "ZZ Min Legs Global", "ZZ Q Strong",
                    "ZZ K Local", "ZZ Entry Side", "ZZ Arm Timeout Soft",
                    "ZZ Arm Timeout Hard"):
            total_row[col] = rows[0].get(col)
        rows.append(total_row)

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="filters_summary", index=False)


# Metrics used in the aggregate table (Summary block 2).
_AGGREGATE_METRIC_KEYS = [
    "sum_pnl_pct",
    "sharpe",
    "sortino",
    "max_drawdown",
    "cagr",
    "win_rate",
    "num_trades",
    "profit_factor",
    "avg_trade",
]


def export_equal_blocks_results(
    segment_results: List[SegmentResult],
    output_path: str,
) -> str:
    """
    Export equal_blocks segmentation results to Excel.

    Sheet layout:
    - Summary:        Two tables with an empty-row separator.
                      Table 1: one row per segment with ATR Period, Multiplier, Mode
                      and key metrics.
                      Table 2: aggregate stats (Mean/Std/Min/Max/Median)
                      across segments, INVALID_METRIC_VALUE excluded.
    - Metrics_S1..SN: Detailed metrics per segment (one row each).
    - Trades_S1..SN:  Trades per segment (from segment_trades_df).

    Filename pattern: test_<stem>_eqblk_<timestamp>.xlsx

    Args:
        segment_results: List of SegmentResult from run_equal_blocks().
        output_path: Path to output Excel file.

    Returns:
        Actual output path used (with timestamp).
    """
    output_path = _add_eqblk_timestamp_to_filename(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Summary Table 1: per-segment rows ──
        summary_rows = []
        for seg in segment_results:
            m = seg.segment_metrics
            row = {
                "Segment": seg.segment_label,
                "Range": seg.range_label,
                "Bars": seg.n_bars,
                "ATR Period": seg.atr_period,
                "Multiplier": seg.multiplier,
                "Mode": seg.trade_mode,
                "Start Date": seg.start_date,
                "End Date": seg.end_date,
                "Prepend Bars": seg.prepend_bars,
                "Sum PnL %": m.get("sum_pnl_pct", 0),
                "Sharpe": m.get("sharpe", 0),
                "Sortino": m.get("sortino", 0),
                "Max Drawdown": m.get("max_drawdown", 0),
                "Win Rate": m.get("win_rate", 0),
                "Num Trades": m.get("num_trades", 0),
                "Profit Factor": m.get("profit_factor", 0),
                "Avg Trade": m.get("avg_trade", 0),
            }
            summary_rows.append(row)

        summary_df = pd.DataFrame(summary_rows)

        # Strip timezone from date columns so Excel can write them
        for date_col in ("Start Date", "End Date"):
            if date_col in summary_df.columns:
                col = summary_df[date_col]
                if pd.api.types.is_datetime64_any_dtype(col):
                    if col.dt.tz is not None:
                        summary_df[date_col] = col.dt.tz_localize(None)
                else:
                    # Object dtype: strip tz from individual Timestamp values
                    def _strip_tz(v):
                        if isinstance(v, pd.Timestamp) and v.tzinfo is not None:
                            return v.tz_localize(None)
                        return v
                    summary_df[date_col] = summary_df[date_col].map(_strip_tz)

        # ── Summary Table 2: aggregate stats ──
        agg_rows = []
        for key in _AGGREGATE_METRIC_KEYS:
            display_name = METRICS_DISPLAY_NAMES.get(key, key)
            values = []
            for seg in segment_results:
                v = seg.segment_metrics.get(key)
                if v is not None and v != INVALID_METRIC_VALUE and np.isfinite(v):
                    values.append(float(v))

            if values:
                arr = np.array(values)
                agg_rows.append({
                    "Metric": display_name,
                    "Mean": float(np.mean(arr)),
                    "Std": float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0,
                    "Min": float(np.min(arr)),
                    "Max": float(np.max(arr)),
                    "Median": float(np.median(arr)),
                })
            else:
                agg_rows.append({
                    "Metric": display_name,
                    "Mean": None,
                    "Std": None,
                    "Min": None,
                    "Max": None,
                    "Median": None,
                })

        agg_df = pd.DataFrame(agg_rows)

        # Write Table 1
        summary_df = format_excel_export_df(summary_df)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        # Write Table 2 after an empty row
        start_row = len(summary_df) + 2  # +1 header + 1 empty row (0-indexed)
        agg_df = format_excel_export_df(agg_df)
        agg_df.to_excel(
            writer, sheet_name="Summary", index=False, startrow=start_row
        )

        # ── Per-segment sheets ──
        for seg in segment_results:
            label = seg.segment_label  # "S1", "S2", ...

            # Metrics sheet
            metrics_sheet = f"Metrics_{label}"
            if len(metrics_sheet) > 31:
                metrics_sheet = metrics_sheet[:31]

            metrics_row = {
                "Segment": seg.segment_label,
                "Range": seg.range_label,
                "ATR Period": seg.atr_period,
                "Multiplier": seg.multiplier,
                "Mode": seg.trade_mode,
                "Commission": seg.commission,
                "Prepend Bars": seg.prepend_bars,
            }
            for key, display_name in METRICS_DISPLAY_NAMES.items():
                metrics_row[display_name] = seg.segment_metrics.get(key, 0)

            pd.DataFrame([metrics_row]).pipe(format_excel_export_df).to_excel(
                writer, sheet_name=metrics_sheet, index=False
            )

            # Trades sheet
            trades_sheet = f"Trades_{label}"
            if len(trades_sheet) > 31:
                trades_sheet = trades_sheet[:31]

            _write_trades_sheet(writer, trades_sheet, seg.segment_trades_df)

        # ZigZag Legs sheet (plan §3.8.2) — only for zz modes.
        # For equal_blocks: merge legs from all segments (combine tuple legs,
        # deduplicate by leg_id since each segment processes an ext slice).
        _first_fd = segment_results[0].filter_diagnostics if segment_results else None
        if _first_fd and _first_fd.get("mode") in ("zigzag", "zigzag_and_volume"):
            from supertrend_optimizer.utils.enums import ExecutionModel as _EM
            # Collect all legs across segments (may overlap in prepend windows).
            # We rely on deduplication by leg_id (keep highest confirm_bar on tie).
            _all_legs_by_id: dict = {}
            for seg in segment_results:
                _seg_fd = seg.filter_diagnostics or {}
                for lg in _seg_fd.get("zz_legs") or ():
                    _all_legs_by_id[int(lg.leg_id)] = lg
            _merged_legs = sorted(_all_legs_by_id.values(), key=lambda lg: lg.confirm_bar)
            # D1+D2: read execution_model and zz_index from first segment diagnostics.
            _seg0_fd = segment_results[0].filter_diagnostics or {}
            _em_val = _seg0_fd.get("execution_model", "open_to_open")
            _exec_model = _EM(_em_val)
            _zz_index = _seg0_fd.get("zz_index", pd.Index([]))
            _write_legs_sheet(writer, _merged_legs, _zz_index, _exec_model)

        # filters_summary sheet (plan §7.10) — one row per segment + total
        fd_list = [seg.filter_diagnostics for seg in segment_results]
        labels_fs = [seg.segment_label for seg in segment_results]
        _write_filters_summary_sheet(
            writer, fd_list, labels_fs, include_total_row=True
        )

    return output_path