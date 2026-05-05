"""
Excel export functionality for SuperTrend Tester.

This module exports tester results to Excel format with metrics and trades.
Phase 2 (WP-T7): Conditional filter columns in Signals/Trades, optional
FilterDiagnostics_100 / ZigZag_Trigger_Events / filters_summary sheets,
Summary filter block — all gated on trade_filter_config.enabled and
diagnostics flags.  Disabled path is bit-identical to the pre-Phase-2 baseline.
"""

from typing import Any, Dict, List, Mapping, Optional, Tuple
from datetime import datetime
from pathlib import Path
import numpy as np
import pandas as pd

from supertrend_optimizer.testing.runner import PeriodResult, SegmentResult
from supertrend_optimizer.utils.constants import INVALID_METRIC_VALUE
from supertrend_optimizer.io.excel_format_helpers import format_excel_export_df
from supertrend_optimizer.core.zigzag_st_filter import (
    compute_confirmed_legs_reset_aware,
    compute_zigzag_per_bar,
)


# ---------------------------------------------------------------------------
# Display name maps
# ---------------------------------------------------------------------------

# Metrics column display names
METRICS_DISPLAY_NAMES = {
    "sum_pnl_pct": "Sum PnL %",
    "sharpe": "Sharpe",
    "sortino": "Sortino",
    "max_drawdown": "Max Drawdown",
    "cagr": "CAGR",
    "win_rate": "Win Rate",
    "num_trades": "Num Trades",
    "profit_factor": "Profit Factor",
    "avg_trade": "Avg Trade"
}

# Signals column display names (base + WP-T7 filter columns)
SIGNALS_DISPLAY_NAMES = {
    "signal_time":                    "Signal Time",
    "signal_bar_index":               "Signal Bar Index",
    "event_type":                     "Event Type",
    "direction":                      "Direction",
    "st_color_before":                "ST Color Before",
    "st_color_after":                 "ST Color After",
    "is_reversal":                    "Is Reversal",
    # WP-T7: 4 filter columns (plan §9.5 — after Is Reversal, before Exec Price)
    "filter_state_at_signal":         "Filter State at Signal",
    "filter_decision":                "Filter Decision",
    "filter_block_reason":            "Filter Block Reason",
    "filter_trigger_source":          "Filter Trigger Source",
    # base continues
    "exec_price":                     "Exec Price",
    "signal_open_price":              "Signal Open Price",
    "signal_close_price":             "Signal Close Price",
    "signal_body_pct":                "Signal Body %",
    "signal_range_pct":               "Signal Range %",
    "signal_body_atr":                "Signal Body ATR",
    "signal_range_atr":               "Signal Range ATR",
    "t1_return_pct":                  "T+1 Return %",
    "t2_return_pct":                  "T+2 Return %",
    "t3_return_pct":                  "T+3 Return %",
    "signal_body_pct_median_ratio":   "Signal Body % / Median",
    "signal_range_pct_median_ratio":  "Signal Range % / Median",
}

# Trades column display names (canonical + WP-T7 filter diagnostics columns)
TRADES_DISPLAY_NAMES = {
    "trade_id":            "Trade ID",
    "direction":           "Direction",
    "entry_time":          "Entry Time",
    "entry_index":         "Entry Index",
    "entry_price":         "Entry Price",
    "exit_time":           "Exit Time",
    "exit_index":          "Exit Index",
    "exit_price":          "Exit Price",
    "bars_held":           "Bars Held",
    "gross_pnl_pct":       "Gross PnL %",
    "commission_pct":      "Commission %",
    "net_pnl_pct":         "Net PnL %",
    "supertrend_color":    "SuperTrend Color",
    # WP-T7: filter trade diagnostics (plan §9.5 — after canonical donor columns)
    "entry_filter_state":  "Entry Filter State",
    "entry_trigger_source":"Entry Trigger Source",
    "exit_reason":         "Exit Reason",
}

# Internal keys of filter-specific trade columns (used to separate canonical vs filter cols)
_FILTER_TRADE_COLUMN_KEYS: frozenset = frozenset({
    "entry_filter_state", "entry_trigger_source", "exit_reason",
})

# Canonical trade column display names (no filter cols — for disabled-path empty header)
_CANONICAL_TRADE_DISPLAY_COLS: List[str] = [
    v for k, v in TRADES_DISPLAY_NAMES.items() if k not in _FILTER_TRADE_COLUMN_KEYS
]

# Per-bar filter diagnostics display names (for FilterDiagnostics_100 sheet, plan §3.3.1)
FILTER_DIAGNOSTICS_100_DISPLAY_NAMES: Dict[str, str] = {
    "trade_filter_enabled":         "Filter Enabled",
    "trade_filter_state":           "Filter State",
    "trade_filter_trigger_source":  "Trigger Source",
    "zigzag_reversal_threshold":    "Reversal Threshold",
    "candidate_height_pct":         "Candidate Height %",
    "candidate_trigger_threshold":  "Candidate Trigger Threshold",
    # WP-V3-8: new v3 display columns (§11.1), ordered after Candidate Trigger Threshold
    "zigzag_mode":                              "ZigZag Mode",
    "candidate_age_bars":                       "Candidate Age Bars",
    "candidate_leg_direction":                  "Candidate Leg Direction",
    "candidate_duration_gate_enabled":          "Candidate Duration Gate Enabled",
    "candidate_duration_max_bars":              "Candidate Duration Max Bars",
    "candidate_duration_gate_passed":           "Candidate Duration Gate Passed",
    "candidate_threshold_ok":                   "Candidate Threshold OK",
    "candidate_component_ok":                   "Candidate Component OK",
    "confirmed_median_ok":                      "Confirmed Median OK",
    "b_component_ok":                           "B Component OK",
    "immediate_allowed":                        "Immediate Allowed",
    "immediate_candidate_entry_used":           "Immediate Candidate Entry Used",
    "immediate_candidate_entry_block_reason":   "Immediate Candidate Entry Block Reason",
    "local_median_N":               "Local Median N",
    "local_median_available":       "Local Median Available",
    "local_window":                 "Local Window",
    "global_median":                "Global Median",
    "global_stats_available":       "Global Stats Available",
    "confirmed_legs_since_start":   "Confirmed Legs Since Start",
    "freeze_confirmed_legs":        "Freeze Confirmed Legs",
    "median_stop_triggered":        "Median Stop Triggered",
    "stopping_started_at_index":    "Stopping Started At Index",
    "exit_off_mode":                "Exit-OFF Mode",
    "exit_off_zz_leg_count":        "Exit-OFF ZZ Leg Count",
    "zz_legs_since_lifecycle_start": "ZZ Legs Since Start",
    "zz_leg_stop_triggered":        "ZZ Leg Stop Triggered",
    "filter_allowed_entry":         "Filter Allowed Entry",
    "filter_block_reason":          "Filter Block Reason",
    "trade_filter_state_code":      "Filter State Code",
    "st_flip_dir":                  "ST Flip Direction",
}

# ZigZag_Trigger_Events sheet column order (plan §9.2, extended by WP-V3-8 §11.2)
_TRIGGER_EVENTS_COLUMNS = (
    "Trigger ID",
    "Trigger Bar",
    "Trigger Time",
    "Trigger Source",
    "Threshold Used",
    "Quantile Used",
    "Global Median",
    "Local Median N",
    "Candidate Height %",
    "Triggered Lifecycle Start",
    "Linked Trade ID",
    # WP-V3-8: new §11.2 columns
    "ZigZag Mode",
    "Immediate Candidate Entry Used",
    "Immediate Candidate Entry Block Reason",
    "Candidate Age Bars",
    "Candidate Leg Direction",
    "Candidate Duration Gate Passed",
)

# Cycle sheet contract (plan_xlsx_cycle_sheet_implementation §4, §11).
CYCLE_SHEET_NAME = "cycle"
CYCLE_SHEET_COLUMNS: Tuple[str, ...] = (
    "Начало цикла",
    "Конец цикла",
    "Направление цикла",
    "Баров в цикле",
    "Ног ZigZag в цикле",
    "Медиана ног",
    "Ног выше порога триггера кандидата",
    "Размер цикла, %",
    "ID цикла",
    "Start bar index",
    "End bar index",
    "Цена начала",
    "Цена конца",
    "High цикла",
    "Low цикла",
    "Макс. движение по циклу, %",
    "Макс. просадка внутри цикла, %",
    "Причина завершения",
    "Макс. высота ноги",
    "Доля ног выше порога, %",
    "Сделок в цикле",
    "Фин результат цикла, %",
    "% сделок с положительным фин результатом в цикле",
)

# Legacy export: diagnostic sheet for very short trades (100% slice only)
FALSE_START_SHEET_NAME = "false start"

# Canonical false-start columns (no filter; filter_block_reason col added conditionally)
FALSE_START_COLUMNS: Tuple[str, ...] = (
    TRADES_DISPLAY_NAMES["trade_id"],
    TRADES_DISPLAY_NAMES["direction"],
    TRADES_DISPLAY_NAMES["entry_time"],
    TRADES_DISPLAY_NAMES["entry_price"],
    TRADES_DISPLAY_NAMES["exit_time"],
    TRADES_DISPLAY_NAMES["exit_price"],
    TRADES_DISPLAY_NAMES["bars_held"],
    TRADES_DISPLAY_NAMES["net_pnl_pct"],
    TRADES_DISPLAY_NAMES["commission_pct"],
    SIGNALS_DISPLAY_NAMES["signal_time"],
    SIGNALS_DISPLAY_NAMES["event_type"],
    SIGNALS_DISPLAY_NAMES["is_reversal"],
    SIGNALS_DISPLAY_NAMES["exec_price"],
    SIGNALS_DISPLAY_NAMES["signal_body_pct_median_ratio"],
    SIGNALS_DISPLAY_NAMES["signal_range_pct_median_ratio"],
    SIGNALS_DISPLAY_NAMES["t1_return_pct"],
    SIGNALS_DISPLAY_NAMES["t2_return_pct"],
    SIGNALS_DISPLAY_NAMES["t3_return_pct"],
    "False Start Type",
)

# Extended false-start columns with filter block reason (enabled + export_state_columns=True)
FALSE_START_COLUMNS_WITH_FILTER: Tuple[str, ...] = FALSE_START_COLUMNS + (
    "Filter Block Reason at Signal",
)


# ---------------------------------------------------------------------------
# Low-level helpers (unchanged from pre-Phase-2)
# ---------------------------------------------------------------------------

def _normalize_timestamp_naive(ts: Any) -> pd.Timestamp:
    """Strip timezone for comparisons / Excel compatibility."""
    if pd.isna(ts):
        return pd.NaT
    t = pd.Timestamp(ts)
    if t.tzinfo is not None:
        return t.tz_localize(None)
    return t


def _excel_safe_datetime_value(value: Any) -> Any:
    """Return a timezone-naive datetime-like value suitable for Excel."""
    if value is None or pd.isna(value):
        return np.nan
    if isinstance(value, pd.Timestamp):
        return value.tz_localize(None) if value.tzinfo is not None else value
    if isinstance(value, np.datetime64):
        return pd.Timestamp(value)
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _materialize_cycle_float_scalar(values: Any, n: int) -> Optional[float]:
    """Materialize a uniform finite scalar from the aligned diagnostics prefix."""
    if values is None or n <= 0:
        return None
    arr = np.asarray(values, dtype=object)
    if arr.ndim == 0:
        arr = arr.reshape(1)

    finite_values: List[float] = []
    for value in arr[:n]:
        if isinstance(value, (bool, np.bool_)):
            return None
        try:
            value_f = float(value)
        except (TypeError, ValueError):
            if pd.isna(value):
                continue
            return None
        if np.isfinite(value_f):
            finite_values.append(value_f)

    if not finite_values:
        return None
    first = finite_values[0]
    if any(value != first for value in finite_values):
        return None
    return float(first)


def _materialize_cycle_candidate_threshold(values: Any, n: int) -> Optional[float]:
    return _materialize_cycle_float_scalar(values, n)


def _materialize_cycle_reversal_threshold(values: Any, n: int) -> Optional[float]:
    scalar = _materialize_cycle_float_scalar(values, n)
    if scalar is None or scalar <= 0.0 or scalar >= 1.0:
        return None
    return scalar


def _materialize_cycle_local_window(values: Any, n: int) -> Optional[int]:
    scalar = _materialize_cycle_float_scalar(values, n)
    if scalar is None or not float(scalar).is_integer():
        return None
    window = int(scalar)
    if window < 1:
        return None
    return window


def _cycle_direction_symbol(direction: Any) -> str:
    try:
        direction_i = int(direction)
    except (TypeError, ValueError):
        return ""
    if direction_i > 0:
        return "+"
    if direction_i < 0:
        return "-"
    return ""


def _resolve_cycle_direction(
    trigger_source: Any,
    candidate_leg_direction: Any,
    start_bar: int,
    confirmed_legs: List[Any],
) -> str:
    source = str(trigger_source)
    if source in ("candidate_threshold", "both"):
        return _cycle_direction_symbol(candidate_leg_direction)
    if source == "confirmed_median":
        for leg in confirmed_legs:
            if getattr(leg, "confirm_bar", None) == start_bar:
                return _cycle_direction_symbol(getattr(leg, "direction", None))
    return ""


def _empty_cycle_sheet_df() -> pd.DataFrame:
    return pd.DataFrame(columns=CYCLE_SHEET_COLUMNS)


def _completed_cycle_segments(state_arr: np.ndarray) -> List[Tuple[int, int]]:
    segments: List[Tuple[int, int]] = []
    start: Optional[int] = None
    for idx, state in enumerate(state_arr):
        active = str(state) != "OFF"
        if active and start is None:
            start = idx
        elif not active and start is not None:
            segments.append((start, idx - 1))
            start = None
    return segments


def _cycle_trades_for_segment(
    trades_df: Optional[pd.DataFrame],
    start_bar: int,
    end_bar: int,
) -> List[Any]:
    if trades_df is None or len(trades_df) == 0 or "entry_index" not in trades_df.columns:
        return []

    rows: List[Any] = []
    for row in trades_df.itertuples(index=False):
        entry_index = getattr(row, "entry_index", np.nan)
        if pd.isna(entry_index):
            continue
        try:
            entry_signal_idx = max(int(entry_index) - 1, 0)
        except (TypeError, ValueError):
            continue
        if start_bar <= entry_signal_idx <= end_bar:
            rows.append(row)
    return rows


def _cycle_positive_trades_pct(
    trades_df: Optional[pd.DataFrame],
    cycle_trades: List[Any],
) -> float:
    if not cycle_trades:
        return float("nan")
    if trades_df is None or "net_pnl_pct" not in trades_df.columns:
        return float("nan")

    positive = 0
    for row in cycle_trades:
        pnl = getattr(row, "net_pnl_pct", np.nan)
        try:
            pnl_f = float(pnl)
        except (TypeError, ValueError):
            return float("nan")
        if not np.isfinite(pnl_f):
            return float("nan")
        if pnl_f > 0.0:
            positive += 1
    return positive / len(cycle_trades) * 100.0


def _cycle_final_result_pct(
    trades_df: Optional[pd.DataFrame],
    cycle_trades: List[Any],
) -> float:
    if not cycle_trades:
        return float("nan")
    if trades_df is None or "net_pnl_pct" not in trades_df.columns:
        return float("nan")

    total = 0.0
    for row in cycle_trades:
        pnl = getattr(row, "net_pnl_pct", np.nan)
        try:
            pnl_f = float(pnl)
        except (TypeError, ValueError):
            return float("nan")
        if not np.isfinite(pnl_f):
            return float("nan")
        total += pnl_f
    return total


def _build_cycle_sheet_df(
    filter_diagnostics: Dict[str, np.ndarray],
    df: Optional[pd.DataFrame],
    trades_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    result_empty = _empty_cycle_sheet_df()
    required_keys = (
        "trade_filter_state",
        "trade_filter_trigger_source",
        "candidate_trigger_threshold",
        "zigzag_reversal_threshold",
        "local_window",
        "daily_reset_event",
        "candidate_leg_direction",
    )
    if not isinstance(filter_diagnostics, dict):
        return result_empty
    if df is None or not {"close", "high", "low"}.issubset(df.columns):
        return result_empty
    if any(key not in filter_diagnostics for key in required_keys):
        return result_empty

    try:
        close = np.asarray(df["close"].to_numpy(), dtype=np.float64)
        high = np.asarray(df["high"].to_numpy(), dtype=np.float64)
        low = np.asarray(df["low"].to_numpy(), dtype=np.float64)
    except (TypeError, ValueError):
        return result_empty

    arrays: Dict[str, np.ndarray] = {
        key: np.asarray(filter_diagnostics[key]) for key in required_keys
    }
    n = min(
        len(df),
        len(arrays["trade_filter_state"]),
        len(close),
        len(high),
        len(low),
        len(arrays["trade_filter_trigger_source"]),
        len(arrays["candidate_leg_direction"]),
        len(arrays["candidate_trigger_threshold"]),
        len(arrays["zigzag_reversal_threshold"]),
        len(arrays["local_window"]),
        len(arrays["daily_reset_event"]),
    )
    if n == 0:
        return result_empty

    close = close[:n]
    high = high[:n]
    low = low[:n]
    state_arr = arrays["trade_filter_state"][:n]
    trigger_source_arr = arrays["trade_filter_trigger_source"][:n]
    candidate_dir_arr = arrays["candidate_leg_direction"][:n]
    daily_reset_arr = arrays["daily_reset_event"][:n]

    candidate_threshold = _materialize_cycle_candidate_threshold(
        arrays["candidate_trigger_threshold"], n
    )
    reversal_threshold = _materialize_cycle_reversal_threshold(
        arrays["zigzag_reversal_threshold"], n
    )
    if reversal_threshold is None:
        return result_empty
    local_window = _materialize_cycle_local_window(arrays["local_window"], n)
    if local_window is None:
        return result_empty

    per_bar = compute_zigzag_per_bar(
        close,
        reversal_threshold,
        local_window,
        daily_reset_event=daily_reset_arr,
    )
    confirmed_legs = compute_confirmed_legs_reset_aware(
        close,
        reversal_threshold,
        daily_reset_event=daily_reset_arr,
    )

    index = df.iloc[:n].index
    rows: List[Dict[str, Any]] = []
    for start_bar, end_bar in _completed_cycle_segments(state_arr):
        interval = slice(start_bar, end_bar + 1)
        direction = _resolve_cycle_direction(
            trigger_source_arr[start_bar],
            candidate_dir_arr[start_bar],
            start_bar,
            confirmed_legs,
        )

        leg_mask = np.asarray(per_bar.confirm_event[interval]) == 1
        leg_heights = np.asarray(per_bar.last_confirmed_leg_height_pct[interval], dtype=np.float64)[leg_mask]
        leg_heights = leg_heights[np.isfinite(leg_heights)]
        legs_count = int(len(leg_heights))
        if legs_count:
            median_legs = float(np.median(leg_heights))
            max_leg_height = float(np.max(leg_heights))
        else:
            median_legs = float("nan")
            max_leg_height = float("nan")

        if candidate_threshold is None:
            legs_above_threshold: Any = float("nan")
            share_above_threshold = float("nan")
        else:
            legs_above_threshold = int(np.sum(leg_heights > candidate_threshold))
            share_above_threshold = (
                float(legs_above_threshold) / legs_count * 100.0
                if legs_count > 0 else float("nan")
            )

        interval_close = close[interval]
        interval_high = high[interval]
        interval_low = low[interval]
        ohlc_valid = bool(
            np.all(np.isfinite(interval_close))
            and np.all(np.isfinite(interval_high))
            and np.all(np.isfinite(interval_low))
        )
        close_start = float(close[start_bar]) if ohlc_valid else float("nan")
        close_end = float(close[end_bar]) if ohlc_valid else float("nan")
        high_cycle = float(np.max(interval_high)) if ohlc_valid else float("nan")
        low_cycle = float(np.min(interval_low)) if ohlc_valid else float("nan")

        if ohlc_valid and close_start > 0.0:
            cycle_size_pct = (close_end - close_start) / close_start * 100.0
            if direction == "+":
                max_move_pct = (high_cycle - close_start) / close_start * 100.0
                max_drawdown_pct = (close_start - low_cycle) / close_start * 100.0
            elif direction == "-":
                max_move_pct = (close_start - low_cycle) / close_start * 100.0
                max_drawdown_pct = (high_cycle - close_start) / close_start * 100.0
            else:
                max_move_pct = float("nan")
                max_drawdown_pct = float("nan")
        else:
            cycle_size_pct = float("nan")
            max_move_pct = float("nan")
            max_drawdown_pct = float("nan")

        off_bar = end_bar + 1
        end_reason = (
            "daily_reset"
            if off_bar < n and int(daily_reset_arr[off_bar]) == 1
            else "FSM_OFF"
        )
        cycle_trades = _cycle_trades_for_segment(trades_df, start_bar, end_bar)

        rows.append({
            "Начало цикла": _excel_safe_datetime_value(index[start_bar]),
            "Конец цикла": _excel_safe_datetime_value(index[end_bar]),
            "Направление цикла": direction,
            "Баров в цикле": end_bar - start_bar + 1,
            "Ног ZigZag в цикле": legs_count,
            "Медиана ног": median_legs,
            "Ног выше порога триггера кандидата": legs_above_threshold,
            "Размер цикла, %": cycle_size_pct,
            "ID цикла": 0,
            "Start bar index": start_bar,
            "End bar index": end_bar,
            "Цена начала": close_start,
            "Цена конца": close_end,
            "High цикла": high_cycle,
            "Low цикла": low_cycle,
            "Макс. движение по циклу, %": max_move_pct,
            "Макс. просадка внутри цикла, %": max_drawdown_pct,
            "Причина завершения": end_reason,
            "Макс. высота ноги": max_leg_height,
            "Доля ног выше порога, %": share_above_threshold,
            "Сделок в цикле": len(cycle_trades),
            "Фин результат цикла, %": (
                _cycle_final_result_pct(trades_df, cycle_trades)
            ),
            "% сделок с положительным фин результатом в цикле": (
                _cycle_positive_trades_pct(trades_df, cycle_trades)
            ),
        })

    if not rows:
        return result_empty

    result = pd.DataFrame(rows, columns=CYCLE_SHEET_COLUMNS)
    result = result.sort_values("Start bar index", kind="stable").reset_index(drop=True)
    result["ID цикла"] = np.arange(1, len(result) + 1, dtype=np.int64)
    return result.loc[:, list(CYCLE_SHEET_COLUMNS)]


def _write_cycle_sheet(writer: pd.ExcelWriter, cycle_df: pd.DataFrame) -> None:
    cycle_df.to_excel(writer, sheet_name=CYCLE_SHEET_NAME, index=False)

    ws = writer.sheets[CYCLE_SHEET_NAME]
    from openpyxl.utils import get_column_letter

    n_cols = len(cycle_df.columns)
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    datetime_columns = {"Начало цикла", "Конец цикла"}
    for col_idx, col_name in enumerate(cycle_df.columns, start=1):
        if col_name not in datetime_columns:
            continue
        for row_idx in range(2, len(cycle_df) + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = "YYYY-MM-DD HH:MM:SS"


def _price_matches_exec_entry(exec_price: Any, entry_price: Any, rtol: float = 1e-5, atol: float = 1e-8) -> bool:
    if pd.isna(exec_price) or pd.isna(entry_price):
        return False
    try:
        return bool(np.isclose(float(exec_price), float(entry_price), rtol=rtol, atol=atol))
    except (TypeError, ValueError):
        return False


def _match_trade_to_signal_row(
    trade_row: pd.Series,
    signals_prep: Optional[pd.DataFrame],
) -> Optional[pd.Series]:
    """
    Deterministic trade → signal match for legacy false-start diagnostics.

    1) Same Direction, signal_time <= entry_time, Exec Price matches Entry Price
       → choose the signal with the maximum signal_time (closest preceding among price hits).
    2) Else: same Direction, signal_time <= entry_time → nearest preceding (max signal_time).
    3) Else: None (caller fills NaNs).
    """
    if signals_prep is None or len(signals_prep) == 0:
        return None

    t_dir = trade_row.get("direction")
    if pd.isna(t_dir):
        return None

    entry_t = _normalize_timestamp_naive(trade_row.get("entry_time"))
    if pd.isna(entry_t):
        return None

    entry_price = trade_row.get("entry_price")

    same_dir = signals_prep[signals_prep["direction"] == t_dir]
    if same_dir.empty:
        return None

    before = same_dir[same_dir["signal_time"] <= entry_t]
    if before.empty:
        return None

    # Step 1 — exec price matches entry price
    price_hits = before[
        before.apply(
            lambda r: _price_matches_exec_entry(r.get("exec_price"), entry_price),
            axis=1,
        )
    ]
    if len(price_hits) > 0:
        best_idx = price_hits["signal_time"].idxmax()
        return price_hits.loc[best_idx]

    # Step 2 — nearest preceding (max signal_time)
    best_idx = before["signal_time"].idxmax()
    return before.loc[best_idx]


def _prepare_signals_lookup(signals_df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    """Copy signals with naive signal_time for matching (internal column names)."""
    if signals_df is None or len(signals_df) == 0:
        return None
    df = signals_df.copy()
    if "signal_time" not in df.columns or "direction" not in df.columns:
        return None
    st = pd.to_datetime(df["signal_time"])
    if pd.api.types.is_datetime64_any_dtype(st):
        if st.dt.tz is not None:
            st = st.dt.tz_localize(None)
    df["signal_time"] = st
    return df


def _classify_false_start_type(
    gross: Any,
    net: Any,
    bars_held: Any,
    t1: Any,
    t2: Any,
    t3: Any,
) -> str:
    """Rule-based False Start Type (informational only; priority order)."""
    if pd.notna(gross) and pd.notna(net) and gross > 0 and net <= 0:
        return "commission_drag"

    if pd.notna(net) and pd.notna(bars_held):
        try:
            b = float(bars_held)
        except (TypeError, ValueError):
            b = float("nan")
        if not np.isnan(b) and b == 1.0 and net < 0:
            return "instant_loss"

    if pd.notna(net) and net <= 0:
        t2_pos = pd.notna(t2) and t2 > 0
        t3_pos = pd.notna(t3) and t3 > 0
        if t2_pos or t3_pos:
            return "missed_followthrough"

    if pd.notna(net) and net <= 0:
        available = [x for x in (t1, t2, t3) if pd.notna(x)]
        if len(available) > 0 and all(x <= 0 for x in available):
            return "correct_early_exit"

    return "flat_noise"


def _build_false_start_sheet_df(
    trades_df_raw: Optional[pd.DataFrame],
    signals_df: Optional[pd.DataFrame],
    false_start_max_bars: int = 4,
    include_filter_block_reason: bool = False,
) -> pd.DataFrame:
    """
    Build the false start DataFrame; may be empty with headers only.

    ``include_filter_block_reason=True`` adds "Filter Block Reason at Signal"
    column after "False Start Type" (plan §9.2.2; gated on
    ``diagnostics.export_state_columns=True`` AND ``enabled``).
    """
    col_spec = FALSE_START_COLUMNS_WITH_FILTER if include_filter_block_reason else FALSE_START_COLUMNS
    empty = pd.DataFrame(columns=list(col_spec))
    if trades_df_raw is None or len(trades_df_raw) == 0:
        return empty
    if "bars_held" not in trades_df_raw.columns:
        return empty

    short_tr = trades_df_raw[trades_df_raw["bars_held"] < false_start_max_bars].copy()
    if len(short_tr) == 0:
        return empty

    need_gross = "gross_pnl_pct" in short_tr.columns

    signals_lookup = _prepare_signals_lookup(signals_df)

    rows_out: List[dict] = []
    for _, trow in short_tr.iterrows():
        gross = trow["gross_pnl_pct"] if need_gross else float("nan")
        net = trow.get("net_pnl_pct")
        bars = trow.get("bars_held")

        matched = _match_trade_to_signal_row(trow, signals_lookup)

        sig_time = matched.get("signal_time") if matched is not None else np.nan
        evt = matched.get("event_type") if matched is not None else np.nan
        rev = matched.get("is_reversal") if matched is not None else np.nan
        ex_px = matched.get("exec_price") if matched is not None else np.nan
        body_m = matched.get("signal_body_pct_median_ratio") if matched is not None else np.nan
        rng_m = matched.get("signal_range_pct_median_ratio") if matched is not None else np.nan
        t1 = matched.get("t1_return_pct") if matched is not None else np.nan
        t2 = matched.get("t2_return_pct") if matched is not None else np.nan
        t3 = matched.get("t3_return_pct") if matched is not None else np.nan

        ftype = _classify_false_start_type(gross, net, bars, t1, t2, t3)

        row_dict = {
            TRADES_DISPLAY_NAMES["trade_id"]: trow.get("trade_id"),
            TRADES_DISPLAY_NAMES["direction"]: trow.get("direction"),
            TRADES_DISPLAY_NAMES["entry_time"]: trow.get("entry_time"),
            TRADES_DISPLAY_NAMES["entry_price"]: trow.get("entry_price"),
            TRADES_DISPLAY_NAMES["exit_time"]: trow.get("exit_time"),
            TRADES_DISPLAY_NAMES["exit_price"]: trow.get("exit_price"),
            TRADES_DISPLAY_NAMES["bars_held"]: trow.get("bars_held"),
            TRADES_DISPLAY_NAMES["net_pnl_pct"]: trow.get("net_pnl_pct"),
            TRADES_DISPLAY_NAMES["commission_pct"]: trow.get("commission_pct"),
            SIGNALS_DISPLAY_NAMES["signal_time"]: sig_time,
            SIGNALS_DISPLAY_NAMES["event_type"]: evt,
            SIGNALS_DISPLAY_NAMES["is_reversal"]: rev,
            SIGNALS_DISPLAY_NAMES["exec_price"]: ex_px,
            SIGNALS_DISPLAY_NAMES["signal_body_pct_median_ratio"]: body_m,
            SIGNALS_DISPLAY_NAMES["signal_range_pct_median_ratio"]: rng_m,
            SIGNALS_DISPLAY_NAMES["t1_return_pct"]: t1,
            SIGNALS_DISPLAY_NAMES["t2_return_pct"]: t2,
            SIGNALS_DISPLAY_NAMES["t3_return_pct"]: t3,
            "False Start Type": ftype,
        }
        if include_filter_block_reason:
            # filter_block_reason is an internal column in signals_df (WP-T6)
            fbr = matched.get("filter_block_reason") if matched is not None else np.nan
            row_dict["Filter Block Reason at Signal"] = fbr
        rows_out.append(row_dict)

    return pd.DataFrame(rows_out, columns=list(col_spec))


def _strip_tz_from_false_start_datetimes(df: pd.DataFrame) -> None:
    """In-place: remove tz from Entry/Exit/Signal Time for Excel."""
    for col in (
        TRADES_DISPLAY_NAMES["entry_time"],
        TRADES_DISPLAY_NAMES["exit_time"],
        SIGNALS_DISPLAY_NAMES["signal_time"],
    ):
        if col not in df.columns:
            continue
        s = df[col]
        if pd.api.types.is_datetime64_any_dtype(s) and s.dt.tz is not None:
            df[col] = s.dt.tz_localize(None)


def _build_false_start_summary_df(
    trades_df_raw: Optional[pd.DataFrame],
    false_start_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Summary metrics for the ``false start`` sheet (Metric / Value, 7 rows).
    """
    n_trades = len(trades_df_raw) if trades_df_raw is not None else 0
    n_fs = len(false_start_df)

    col_bars = TRADES_DISPLAY_NAMES["bars_held"]
    col_net = TRADES_DISPLAY_NAMES["net_pnl_pct"]
    col_rev = SIGNALS_DISPLAY_NAMES["is_reversal"]

    total = float(n_fs)
    ratio_all = np.nan if n_trades == 0 else n_fs / n_trades

    if n_fs == 0:
        med_bars = med_net = losing_share = bars1_share = np.nan
    else:
        med_bars = false_start_df[col_bars].median() if col_bars in false_start_df.columns else np.nan
        med_net = false_start_df[col_net].median() if col_net in false_start_df.columns else np.nan
        net_s = false_start_df[col_net] if col_net in false_start_df.columns else pd.Series(dtype=float)
        losing_share = float((net_s < 0).sum()) / n_fs
        bh = false_start_df[col_bars] if col_bars in false_start_df.columns else pd.Series(dtype=float)
        bars1_share = float((bh == 1).sum()) / n_fs

    if n_fs == 0 or col_rev not in false_start_df.columns:
        rev_share = np.nan
    else:
        rev_s = false_start_df[col_rev]
        valid_rev = rev_s.notna()
        n_valid = int(valid_rev.sum())
        if n_valid == 0:
            rev_share = np.nan
        else:
            rv = rev_s[valid_rev]
            is_rev = (rv == True) | (rv == 1)
            rev_share = float(is_rev.sum()) / n_valid

    metrics = [
        "Total false starts",
        "False starts / all Trades_100",
        "Median Bars Held",
        "Median Net PnL %",
        "Losing false starts share",
        "Bars Held = 1 share",
        "Reversal-signal share among false starts",
    ]
    values = [total, ratio_all, med_bars, med_net, losing_share, bars1_share, rev_share]
    return pd.DataFrame({"Metric": metrics, "Value": values})


def _write_false_start_sheet(
    writer: pd.ExcelWriter,
    trades_df_raw: Optional[pd.DataFrame],
    false_start_df: pd.DataFrame,
) -> None:
    """Write ``false start`` sheet: summary block, main table, autofilter, datetimes."""
    from openpyxl.utils import get_column_letter

    _strip_tz_from_false_start_datetimes(false_start_df)

    summary_df = _build_false_start_summary_df(trades_df_raw, false_start_df)
    main_startrow = len(summary_df) + 2

    summary_df.to_excel(writer, sheet_name=FALSE_START_SHEET_NAME, index=False)
    false_start_df.to_excel(
        writer, sheet_name=FALSE_START_SHEET_NAME, index=False, startrow=main_startrow
    )
    ws = writer.sheets[FALSE_START_SHEET_NAME]
    n_cols = len(false_start_df.columns)
    if n_cols > 0:
        header_excel_row = main_startrow + 1
        ws.auto_filter.ref = (
            f"A{header_excel_row}:{get_column_letter(n_cols)}{header_excel_row}"
        )

    if len(false_start_df) == 0:
        return

    for col_name in (
        TRADES_DISPLAY_NAMES["entry_time"],
        TRADES_DISPLAY_NAMES["exit_time"],
        SIGNALS_DISPLAY_NAMES["signal_time"],
    ):
        if col_name not in false_start_df.columns:
            continue
        series = false_start_df[col_name]
        if not pd.api.types.is_datetime64_any_dtype(series):
            continue
        drop = series.dropna()
        if len(drop) == 0:
            continue
        first_val = drop.iloc[0]
        if isinstance(first_val, pd.Timestamp) and (
            first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
        ):
            col_idx = list(false_start_df.columns).index(col_name) + 1
            first_data_excel_row = main_startrow + 2
            last_data_excel_row = main_startrow + 1 + len(false_start_df)
            for row_idx in range(first_data_excel_row, last_data_excel_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "YYYY-MM-DD HH:MM:SS"


# ---------------------------------------------------------------------------
# WP-T7 helpers for optional filter sheets
# ---------------------------------------------------------------------------

def _build_filter_summary_block_df(period_results: List[PeriodResult]) -> Optional[pd.DataFrame]:
    """Build the Filter Diagnostics Summary block for the Summary sheet (plan §9.2).

    Returns None if no period has filter_diagnostics_summary (disabled path).
    """
    rows = []
    for pr in period_results:
        s = pr.filter_diagnostics_summary
        if s is None:
            return None
        ctr = s.get("counters", {})
        bis = s.get("bars_in_state", {})
        rows.append({
            "Period":              pr.period_label,
            "ZigZag Mode":         s.get("zigzag_mode", s.get("mode", "")),
            "Candidate Duration Gate Enabled": s.get("candidate_duration_gate_enabled", ""),
            "Candidate Duration Max Bars": s.get("candidate_duration_max_bars", ""),
            "Bars OFF":            bis.get("OFF", 0),
            "Bars WAIT":           bis.get("WAIT_FIRST_ST_FLIP", 0),
            "Bars FREEZE":         bis.get("ST_ACTIVE_FREEZE", 0),
            "Bars MONITORING":     bis.get("ST_ACTIVE_MONITORING", 0),
            "Bars COUNTING ZZ":    bis.get("ST_COUNTING_ZZ_LEGS", 0),
            "Bars STOPPING":       bis.get("ST_STOPPING", 0),
            "Lifecycle Starts":    ctr.get("lifecycle_starts", s.get("lifecycle_starts_count", 0)),
            "Median Stop Events":  ctr.get("median_stop_triggered", s.get("median_stop_triggered_count", 0)),
            "ZZ Leg Stop Events":  ctr.get("zz_leg_stop_triggered", s.get("zz_leg_stop_triggered_count", 0)),
            "Raw ST Flips":        ctr.get("raw_st_flips", 0),
            "Entries Allowed":     ctr.get("passed_entry_signals", 0),
            "Entries Blocked":     ctr.get("blocked_entry_signals", 0),
            "Exits Opposite Flip": ctr.get("exits_opposite_flip", 0),
            "Immediate Entries Count": s.get(
                "immediate_entries_count",
                ctr.get("immediate_entries_count", 0),
            ),
            "Immediate Entries Blocked Count": s.get(
                "immediate_entries_blocked_count",
                ctr.get("immediate_entries_blocked_count", 0),
            ),
        })
    if not rows:
        return None
    return pd.DataFrame(rows)


def _write_filter_diagnostics_100_sheet(
    writer: pd.ExcelWriter,
    filter_diagnostics: Optional[Dict[str, np.ndarray]],
) -> None:
    """Write FilterDiagnostics_100 sheet (plan §9.2, gated export_state_columns=True)."""
    if filter_diagnostics is None:
        return
    n = len(next(iter(filter_diagnostics.values())))
    row_data: Dict[str, Any] = {"Bar Index": np.arange(n)}
    for key, arr in filter_diagnostics.items():
        display = FILTER_DIAGNOSTICS_100_DISPLAY_NAMES.get(key, key)
        row_data[display] = arr
    fd_df = pd.DataFrame(row_data)
    fd_df.to_excel(writer, sheet_name="FilterDiagnostics_100", index=False)


def _two_step_trade_trigger_link(
    trigger_bars: List[int],
    state_arr: Optional[np.ndarray],
    trades_df: Optional[pd.DataFrame],
) -> Dict[int, Any]:
    """Two-step trade <-> trigger linker (plan §9.5.1).

    Returns dict: trigger_bar → linked_trade_id (or "N/A").

    Algorithm (per-trade, then inverted to per-trigger):

    Step 1 — exact match: find trigger_bar == entry_signal_bar (decision bar).
      Covers 100% of same-bar trigger + flip cases.

    Step 2 — backward search: find the last trigger_bar < entry_signal_bar
      where no "OFF" state exists in state_arr[trigger_bar+1 : entry_signal_bar+1].
      "No OFF" means the lifecycle that started at trigger_bar is still active when
      the trade entered.  Covers trigger at t1, first allowed flip at t2 > t1 (freeze
      period holds FSM in WAIT while waiting for eligible flip).

    If neither step finds a match → "N/A" (unexpected for normal enabled runs;
    test failure per plan §9.5.1).
    """
    result: Dict[int, Any] = {t: "N/A" for t in trigger_bars}
    if not trigger_bars or trades_df is None or len(trades_df) == 0:
        return result
    if "entry_index" not in trades_df.columns or "trade_id" not in trades_df.columns:
        return result

    trigger_set = set(trigger_bars)

    # Build (entry_signal_bar, trade_id) sorted by entry_signal_bar ascending
    # (process earlier trades first so that the first trade per lifecycle gets the link)
    trade_pairs = []
    for _, trow in trades_df.iterrows():
        ei = trow.get("entry_index")
        tid = trow.get("trade_id")
        if pd.notna(ei):
            trade_pairs.append((max(int(ei) - 1, 0), tid))
    trade_pairs.sort(key=lambda x: x[0])

    for signal_bar, tid in trade_pairs:
        # Step 1: exact match — trigger at the same decision bar as the entry
        if signal_bar in trigger_set:
            if result.get(signal_bar) == "N/A":  # first trade wins for this trigger
                result[signal_bar] = tid
            continue

        # Step 2: backward search — latest trigger before signal_bar in the same lifecycle
        candidates = sorted(
            [t for t in trigger_bars if t < signal_bar], reverse=True
        )
        for t_trig in candidates:
            # Check that no OFF state exists between trigger_bar and entry_signal_bar
            if state_arr is not None:
                seg_start = t_trig + 1
                seg_end = signal_bar + 1  # inclusive of signal_bar
                if seg_start < seg_end:
                    segment = state_arr[seg_start:seg_end]
                    if any(str(s) == "OFF" for s in segment):
                        continue  # lifecycle ended between trigger and entry → wrong cycle
            if result.get(t_trig) == "N/A":
                result[t_trig] = tid
            break  # closest (most recent) eligible trigger wins; move to next trade

    return result


def _build_zigzag_trigger_events_df(
    filter_diagnostics: Optional[Dict[str, np.ndarray]],
    filter_diagnostics_summary: Optional[Dict[str, Any]] = None,
    df_index: Optional[Any] = None,
    trades_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Build ZigZag_Trigger_Events DataFrame (plan §9.2.1 reconstruction formula).

    Reconstruction: trigger_event_rows = [t for t in range(n) if trigger_source[t] != "none"]
    Per plan §9.2.1 audit-fix v0.5: trigger_source[t] != "none" is the exact marker;
    no heuristic via confirmed_legs_since_start delta.

    Triggered Lifecycle Start semantics (WP-V3-8):
    - Mode A/B/A+B: trigger bar state = WAIT; scan forward through WAIT until FREEZE.
    - Mode C / C+B immediate entry: trigger bar state is already FREEZE (OFF→FREEZE
      same bar); no WAIT transition — detected by checking state at t directly.
    - C+B B-rescue WAIT start: treated identically to A/B (state at t = WAIT, scan forward).

    Linked Trade ID: two-step linker per plan §9.5.1 (exact match, then backward search
    through same FSM lifecycle — no "OFF" between trigger_bar and entry_signal_bar).
    Absence of a link for an enabled-run trade row is a test failure (plan §9.5.1).
    """
    empty = pd.DataFrame(columns=list(_TRIGGER_EVENTS_COLUMNS))
    if filter_diagnostics is None:
        return empty

    trigger_source_arr = filter_diagnostics.get("trade_filter_trigger_source")
    if trigger_source_arr is None:
        return empty

    n = len(trigger_source_arr)
    state_arr = filter_diagnostics.get("trade_filter_state")
    threshold_arr = filter_diagnostics.get("candidate_trigger_threshold")
    global_median_arr = filter_diagnostics.get("global_median")
    local_median_n_arr = filter_diagnostics.get("local_median_N")
    candidate_height_arr = filter_diagnostics.get("candidate_height_pct")
    # WP-V3-8: new §11.2 arrays
    zigzag_mode_arr = filter_diagnostics.get("zigzag_mode")
    imm_used_arr = filter_diagnostics.get("immediate_candidate_entry_used")
    imm_reason_arr = filter_diagnostics.get("immediate_candidate_entry_block_reason")
    cand_age_arr = filter_diagnostics.get("candidate_age_bars")
    cand_dir_arr = filter_diagnostics.get("candidate_leg_direction")
    gate_passed_arr = filter_diagnostics.get("candidate_duration_gate_passed")

    quantile_used = None
    if filter_diagnostics_summary is not None:
        # Support both nested format (legacy donor) and flat format (step_executor)
        thr = filter_diagnostics_summary.get("thresholds", {})
        quantile_used = thr.get("candidate_trigger_quantile")

    # Collect trigger bars in bar order (plan §9.2.1)
    trigger_bars: List[int] = [
        t for t in range(n) if str(trigger_source_arr[t]) != "none"
    ]

    # Two-step Linked Trade ID (plan §9.5.1)
    linked_trade_map = _two_step_trade_trigger_link(trigger_bars, state_arr, trades_df)

    rows = []
    trigger_id = 1
    for t in trigger_bars:
        src = str(trigger_source_arr[t])

        # Triggered Lifecycle Start (WP-V3-8):
        # Mode C immediate entry: state at trigger bar is already ST_ACTIVE_FREEZE
        # (OFF→FREEZE same bar, no WAIT).  All other modes (A/B/A+B/C+B B-rescue)
        # enter WAIT first — scan forward from t+1 through WAIT until FREEZE or other.
        triggered_lc_start = False
        if state_arr is not None:
            state_at_t = str(state_arr[t])
            if state_at_t in ("ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING"):
                # Mode C same-bar immediate entry (already in lifecycle at bar t)
                triggered_lc_start = True
            else:
                # A/B/A+B/C+B-rescue: scan forward through WAIT until lifecycle
                # starts.  freeze_confirmed_legs=0 skips ST_ACTIVE_FREEZE and
                # lands directly in ST_ACTIVE_MONITORING, so both states count.
                for t2 in range(t + 1, min(t + 500, n)):
                    s2 = str(state_arr[t2])
                    if s2 == "WAIT_FIRST_ST_FLIP":
                        continue
                    triggered_lc_start = (s2 in ("ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING"))
                    break

        trigger_time = (
            _excel_safe_datetime_value(df_index[t])
            if df_index is not None
            else np.nan
        )
        threshold = float(threshold_arr[t]) if threshold_arr is not None else np.nan
        g_median = float(global_median_arr[t]) if global_median_arr is not None else np.nan
        l_median_n = float(local_median_n_arr[t]) if local_median_n_arr is not None else np.nan
        cand_ht = float(candidate_height_arr[t]) if candidate_height_arr is not None else np.nan

        rows.append({
            "Trigger ID":                trigger_id,
            "Trigger Bar":               t,
            "Trigger Time":              trigger_time,
            "Trigger Source":            src,
            "Threshold Used":            threshold,
            "Quantile Used":             quantile_used if quantile_used is not None else "N/A",
            "Global Median":             g_median,
            "Local Median N":            l_median_n,
            "Candidate Height %":        cand_ht,
            "Triggered Lifecycle Start": triggered_lc_start,
            "Linked Trade ID":           linked_trade_map.get(t, "N/A"),
            # WP-V3-8: new §11.2 columns
            "ZigZag Mode":                              str(zigzag_mode_arr[t]) if zigzag_mode_arr is not None else "",
            "Immediate Candidate Entry Used":           int(imm_used_arr[t]) if imm_used_arr is not None else 0,
            "Immediate Candidate Entry Block Reason":   str(imm_reason_arr[t]) if imm_reason_arr is not None else "",
            "Candidate Age Bars":                       int(cand_age_arr[t]) if cand_age_arr is not None else -1,
            "Candidate Leg Direction":                  int(cand_dir_arr[t]) if cand_dir_arr is not None else 0,
            "Candidate Duration Gate Passed":           int(gate_passed_arr[t]) if gate_passed_arr is not None else 0,
        })
        trigger_id += 1

    if rows:
        return pd.DataFrame(rows, columns=list(_TRIGGER_EVENTS_COLUMNS))
    return empty


def _write_zigzag_trigger_events_sheet(
    writer: pd.ExcelWriter,
    trigger_events_df: pd.DataFrame,
) -> None:
    """Write ZigZag_Trigger_Events sheet (plan §9.2, gated export_trigger_columns=True)."""
    trigger_events_df.to_excel(writer, sheet_name="ZigZag_Trigger_Events", index=False)


def _build_filters_summary_df(period_results: List[PeriodResult]) -> Optional[pd.DataFrame]:
    """Build the filters_summary sheet content (plan §9.2, extended WP-V3-8 §11.3).

    Returns (params_df, per_period_df) or None if disabled.
    Structure: two sections.
    Section (a): parameter table (ZigZag Mode, thresholds, gate config).
    Section (b): one row per period with counters + bars_in_state +
                 immediate entries counts (WP-V3-8).

    Reads from the flat dict produced by _compute_filter_diagnostics_summary
    (step_executor.py).  Legacy nested "counters"/"bars_in_state" sub-dicts
    are supported as a fallback for backward compatibility.
    """
    # Check if any result has summary
    for pr in period_results:
        if pr.filter_diagnostics_summary is not None:
            break
    else:
        return None  # all disabled

    # Get params from first period (they're the same for all periods)
    pr0 = next((pr for pr in period_results if pr.filter_diagnostics_summary is not None), None)
    if pr0 is None:
        return None
    s0 = pr0.filter_diagnostics_summary
    # Support both nested (legacy) and flat (step_executor) formats
    thr = s0.get("thresholds", {})

    params_rows = [
        # WP-V3-8: "ZigZag Mode" replaces legacy "Mode"; fall back for older summaries
        {"Parameter": "ZigZag Mode",                   "Value": s0.get("zigzag_mode", s0.get("mode", ""))},
        {"Parameter": "Reversal Threshold",            "Value": thr.get("reversal_threshold", "")},
        {"Parameter": "Candidate Trigger Threshold",   "Value": thr.get("candidate_trigger_threshold", "")},
        {"Parameter": "Candidate Trigger Quantile",    "Value": thr.get("candidate_trigger_quantile", "")},
        {"Parameter": "Candidate Trigger Source",      "Value": thr.get("candidate_trigger_source", "")},
        {"Parameter": "Global Median",                 "Value": thr.get("global_median", "")},
        {"Parameter": "Local Window",                  "Value": thr.get("local_window", "")},
        {"Parameter": "Freeze Confirmed Legs",         "Value": thr.get("freeze_confirmed_legs", "")},
        # WP-V3-8: gate params (§11.3)
        {"Parameter": "Candidate Duration Gate Enabled",  "Value": s0.get("candidate_duration_gate_enabled", thr.get("candidate_duration_gate_enabled", ""))},
        {"Parameter": "Candidate Duration Max Bars",      "Value": s0.get("candidate_duration_max_bars", thr.get("candidate_duration_max_bars", ""))},
        # exit-off modes (plan_exit_off_modes_v2.txt §8.2)
        {"Parameter": "Exit-OFF Mode",         "Value": thr.get("exit_off_mode", s0.get("exit_off_mode", ""))},
        {"Parameter": "Exit-OFF ZZ Leg Count", "Value": thr.get("exit_off_zz_leg_count", s0.get("exit_off_zz_leg_count", -1))},
    ]
    params_df = pd.DataFrame(params_rows)

    # Section (b): per-period aggregates
    period_rows = []
    for pr in period_results:
        s = pr.filter_diagnostics_summary
        if s is None:
            period_rows.append({"Period": pr.period_label})
            continue
        # Support both nested (legacy) and flat (step_executor) formats
        ctr = s.get("counters", {})
        bis = s.get("bars_in_state", {})
        period_rows.append({
            "Period":              pr.period_label,
            "Raw ST Flips":        ctr.get("raw_st_flips", 0),
            "Entries Allowed":     ctr.get("passed_entry_signals", 0),
            "Entries Blocked":     ctr.get("blocked_entry_signals", 0),
            "Blocked Filter Off":  ctr.get("blocked_filter_off", 0),
            "Blocked Waiting":     ctr.get("blocked_waiting_first", 0),
            "Blocked Trade Mode":  ctr.get("blocked_trade_mode", 0),
            "Blocked Local Med":   ctr.get("blocked_local_median", 0),
            "Blocked Invalid Stats": ctr.get("blocked_invalid_stats", 0),
            "Blocked Stopping":    ctr.get("blocked_stopping", 0),
            "Lifecycle Starts":    ctr.get("lifecycle_starts", s.get("lifecycle_starts_count", 0)),
            "Median Stops":        ctr.get("median_stop_triggered", s.get("median_stop_triggered_count", 0)),
            "ZZ Leg Stops":        ctr.get("zz_leg_stop_triggered", s.get("zz_leg_stop_triggered_count", 0)),
            "Exits Opp Flip":      ctr.get("exits_opposite_flip", 0),
            "Bars OFF":            bis.get("OFF", s.get("n_bars_in_off", 0)),
            "Bars WAIT":           bis.get("WAIT_FIRST_ST_FLIP", s.get("n_bars_in_wait_first_st_flip", 0)),
            "Bars FREEZE":         bis.get("ST_ACTIVE_FREEZE", s.get("n_bars_in_freeze", 0)),
            "Bars MONITORING":     bis.get("ST_ACTIVE_MONITORING", s.get("n_bars_in_monitoring", 0)),
            "Bars COUNTING ZZ":    bis.get("ST_COUNTING_ZZ_LEGS", s.get("n_bars_in_counting_zz_legs", 0)),
            "Bars STOPPING":       bis.get("ST_STOPPING", s.get("n_bars_in_stopping", 0)),
            # WP-V3-8: immediate entries (§11.3)
            "Immediate Entries Count":         s.get("immediate_entries_count", ctr.get("immediate_entries_count", 0)),
            "Immediate Entries Blocked Count": s.get("immediate_entries_blocked_count", ctr.get("immediate_entries_blocked_count", 0)),
        })
    period_df = pd.DataFrame(period_rows)
    return (params_df, period_df)  # type: ignore[return-value]


def _write_filters_summary_sheet(
    writer: pd.ExcelWriter,
    period_results: List[PeriodResult],
) -> None:
    """Write filters_summary sheet (plan §9.2, gated export_state_columns=True)."""
    result = _build_filters_summary_df(period_results)
    if result is None:
        return
    params_df, period_df = result  # type: ignore[misc]
    sheet = "filters_summary"
    params_df.to_excel(writer, sheet_name=sheet, index=False)
    period_start = len(params_df) + 2
    period_df.to_excel(writer, sheet_name=sheet, index=False, startrow=period_start)


# ---------------------------------------------------------------------------
# Sheet writers (updated for WP-T7)
# ---------------------------------------------------------------------------

def add_test_timestamp_to_filename(filepath: str, timestamp_format: str = "%Y%m%d_%H%M%S") -> str:
    """Add timestamp with "TEST" prefix to filename before extension."""
    path = Path(filepath)
    timestamp = datetime.now().strftime(timestamp_format)
    new_name = f"test_{path.stem}_{timestamp}{path.suffix}"
    return str(path.parent / new_name)


def _write_signals_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    signals_df: "pd.DataFrame | None",
) -> None:
    """Write the Signals sheet with autofilter and conditional formatting.

    WP-T7 fix: empty-result case preserves the incoming column structure.
    Disabled signals_df (19 cols) → 19 display cols; enabled (23 cols) →
    23 display cols (filter cols auto-renamed via SIGNALS_DISPLAY_NAMES).
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles import PatternFill

    has_data = signals_df is not None and len(signals_df) > 0

    if signals_df is not None:
        # Rename only columns present in the DataFrame (filter cols included when enabled)
        signals_df = signals_df.rename(columns=SIGNALS_DISPLAY_NAMES)
        if has_data:
            time_col = SIGNALS_DISPLAY_NAMES["signal_time"]
            if time_col in signals_df.columns:
                if pd.api.types.is_datetime64_any_dtype(signals_df[time_col]):
                    if signals_df[time_col].dt.tz is not None:
                        signals_df[time_col] = signals_df[time_col].dt.tz_localize(None)
    else:
        # Fallback: signals_df not provided (legacy path without WP-T6 wiring)
        signals_df = pd.DataFrame(columns=[
            v for k, v in SIGNALS_DISPLAY_NAMES.items()
            if k not in {"filter_state_at_signal", "filter_decision",
                         "filter_block_reason", "filter_trigger_source"}
        ])

    signals_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    n_cols = len(signals_df.columns)

    # Autofilter on header row (always, even for empty sheet)
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if has_data:
        n_rows = len(signals_df)
        cols = list(signals_df.columns)

        # Datetime formatting for Signal Time
        time_col = SIGNALS_DISPLAY_NAMES["signal_time"]
        if time_col in cols:
            first_vals = signals_df[time_col].dropna()
            if len(first_vals) > 0:
                first_val = first_vals.iloc[0]
                if isinstance(first_val, pd.Timestamp) and (
                    first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
                ):
                    col_idx = cols.index(time_col) + 1
                    for row_idx in range(2, n_rows + 2):
                        ws.cell(row=row_idx, column=col_idx).number_format = (
                            "YYYY-MM-DD HH:MM:SS"
                        )

        # Conditional formatting: green→yellow→red on Signal Body/Range %
        for internal_key in ("signal_body_pct", "signal_range_pct"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                ws.conditional_formatting.add(
                    cf_range,
                    ColorScaleRule(
                        start_type="min", start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="F8696B",
                    ),
                )

        _fill_ratio_strong = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
        _fill_ratio_soft   = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
        for internal_key in ("signal_body_pct_median_ratio", "signal_range_pct_median_ratio"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="greaterThan", formula=["2"],
                               fill=_fill_ratio_strong, stopIfTrue=True),
                )
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="greaterThan", formula=["1.6"],
                               fill=_fill_ratio_soft),
                )

        _fill_neg = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for internal_key in ("t1_return_pct", "t2_return_pct", "t3_return_pct"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            if display_name in cols:
                col_idx = cols.index(display_name) + 1
                col_letter = get_column_letter(col_idx)
                cf_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
                ws.conditional_formatting.add(
                    cf_range,
                    CellIsRule(operator="lessThan", formula=["0"], fill=_fill_neg),
                )


# Metrics used in the aggregate table (Summary block 2).
_AGGREGATE_METRIC_KEYS = [
    "sum_pnl_pct",
    "sharpe",
    "sortino",
    "max_drawdown",
    "cagr",
    "win_rate",
    "num_trades",
    "profit_factor",
    "avg_trade",
]


def _prepare_legacy_trades_df(
    trades_df: Optional[pd.DataFrame],
    filter_enabled: bool,
    filter_diagnostics_available: bool,
) -> pd.DataFrame:
    """Rename and prepare trades_df for the legacy export, handling filter columns.

    Disabled path → canonical columns only.
    Enabled path with data → canonical + filter columns renamed.
    Enabled path with 0 trades → canonical + filter column headers (plan §9.4).
    """
    if trades_df is not None and len(trades_df) > 0:
        df = trades_df.rename(columns=TRADES_DISPLAY_NAMES)
        # Strip timezone from datetime columns
        for col_key in ("entry_time", "exit_time"):
            col_name = TRADES_DISPLAY_NAMES[col_key]
            if col_name in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col_name]):
                    if df[col_name].dt.tz is not None:
                        df[col_name] = df[col_name].dt.tz_localize(None)
        return df
    else:
        # Empty trades — choose header set based on filter state (plan §9.4)
        if filter_enabled and filter_diagnostics_available:
            return pd.DataFrame(columns=list(TRADES_DISPLAY_NAMES.values()))
        else:
            return pd.DataFrame(columns=_CANONICAL_TRADE_DISPLAY_COLS)


def _format_trades_datetime(
    trades_df: pd.DataFrame,
    worksheet: Any,
) -> None:
    """Apply YYYY-MM-DD HH:MM:SS format to entry/exit time cells if they include time."""
    entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
    exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
    if entry_time_col not in trades_df.columns:
        return
    first_entry = trades_df[entry_time_col].dropna()
    if len(first_entry) == 0:
        return
    first_val = first_entry.iloc[0]
    if not (isinstance(first_val, pd.Timestamp) and (
        first_val.hour != 0 or first_val.minute != 0 or first_val.second != 0
    )):
        return
    cols = list(trades_df.columns)
    entry_col_idx = cols.index(entry_time_col) + 1
    exit_col_idx = cols.index(exit_time_col) + 1 if exit_time_col in cols else None
    for row_idx in range(2, len(trades_df) + 2):
        worksheet.cell(row=row_idx, column=entry_col_idx).number_format = "YYYY-MM-DD HH:MM:SS"
        if exit_col_idx is not None:
            worksheet.cell(row=row_idx, column=exit_col_idx).number_format = "YYYY-MM-DD HH:MM:SS"


# ---------------------------------------------------------------------------
# Main export functions
# ---------------------------------------------------------------------------

TESTER_CONFIG_SHEET_NAME = "Tester_Config"


def _format_tester_config_value(value: Any) -> str:
    """Render YAML/run metadata scalar to a stable Excel-friendly string."""
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        dt = value.replace(tzinfo=None) if value.tzinfo is not None else value
        return dt.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, pd.Timestamp):
        ts = value.tz_localize(None) if value.tzinfo is not None else value
        return ts.isoformat()
    if isinstance(value, (np.integer, np.floating)):
        return str(value.item())
    return str(value)


def _flatten_mapping_rows(
    payload: Any,
    *,
    section: str,
    prefix: str = "",
) -> List[Dict[str, str]]:
    """
    Flatten mapping/list payload into Section/Parameter/Value rows.

    Parameter path format:
      - mapping nesting: ``a.b.c``
      - list nesting: ``a.items[0]``
    """
    rows: List[Dict[str, str]] = []
    if isinstance(payload, Mapping):
        for key in payload:
            key_str = str(key)
            nested_prefix = f"{prefix}.{key_str}" if prefix else key_str
            rows.extend(
                _flatten_mapping_rows(payload[key], section=section, prefix=nested_prefix)
            )
        return rows

    if isinstance(payload, list):
        for idx, item in enumerate(payload):
            nested_prefix = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
            rows.extend(
                _flatten_mapping_rows(item, section=section, prefix=nested_prefix)
            )
        if not payload and prefix:
            rows.append({
                "Section": section,
                "Parameter": prefix,
                "Value": "[]",
            })
        return rows

    rows.append({
        "Section": section,
        "Parameter": prefix or "_root",
        "Value": _format_tester_config_value(payload),
    })
    return rows


def _write_tester_config_sheet(
    writer: pd.ExcelWriter,
    config_yaml_snapshot: Optional[Dict[str, Any]] = None,
    run_metadata: Optional[Dict[str, Any]] = None,
) -> None:
    """Write first workbook sheet with config snapshot + run metadata."""
    from openpyxl.utils import get_column_letter

    rows: List[Dict[str, str]] = []
    if config_yaml_snapshot is not None:
        if len(config_yaml_snapshot) == 0:
            rows.append({
                "Section": "config_file",
                "Parameter": "note",
                "Value": "YAML root mapping is empty (no keys)",
            })
        else:
            rows.extend(
                _flatten_mapping_rows(config_yaml_snapshot, section="config_file")
            )
    else:
        rows.append({
            "Section": "config_file",
            "Parameter": "note",
            "Value": "config file was not provided; built-in defaults were used",
        })

    if run_metadata:
        rows.extend(_flatten_mapping_rows(run_metadata, section="run"))

    cfg_df = pd.DataFrame(rows, columns=["Section", "Parameter", "Value"])
    if not cfg_df.empty:
        cfg_df = cfg_df.sort_values(["Section", "Parameter"], kind="mergesort")
    cfg_df.to_excel(writer, sheet_name=TESTER_CONFIG_SHEET_NAME, index=False)

    ws = writer.sheets[TESTER_CONFIG_SHEET_NAME]
    n_cols = len(cfg_df.columns)
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


def export_tester_results(
    period_results: List[PeriodResult],
    output_path: str,
    signals_df: Optional[pd.DataFrame] = None,
    false_start_max_bars: int = 4,
    # WP-T7: optional filter config for conditional sheets/columns
    trade_filter_config: Any = None,
    # WP-T7: optional original df for Trigger Time column in ZigZag_Trigger_Events
    df: Optional[pd.DataFrame] = None,
    # YAML snapshot as loaded from file before normalization (flattened to Tester_Config)
    config_yaml_snapshot: Optional[Dict[str, Any]] = None,
    # Extra runtime metadata (resolved values, paths, effective warmup, etc.)
    run_metadata: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Export tester results to Excel file.

    WP-T7 additions (plan §9.2):
    - Trades_<period>: includes Entry Filter State / Entry Trigger Source / Exit Reason
      when ``trade_filter_config.enabled=True`` and data is present.
    - Summary: adds Filter Diagnostics Summary block after standard metrics (enabled only).
    - Signals: filter columns (Filter State at Signal / Filter Decision / Filter Block
      Reason / Filter Trigger Source) automatically included when ``signals_df`` already
      carries them (from build_signal_events WP-T6).
    - Optional FilterDiagnostics_100 sheet (gated export_state_columns=True).
    - Optional ZigZag_Trigger_Events sheet (gated export_trigger_columns=True).
    - Optional filters_summary sheet (gated export_state_columns=True).
    - False start sheet: optional "Filter Block Reason at Signal" column (gated same).
    Disabled path is bit-identical to the pre-Phase-2 baseline.

    Args:
        period_results: List of PeriodResult for 100%, 75%, 50%, 33%, 25%.
        output_path: Path to output Excel file (timestamp will be added).
        signals_df: Optional DataFrame from ``build_signal_events()``.
        false_start_max_bars: False-start threshold (bars_held < N).
        trade_filter_config: Optional TradeFilterConfig. None → disabled path.
        df: Optional original OHLC DataFrame for Trigger Time in ZigZag_Trigger_Events.
        config_yaml_snapshot: Raw mapping from ``load_config`` (sheet ``Tester_Config`` / ``config_file``).
        run_metadata: Runtime fields (paths, ``resolved_periods_per_year``,
            ``warmup_period_resolved``, ``warmup_period_effective``, …). The exporter appends ``output_path_actual``.

    Returns:
        Actual output path used (with TEST timestamp).
    """
    output_path = add_test_timestamp_to_filename(output_path)
    run_metadata_payload: Dict[str, Any] = dict(run_metadata or {})
    run_metadata_payload["output_path_actual"] = output_path

    # Determine filter mode (plan §9.1)
    filter_enabled = (trade_filter_config is not None and trade_filter_config.enabled)
    export_state_cols = (
        filter_enabled
        and trade_filter_config.diagnostics is not None
        and trade_filter_config.diagnostics.export_state_columns
    )
    export_trigger_cols = (
        filter_enabled
        and trade_filter_config.diagnostics is not None
        and trade_filter_config.diagnostics.export_trigger_columns
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Tester_Config (always first) ──
        _write_tester_config_sheet(
            writer,
            config_yaml_snapshot=config_yaml_snapshot,
            run_metadata=run_metadata_payload,
        )

        # ── Summary sheet ──
        summary_data = []
        for pr in period_results:
            row = {
                "Period": pr.period_label,
                "ATR Period": pr.atr_period,
                "Multiplier": pr.multiplier,
                "Mode": pr.trade_mode,
                "Sum PnL %": pr.metrics.get("sum_pnl_pct", 0),
                "Sharpe": pr.metrics.get("sharpe", 0),
                "Sortino": pr.metrics.get("sortino", 0),
                "Max Drawdown": pr.metrics.get("max_drawdown", 0),
                "Win Rate": pr.metrics.get("win_rate", 0),
                "Num Trades": pr.metrics.get("num_trades", 0),
            }
            summary_data.append(row)

        summary_df = pd.DataFrame(summary_data)
        summary_df = format_excel_export_df(summary_df)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # WP-T7: Filter Diagnostics Summary block after standard metrics (plan §9.2, §9.4)
        if filter_enabled:
            filter_block_df = _build_filter_summary_block_df(period_results)
            if filter_block_df is not None:
                filter_start_row = len(summary_df) + 2  # +1 header + 1 blank row
                filter_block_df.to_excel(
                    writer, sheet_name="Summary", index=False, startrow=filter_start_row
                )

        # ── Metrics and Trades sheets for each period ──
        for pr in period_results:
            period_num = pr.period_label.replace("%", "")

            # Metrics sheet
            metrics_sheet = f"Metrics_{period_num}"[:31]
            metrics_row = {
                "Period": pr.period_label,
                "ATR Period": pr.atr_period,
                "Multiplier": pr.multiplier,
                "Mode": pr.trade_mode,
                "Commission": pr.commission,
                "Warmup (requested)": pr.warmup,
                "Warmup (effective)": pr.result.effective_warmup,
            }
            for key, display_name in METRICS_DISPLAY_NAMES.items():
                metrics_row[display_name] = pr.metrics.get(key, 0)
            metrics_df = format_excel_export_df(pd.DataFrame([metrics_row]))
            metrics_df.to_excel(writer, sheet_name=metrics_sheet, index=False)

            # Trades sheet (WP-T7: conditional filter columns)
            trades_sheet = f"Trades_{period_num}"[:31]
            trades_df = _prepare_legacy_trades_df(
                pr.trades_df,
                filter_enabled=filter_enabled,
                filter_diagnostics_available=(pr.filter_diagnostics is not None),
            )
            trades_df.to_excel(writer, sheet_name=trades_sheet, index=False)
            if len(trades_df) > 0:
                _format_trades_datetime(trades_df, writer.sheets[trades_sheet])

        # ── Signals sheet ──
        if signals_df is not None:
            _write_signals_sheet(writer, "Signals", signals_df)

        # ── False start sheet ──
        pr_100 = next((pr for pr in period_results if pr.period_label == "100%"), None)
        trades_100_raw = pr_100.trades_df if pr_100 is not None else None
        include_fbr = (export_state_cols and signals_df is not None
                       and "filter_block_reason" in (signals_df.columns if signals_df is not None else []))
        false_start_df = _build_false_start_sheet_df(
            trades_100_raw,
            signals_df,
            false_start_max_bars=false_start_max_bars,
            include_filter_block_reason=include_fbr,
        )
        _write_false_start_sheet(writer, trades_100_raw, false_start_df)

        # ── Optional filter sheets (gated on enabled + diagnostics flags) ──
        if filter_enabled and pr_100 is not None:
            fd_100 = pr_100.filter_diagnostics

            # FilterDiagnostics_100 (gated: export_state_columns=True)
            if export_state_cols and fd_100 is not None:
                _write_filter_diagnostics_100_sheet(writer, fd_100)

            # ZigZag_Trigger_Events (gated: export_trigger_columns=True)
            if export_trigger_cols and fd_100 is not None:
                df_index = df.index if df is not None else None
                trigger_df = _build_zigzag_trigger_events_df(
                    filter_diagnostics=fd_100,
                    filter_diagnostics_summary=pr_100.filter_diagnostics_summary,
                    df_index=df_index,
                    trades_df=trades_100_raw,
                )
                _write_zigzag_trigger_events_sheet(writer, trigger_df)

            # filters_summary (gated: export_state_columns=True)
            if export_state_cols:
                _write_filters_summary_sheet(writer, period_results)

            # cycle (independent of diagnostic export flags; requires enabled + diagnostics)
            if fd_100 is not None:
                cycle_df = _build_cycle_sheet_df(fd_100, df, trades_100_raw)
                _write_cycle_sheet(writer, cycle_df)

    return output_path


# ---------------------------------------------------------------------------
# Equal-blocks export (unchanged — equal_blocks enabled is rejected upstream)
# ---------------------------------------------------------------------------

def _add_eqblk_timestamp_to_filename(
    filepath: str, timestamp_format: str = "%Y%m%d_%H%M%S"
) -> str:
    """Add timestamp with 'test_' prefix and '_eqblk' suffix to filename."""
    path = Path(filepath)
    timestamp = datetime.now().strftime(timestamp_format)
    new_name = f"test_{path.stem}_eqblk_{timestamp}{path.suffix}"
    return str(path.parent / new_name)


def _write_trades_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    trades_df: "pd.DataFrame | None",
) -> None:
    """Write a single trades sheet (equal_blocks path; no filter columns)."""
    if trades_df is not None and len(trades_df) > 0:
        trades_df = trades_df.rename(columns=TRADES_DISPLAY_NAMES)
        entry_time_col = TRADES_DISPLAY_NAMES["entry_time"]
        exit_time_col = TRADES_DISPLAY_NAMES["exit_time"]
        for col in (entry_time_col, exit_time_col):
            if col in trades_df.columns:
                if pd.api.types.is_datetime64_any_dtype(trades_df[col]):
                    if trades_df[col].dt.tz is not None:
                        trades_df[col] = trades_df[col].dt.tz_localize(None)
    else:
        trades_df = pd.DataFrame(columns=_CANONICAL_TRADE_DISPLAY_COLS)

    trades_df.to_excel(writer, sheet_name=sheet_name, index=False)

    if len(trades_df) > 0:
        _format_trades_datetime(trades_df, writer.sheets[sheet_name])


def export_equal_blocks_results(
    segment_results: List[SegmentResult],
    output_path: str,
    config_yaml_snapshot: Optional[Dict[str, Any]] = None,
    run_metadata: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Export equal_blocks segmentation results to Excel.

    Phase 2: filter is not supported in equal_blocks (rejected upstream by
    run_equal_blocks / config gate). Sheet ``Tester_Config`` (first) carries
    the raw YAML snapshot and ``run_metadata`` when provided.
    """
    output_path = _add_eqblk_timestamp_to_filename(output_path)
    run_metadata_payload: Dict[str, Any] = dict(run_metadata or {})
    run_metadata_payload["output_path_actual"] = output_path

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Tester_Config (always first) ──
        _write_tester_config_sheet(
            writer,
            config_yaml_snapshot=config_yaml_snapshot,
            run_metadata=run_metadata_payload,
        )

        # ── Summary Table 1: per-segment rows ──
        summary_rows = []
        for seg in segment_results:
            m = seg.segment_metrics
            row = {
                "Segment": seg.segment_label,
                "Range": seg.range_label,
                "Bars": seg.n_bars,
                "ATR Period": seg.atr_period,
                "Multiplier": seg.multiplier,
                "Mode": seg.trade_mode,
                "Start Date": seg.start_date,
                "End Date": seg.end_date,
                "Prepend Bars": seg.prepend_bars,
                "Sum PnL %": m.get("sum_pnl_pct", 0),
                "Sharpe": m.get("sharpe", 0),
                "Sortino": m.get("sortino", 0),
                "Max Drawdown": m.get("max_drawdown", 0),
                "Win Rate": m.get("win_rate", 0),
                "Num Trades": m.get("num_trades", 0),
                "Profit Factor": m.get("profit_factor", 0),
                "Avg Trade": m.get("avg_trade", 0),
            }
            summary_rows.append(row)

        summary_df = pd.DataFrame(summary_rows)

        for date_col in ("Start Date", "End Date"):
            if date_col in summary_df.columns:
                col = summary_df[date_col]
                if pd.api.types.is_datetime64_any_dtype(col):
                    if col.dt.tz is not None:
                        summary_df[date_col] = col.dt.tz_localize(None)
                else:
                    def _strip_tz(v):
                        if isinstance(v, pd.Timestamp) and v.tzinfo is not None:
                            return v.tz_localize(None)
                        return v
                    summary_df[date_col] = summary_df[date_col].map(_strip_tz)

        # ── Summary Table 2: aggregate stats ──
        agg_rows = []
        for key in _AGGREGATE_METRIC_KEYS:
            display_name = METRICS_DISPLAY_NAMES.get(key, key)
            values = []
            for seg in segment_results:
                v = seg.segment_metrics.get(key)
                if v is not None and v != INVALID_METRIC_VALUE and np.isfinite(v):
                    values.append(float(v))

            if values:
                arr = np.array(values)
                agg_rows.append({
                    "Metric": display_name,
                    "Mean": float(np.mean(arr)),
                    "Std": float(np.std(arr, ddof=1)) if len(arr) > 1 else 0.0,
                    "Min": float(np.min(arr)),
                    "Max": float(np.max(arr)),
                    "Median": float(np.median(arr)),
                })
            else:
                agg_rows.append({
                    "Metric": display_name,
                    "Mean": None, "Std": None, "Min": None, "Max": None, "Median": None,
                })

        agg_df = pd.DataFrame(agg_rows)

        summary_df = format_excel_export_df(summary_df)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        start_row = len(summary_df) + 2
        agg_df = format_excel_export_df(agg_df)
        agg_df.to_excel(writer, sheet_name="Summary", index=False, startrow=start_row)

        # ── Per-segment sheets ──
        for seg in segment_results:
            label = seg.segment_label
            metrics_sheet = f"Metrics_{label}"[:31]
            metrics_row = {
                "Segment": seg.segment_label,
                "Range": seg.range_label,
                "ATR Period": seg.atr_period,
                "Multiplier": seg.multiplier,
                "Mode": seg.trade_mode,
                "Commission": seg.commission,
                "Prepend Bars": seg.prepend_bars,
            }
            for key, display_name in METRICS_DISPLAY_NAMES.items():
                metrics_row[display_name] = seg.segment_metrics.get(key, 0)
            pd.DataFrame([metrics_row]).pipe(format_excel_export_df).to_excel(
                writer, sheet_name=metrics_sheet, index=False
            )

            trades_sheet = f"Trades_{label}"[:31]
            _write_trades_sheet(writer, trades_sheet, seg.segment_trades_df)

    return output_path
