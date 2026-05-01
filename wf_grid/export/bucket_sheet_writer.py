"""
Excel sheet writer for BucketMatrix_Median.

Public API
----------
write_bucket_matrix_median_sheet(bucket_matrix_df, writer, atr_bucket_step, mult_bucket_step)
"""

from __future__ import annotations

from typing import Optional

import numpy as np
import pandas as pd

from wf_grid.bucket.heatmap_builder import build_risk_heatmap, build_stability_heatmap

# ---------------------------------------------------------------------------
# Local formatting (not format_excel_export_df — uses own frozensets)
# ---------------------------------------------------------------------------

_BUCKET_INTEGER_COLS = frozenset({
    "wins_count", "top3_count", "presence_count",
    "above_median_count", "eligible_median_steps_count",
    "bucket_size", "atr_bucket", "mult_bucket_ticks",
})

_BUCKET_SCORE_COLS = frozenset({
    "bucket_stability_score", "zone_dominance_score", "bucket_balanced_score",
})

_BUCKET_METRIC_COLS = frozenset({
    "mean_oos_pnl", "median_oos_pnl", "std_bucket", "pct_params_positive_pnl",
    "above_median_ratio", "above_median_ratio_present",
    "above_median_ratio_eligible",
    "max_drawdown_Median", "max_drawdown_Min",
})


def _safe_to_int64(s: pd.Series) -> pd.Series:
    """Convert a series to Int64 (nullable) safely."""
    try:
        return pd.to_numeric(s, errors="coerce").round().astype("Int64")
    except (TypeError, ValueError):
        return s


def _format_bucket_df(df: pd.DataFrame) -> pd.DataFrame:
    """Format bucket matrix DataFrame for Excel export.

    Own frozensets — does NOT call ``format_excel_export_df``.
    Scores and metrics → round(6), integer cols → Int64, Step cols → round(6).
    """
    df = df.copy()

    for col in df.columns:
        if col in _BUCKET_INTEGER_COLS:
            df[col] = _safe_to_int64(df[col])
        elif col in _BUCKET_SCORE_COLS:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(6)
        elif col in _BUCKET_METRIC_COLS:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(6)
        elif col.startswith("DD_"):
            df[col] = pd.to_numeric(df[col], errors="coerce").round(6)
        elif col.startswith("Step"):
            df[col] = pd.to_numeric(df[col], errors="coerce").round(6)

    return df


# ---------------------------------------------------------------------------
# Placeholder column headers
# ---------------------------------------------------------------------------

_PLACEHOLDER_COLS = [
    "bucket_param", "bucket_key", "atr_bucket", "mult_bucket_ticks",
    "bucket_size",
    "bucket_presence_steps", "mean_oos_pnl", "median_oos_pnl", "std_bucket",
    "pct_params_positive_pnl",
    "wins_count", "win_steps", "top3_count",
    "above_median_count", "above_median_ratio",
    "presence_count", "above_median_ratio_present",
    "eligible_median_steps_count", "above_median_ratio_eligible",
    "bucket_stability_score", "zone_dominance_score", "bucket_balanced_score",
    "max_drawdown_Median", "max_drawdown_Min",
]

_SHEET_NAME = "BucketMatrix_Median"

# ---------------------------------------------------------------------------
# Column widths (donor-compatible)
# ---------------------------------------------------------------------------

_FIXED_WIDTHS = {
    "bucket_param": 28,
    "bucket_key": 16,
    "atr_bucket": 12,
    "mult_bucket_ticks": 16,
    "bucket_size": 12,
    "bucket_presence_steps": 20,
    "mean_oos_pnl": 14,
    "median_oos_pnl": 16,
    "std_bucket": 14,
    "pct_params_positive_pnl": 22,
    "wins_count": 10,
    "win_steps": 16,
    "top3_count": 10,
    "above_median_count": 16,
    "above_median_ratio": 18,
    "presence_count": 14,
    "above_median_ratio_present": 22,
    "eligible_median_steps_count": 24,
    "above_median_ratio_eligible": 24,
    "bucket_stability_score": 22,
    "zone_dominance_score": 22,
    "bucket_balanced_score": 22,
    "max_drawdown_Median": 22,
    "max_drawdown_Min": 20,
}

_STEP_COL_WIDTH = 12

# Risk heatmap fixed color scale anchors (tied to gates threshold)
_RISK_HM_DD_WORST = -0.50
_RISK_HM_DD_MID = -0.15
_RISK_HM_DD_BEST = 0.0


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_bucket_matrix_median_sheet(
    bucket_matrix_df: Optional[pd.DataFrame],
    writer: pd.ExcelWriter,
    atr_bucket_step: int,
    mult_bucket_step: float,
) -> None:
    """Write BucketMatrix_Median sheet to the Excel workbook.

    Truth-table (§3.7):
        None or empty df → placeholder (headers + freeze panes, 0 data rows)
        Full-grid with NaN/0 → normal sheet
        Normal df → normal sheet

    Parameters
    ----------
    bucket_matrix_df:
        Output of ``build_median_bucket_matrix``, or ``None``.
    writer:
        Active ``pd.ExcelWriter`` (openpyxl engine).
    atr_bucket_step:
        ATR bucket width for heatmap label generation.
    mult_bucket_step:
        Multiplier bucket width for heatmap label generation.
    """
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    # ── Placeholder mode ─────────────────────────────────────────────────
    if bucket_matrix_df is None or len(bucket_matrix_df) == 0:
        pd.DataFrame(columns=_PLACEHOLDER_COLS).to_excel(
            writer, sheet_name=_SHEET_NAME, index=False, startrow=0,
        )
        ws = writer.sheets[_SHEET_NAME]
        ws.freeze_panes = ws.cell(row=2, column=5)
        n_ph = len(_PLACEHOLDER_COLS)
        ws.auto_filter.ref = f"A1:{get_column_letter(n_ph)}1"
        return

    # ── Normal mode ──────────────────────────────────────────────────────
    display_df = _format_bucket_df(bucket_matrix_df)

    start_main = 0
    display_df.to_excel(
        writer, sheet_name=_SHEET_NAME, index=False, startrow=start_main,
    )

    # ── Heatmap section ──────────────────────────────────────────────────
    heatmap_df = build_stability_heatmap(
        bucket_matrix_df, atr_bucket_step, mult_bucket_step,
    )
    start_title_hm = start_main + len(display_df) + 2
    start_hm = start_title_hm + 2

    if not heatmap_df.empty:
        title_hm_df = pd.DataFrame(
            [{"ATR \\ MULT": "ZONE STRENGTH HEATMAP (MEDIAN-BASED, ATR \u00d7 MULT BUCKET)"}]
        )
        title_hm_df.to_excel(
            writer, sheet_name=_SHEET_NAME, index=False, header=False,
            startrow=start_title_hm,
        )
        heatmap_df.to_excel(
            writer, sheet_name=_SHEET_NAME, index=True, startrow=start_hm,
        )

    ws = writer.sheets[_SHEET_NAME]

    # ── Freeze panes: E2 ────────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=2, column=5)

    # ── Autofilter: header-only (3.0 pattern) ────────────────────────────
    n_cols = len(display_df.columns)
    if n_cols > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    # ── Column widths ────────────────────────────────────────────────────
    n_data_rows = len(display_df)
    for ci, col_name in enumerate(display_df.columns, start=1):
        col_letter = get_column_letter(ci)
        if col_name in _FIXED_WIDTHS:
            width = _FIXED_WIDTHS[col_name]
        elif col_name.startswith("DD_"):
            width = _STEP_COL_WIDTH
        elif col_name.startswith("Step"):
            width = _STEP_COL_WIDTH
        else:
            max_len = len(str(col_name)) + 2
            for ri in range(start_main + 2, start_main + 2 + n_data_rows):
                v = ws.cell(row=ri, column=ci).value
                if v is not None:
                    max_len = max(max_len, len(str(v)) + 2)
            width = min(max_len, 40)
        ws.column_dimensions[col_letter].width = width

    # ── Number formats ───────────────────────────────────────────────────
    _dd_step_cols_set = {c for c in display_df.columns if c.startswith("DD_")}
    _dd_summary_cols_set = {"max_drawdown_Median", "max_drawdown_Min"}
    _step_cols_set = {c for c in display_df.columns if c.startswith("Step")}
    _score_cols_set = {
        "above_median_ratio", "bucket_stability_score",
        "zone_dominance_score", "bucket_balanced_score",
        "pct_params_positive_pnl",
        "above_median_ratio_present", "above_median_ratio_eligible",
    }
    _pnl_cols_set = {"mean_oos_pnl", "median_oos_pnl", "std_bucket"}
    _int_cols_set = {
        "atr_bucket", "mult_bucket_ticks", "bucket_size",
        "wins_count", "top3_count", "above_median_count",
        "presence_count", "eligible_median_steps_count",
    }

    data_row_start = start_main + 2
    data_row_end = start_main + 1 + n_data_rows

    for ci, col_name in enumerate(display_df.columns, start=1):
        if col_name in _dd_step_cols_set or col_name in _dd_summary_cols_set:
            fmt = "0.00%"
        elif col_name in _step_cols_set:
            fmt = "0.000000"
        elif col_name in _score_cols_set:
            fmt = "0.000000"
        elif col_name in _pnl_cols_set:
            fmt = "0.000000"
        elif col_name in _int_cols_set:
            fmt = "0"
        else:
            continue
        for ri in range(data_row_start, data_row_end + 1):
            cell = ws.cell(row=ri, column=ci)
            if cell.value is not None:
                cell.number_format = fmt

    # ── Heatmap title styling ────────────────────────────────────────────
    if not heatmap_df.empty:
        title_excel_row = start_title_hm + 1
        title_cell = ws.cell(row=title_excel_row, column=1)
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="left")

    # ── Heatmap conditional formatting (white → yellow → green) ──────────
    if not heatmap_df.empty:
        n_hm_rows = len(heatmap_df)
        n_hm_cols = len(heatmap_df.columns)
        hm_data_row_start = start_hm + 2
        hm_data_row_end = start_hm + 1 + n_hm_rows
        hm_col_start = 2
        hm_col_end = 1 + n_hm_cols

        hm_range = (
            f"{get_column_letter(hm_col_start)}{hm_data_row_start}:"
            f"{get_column_letter(hm_col_end)}{hm_data_row_end}"
        )
        ws.conditional_formatting.add(
            hm_range,
            ColorScaleRule(
                start_type="min", start_color="FFFFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF99",
                end_type="max", end_color="FF00AA00",
            ),
        )
        for ri in range(start_hm + 1, start_hm + 2 + n_hm_rows):
            for ci in range(1, hm_col_end + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(
                    horizontal="center"
                )
        for ci in range(hm_col_start, hm_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11

    # ── Risk heatmap section ──────────────────────────────────────────────
    risk_heatmap_df = build_risk_heatmap(
        bucket_matrix_df, atr_bucket_step, mult_bucket_step,
    )

    if not risk_heatmap_df.empty:
        if not heatmap_df.empty:
            start_title_hm2 = start_hm + len(heatmap_df) + 3
        else:
            start_title_hm2 = start_main + len(display_df) + 2
        start_hm2 = start_title_hm2 + 2

        title_hm2_df = pd.DataFrame(
            [{"ATR \\ MULT": "RISK HEATMAP: MEDIAN MAX DRAWDOWN (ATR \u00d7 MULT BUCKET)"}]
        )
        title_hm2_df.to_excel(
            writer, sheet_name=_SHEET_NAME, index=False, header=False,
            startrow=start_title_hm2,
        )
        risk_heatmap_df.to_excel(
            writer, sheet_name=_SHEET_NAME, index=True, startrow=start_hm2,
        )

        # ── Risk heatmap title styling ────────────────────────────────
        title2_excel_row = start_title_hm2 + 1
        title2_cell = ws.cell(row=title2_excel_row, column=1)
        title2_cell.font = Font(bold=True, size=12)
        title2_cell.alignment = Alignment(horizontal="left")

        # ── Risk heatmap conditional formatting (fixed absolute scale) ─
        n_hm2_rows = len(risk_heatmap_df)
        n_hm2_cols = len(risk_heatmap_df.columns)
        hm2_data_row_start = start_hm2 + 2
        hm2_data_row_end = start_hm2 + 1 + n_hm2_rows
        hm2_col_start = 2
        hm2_col_end = 1 + n_hm2_cols

        hm2_range = (
            f"{get_column_letter(hm2_col_start)}{hm2_data_row_start}:"
            f"{get_column_letter(hm2_col_end)}{hm2_data_row_end}"
        )
        ws.conditional_formatting.add(
            hm2_range,
            ColorScaleRule(
                start_type="num", start_value=_RISK_HM_DD_WORST,
                start_color="FFFF0000",
                mid_type="num", mid_value=_RISK_HM_DD_MID,
                mid_color="FFFFFF99",
                end_type="num", end_value=_RISK_HM_DD_BEST,
                end_color="FF00AA00",
            ),
        )

        # ── Risk heatmap alignment + column widths ────────────────────
        for ri in range(start_hm2 + 1, start_hm2 + 2 + n_hm2_rows):
            for ci in range(1, hm2_col_end + 1):
                ws.cell(row=ri, column=ci).alignment = Alignment(
                    horizontal="center"
                )
        for ci in range(hm2_col_start, hm2_col_end + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11
