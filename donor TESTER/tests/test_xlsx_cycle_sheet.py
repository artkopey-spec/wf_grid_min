from __future__ import annotations

import inspect
import math
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest

import supertrend_optimizer.core.zigzag_st_filter as zzmod
from supertrend_optimizer.core.zigzag_st_filter import (
    ConfirmedLeg,
    compute_confirmed_legs_reset_aware,
    compute_zigzag_per_bar,
)


def _expected_cycle_columns() -> list[str]:
    return [
        "Начало цикла",
        "Конец цикла",
        "Направление цикла",
        "Баров в цикле",
        "Ног ZigZag в цикле",
        "Медиана ног",
        "Ног выше порога триггера кандидата",
        "Размер цикла, %",
        "ID цикла",
        "Start bar index",
        "End bar index",
        "Цена начала",
        "Цена конца",
        "High цикла",
        "Low цикла",
        "Макс. движение по циклу, %",
        "Макс. просадка внутри цикла, %",
        "Причина завершения",
        "Макс. высота ноги",
        "Доля ног выше порога, %",
        "Сделок в цикле",
        "Фин результат цикла, %",
        "% сделок с положительным фин результатом в цикле",
    ]


def _excel_tester_module():
    import supertrend_optimizer.io.excel_tester as excel_tester

    required = ["CYCLE_SHEET_NAME", "CYCLE_SHEET_COLUMNS", "_build_cycle_sheet_df"]
    missing = [name for name in required if not hasattr(excel_tester, name)]
    assert missing == [], f"cycle sheet API missing from excel_tester: {missing}"
    return excel_tester


def _df(n: int) -> pd.DataFrame:
    close = np.array([100.0 + i for i in range(n)], dtype=np.float64)
    idx = pd.date_range("2026-01-01", periods=n, freq="h")
    return pd.DataFrame(
        {
            "close": close,
            "high": close + 2.0,
            "low": close - 2.0,
        },
        index=idx,
    )


def _diagnostics(
    states: list[str],
    *,
    trigger_sources: list[str] | None = None,
    candidate_dirs: list[int] | None = None,
    candidate_threshold=0.03,
    reversal_threshold=0.02,
    local_window=1,
    daily_reset_event: list[int] | np.ndarray | None = None,
) -> dict[str, np.ndarray]:
    n = len(states)
    if trigger_sources is None:
        trigger_sources = ["none"] * n
    if candidate_dirs is None:
        candidate_dirs = [0] * n
    if daily_reset_event is None:
        daily_reset_event = np.zeros(n, dtype=np.int8)

    return {
        "trade_filter_state": np.array(states, dtype=object),
        "trade_filter_trigger_source": np.array(trigger_sources, dtype=object),
        "candidate_trigger_threshold": np.full(n, candidate_threshold, dtype=np.float64)
        if np.isscalar(candidate_threshold)
        else np.asarray(candidate_threshold),
        "zigzag_reversal_threshold": np.full(n, reversal_threshold, dtype=np.float64)
        if np.isscalar(reversal_threshold)
        else np.asarray(reversal_threshold),
        "local_window": np.full(n, local_window, dtype=np.float64)
        if np.isscalar(local_window)
        else np.asarray(local_window),
        "daily_reset_event": np.asarray(daily_reset_event, dtype=np.int8),
        "candidate_leg_direction": np.asarray(candidate_dirs, dtype=np.int8),
        "st_flip_dir": np.full(n, -1, dtype=np.int8),
    }


def _volume_diagnostics(
    states: list[str],
    *,
    daily_reset_event: list[int] | np.ndarray | None = None,
    time_filter_reset_event: list[int] | np.ndarray | None = None,
    volume_regime: list[object] | np.ndarray | None = None,
    median_relative_volume: list[float] | np.ndarray | None = None,
) -> dict[str, np.ndarray]:
    diag: dict[str, np.ndarray] = {
        "trade_filter_state": np.asarray(states, dtype=object),
    }
    if daily_reset_event is not None:
        diag["daily_reset_event"] = np.asarray(daily_reset_event)
    if time_filter_reset_event is not None:
        diag["time_filter_reset_event"] = np.asarray(time_filter_reset_event)
    if volume_regime is not None:
        diag["volume_regime"] = np.asarray(volume_regime, dtype=object)
    if median_relative_volume is not None:
        diag["median_relative_volume"] = np.asarray(median_relative_volume, dtype=np.float64)
    return diag


def _patch_no_legs(monkeypatch, excel_tester):
    n_holder: dict[str, int] = {}

    def fake_per_bar(close, reversal_threshold, local_window, *, daily_reset_event):
        n_holder["n"] = len(close)
        return SimpleNamespace(
            confirm_event=np.zeros(len(close), dtype=np.int8),
            last_confirmed_leg_height_pct=np.full(len(close), np.nan),
        )

    monkeypatch.setattr(excel_tester, "compute_zigzag_per_bar", fake_per_bar, raising=False)
    monkeypatch.setattr(
        excel_tester,
        "compute_confirmed_legs_reset_aware",
        lambda close, reversal_threshold, *, daily_reset_event: [],
        raising=False,
    )
    return n_holder


def _patch_leg_outputs(monkeypatch, excel_tester, *, confirm_bar: int, height: float, direction: int = 1):
    def fake_per_bar(close, reversal_threshold, local_window, *, daily_reset_event):
        confirm_event = np.zeros(len(close), dtype=np.int8)
        heights = np.full(len(close), np.nan)
        confirm_event[confirm_bar] = 1
        heights[confirm_bar] = height
        return SimpleNamespace(
            confirm_event=confirm_event,
            last_confirmed_leg_height_pct=heights,
        )

    monkeypatch.setattr(excel_tester, "compute_zigzag_per_bar", fake_per_bar, raising=False)
    monkeypatch.setattr(
        excel_tester,
        "compute_confirmed_legs_reset_aware",
        lambda close, reversal_threshold, *, daily_reset_event: [
            ConfirmedLeg(0, confirm_bar - 1, confirm_bar, 100.0, 103.0, direction, height)
        ],
        raising=False,
    )


def _period_result(filter_diagnostics, trades_df=None):
    from supertrend_optimizer.engine.result import BacktestResult
    from supertrend_optimizer.testing.runner import PeriodResult

    n = len(next(iter(filter_diagnostics.values()))) if filter_diagnostics else 4
    result = BacktestResult(
        atr_period=14,
        multiplier=3.0,
        trade_mode="revers",
        commission=0.001,
        warmup=0,
        returns=np.zeros(n - 1, dtype=np.float64),
        equity_curve=np.ones(n, dtype=np.float64),
        positions=np.zeros(n, dtype=np.int8),
        trend=np.zeros(n, dtype=np.int8),
        metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
        early_exit=False,
        exit_bar=None,
        exit_drawdown=None,
        trades_df=trades_df if trades_df is not None else pd.DataFrame(),
        n_bars_original=n,
        period_label="100%",
        filter_diagnostics=filter_diagnostics,
    )
    return PeriodResult(
        period_label="100%",
        n_bars=n,
        result=result,
        filter_diagnostics=filter_diagnostics,
        filter_diagnostics_summary=None,
    )


def _enabled_cfg(*, export_state_columns=True, export_trigger_columns=True):
    return SimpleNamespace(
        enabled=True,
        zigzag=SimpleNamespace(enabled=True),
        diagnostics=SimpleNamespace(
            export_state_columns=export_state_columns,
            export_trigger_columns=export_trigger_columns,
        ),
    )


def test_compute_confirmed_legs_reset_aware_public_contract():
    assert "compute_confirmed_legs_reset_aware" in zzmod.__all__

    sig = inspect.signature(compute_confirmed_legs_reset_aware)
    assert list(sig.parameters.keys()) == [
        "close",
        "reversal_threshold",
        "daily_reset_event",
    ]
    assert sig.parameters["daily_reset_event"].kind is inspect.Parameter.KEYWORD_ONLY


def test_reset_aware_confirmed_legs_match_per_bar_confirm_events_and_heights():
    close = np.array(
        [100.0, 105.0, 100.0, 106.0, 100.0, 107.0, 100.0, 108.0, 100.0],
        dtype=np.float64,
    )
    daily_reset_event = np.array([0, 0, 0, 0, 0, 1, 0, 0, 0], dtype=np.int8)

    legs = compute_confirmed_legs_reset_aware(
        close,
        0.01,
        daily_reset_event=daily_reset_event,
    )
    per_bar = compute_zigzag_per_bar(
        close,
        0.01,
        3,
        daily_reset_event=daily_reset_event,
    )

    confirm_bars = [leg.confirm_bar for leg in legs]
    np.testing.assert_array_equal(
        np.flatnonzero(per_bar.confirm_event == 1),
        np.array(confirm_bars, dtype=np.int64),
    )
    np.testing.assert_allclose(
        per_bar.last_confirmed_leg_height_pct[confirm_bars],
        np.array([leg.height_pct for leg in legs], dtype=np.float64),
        rtol=0,
        atol=1e-15,
    )


def test_cycle_sheet_constants_contract():
    import supertrend_optimizer.io.excel_tester as excel_tester

    assert excel_tester.CYCLE_SHEET_NAME == "cycle"
    assert list(excel_tester.CYCLE_SHEET_COLUMNS) == _expected_cycle_columns()
    assert len(excel_tester.CYCLE_SHEET_COLUMNS) == 23


def test_cycle_float_scalar_materialize_contract():
    import supertrend_optimizer.io.excel_tester as excel_tester

    assert excel_tester._materialize_cycle_candidate_threshold(
        np.array([np.nan, 0.03, 0.03, np.nan]), 4
    ) == 0.03
    assert excel_tester._materialize_cycle_candidate_threshold(
        np.array([0.03, 0.04, 0.03]), 3
    ) is None
    assert excel_tester._materialize_cycle_candidate_threshold(
        np.array([np.nan, np.inf]), 2
    ) is None
    assert excel_tester._materialize_cycle_candidate_threshold(
        np.array([0.03, 0.04]), 1
    ) == 0.03


def test_cycle_reversal_threshold_materialize_contract():
    import supertrend_optimizer.io.excel_tester as excel_tester

    assert excel_tester._materialize_cycle_reversal_threshold(
        np.array([0.02, 0.02]), 2
    ) == 0.02
    for values in (
        np.array([np.nan, np.nan]),
        np.array([0.0, 0.0]),
        np.array([1.0, 1.0]),
        np.array([0.02, 0.03]),
    ):
        assert excel_tester._materialize_cycle_reversal_threshold(values, len(values)) is None


@pytest.mark.parametrize(
    "values,n,expected",
    [
        ([20, 20, 20], 3, 20),
        ([np.nan, 20.0, 20.0], 3, 20),
        ([0, 20, 20], 3, None),
        ([20.5, 20.5], 2, None),
        ([20, 21, 20], 3, None),
        ([0, 0, 0], 3, None),
        ([True, True], 2, None),
        ([20, 21], 1, 20),
    ],
)
def test_cycle_local_window_materialize_contract(values, n, expected):
    import supertrend_optimizer.io.excel_tester as excel_tester

    assert excel_tester._materialize_cycle_local_window(np.array(values), n) == expected


@pytest.mark.parametrize(
    "source,candidate_dir,legs,expected",
    [
        ("candidate_threshold", 1, [], "+"),
        ("candidate_threshold", -1, [], "-"),
        ("candidate_threshold", 0, [], ""),
        ("both", -1, [ConfirmedLeg(0, 1, 2, 100.0, 101.0, 1, 0.01)], "-"),
        ("confirmed_median", 1, [ConfirmedLeg(0, 1, 2, 100.0, 101.0, -1, 0.01)], "-"),
        ("confirmed_median", 1, [ConfirmedLeg(0, 1, 3, 100.0, 101.0, 1, 0.01)], ""),
        ("none", 1, [ConfirmedLeg(0, 1, 2, 100.0, 101.0, 1, 0.01)], ""),
        ("unknown", -1, [ConfirmedLeg(0, 1, 2, 100.0, 101.0, -1, 0.01)], ""),
    ],
)
def test_cycle_direction_resolver_contract(source, candidate_dir, legs, expected):
    import supertrend_optimizer.io.excel_tester as excel_tester

    assert excel_tester._resolve_cycle_direction(source, candidate_dir, 2, legs) == expected


def test_headers_contract_and_empty_builder_output(monkeypatch):
    excel_tester = _excel_tester_module()
    assert excel_tester.CYCLE_SHEET_NAME == "cycle"
    assert list(excel_tester.CYCLE_SHEET_COLUMNS) == _expected_cycle_columns()
    assert len(excel_tester.CYCLE_SHEET_COLUMNS) == 23

    result = excel_tester._build_cycle_sheet_df({}, _df(3))
    assert list(result.columns) == _expected_cycle_columns()
    assert result.empty


def test_completed_cycles_skip_unfinished_tail(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    states = ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "OFF", "WAIT_FIRST_ST_FLIP"]
    diag = _diagnostics(
        states,
        trigger_sources=["none", "candidate_threshold", "none", "none", "candidate_threshold"],
        candidate_dirs=[0, 1, 0, 0, -1],
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(len(states)))

    assert len(result) == 1
    row = result.iloc[0]
    assert row["ID цикла"] == 1
    assert row["Start bar index"] == 1
    assert row["End bar index"] == 2
    assert row["Баров в цикле"] == 2


def test_post_close_fsm_boundary_uses_off_bar_only_for_reason(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_leg_outputs(monkeypatch, excel_tester, confirm_bar=6, height=0.05)
    states = ["OFF", "OFF"] + ["ST_ACTIVE_MONITORING"] * 5 + ["OFF"]
    diag = _diagnostics(
        states,
        trigger_sources=["none", "none", "candidate_threshold"] + ["none"] * 5,
        candidate_dirs=[0, 0, 1, 0, 0, 0, 0, 0],
        daily_reset_event=[0, 0, 0, 0, 0, 0, 0, 1],
    )
    trades = pd.DataFrame(
        {
            "entry_index": [3, 7, 8, np.nan],
            "net_pnl_pct": [1.0, -1.0, 100.0, 5.0],
        }
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(len(states)), trades)

    row = result.iloc[0]
    assert row["Start bar index"] == 2
    assert row["End bar index"] == 6
    assert row["Причина завершения"] == "daily_reset"
    assert row["Цена начала"] == 102.0
    assert row["Цена конца"] == 106.0
    assert row["High цикла"] == 108.0
    assert row["Low цикла"] == 100.0
    assert row["Ног ZigZag в цикле"] == 1
    assert row["Сделок в цикле"] == 2


@pytest.mark.parametrize(
    "mutation",
    [
        lambda d: d.pop("daily_reset_event"),
        lambda d: d.update({"zigzag_reversal_threshold": np.full(4, np.nan)}),
        lambda d: d.update({"zigzag_reversal_threshold": np.full(4, 1.0)}),
        lambda d: d.update({"local_window": np.array([0, 20, 20, 20], dtype=float)}),
        lambda d: d.update({"trade_filter_state": np.array([], dtype=object)}),
    ],
)
def test_scalar_guards_return_empty_and_do_not_call_zigzag_helpers(monkeypatch, mutation):
    excel_tester = _excel_tester_module()
    diag = _diagnostics(["OFF", "ST_ACTIVE_FREEZE", "OFF", "OFF"])
    mutation(diag)

    def fail(*args, **kwargs):
        raise AssertionError("ZigZag helper must not be called for invalid cycle inputs")

    monkeypatch.setattr(excel_tester, "compute_zigzag_per_bar", fail, raising=False)
    monkeypatch.setattr(excel_tester, "compute_confirmed_legs_reset_aware", fail, raising=False)

    result = excel_tester._build_cycle_sheet_df(diag, _df(4))
    assert list(result.columns) == _expected_cycle_columns()
    assert result.empty


@pytest.mark.parametrize(
    "local_window, expected_rows",
    [
        ([20, 20, 20, 20], 1),
        ([0, 20, 20, 20], 0),
        ([20.5, 20.5, 20.5, 20.5], 0),
        ([20, 21, 20, 20], 0),
    ],
)
def test_local_window_materialize_contract(monkeypatch, local_window, expected_rows):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none"],
        candidate_dirs=[0, 1, 0, 0],
        local_window=local_window,
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(4))
    assert len(result) == expected_rows


@pytest.mark.parametrize(
    "source,candidate_dir,leg_dir,expected",
    [
        ("candidate_threshold", 1, -1, "+"),
        ("both", -1, 1, "-"),
        ("confirmed_median", 1, -1, "-"),
        ("none", 1, -1, ""),
        ("unexpected", -1, 1, ""),
    ],
)
def test_direction_matrix_ignores_st_flip_fallback(monkeypatch, source, candidate_dir, leg_dir, expected):
    excel_tester = _excel_tester_module()
    _patch_leg_outputs(monkeypatch, excel_tester, confirm_bar=1, height=0.05, direction=leg_dir)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", source, "none"],
        candidate_dirs=[0, candidate_dir, 0],
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(3))

    assert result.iloc[0]["Направление цикла"] == expected


def test_trades_mapping_uses_entry_signal_idx_and_skips_nan(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none", "none"],
        candidate_dirs=[0, 1, 0, 0, 0],
    )
    trades = pd.DataFrame({"entry_index": [1, 2, 4, 5, np.nan]})

    result = excel_tester._build_cycle_sheet_df(diag, _df(5), trades)

    assert result.iloc[0]["Сделок в цикле"] == 2


def test_positive_trades_percent_scale_and_non_finite_guard(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none", "none"],
        candidate_dirs=[0, 1, 0, 0, 0],
    )
    trades = pd.DataFrame({"entry_index": [2, 3, 4], "net_pnl_pct": [2.5, 0.0, -1.0]})
    result = excel_tester._build_cycle_sheet_df(diag, _df(5), trades)
    assert result.iloc[0]["Фин результат цикла, %"] == pytest.approx(1.5)
    assert result.iloc[0]["% сделок с положительным фин результатом в цикле"] == pytest.approx(100.0 / 3.0)

    bad_trades = pd.DataFrame({"entry_index": [2, 3], "net_pnl_pct": [1.0, math.inf]})
    bad = excel_tester._build_cycle_sheet_df(diag, _df(5), bad_trades)
    assert pd.isna(bad.iloc[0]["Фин результат цикла, %"])
    assert pd.isna(bad.iloc[0]["% сделок с положительным фин результатом в цикле"])


def test_cycle_final_result_is_nan_when_cycle_has_no_trades(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none"],
        candidate_dirs=[0, 1, 0, 0],
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(4))

    assert result.iloc[0]["Сделок в цикле"] == 0
    assert pd.isna(result.iloc[0]["Фин результат цикла, %"])


def test_cycle_final_result_nan_pnl_is_guarded(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none"],
        candidate_dirs=[0, 1, 0, 0],
    )
    trades = pd.DataFrame({"entry_index": [2], "net_pnl_pct": [math.nan]})

    result = excel_tester._build_cycle_sheet_df(diag, _df(4), trades)

    assert result.iloc[0]["Сделок в цикле"] == 1
    assert pd.isna(result.iloc[0]["Фин результат цикла, %"])


def test_cycle_final_result_missing_pnl_column_is_nan(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none"],
        candidate_dirs=[0, 1, 0, 0],
    )
    trades = pd.DataFrame({"entry_index": [2]})

    result = excel_tester._build_cycle_sheet_df(diag, _df(4), trades)

    assert result.iloc[0]["Сделок в цикле"] == 1
    assert pd.isna(result.iloc[0]["Фин результат цикла, %"])


def test_acceptance_active_starts_at_zero_and_adjacent_cycles(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["ST_ACTIVE_FREEZE", "OFF", "WAIT_FIRST_ST_FLIP", "OFF"],
        trigger_sources=["candidate_threshold", "none", "candidate_threshold", "none"],
        candidate_dirs=[1, 0, -1, 0],
    )

    result = excel_tester._build_cycle_sheet_df(diag, _df(4))

    assert list(result["Start bar index"]) == [0, 2]
    assert list(result["End bar index"]) == [0, 2]
    assert list(result["ID цикла"]) == [1, 2]


def test_acceptance_confirmed_leg_membership_and_threshold_strictness(monkeypatch):
    excel_tester = _excel_tester_module()

    def fake_per_bar(close, reversal_threshold, local_window, *, daily_reset_event):
        confirm_event = np.zeros(len(close), dtype=np.int8)
        heights = np.full(len(close), np.nan)
        confirm_event[[1, 2, 3]] = 1
        heights[1] = 0.03
        heights[2] = 0.031
        heights[3] = 0.05
        return SimpleNamespace(
            confirm_event=confirm_event,
            last_confirmed_leg_height_pct=heights,
        )

    monkeypatch.setattr(excel_tester, "compute_zigzag_per_bar", fake_per_bar, raising=False)
    monkeypatch.setattr(
        excel_tester,
        "compute_confirmed_legs_reset_aware",
        lambda close, reversal_threshold, *, daily_reset_event: [],
        raising=False,
    )
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none"],
        candidate_dirs=[0, 1, 0, 0],
        candidate_threshold=0.03,
    )

    row = excel_tester._build_cycle_sheet_df(diag, _df(4)).iloc[0]

    assert row["Ног ZigZag в цикле"] == 2
    assert row["Ног выше порога триггера кандидата"] == 1
    assert row["Доля ног выше порога, %"] == 50.0
    assert row["Медиана ног"] == pytest.approx(0.0305)
    assert row["Макс. высота ноги"] == pytest.approx(0.031)


def test_acceptance_non_uniform_candidate_threshold_preserves_rows(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_leg_outputs(monkeypatch, excel_tester, confirm_bar=1, height=0.05)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none"],
        candidate_dirs=[0, 1, 0],
        candidate_threshold=[0.03, 0.04, 0.03],
    )

    row = excel_tester._build_cycle_sheet_df(diag, _df(3)).iloc[0]

    assert row["Ног ZigZag в цикле"] == 1
    assert pd.isna(row["Ног выше порога триггера кандидата"])
    assert pd.isna(row["Доля ног выше порога, %"])


def test_acceptance_empty_legs_metrics(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none"],
        candidate_dirs=[0, 1, 0],
    )

    row = excel_tester._build_cycle_sheet_df(diag, _df(3)).iloc[0]

    assert row["Ног ZigZag в цикле"] == 0
    assert row["Ног выше порога триггера кандидата"] == 0
    assert pd.isna(row["Медиана ног"])
    assert pd.isna(row["Макс. высота ноги"])
    assert pd.isna(row["Доля ног выше порога, %"])


def test_acceptance_percent_metrics_direction_variants(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    df = pd.DataFrame(
        {
            "close": [100.0, 100.0, 110.0, 120.0, 0.0, 5.0],
            "high": [101.0, 130.0, 132.0, 121.0, 6.0, 7.0],
            "low": [99.0, 90.0, 88.0, 100.0, 0.0, 4.0],
        },
        index=pd.date_range("2026-01-01", periods=6, freq="h"),
    )

    plus = excel_tester._build_cycle_sheet_df(
        _diagnostics(
            ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
            trigger_sources=["none", "candidate_threshold", "none"],
            candidate_dirs=[0, 1, 0],
        ),
        df.iloc[:3],
    ).iloc[0]
    minus = excel_tester._build_cycle_sheet_df(
        _diagnostics(
            ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
            trigger_sources=["none", "candidate_threshold", "none"],
            candidate_dirs=[0, -1, 0],
        ),
        df.iloc[:3],
    ).iloc[0]
    empty_direction = excel_tester._build_cycle_sheet_df(
        _diagnostics(["OFF", "ST_ACTIVE_FREEZE", "OFF"]),
        df.iloc[:3],
    ).iloc[0]
    non_positive_start = excel_tester._build_cycle_sheet_df(
        _diagnostics(
            ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
            trigger_sources=["none", "candidate_threshold", "none"],
            candidate_dirs=[0, 1, 0],
        ),
        df.iloc[3:6],
    ).iloc[0]

    assert plus["Размер цикла, %"] == 0.0
    assert plus["Макс. движение по циклу, %"] == 30.0
    assert plus["Макс. просадка внутри цикла, %"] == 10.0
    assert minus["Макс. движение по циклу, %"] == 10.0
    assert minus["Макс. просадка внутри цикла, %"] == 30.0
    assert pd.isna(empty_direction["Макс. движение по циклу, %"])
    assert pd.isna(empty_direction["Макс. просадка внутри цикла, %"])
    assert pd.isna(non_positive_start["Размер цикла, %"])


def test_acceptance_alignment_mismatch_uses_aligned_prefix(monkeypatch):
    excel_tester = _excel_tester_module()
    n_holder = _patch_no_legs(monkeypatch, excel_tester)
    df = _df(6)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF", "ST_ACTIVE_FREEZE"],
        trigger_sources=["none", "candidate_threshold", "none", "candidate_threshold"],
        candidate_dirs=[0, 1, 0, -1],
    )

    result = excel_tester._build_cycle_sheet_df(diag, df)

    assert n_holder["n"] == 4
    assert len(result) == 1
    assert result.iloc[0]["Start bar index"] == 1
    assert result.iloc[0]["Начало цикла"] == df.index[1]


def test_volume_cycle_alignment_uses_shortest_optional_arrays_without_index_errors():
    excel_tester = _excel_tester_module()
    df = _df(6)
    diag = _volume_diagnostics(
        ["OFF", "ACTIVE_LONG", "ACTIVE_LONG", "OFF", "ACTIVE_SHORT", "OFF"],
        daily_reset_event=[0, 0, 0, 0, 0, 0],
        time_filter_reset_event=[0, 0, 0, 1],
        volume_regime=[np.nan, "HIGH", "LOW", "MID"],
        median_relative_volume=[np.nan, 1.0, 2.0, np.nan],
    )

    result = excel_tester._build_volume_cycle_sheet_df(diag, df)

    assert list(result["Start bar index"]) == [1]
    assert list(result["End bar index"]) == [2]
    assert result.iloc[0]["Причина завершения"] == "time_filter_reset"
    assert result.iloc[0]["Режим объёма (старт)"] == "HIGH"
    assert result.iloc[0]["Ср. медиана объёма"] == pytest.approx(1.5)


def test_volume_cycle_volume_regime_nan_is_not_string_nan():
    excel_tester = _excel_tester_module()
    diag = _volume_diagnostics(
        ["OFF", "ACTIVE_LONG", "OFF"],
        daily_reset_event=[0, 0, 0],
        volume_regime=[np.nan, np.nan, np.nan],
        median_relative_volume=[1.0, np.nan, 2.0],
    )

    result = excel_tester._build_volume_cycle_sheet_df(diag, _df(3))

    assert len(result) == 1
    assert pd.isna(result.iloc[0]["Режим объёма (старт)"])


def test_acceptance_daily_reset_inside_cycle_does_not_end_cycle(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_no_legs(monkeypatch, excel_tester)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none", "none"],
        candidate_dirs=[0, 1, 0, 0, 0],
        daily_reset_event=[0, 0, 1, 0, 0],
    )

    row = excel_tester._build_cycle_sheet_df(diag, _df(5)).iloc[0]

    assert row["Start bar index"] == 1
    assert row["End bar index"] == 3
    assert row["Причина завершения"] == "FSM_OFF"


def test_acceptance_both_uses_candidate_direction_not_confirmed_leg(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_leg_outputs(monkeypatch, excel_tester, confirm_bar=1, height=0.05, direction=1)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", "both", "none"],
        candidate_dirs=[0, -1, 0],
    )

    row = excel_tester._build_cycle_sheet_df(diag, _df(3)).iloc[0]

    assert row["Направление цикла"] == "-"


def test_acceptance_builder_reuses_both_zigzag_helpers(monkeypatch):
    excel_tester = _excel_tester_module()
    calls = {"per_bar": 0, "legs": 0}

    def fake_per_bar(close, reversal_threshold, local_window, *, daily_reset_event):
        calls["per_bar"] += 1
        return SimpleNamespace(
            confirm_event=np.zeros(len(close), dtype=np.int8),
            last_confirmed_leg_height_pct=np.full(len(close), np.nan),
        )

    def fake_legs(close, reversal_threshold, *, daily_reset_event):
        calls["legs"] += 1
        return []

    monkeypatch.setattr(excel_tester, "compute_zigzag_per_bar", fake_per_bar)
    monkeypatch.setattr(excel_tester, "compute_confirmed_legs_reset_aware", fake_legs)
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none"],
        candidate_dirs=[0, 1, 0],
    )

    excel_tester._build_cycle_sheet_df(diag, _df(3))

    assert calls == {"per_bar": 1, "legs": 1}


def test_acceptance_non_finite_ohlc_preserves_row_with_nan_metrics(monkeypatch):
    excel_tester = _excel_tester_module()
    _patch_leg_outputs(monkeypatch, excel_tester, confirm_bar=1, height=0.05)
    df = _df(3)
    df.loc[df.index[1], "high"] = np.nan
    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none"],
        candidate_dirs=[0, 1, 0],
    )

    row = excel_tester._build_cycle_sheet_df(diag, df).iloc[0]

    assert row["Ног ZigZag в цикле"] == 1
    assert pd.isna(row["Цена начала"])
    assert pd.isna(row["High цикла"])
    assert pd.isna(row["Размер цикла, %"])
    assert pd.isna(row["Макс. движение по циклу, %"])


def test_acceptance_enabled_export_without_daily_reset_event_writes_empty_cycle(tmp_path):
    import openpyxl
    import supertrend_optimizer.io.excel_tester as excel_tester

    diag = _diagnostics(["OFF", "ST_ACTIVE_FREEZE", "OFF"])
    diag.pop("daily_reset_event")
    output = excel_tester.export_tester_results(
        [_period_result(diag)],
        str(tmp_path / "cycle_missing_reset.xlsx"),
        trade_filter_config=_enabled_cfg(),
        df=_df(3),
    )

    ws = openpyxl.load_workbook(output)["cycle"]
    assert [cell.value for cell in ws[1]] == _expected_cycle_columns()
    assert ws.max_row == 1
    assert ws.auto_filter.ref == "A1:W1"


def test_acceptance_enabled_filter_diagnostics_none_does_not_call_builder(monkeypatch, tmp_path):
    import openpyxl
    import supertrend_optimizer.io.excel_tester as excel_tester

    def fail(*args, **kwargs):
        raise AssertionError("_build_cycle_sheet_df must not be called without diagnostics")

    monkeypatch.setattr(excel_tester, "_build_cycle_sheet_df", fail)
    output = excel_tester.export_tester_results(
        [_period_result(None)],
        str(tmp_path / "cycle_no_diag.xlsx"),
        trade_filter_config=_enabled_cfg(),
        df=_df(3),
    )

    assert "cycle" not in openpyxl.load_workbook(output).sheetnames


def test_write_cycle_sheet_autofilter_and_datetime_format(tmp_path):
    import openpyxl
    import supertrend_optimizer.io.excel_tester as excel_tester

    cycle_df = pd.DataFrame(
        [
            {
                column: None
                for column in excel_tester.CYCLE_SHEET_COLUMNS
            }
        ],
        columns=excel_tester.CYCLE_SHEET_COLUMNS,
    )
    cycle_df.loc[0, "Начало цикла"] = pd.Timestamp("2026-01-01 01:02:03")
    cycle_df.loc[0, "Конец цикла"] = pd.Timestamp("2026-01-01 03:04:05")

    path = tmp_path / "cycle_writer.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        excel_tester._write_cycle_sheet(writer, cycle_df)

    wb = openpyxl.load_workbook(path)
    ws = wb["cycle"]
    assert [cell.value for cell in ws[1]] == list(excel_tester.CYCLE_SHEET_COLUMNS)
    assert ws.auto_filter.ref == "A1:W1"
    assert ws["A2"].number_format == "YYYY-MM-DD HH:MM:SS"
    assert ws["B2"].number_format == "YYYY-MM-DD HH:MM:SS"


def test_write_empty_cycle_sheet_keeps_autofilter(tmp_path):
    import openpyxl
    import supertrend_optimizer.io.excel_tester as excel_tester

    path = tmp_path / "cycle_writer_empty.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        excel_tester._write_cycle_sheet(
            writer,
            pd.DataFrame(columns=excel_tester.CYCLE_SHEET_COLUMNS),
        )

    wb = openpyxl.load_workbook(path)
    assert wb["cycle"].auto_filter.ref == "A1:W1"


def test_enabled_export_writes_cycle_headers_and_autofilter(monkeypatch, tmp_path):
    excel_tester = _excel_tester_module()
    import openpyxl

    diag = _diagnostics(["OFF", "ST_ACTIVE_FREEZE", "OFF", "OFF"])
    monkeypatch.setattr(
        excel_tester,
        "_build_cycle_sheet_df",
        lambda filter_diagnostics, df, trades_df=None: pd.DataFrame(columns=_expected_cycle_columns()),
        raising=False,
    )

    output = excel_tester.export_tester_results(
        [_period_result(diag)],
        str(tmp_path / "cycle_enabled.xlsx"),
        trade_filter_config=_enabled_cfg(),
        df=_df(4),
    )
    wb = openpyxl.load_workbook(output)

    assert "cycle" in wb.sheetnames
    assert [cell.value for cell in wb["cycle"][1]] == _expected_cycle_columns()
    assert wb["cycle"].auto_filter.ref


def test_enabled_export_writes_cycle_when_diagnostic_flags_off(monkeypatch, tmp_path):
    excel_tester = _excel_tester_module()
    import openpyxl

    diag = _diagnostics(["OFF", "ST_ACTIVE_FREEZE", "OFF", "OFF"])
    monkeypatch.setattr(
        excel_tester,
        "_build_cycle_sheet_df",
        lambda filter_diagnostics, df, trades_df=None: pd.DataFrame(columns=_expected_cycle_columns()),
        raising=False,
    )

    output = excel_tester.export_tester_results(
        [_period_result(diag)],
        str(tmp_path / "cycle_flags_off.xlsx"),
        trade_filter_config=_enabled_cfg(export_state_columns=False, export_trigger_columns=False),
        df=_df(4),
    )
    wb = openpyxl.load_workbook(output)

    assert "FilterDiagnostics_100" not in wb.sheetnames
    assert "ZigZag_Trigger_Events" not in wb.sheetnames
    assert "cycle" in wb.sheetnames


def test_enabled_export_writes_cycle_final_result_numeric_to_xlsx(tmp_path):
    excel_tester = _excel_tester_module()
    import openpyxl

    diag = _diagnostics(
        ["OFF", "ST_ACTIVE_FREEZE", "ST_ACTIVE_MONITORING", "ST_ACTIVE_MONITORING", "OFF"],
        trigger_sources=["none", "candidate_threshold", "none", "none", "none"],
        candidate_dirs=[0, 1, 0, 0, 0],
        local_window=1,
    )
    trades = pd.DataFrame(
        {
            "entry_index": [2, 3, 5],
            "net_pnl_pct": [0.5, -1.5, 99.0],
        }
    )

    output = excel_tester.export_tester_results(
        [_period_result(diag, trades)],
        str(tmp_path / "cycle_final_result.xlsx"),
        trade_filter_config=_enabled_cfg(export_state_columns=False, export_trigger_columns=False),
        df=_df(5),
    )
    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb["cycle"]
    headers = [cell.value for cell in ws[1]]
    col_idx = headers.index("Фин результат цикла, %") + 1
    value = ws.cell(row=2, column=col_idx).value

    assert isinstance(value, (int, float))
    assert value == pytest.approx(-1.0)


def test_disabled_export_does_not_write_cycle_or_call_builder(monkeypatch, tmp_path):
    excel_tester = _excel_tester_module()
    import openpyxl

    called = {"value": False}

    def spy(*args, **kwargs):
        called["value"] = True
        raise AssertionError("_build_cycle_sheet_df must not be called when filter is disabled")

    monkeypatch.setattr(excel_tester, "_build_cycle_sheet_df", spy, raising=False)

    output = excel_tester.export_tester_results(
        [_period_result(None)],
        str(tmp_path / "cycle_disabled.xlsx"),
        trade_filter_config=SimpleNamespace(enabled=False, diagnostics=None),
        df=_df(4),
    )
    wb = openpyxl.load_workbook(output)

    assert "cycle" not in wb.sheetnames
    assert called["value"] is False
