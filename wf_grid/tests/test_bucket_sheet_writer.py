"""
Unit tests for Этап 6: wf_grid/export/bucket_sheet_writer.py
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from wf_grid.export.bucket_sheet_writer import (
    _PLACEHOLDER_COLS,
    _format_bucket_df,
    write_bucket_matrix_median_sheet,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_bucket_matrix_df(n_buckets=3, n_steps=2):
    """Create a minimal bucket matrix DataFrame for testing.

    Includes interleaved Step*/DD_Step* columns and risk summary columns
    to match the 24+2N layout produced by build_median_bucket_matrix.
    """
    rows = []
    for i in range(n_buckets):
        row = {
            "bucket_param": f"ATR {10+i*2}\u201312, M 2.0\u20132.2",
            "bucket_key": f"{10+i*2}_{10+i}",
            "atr_bucket": 10 + i * 2,
            "mult_bucket_ticks": 10 + i,
            "bucket_size": 4,
        }
        for s in range(1, n_steps + 1):
            row[f"Step{s}"] = float(10 + i + s * 0.5)
            row[f"DD_Step{s}"] = float(-0.05 - s * 0.01 - i * 0.005)
        row.update({
            "max_drawdown_Median": float(-0.10 - i * 0.02),
            "max_drawdown_Min": float(-0.15 - i * 0.03),
            "bucket_presence_steps": ",".join(f"Step{s}" for s in range(1, n_steps + 1)),
            "mean_oos_pnl": float(10 + i),
            "median_oos_pnl": float(9.5 + i),
            "std_bucket": float(2.0 + i * 0.1),
            "pct_params_positive_pnl": 0.75,
            "wins_count": 1 if i == 0 else 0,
            "win_steps": "Step1" if i == 0 else "",
            "top3_count": 1,
            "above_median_count": 1,
            "above_median_ratio": 0.5,
            "presence_count": n_steps,
            "above_median_ratio_present": 0.5,
            "eligible_median_steps_count": n_steps,
            "above_median_ratio_eligible": 0.5,
            "bucket_stability_score": round(0.8 - i * 0.1, 6),
            "zone_dominance_score": round(0.7 - i * 0.1, 6),
            "bucket_balanced_score": round(0.6 - i * 0.1, 6),
        })
        rows.append(row)
    return pd.DataFrame(rows)


def _write_and_load(bucket_df, atr_step=2, mult_step=0.2) -> openpyxl.Workbook:
    """Write bucket sheet to temp file and load with openpyxl."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        write_bucket_matrix_median_sheet(
            bucket_df, writer, atr_step, mult_step,
        )
    wb = openpyxl.load_workbook(path)
    Path(path).unlink(missing_ok=True)
    return wb


# ===========================================================================
# _format_bucket_df
# ===========================================================================

class TestFormatBucketDf:
    def test_integer_cols_int64(self):
        df = _make_bucket_matrix_df()
        formatted = _format_bucket_df(df)
        for col in ["wins_count", "top3_count", "presence_count",
                     "above_median_count", "eligible_median_steps_count",
                     "bucket_size", "atr_bucket", "mult_bucket_ticks"]:
            assert formatted[col].dtype.name == "Int64", (
                f"{col} should be Int64, got {formatted[col].dtype}"
            )

    def test_score_precision_6(self):
        df = _make_bucket_matrix_df()
        formatted = _format_bucket_df(df)
        for col in ["bucket_stability_score", "zone_dominance_score"]:
            for val in formatted[col].dropna():
                s = f"{val:.10f}"
                assert len(s.split(".")[1].rstrip("0")) <= 6

    def test_metric_precision_6(self):
        df = _make_bucket_matrix_df()
        df["mean_oos_pnl"] = 12.1234567890
        formatted = _format_bucket_df(df)
        val = formatted["mean_oos_pnl"].iloc[0]
        assert val == pytest.approx(12.123457, abs=1e-7)

    def test_step_precision_6(self):
        df = _make_bucket_matrix_df()
        df["Step1"] = 1.23456789
        formatted = _format_bucket_df(df)
        val = formatted["Step1"].iloc[0]
        assert val == pytest.approx(1.234568, abs=1e-7)

    def test_does_not_mutate_input(self):
        df = _make_bucket_matrix_df()
        original_val = df["bucket_stability_score"].iloc[0]
        _format_bucket_df(df)
        assert df["bucket_stability_score"].iloc[0] == original_val

    def test_dd_step_columns_rounded(self):
        """DD_Step* округлены до 6 знаков."""
        df = _make_bucket_matrix_df()
        df["DD_Step1"] = -0.123456789
        formatted = _format_bucket_df(df)
        val = formatted["DD_Step1"].iloc[0]
        assert val == pytest.approx(-0.123457, abs=1e-7)

    def test_risk_summary_columns_rounded(self):
        """max_drawdown_Median/Min округлены до 6 знаков."""
        df = _make_bucket_matrix_df()
        df["max_drawdown_Median"] = -0.123456789
        df["max_drawdown_Min"] = -0.987654321
        formatted = _format_bucket_df(df)
        assert formatted["max_drawdown_Median"].iloc[0] == pytest.approx(-0.123457, abs=1e-7)
        assert formatted["max_drawdown_Min"].iloc[0] == pytest.approx(-0.987654, abs=1e-7)


# ===========================================================================
# Placeholder mode
# ===========================================================================

class TestPlaceholderMode:
    def test_none_input(self):
        wb = _write_and_load(None)
        assert "BucketMatrix_Median" in wb.sheetnames
        ws = wb["BucketMatrix_Median"]
        # Header row exists
        assert ws.cell(row=1, column=1).value is not None
        # No data rows (row 2 should be empty)
        assert ws.cell(row=2, column=1).value is None

    def test_empty_df_input(self):
        wb = _write_and_load(pd.DataFrame())
        ws = wb["BucketMatrix_Median"]
        assert ws.cell(row=1, column=1).value is not None
        assert ws.cell(row=2, column=1).value is None

    def test_placeholder_freeze_panes(self):
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        assert ws.freeze_panes == "E2"

    def test_placeholder_autofilter(self):
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref.startswith("A1:")


# ===========================================================================
# Normal mode
# ===========================================================================

class TestNormalMode:
    def test_sheet_exists(self):
        df = _make_bucket_matrix_df()
        wb = _write_and_load(df)
        assert "BucketMatrix_Median" in wb.sheetnames

    def test_freeze_panes_e2(self):
        df = _make_bucket_matrix_df()
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        assert ws.freeze_panes == "E2"

    def test_autofilter_header_only(self):
        df = _make_bucket_matrix_df()
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        ref = ws.auto_filter.ref
        assert ref is not None
        # Header-only: row range is just "1"
        assert ref.endswith("1")

    def test_data_rows_written(self):
        df = _make_bucket_matrix_df(n_buckets=3)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        # Row 1 = header, rows 2-4 = data
        assert ws.cell(row=2, column=1).value is not None
        assert ws.cell(row=4, column=1).value is not None
        assert ws.cell(row=5, column=1).value is None

    def test_heatmap_exists(self):
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        # Heatmap title should be somewhere below data
        found_heatmap_title = False
        for row in ws.iter_rows(min_row=5, max_row=20, max_col=1, values_only=True):
            if row[0] and "HEATMAP" in str(row[0]).upper():
                found_heatmap_title = True
                break
        assert found_heatmap_title, "Heatmap title row not found"

    def test_heatmap_does_not_overwrite_main_table(self):
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        # First data row should still have data
        assert ws.cell(row=2, column=1).value is not None

    def test_conditional_formatting_exists(self):
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        assert len(ws.conditional_formatting._cf_rules) > 0

    def test_full_grid_nan_scores_is_normal_not_placeholder(self):
        """Full grid with NaN/0 scores → normal sheet (not placeholder)."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        df["Step1"] = np.nan
        df["Step2"] = np.nan
        df["bucket_stability_score"] = 0.0
        df["zone_dominance_score"] = 0.0
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        # Data rows should exist (row 2 has data)
        assert ws.cell(row=2, column=1).value is not None


# ===========================================================================
# Placeholder column contract
# ===========================================================================

class TestPlaceholderCols:
    def test_placeholder_has_24_columns(self):
        """Placeholder содержит ровно 24 колонки."""
        assert len(_PLACEHOLDER_COLS) == 24

    def test_placeholder_includes_risk_summary_cols(self):
        """Placeholder содержит max_drawdown_Median и max_drawdown_Min."""
        assert "max_drawdown_Median" in _PLACEHOLDER_COLS
        assert "max_drawdown_Min" in _PLACEHOLDER_COLS

    def test_placeholder_excludes_dd_step_cols(self):
        """Placeholder не содержит DD_Step* колонок."""
        for col in _PLACEHOLDER_COLS:
            assert not col.startswith("DD_Step"), (
                f"Placeholder must not contain DD_Step* cols, found: {col}"
            )

    def test_placeholder_excludes_step_cols(self):
        """Placeholder не содержит Step* колонок (динамические)."""
        for col in _PLACEHOLDER_COLS:
            assert not col.startswith("Step"), (
                f"Placeholder must not contain Step* cols, found: {col}"
            )

    def test_placeholder_written_header_count(self):
        """При None input Excel лист содержит ровно 24 заголовка."""
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        # Count non-None cells in header row
        header_vals = [ws.cell(row=1, column=c).value for c in range(1, 32)]
        headers = [v for v in header_vals if v is not None]
        assert len(headers) == 24

    def test_placeholder_risk_summary_in_header(self):
        """Excel header строка placeholder содержит оба risk summary."""
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 30)]
        assert "max_drawdown_Median" in headers
        assert "max_drawdown_Min" in headers

    def test_placeholder_no_dd_step_in_header(self):
        """Excel header строка placeholder не содержит DD_Step*."""
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 30)]
        for h in headers:
            if h is not None:
                assert not str(h).startswith("DD_Step"), (
                    f"Placeholder header must not contain DD_Step*, found: {h}"
                )

    def test_placeholder_no_heatmaps(self):
        """Placeholder → 0 heatmaps (no conditional formatting)."""
        wb = _write_and_load(None)
        ws = wb["BucketMatrix_Median"]
        assert len(ws.conditional_formatting._cf_rules) == 0

    def test_placeholder_exact_column_order(self):
        """_PLACEHOLDER_COLS matches the exact expected 24-column ordered list.

        Expected layout (no Step*/DD_Step* — those are dynamic):
            Block A (5):  bucket_param, bucket_key, atr_bucket, mult_bucket_ticks, bucket_size
            Block C (17): bucket_presence_steps, mean_oos_pnl, median_oos_pnl,
                          std_bucket, pct_params_positive_pnl, wins_count, win_steps,
                          top3_count, above_median_count, above_median_ratio,
                          presence_count, above_median_ratio_present,
                          eligible_median_steps_count, above_median_ratio_eligible,
                          bucket_stability_score, zone_dominance_score,
                          bucket_balanced_score
            Block E (2):  max_drawdown_Median, max_drawdown_Min
        """
        expected = [
            # Block A
            "bucket_param",
            "bucket_key",
            "atr_bucket",
            "mult_bucket_ticks",
            "bucket_size",
            # Block C
            "bucket_presence_steps",
            "mean_oos_pnl",
            "median_oos_pnl",
            "std_bucket",
            "pct_params_positive_pnl",
            "wins_count",
            "win_steps",
            "top3_count",
            "above_median_count",
            "above_median_ratio",
            "presence_count",
            "above_median_ratio_present",
            "eligible_median_steps_count",
            "above_median_ratio_eligible",
            "bucket_stability_score",
            "zone_dominance_score",
            "bucket_balanced_score",
            # Block E (risk summary, always present regardless of N)
            "max_drawdown_Median",
            "max_drawdown_Min",
        ]
        assert list(_PLACEHOLDER_COLS) == expected, (
            f"Placeholder column order mismatch.\n"
            f"Expected: {expected}\n"
            f"Got:      {list(_PLACEHOLDER_COLS)}"
        )


# ===========================================================================
# Number format contract tests (3.15)
# ===========================================================================

def _find_col_index(ws, col_name):
    """Find 1-based column index by header name in row 1."""
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == col_name:
            return ci
    return None


class TestNumberFormats:
    def test_dd_step_number_format_is_percent(self):
        """DD_Step* ячейки имеют формат '0.00%'."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "DD_Step1")
        assert ci is not None, "DD_Step1 column not found in header"
        cell = ws.cell(row=2, column=ci)
        assert cell.number_format == "0.00%", (
            f"DD_Step1 format should be '0.00%', got '{cell.number_format}'"
        )

    def test_dd_step2_number_format_is_percent(self):
        """DD_Step2 ячейки имеют формат '0.00%'."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "DD_Step2")
        assert ci is not None
        cell = ws.cell(row=2, column=ci)
        assert cell.number_format == "0.00%"

    def test_pnl_step_number_format_unchanged(self):
        """Step* ячейки по-прежнему '0.000000'."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "Step1")
        assert ci is not None, "Step1 column not found"
        cell = ws.cell(row=2, column=ci)
        assert cell.number_format == "0.000000", (
            f"Step1 format should be '0.000000', got '{cell.number_format}'"
        )

    def test_risk_summary_number_format_is_percent(self):
        """max_drawdown_Median и max_drawdown_Min имеют формат '0.00%'."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        for col_name in ("max_drawdown_Median", "max_drawdown_Min"):
            ci = _find_col_index(ws, col_name)
            assert ci is not None, f"{col_name} column not found"
            cell = ws.cell(row=2, column=ci)
            assert cell.number_format == "0.00%", (
                f"{col_name} format should be '0.00%', got '{cell.number_format}'"
            )

    def test_stability_score_format_unchanged(self):
        """bucket_stability_score по-прежнему '0.000000'."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        ci = _find_col_index(ws, "bucket_stability_score")
        assert ci is not None
        cell = ws.cell(row=2, column=ci)
        assert cell.number_format == "0.000000"


# ===========================================================================
# Second heatmap tests (3.16)
# ===========================================================================

class TestSecondHeatmap:
    def test_second_heatmap_written(self):
        """Вторая heatmap присутствует на листе."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        titles = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and "HEATMAP" in str(row[0]).upper():
                titles.append(str(row[0]))
        assert len(titles) >= 2, f"Expected 2 heatmap titles, found {len(titles)}: {titles}"

    def test_second_heatmap_title(self):
        """Вторая heatmap содержит 'RISK HEATMAP' в заголовке."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        found_risk = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] and "RISK HEATMAP" in str(row[0]).upper():
                found_risk = True
                break
        assert found_risk, "Risk heatmap title not found"

    def test_second_heatmap_position(self):
        """Вторая heatmap ниже первой (startrow > first heatmap)."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        heatmap_rows = []
        for ri in range(1, ws.max_row + 1):
            val = ws.cell(row=ri, column=1).value
            if val and "HEATMAP" in str(val).upper():
                heatmap_rows.append(ri)
        assert len(heatmap_rows) >= 2, "Expected at least 2 heatmap title rows"
        assert heatmap_rows[1] > heatmap_rows[0], (
            f"Risk heatmap (row {heatmap_rows[1]}) must be below "
            f"stability heatmap (row {heatmap_rows[0]})"
        )

    def test_risk_heatmap_color_scale_is_fixed(self):
        """Risk heatmap ColorScaleRule: start_type='num', start_value=-0.50."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        assert len(cf_keys) >= 2, f"Expected >= 2 CF rules, got {len(cf_keys)}"
        risk_cf = cf_keys[1]
        rule = risk_cf.rules[0]
        cs = rule.colorScale
        assert cs.cfvo[0].type == "num", (
            f"Risk heatmap start_type should be 'num', got '{cs.cfvo[0].type}'"
        )
        assert float(cs.cfvo[0].val) == pytest.approx(-0.50)
        assert float(cs.cfvo[1].val) == pytest.approx(-0.15)
        assert float(cs.cfvo[2].val) == pytest.approx(0.0)

    def test_risk_heatmap_colors(self):
        """Risk heatmap цвета: red → yellow → green."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        risk_cf = cf_keys[1]
        colors = risk_cf.rules[0].colorScale.color
        assert colors[0].rgb == "FFFF0000"  # red
        assert colors[1].rgb == "FFFFFF99"  # yellow
        assert colors[2].rgb == "FF00AA00"  # green

    def test_first_heatmap_not_broken(self):
        """Первая heatmap всё ещё имеет auto-scale (start_type='min')."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        cf_keys = list(ws.conditional_formatting._cf_rules.keys())
        assert len(cf_keys) >= 2
        stability_cf = cf_keys[0]
        rule = stability_cf.rules[0]
        cs = rule.colorScale
        assert cs.cfvo[0].type == "min", (
            f"Stability heatmap start_type should be 'min', got '{cs.cfvo[0].type}'"
        )

    def test_two_conditional_formatting_rules(self):
        """Normal mode → ровно 2 conditional formatting rules."""
        df = _make_bucket_matrix_df(n_buckets=3, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        assert len(ws.conditional_formatting._cf_rules) == 2


# ===========================================================================
# New columns: median_oos_pnl and bucket_balanced_score
# ===========================================================================

class TestMedianOosPnlInWriter:
    def test_median_oos_pnl_in_placeholder(self):
        """median_oos_pnl присутствует в _PLACEHOLDER_COLS."""
        assert "median_oos_pnl" in _PLACEHOLDER_COLS

    def test_median_oos_pnl_in_bucket_metric_cols(self):
        from wf_grid.export.bucket_sheet_writer import _BUCKET_METRIC_COLS
        assert "median_oos_pnl" in _BUCKET_METRIC_COLS

    def test_median_oos_pnl_number_format(self):
        """median_oos_pnl ячейки имеют формат 0.000000."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "median_oos_pnl" in cols:
            ci = cols.index("median_oos_pnl") + 1
            cell = ws.cell(row=2, column=ci)
            assert cell.number_format == "0.000000"

    def test_median_oos_pnl_rounded_in_format(self):
        """_format_bucket_df rounds median_oos_pnl to 6 decimal places."""
        df = _make_bucket_matrix_df()
        df["median_oos_pnl"] = 3.123456789123
        formatted = _format_bucket_df(df)
        val = formatted["median_oos_pnl"].iloc[0]
        assert round(val, 6) == val


class TestBucketBalancedScoreInWriter:
    def test_bucket_balanced_score_in_placeholder(self):
        """bucket_balanced_score присутствует в _PLACEHOLDER_COLS."""
        assert "bucket_balanced_score" in _PLACEHOLDER_COLS

    def test_bucket_balanced_score_in_score_cols(self):
        from wf_grid.export.bucket_sheet_writer import _BUCKET_SCORE_COLS
        assert "bucket_balanced_score" in _BUCKET_SCORE_COLS

    def test_bucket_balanced_score_number_format(self):
        """bucket_balanced_score ячейки имеют формат 0.000000."""
        df = _make_bucket_matrix_df(n_buckets=2, n_steps=2)
        wb = _write_and_load(df)
        ws = wb["BucketMatrix_Median"]
        cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "bucket_balanced_score" in cols:
            ci = cols.index("bucket_balanced_score") + 1
            cell = ws.cell(row=2, column=ci)
            assert cell.number_format == "0.000000"

    def test_bucket_balanced_score_rounded_in_format(self):
        """_format_bucket_df rounds bucket_balanced_score to 6 decimal places."""
        df = _make_bucket_matrix_df()
        df["bucket_balanced_score"] = 0.123456789123
        formatted = _format_bucket_df(df)
        val = formatted["bucket_balanced_score"].iloc[0]
        assert round(val, 6) == val

    def test_placeholder_balanced_score_last_score_col(self):
        """bucket_balanced_score appears after zone_dominance_score in placeholder."""
        zone_idx = _PLACEHOLDER_COLS.index("zone_dominance_score")
        balanced_idx = _PLACEHOLDER_COLS.index("bucket_balanced_score")
        assert balanced_idx == zone_idx + 1
