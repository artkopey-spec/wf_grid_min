"""
Tests for build_signal_events() — 17 test cases per plan §15.

Tests 1–15:  builder logic (signal_events.py)
Tests 16–17: Excel integration (excel_tester.py)

Synthetic OHLC helper
---------------------
All helpers produce DataFrames with a RangeIndex (no DatetimeIndex) which is
fine — build_signal_events() only requires ``df["open/high/low/close"]`` and
``df.index`` (used as ``signal_time``).
"""

import math
import io
import tempfile
import os

import numpy as np
import pandas as pd
import pytest

from supertrend_optimizer.testing.signal_events import build_signal_events
from supertrend_optimizer.utils.enums import ExecutionModel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_flat_ohlc(n: int, price: float = 100.0) -> pd.DataFrame:
    """All bars identical (close = open = high = low = price)."""
    return pd.DataFrame(
        {
            "open": np.full(n, price),
            "high": np.full(n, price),
            "low": np.full(n, price),
            "close": np.full(n, price),
        }
    )


def _make_ohlc(
    n: int,
    open_: float = 100.0,
    close_: float = 101.0,
    high_delta: float = 2.0,
    low_delta: float = 2.0,
) -> pd.DataFrame:
    """All bars have the same OHLC values for predictable formulas."""
    high = max(open_, close_) + high_delta
    low = min(open_, close_) - low_delta
    return pd.DataFrame(
        {
            "open": np.full(n, open_),
            "high": np.full(n, high),
            "low": np.full(n, low),
            "close": np.full(n, close_),
        }
    )


def _trend_from_list(values: list) -> np.ndarray:
    return np.array(values, dtype=np.int8)


ATR_PERIOD = 3  # small so tests are tractable


# ---------------------------------------------------------------------------
# Test 1: long mode — green→open_signal, red→close_signal
# ---------------------------------------------------------------------------

def test_signal_events_long():
    """long mode: green → close_signal; red → open_signal (of next bar)."""
    n = 10
    df = _make_ohlc(n)
    # Stable trend: first atr_period-1 bars = 0, then: 2 greens, red, green, red
    # t=0,1 → 0 (neutral)
    # t=2 → 1 (first stable)
    # t=3 → 1
    # t=4 → -1  (flip: green→red at t=4, t-1=green → long close_signal)
    # t=5 → 1   (flip: red→green at t=5, t-1=red   → long open_signal)
    # t=6..9 → 1
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    assert len(result) == 2
    assert list(result["event_type"]) == ["close_signal", "open_signal"]
    assert list(result["direction"]) == ["LONG", "LONG"]
    assert list(result["signal_bar_index"]) == [4, 5]
    assert all(~result["is_reversal"])


# ---------------------------------------------------------------------------
# Test 2: short mode — inversion of long
# ---------------------------------------------------------------------------

def test_signal_events_short():
    """short mode: red → close_signal; green → open_signal."""
    n = 10
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "short", ExecutionModel.CLOSE_TO_CLOSE)

    assert len(result) == 2
    # green→red: short open_signal (was in green = was not in short position, now enter)
    # red→green: short close_signal (exit short)
    assert list(result["event_type"]) == ["open_signal", "close_signal"]
    assert list(result["direction"]) == ["SHORT", "SHORT"]


# ---------------------------------------------------------------------------
# Test 3: revers — 2 rows, is_reversal=True, order close→open
# ---------------------------------------------------------------------------

def test_signal_events_revers():
    """revers: each flip → 2 rows, close before open, both is_reversal=True."""
    n = 8
    df = _make_ohlc(n)
    # t=0,1 → 0; t=2 → 1; t=3 → -1 (flip); t=4..7 → 1
    trend = _trend_from_list([0, 0, 1, -1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "revers", ExecutionModel.CLOSE_TO_CLOSE)

    # 2 flips → 4 rows
    assert len(result) == 4

    # First signal (t=3, green→red): long_close + short_open
    row0 = result.iloc[0]
    row1 = result.iloc[1]
    assert row0["event_type"] == "long_close_signal"
    assert row1["event_type"] == "short_open_signal"
    assert row0["signal_bar_index"] == row1["signal_bar_index"]
    assert row0["is_reversal"] is True or row0["is_reversal"] == True
    assert row1["is_reversal"] is True or row1["is_reversal"] == True


# ---------------------------------------------------------------------------
# Test 4: both treated as revers
# ---------------------------------------------------------------------------

def test_signal_events_both_as_revers():
    """trade_mode='both' must produce the same result as 'revers'."""
    n = 8
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, -1, 1, 1, 1, 1])

    res_revers = build_signal_events(df, trend, ATR_PERIOD, "revers", ExecutionModel.CLOSE_TO_CLOSE)
    res_both = build_signal_events(df, trend, ATR_PERIOD, "both", ExecutionModel.CLOSE_TO_CLOSE)

    pd.testing.assert_frame_equal(
        res_revers.reset_index(drop=True),
        res_both.reset_index(drop=True),
    )


# ---------------------------------------------------------------------------
# Test 5: Body% and Range% exact numeric values
# ---------------------------------------------------------------------------

def test_body_range_formulas():
    """Signal Body % = abs(close-open)/open*100; Range % = (high-low)/open*100."""
    n = 6
    open_ = 100.0
    close_ = 102.0
    high_delta = 3.0   # high = 102 + 3 = 105
    low_delta = 4.0    # low  = 100 - 4 = 96
    df = _make_ohlc(n, open_=open_, close_=close_,
                    high_delta=high_delta, low_delta=low_delta)

    # Signal at t=3 (t-1=2 is stable 1, t=3 is -1)
    trend = _trend_from_list([0, 0, 1, -1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    # Only the flip at t=3 is in the signal bar with the given OHLC
    row = result[result["signal_bar_index"] == 3].iloc[0]

    expected_body = abs(close_ - open_) / open_ * 100.0   # 2/100*100 = 2.0
    expected_range = (df["high"].iloc[3] - df["low"].iloc[3]) / open_ * 100.0

    assert math.isclose(row["signal_body_pct"], expected_body, rel_tol=1e-9)
    assert math.isclose(row["signal_range_pct"], expected_range, rel_tol=1e-9)


# ---------------------------------------------------------------------------
# Test 6: Body/Range denominator is open, not close
# ---------------------------------------------------------------------------

def test_body_range_open_denominator():
    """Body % denominator must be open[t], not close[t]."""
    n = 6
    open_ = 100.0
    close_ = 110.0   # close > open → body_pct differs depending on denominator
    df = _make_ohlc(n, open_=open_, close_=close_, high_delta=1.0, low_delta=1.0)
    trend = _trend_from_list([0, 0, 1, -1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    row = result[result["signal_bar_index"] == 3].iloc[0]

    body_with_open_denom = abs(close_ - open_) / open_ * 100.0   # 10%
    body_with_close_denom = abs(close_ - open_) / close_ * 100.0  # ≈ 9.09%

    # Must match open-denominator formula
    assert math.isclose(row["signal_body_pct"], body_with_open_denom, rel_tol=1e-9)
    assert not math.isclose(row["signal_body_pct"], body_with_close_denom, rel_tol=1e-4)


# ---------------------------------------------------------------------------
# Test 7: ATR uses atr[t-1], not atr[t]
# ---------------------------------------------------------------------------

def test_atr_prev_not_current():
    """Signal Body ATR must use ATR at bar t-1, not t."""
    from supertrend_optimizer.core.calculator import calculate_true_range, calculate_atr_rma

    n = 10
    open_ = 100.0
    close_ = 102.0
    df = _make_ohlc(n, open_=open_, close_=close_, high_delta=2.0, low_delta=2.0)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    # Signal at t=4
    signal_t = 4

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    row = result[result["signal_bar_index"] == signal_t].iloc[0]

    # Compute expected ATR[t-1]
    tr = calculate_true_range(df["high"].values, df["low"].values, df["close"].values)
    atr = calculate_atr_rma(tr, ATR_PERIOD)
    atr_prev = atr[signal_t - 1]

    body = abs(close_ - open_)
    expected_body_atr = body / atr_prev

    assert math.isclose(row["signal_body_atr"], expected_body_atr, rel_tol=1e-9)

    # Must differ from atr[t]
    atr_current = atr[signal_t]
    wrong_body_atr = body / atr_current
    # ATR values differ for varying price series; check guard passes when distinct
    if not math.isclose(atr_prev, atr_current, rel_tol=1e-9):
        assert not math.isclose(row["signal_body_atr"], wrong_body_atr, rel_tol=1e-4)


# ---------------------------------------------------------------------------
# Test 8: forward returns open_to_open
# ---------------------------------------------------------------------------

def test_forward_returns_o2o():
    """T+1 o2o = direction*(close[t+1] - open[t+1]) / open[t+1] * 100."""
    n = 10
    open_ = 100.0
    close_ = 101.0
    df = _make_ohlc(n, open_=open_, close_=close_, high_delta=1.0, low_delta=1.0)
    # Signal at t=4 (prev=1 green→red): long close_signal, direction_val=+1
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.OPEN_TO_OPEN)
    row = result[result["signal_bar_index"] == 4].iloc[0]
    assert row["event_type"] == "close_signal"

    # exec_bar = t+1 = 5; exec_price = open[5] = 100.0
    # T+1 target = exec_bar+1 = 6; close[6] = 101.0
    # direction = +1 (close_signal long = post-exit same side)
    exec_price = open_   # open[5]
    t1_expected = +1 * (close_ - exec_price) / exec_price * 100.0

    assert math.isclose(row["t1_return_pct"], t1_expected, rel_tol=1e-9)


# ---------------------------------------------------------------------------
# Test 9: forward returns close_to_close
# ---------------------------------------------------------------------------

def test_forward_returns_c2c():
    """T+1 c2c = direction*(close[t+1] - close[t]) / close[t] * 100."""
    n = 10
    open_ = 100.0
    close_ = 101.0
    df = _make_ohlc(n, open_=open_, close_=close_, high_delta=1.0, low_delta=1.0)
    # Signal at t=4 (green→red): long close_signal, direction_val=+1
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    row = result[result["signal_bar_index"] == 4].iloc[0]
    assert row["event_type"] == "close_signal"

    # exec_bar = t = 4; exec_price = close[4] = 101.0
    # T+1 target = exec_bar+1 = 5; close[5] = 101.0
    exec_price = close_
    t1_expected = +1 * (close_ - exec_price) / exec_price * 100.0  # 0.0

    assert math.isclose(row["t1_return_pct"], t1_expected, abs_tol=1e-9)


# ---------------------------------------------------------------------------
# Test 10: NaN when not enough bars for exec_price or T+k
# ---------------------------------------------------------------------------

def test_nan_at_end():
    """Signal at last bar: exec_price=NaN (o2o); T+k=NaN when bars missing."""
    n = 5
    df = _make_ohlc(n)
    # Signal at last bar t=4 (prev=1 green→red)
    trend = _trend_from_list([0, 0, 1, 1, -1])

    # open_to_open: exec_bar = 5 >= n → exec_price = NaN, all T+k = NaN
    result_o2o = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.OPEN_TO_OPEN)
    row_o2o = result_o2o[result_o2o["signal_bar_index"] == 4].iloc[0]
    assert math.isnan(row_o2o["exec_price"])
    assert math.isnan(row_o2o["t1_return_pct"])
    assert math.isnan(row_o2o["t2_return_pct"])
    assert math.isnan(row_o2o["t3_return_pct"])

    # close_to_close: exec_bar = 4, exec_price = close[4]; T+1..T+3 = NaN (no future bars)
    result_c2c = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    row_c2c = result_c2c[result_c2c["signal_bar_index"] == 4].iloc[0]
    assert not math.isnan(row_c2c["exec_price"])
    assert math.isnan(row_c2c["t1_return_pct"])
    assert math.isnan(row_c2c["t2_return_pct"])
    assert math.isnan(row_c2c["t3_return_pct"])


# ---------------------------------------------------------------------------
# Test 11: no signals → empty DataFrame with correct columns
# ---------------------------------------------------------------------------

def test_no_signals():
    """Trend that never changes → empty DataFrame with all 19 column headers."""
    n = 8
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, 1, 1, 1, 1])  # no flips

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    assert len(result) == 0
    expected_cols = [
        "signal_time", "signal_bar_index", "event_type", "direction",
        "st_color_before", "st_color_after", "is_reversal", "exec_price",
        "signal_open_price", "signal_close_price",
        "signal_body_pct", "signal_range_pct", "signal_body_atr", "signal_range_atr",
        "t1_return_pct", "t2_return_pct", "t3_return_pct",
        # filter columns added by PR 9
        "atr_pct", "volume_ratio", "volatility_pass", "volume_pass",
        "allow_entry", "filtered_reason", "entry_bar_index",
        # amplitude diagnostic columns (schema v2, patch §J)
        "amp_n", "amp_threshold", "separation", "amp_valid", "amp_ok", "not_dead",
        # zigzag diagnostic columns (v2.0, plan §5.1)
        "zz_leg_direction", "zz_cand_height_pct", "zz_global_median",
        "zz_global_p80", "zz_local_median", "zz_n_legs",
        "zz_regime_state", "zz_armed", "zz_armed_side",
        # Phase 5 (RFC v3.1 §7.4) — readiness + arm_source.
        "zz_ready_a", "zz_ready_b", "zz_arm_source",
        # median ratio columns (appended post-build)
        "signal_body_pct_median_ratio", "signal_range_pct_median_ratio",
    ]
    assert list(result.columns) == expected_cols


# ---------------------------------------------------------------------------
# Test 12: exec_price — open[t+1] for o2o, close[t] for c2c
# ---------------------------------------------------------------------------

def test_exec_price_column():
    """Exec Price: open[t+1] for o2o; close[t] for c2c."""
    n = 8
    df = _make_ohlc(n, open_=100.0, close_=101.0, high_delta=1.0, low_delta=1.0)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1])

    # Signal at t=4
    res_o2o = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.OPEN_TO_OPEN)
    row_o2o = res_o2o[res_o2o["signal_bar_index"] == 4].iloc[0]
    assert math.isclose(row_o2o["exec_price"], df["open"].iloc[5], rel_tol=1e-9)

    res_c2c = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    row_c2c = res_c2c[res_c2c["signal_bar_index"] == 4].iloc[0]
    assert math.isclose(row_c2c["exec_price"], df["close"].iloc[4], rel_tol=1e-9)


# ---------------------------------------------------------------------------
# Test 13: Signal Bar Index equals actual t
# ---------------------------------------------------------------------------

def test_signal_bar_index():
    """Signal Bar Index column must equal the bar position t in df."""
    n = 10
    df = _make_ohlc(n)
    # Flips at t=4 and t=7
    trend = _trend_from_list([0, 0, 1, 1, -1, -1, -1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    assert set(result["signal_bar_index"]) == {4, 7}
    assert result["signal_bar_index"].dtype in (np.int64, np.int32, int, object) or \
           np.issubdtype(result["signal_bar_index"].dtype, np.integer)


# ---------------------------------------------------------------------------
# Test 14: ValueError when len(trend) != len(df)
# ---------------------------------------------------------------------------

def test_assert_length_mismatch():
    """build_signal_events must raise ValueError if len(trend) != len(df)."""
    n = 8
    df = _make_ohlc(n)
    trend_wrong = _trend_from_list([0, 0, 1, 1, -1, 1])  # length 6

    with pytest.raises(ValueError):
        build_signal_events(df, trend_wrong, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)


# ---------------------------------------------------------------------------
# Test 15: ValueError when trend contains values outside {-1, 0, 1}
# ---------------------------------------------------------------------------

def test_assert_invalid_trend_values():
    """build_signal_events must raise ValueError for invalid trend values."""
    n = 8
    df = _make_ohlc(n)
    trend_bad = _trend_from_list([0, 0, 1, 2, -1, 1, 1, 1])  # 2 is invalid

    with pytest.raises(ValueError):
        build_signal_events(df, trend_bad, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)


# ---------------------------------------------------------------------------
# Test 16: Excel — Signals sheet appears in the output file with data
# ---------------------------------------------------------------------------

def test_excel_signals_sheet():
    """export_tester_results with signals_df must create a 'Signals' sheet."""
    from supertrend_optimizer.io.excel_tester import export_tester_results, SIGNALS_DISPLAY_NAMES
    from supertrend_optimizer.testing.runner import PeriodResult
    from supertrend_optimizer.engine.result import BacktestResult

    n = 10
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    signals_df = build_signal_events(
        df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE
    )
    assert len(signals_df) > 0, "Need at least one signal for this test"

    # Build minimal PeriodResult
    dummy_result = _make_dummy_period_result(n, trend)
    period_results = [dummy_result]

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results(period_results, out_path, signals_df=signals_df)

        with pd.ExcelFile(actual_path) as xf:
            assert "Signals" in xf.sheet_names
            assert "false start" in xf.sheet_names
            assert "filters_summary" in xf.sheet_names
            sheet_df = xf.parse("Signals")

        display_cols = list(SIGNALS_DISPLAY_NAMES.values())
        for col in display_cols:
            assert col in sheet_df.columns, f"Missing column: {col}"

        assert len(sheet_df) == len(signals_df)


# ---------------------------------------------------------------------------
# Test 17: no signals_df → no Signals sheet; false start sheet still written
# ---------------------------------------------------------------------------

def test_no_signals_sheet_but_false_start_present():
    """Without signals_df: no 'Signals' sheet; 'false start' is still created (trade cols only)."""
    from supertrend_optimizer.io.excel_tester import export_tester_results

    n = 10
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])
    dummy_result = _make_dummy_period_result(n, trend)
    period_results = [dummy_result]

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results(period_results, out_path)

        with pd.ExcelFile(actual_path) as xf:
            assert "Signals" not in xf.sheet_names
            assert "false start" in xf.sheet_names


# ---------------------------------------------------------------------------
# Test 18: ValueError on unknown trade_mode
# ---------------------------------------------------------------------------

def test_invalid_trade_mode():
    """build_signal_events must raise ValueError for an unknown trade_mode."""
    n = 8
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1])

    for bad_mode in ("scalp", "reverSe", "LONG", "", "both2"):
        with pytest.raises(ValueError, match="trade_mode"):
            build_signal_events(df, trend, ATR_PERIOD, bad_mode, ExecutionModel.CLOSE_TO_CLOSE)


# ---------------------------------------------------------------------------
# Test 19: timezone-aware DatetimeIndex — no crash, naive datetime in Excel
# ---------------------------------------------------------------------------

def test_timezone_aware_signal_time():
    """Signals sheet must not crash and must strip tz from Signal Time column."""
    from supertrend_optimizer.io.excel_tester import export_tester_results, SIGNALS_DISPLAY_NAMES

    n = 10
    tz_index = pd.date_range("2024-01-01", periods=n, freq="1h", tz="UTC")
    df = _make_ohlc(n)
    df.index = tz_index

    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    signals_df = build_signal_events(
        df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE
    )
    assert len(signals_df) > 0

    dummy_result = _make_dummy_period_result(n, trend)

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results(
            [dummy_result], out_path, signals_df=signals_df
        )

        with pd.ExcelFile(actual_path) as xf:
            assert "Signals" in xf.sheet_names
            sheet_df = xf.parse("Signals")

    time_col = SIGNALS_DISPLAY_NAMES["signal_time"]
    assert time_col in sheet_df.columns
    # After export the column must not carry timezone info
    if pd.api.types.is_datetime64_any_dtype(sheet_df[time_col]):
        assert sheet_df[time_col].dt.tz is None


# ---------------------------------------------------------------------------
# Test 20: Signal Open Price and Signal Close Price columns
# ---------------------------------------------------------------------------

def test_signal_open_close_price_columns():
    """signal_open_price = open[t], signal_close_price = close[t] for signal bar."""
    n = 10
    open_ = 100.0
    close_ = 103.0
    df = _make_ohlc(n, open_=open_, close_=close_, high_delta=1.0, low_delta=1.0)
    # Signal at t=4 (green→red)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    for em in (ExecutionModel.OPEN_TO_OPEN, ExecutionModel.CLOSE_TO_CLOSE):
        result = build_signal_events(df, trend, ATR_PERIOD, "long", em)
        row = result[result["signal_bar_index"] == 4].iloc[0]

        # Signal bar prices must equal signal bar OHLC (bar t=4)
        assert math.isclose(row["signal_open_price"], df["open"].iloc[4], rel_tol=1e-9)
        assert math.isclose(row["signal_close_price"], df["close"].iloc[4], rel_tol=1e-9)

        # Exec Price must NOT be changed (o2o → open[5], c2c → close[4])
        if em == ExecutionModel.OPEN_TO_OPEN:
            assert math.isclose(row["exec_price"], df["open"].iloc[5], rel_tol=1e-9)
        else:
            assert math.isclose(row["exec_price"], df["close"].iloc[4], rel_tol=1e-9)


# ---------------------------------------------------------------------------
# Test 21: Autofilter is set on Signals sheet
# ---------------------------------------------------------------------------

def test_signals_sheet_autofilter():
    """Signals sheet must have auto_filter.ref set on the header row."""
    import openpyxl
    from supertrend_optimizer.io.excel_tester import export_tester_results

    n = 10
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])
    signals_df = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    dummy_result = _make_dummy_period_result(n, trend)

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results([dummy_result], out_path, signals_df=signals_df)

        wb = openpyxl.load_workbook(actual_path)
        ws = wb["Signals"]
        assert ws.auto_filter.ref is not None, "auto_filter.ref must be set on Signals sheet"
        wb.close()


def test_signals_sheet_autofilter_empty():
    """Autofilter must also be set when Signals sheet has no data rows."""
    import openpyxl
    from supertrend_optimizer.io.excel_tester import export_tester_results

    n = 8
    df = _make_ohlc(n)
    # Trend never flips → empty signals
    trend = _trend_from_list([0, 0, 1, 1, 1, 1, 1, 1])
    signals_df = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    assert len(signals_df) == 0

    dummy_result = _make_dummy_period_result(n, trend)

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results([dummy_result], out_path, signals_df=signals_df)

        wb = openpyxl.load_workbook(actual_path)
        ws = wb["Signals"]
        assert ws.auto_filter.ref is not None, "auto_filter.ref must be set even on empty Signals sheet"
        wb.close()


# ---------------------------------------------------------------------------
# Test 22: Conditional formatting applied to Body % and Range % columns
# ---------------------------------------------------------------------------

def test_signals_conditional_formatting():
    """Signal Body % and Signal Range % must have conditional formatting rules."""
    import openpyxl
    from supertrend_optimizer.io.excel_tester import export_tester_results, SIGNALS_DISPLAY_NAMES

    n = 10
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])
    signals_df = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    assert len(signals_df) > 0

    dummy_result = _make_dummy_period_result(n, trend)

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results([dummy_result], out_path, signals_df=signals_df)

        wb = openpyxl.load_workbook(actual_path)
        ws = wb["Signals"]

        # Collect all CF range strings (keys of _cf_rules dict, converted to str)
        cf_sqref_strings = [str(k) for k in ws.conditional_formatting._cf_rules.keys()]

        # Find which column letters correspond to Body % and Range %
        from openpyxl.utils import get_column_letter
        cols = list(SIGNALS_DISPLAY_NAMES.values())
        body_col_letter = get_column_letter(cols.index(SIGNALS_DISPLAY_NAMES["signal_body_pct"]) + 1)
        range_col_letter = get_column_letter(cols.index(SIGNALS_DISPLAY_NAMES["signal_range_pct"]) + 1)

        assert any(body_col_letter in s for s in cf_sqref_strings), (
            f"No conditional formatting found for Signal Body % (col {body_col_letter}), "
            f"found ranges: {cf_sqref_strings}"
        )
        assert any(range_col_letter in s for s in cf_sqref_strings), (
            f"No conditional formatting found for Signal Range % (col {range_col_letter}), "
            f"found ranges: {cf_sqref_strings}"
        )
        wb.close()


# ---------------------------------------------------------------------------
# Test 23: Median ratio columns — values, position, NaN / zero-median safety
# ---------------------------------------------------------------------------

def test_median_ratio_columns():
    """signal_body_pct_median_ratio = signal_body_pct / median(signal_body_pct)."""
    n = 12
    df = _make_ohlc(n)
    # Multiple flips to get several signal rows
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    # Columns must exist
    assert "signal_body_pct_median_ratio" in result.columns
    assert "signal_range_pct_median_ratio" in result.columns

    # Columns must be the last two
    assert list(result.columns)[-2:] == [
        "signal_body_pct_median_ratio",
        "signal_range_pct_median_ratio",
    ]

    if len(result) > 0:
        # Body ratio correctness
        body_median = result["signal_body_pct"].median()
        if not pd.isna(body_median) and body_median != 0:
            expected = result["signal_body_pct"] / body_median
            pd.testing.assert_series_equal(
                result["signal_body_pct_median_ratio"].reset_index(drop=True),
                expected.reset_index(drop=True),
                check_names=False, rtol=1e-9,
            )

        # Range ratio correctness
        range_median = result["signal_range_pct"].median()
        if not pd.isna(range_median) and range_median != 0:
            expected = result["signal_range_pct"] / range_median
            pd.testing.assert_series_equal(
                result["signal_range_pct_median_ratio"].reset_index(drop=True),
                expected.reset_index(drop=True),
                check_names=False, rtol=1e-9,
            )


def test_median_ratio_zero_median():
    """When source column is all-zero, median=0 → ratio column is all NaN, no exception."""
    import math
    n = 10
    # open == close → body = 0 → signal_body_pct = 0 → median = 0
    df = _make_ohlc(n, open_=100.0, close_=100.0, high_delta=0.5, low_delta=0.5)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, 1, 1, 1, 1])

    result = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)

    assert len(result) > 0
    # signal_body_pct should be 0 for all rows → median = 0 → ratio = NaN
    assert (result["signal_body_pct"] == 0.0).all()
    assert result["signal_body_pct_median_ratio"].isna().all(), (
        "median=0 must produce NaN ratio, not division-by-zero"
    )


# ---------------------------------------------------------------------------
# Test 24: CF rules for threshold highlighting and negative T+N
# ---------------------------------------------------------------------------

def test_signals_threshold_and_negative_formatting():
    """Threshold CF rules on ratio columns and negative-T+N CF rule must be present."""
    import openpyxl
    from supertrend_optimizer.io.excel_tester import export_tester_results, SIGNALS_DISPLAY_NAMES
    from openpyxl.utils import get_column_letter

    n = 12
    df = _make_ohlc(n)
    trend = _trend_from_list([0, 0, 1, 1, -1, 1, -1, 1, 1, 1, 1, 1])
    signals_df = build_signal_events(df, trend, ATR_PERIOD, "long", ExecutionModel.CLOSE_TO_CLOSE)
    assert len(signals_df) > 0

    dummy_result = _make_dummy_period_result(n, trend)

    with tempfile.TemporaryDirectory() as tmpdir:
        out_path = os.path.join(tmpdir, "result.xlsx")
        actual_path = export_tester_results([dummy_result], out_path, signals_df=signals_df)

        wb = openpyxl.load_workbook(actual_path)
        ws = wb["Signals"]

        cf_sqref_strings = [str(k) for k in ws.conditional_formatting._cf_rules.keys()]

        # Use actual sheet columns for position lookup: the Signals sheet is
        # produced by renaming signals_df via SIGNALS_DISPLAY_NAMES, so the
        # column order in the sheet matches signals_df.columns (post-rename).
        # SIGNALS_DISPLAY_NAMES.values() order may differ from the actual
        # DataFrame order (median-ratio columns are appended post-build and
        # new zz_* columns are appended after amp columns in _COLUMN_NAMES).
        sheet_cols = [cell.value for cell in ws[1]]  # header row

        # Median ratio columns must have CF rules
        for internal_key in ("signal_body_pct_median_ratio", "signal_range_pct_median_ratio"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            col_idx = sheet_cols.index(display_name) + 1  # 1-based
            col_letter = get_column_letter(col_idx)
            assert any(col_letter in s for s in cf_sqref_strings), (
                f"Expected CF on {display_name} (col {col_letter}), "
                f"found: {cf_sqref_strings}"
            )

        # T+N columns must have CF rules (negative highlight)
        for internal_key in ("t1_return_pct", "t2_return_pct", "t3_return_pct"):
            display_name = SIGNALS_DISPLAY_NAMES[internal_key]
            col_idx = sheet_cols.index(display_name) + 1
            col_letter = get_column_letter(col_idx)
            assert any(col_letter in s for s in cf_sqref_strings), (
                f"Expected CF on {display_name} (col {col_letter}), "
                f"found: {cf_sqref_strings}"
            )

        wb.close()


# ---------------------------------------------------------------------------
# Helper: build a minimal PeriodResult for Excel tests
# ---------------------------------------------------------------------------

def _make_dummy_period_result(n: int, trend: np.ndarray):
    """Create a PeriodResult that satisfies export_tester_results()."""
    from supertrend_optimizer.testing.runner import PeriodResult
    from supertrend_optimizer.engine.result import BacktestResult

    returns = np.zeros(n, dtype=np.float64)
    equity = np.ones(n + 1, dtype=np.float64)
    positions = np.zeros(n + 1, dtype=np.int8)

    backtest_result = BacktestResult(
        atr_period=ATR_PERIOD,
        multiplier=3.0,
        trade_mode="long",
        commission=0.0,
        warmup=0,
        returns=returns,
        equity_curve=equity,
        positions=positions,
        trend=trend,
        metrics={
            "sum_pnl_pct": 0.0,
            "sharpe": 0.0,
            "sortino": 0.0,
            "max_drawdown": 0.0,
            "cagr": 0.0,
            "win_rate": 0.0,
            "num_trades": 0,
            "profit_factor": 0.0,
            "avg_trade": 0.0,
        },
        early_exit=False,
        exit_bar=None,
        exit_drawdown=None,
        trades_df=None,
        n_bars_original=n,
        period_label="100%",
        effective_warmup=0,
    )

    return PeriodResult(
        period_label="100%",
        n_bars=n,
        result=backtest_result,
    )
