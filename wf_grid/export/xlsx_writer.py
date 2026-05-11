"""
XLSX workbook export for WF grid search (xlsx spec §2, §10–§11).

export_workbook writes the full book in sheet order (§2.7):
  0. DISCLAIMER  (always first)
  1. WF_Config
  2. summary      (immediately after WF_Config for quick access)
  3. WF_Gates  (if gates_result provided and non-empty)
  4. WF_01 … WF_N
  5. WF_Trades
  6. WF_Train_Trades

Pre-export validation (done before any writing):
  - Row-limit guard: len(trades_oos) <= 1_000_000 and len(trades_train) <= 1_000_000
    Abort with ExportError (silent truncation forbidden, §10.2).
  - Column-order check on summary_wide: Block A cols before Block B cols.
  - Deterministic column snapshot validated against expected order.

UX (§11):
  - Auto-filter on header row of every sheet.
  - Freeze panes: leftmost Block-A columns frozen (up to but not including
    the first column outside the freeze set, §11.2).

Donor reuse:
  - format_excel_export_df, _flatten_config_to_rows, _to_excel_safe
    imported from donor reference excerpt.
"""

from __future__ import annotations

import dataclasses
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd

from wf_grid.config.schema import GridConfig
from wf_grid.export.summary_builder import _BLOCK_A, _BLOCK_B, _BLOCK_TAIL, _parse_grid_point_id
from supertrend_optimizer.core.trade_filter_config import is_volume_enabled

# Donor helpers (reuse per policy)
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "donor"))
from REFERENCE_excel_wf_sheets_excerpt import (  # noqa: E402
    _flatten_config_to_rows,
    _to_excel_safe,
    format_excel_export_df,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_ROW_LIMIT: int = 1_000_000

# Columns frozen at the left edge of the summary sheet (§11.2)
_FREEZE_COLS: frozenset[str] = frozenset(
    _BLOCK_A  # same set as Block A identity + rank + gate columns
)

# Deterministic column order for WF_Trades (OOS)
_TRADES_COLUMN_ORDER: List[str] = [
    "grid_point_id",
    "atr_period",
    "multiplier",
    "trade_mode",
    "wf_step",
    "step_status",
    "test_start_idx",
    "test_end_idx",
    "trade_id",
    "direction",
    "entry_time",
    "entry_index",
    "entry_price",
    "exit_time",
    "exit_index",
    "exit_price",
    "bars_held",
    "gross_pnl_pct",
    "commission_pct",
    "net_pnl_pct",
]

# Deterministic column order for WF_Train_Trades (train window indices differ)
_TRAIN_TRADES_COLUMN_ORDER: List[str] = [
    "grid_point_id",
    "atr_period",
    "multiplier",
    "trade_mode",
    "wf_step",
    "step_status",
    "train_start_idx",
    "train_end_idx",
    "trade_id",
    "direction",
    "entry_time",
    "entry_index",
    "entry_price",
    "exit_time",
    "exit_index",
    "exit_price",
    "bars_held",
    "gross_pnl_pct",
    "commission_pct",
    "net_pnl_pct",
]

# Identity columns derived from grid_point_id
_IDENTITY_COLS: List[str] = ["atr_period", "multiplier", "trade_mode"]


# ---------------------------------------------------------------------------
# Identity column helpers
# ---------------------------------------------------------------------------

def _expand_grid_point_id(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure atr_period, multiplier, trade_mode columns exist in df.

    If all three are already present (propagated upstream), returns df unchanged.
    Otherwise parses them from grid_point_id using _parse_grid_point_id.
    Rows where grid_point_id is not parseable get NaN for the three columns.
    Does NOT modify df in-place; always returns a new DataFrame.
    """
    if all(c in df.columns for c in _IDENTITY_COLS):
        return df

    df = df.copy()
    parsed = df["grid_point_id"].apply(_parse_grid_point_id)
    for col in _IDENTITY_COLS:
        if col not in df.columns:
            df[col] = parsed.apply(lambda x, c=col: x.get(c))
    return df


# ---------------------------------------------------------------------------
# Public error type
# ---------------------------------------------------------------------------

class ExportError(RuntimeError):
    """Raised when export cannot proceed (row limit exceeded, etc.)."""


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_workbook(
    summary_wide: pd.DataFrame,
    step_oos_long: pd.DataFrame,
    trades_oos: pd.DataFrame,
    trades_train: pd.DataFrame,
    config: GridConfig,
    output_path: str | Path,
    wf_slices: Optional[List[Any]] = None,
    gates_result: Optional[Any] = None,
    bucket_matrix_median: Optional[pd.DataFrame] = None,
    step_results_oos: Optional[dict[str, list]] = None,
) -> Path:
    """
    Write full XLSX workbook to output_path.

    Parameters
    ----------
    summary_wide:
        Wide summary table from A11 build_summary_wide.
    step_oos_long:
        OOS step results from A5 collect_oos_steps.
    trades_oos:
        OOS trades from A6 collect_oos_trades.
    trades_train:
        Train trades from A6 collect_train_trades.
    config:
        Validated GridConfig (used for WF_Config sheet and freeze panes).
    output_path:
        Destination file path (.xlsx).
    wf_slices:
        Optional list of WFWindowSlice-like objects for WF_01..WF_N sheets.
    gates_result:
        Optional object with .checks iterable for WF_Gates sheet.
    bucket_matrix_median:
        Output of ``build_median_bucket_matrix`` (or None → placeholder sheet).

    Returns
    -------
    Path
        Absolute path to the written file.

    Raises
    ------
    ExportError
        If row-limit guard is triggered.
    """
    output_path = Path(output_path)

    # --- Pre-export validation ---
    _validate_row_limits(trades_oos, trades_train)
    _validate_summary_column_order(summary_wide)

    config_dict = _config_to_dict(config)
    n_steps = int(step_oos_long["wf_step"].max()) if not step_oos_long.empty else 0

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 0. DISCLAIMER (always first)
        _write_disclaimer_sheet(writer)

        # 1. WF_Config
        _write_config_sheet(writer, config_dict, config)

        # 2. summary — immediately after WF_Config for quick access
        _write_summary_sheet(writer, summary_wide, config)

        # 3. WF_Gates (conditional)
        if gates_result is not None:
            _write_gates_sheet(writer, gates_result)

        # 4. WF_01 … WF_N
        for step_idx in range(1, n_steps + 1):
            sheet_name = f"WF_{step_idx:02d}"
            _write_step_sheet(writer, step_oos_long, step_idx, sheet_name)

        # 5. WF_Trades
        _write_trades_sheet(writer, trades_oos, "WF_Trades")

        # 6. WF_Train_Trades
        _write_trades_sheet(writer, trades_train, "WF_Train_Trades")

        # 7. BucketMatrix_Median
        from wf_grid.export.bucket_sheet_writer import write_bucket_matrix_median_sheet
        write_bucket_matrix_median_sheet(
            bucket_matrix_df=bucket_matrix_median,
            writer=writer,
            atr_bucket_step=config.bucket.atr_bucket_step,
            mult_bucket_step=config.bucket.mult_bucket_step,
        )

        if config.export.retain_per_bar_filter_diagnostics:
            _write_filter_diagnostics_sheet(writer, step_results_oos)

    return output_path.resolve()


# ---------------------------------------------------------------------------
# Pre-export validation
# ---------------------------------------------------------------------------

def _validate_row_limits(trades_oos: pd.DataFrame, trades_train: pd.DataFrame) -> None:
    """Abort if any trades DataFrame exceeds row limit (§10.2)."""
    if len(trades_oos) > _ROW_LIMIT:
        raise ExportError(
            f"WF_Trades has {len(trades_oos):,} rows which exceeds the "
            f"Excel limit of {_ROW_LIMIT:,}. Export aborted."
        )
    if len(trades_train) > _ROW_LIMIT:
        raise ExportError(
            f"WF_Train_Trades has {len(trades_train):,} rows which exceeds "
            f"the Excel limit of {_ROW_LIMIT:,}. Export aborted."
        )


def _validate_summary_column_order(summary_wide: pd.DataFrame) -> None:
    """
    Validate deterministic column order in summary_wide.

    Rules:
    - All present Block A columns must appear before all present Block B columns.
    - This mirrors the invariant enforced by build_summary_wide.
    """
    if summary_wide.empty:
        return

    cols = list(summary_wide.columns)
    if not cols:
        return

    a_present = [c for c in _BLOCK_A if c in summary_wide.columns]
    b_present = [c for c in _BLOCK_B if c in summary_wide.columns]

    if not a_present or not b_present:
        return

    last_a = max(cols.index(c) for c in a_present)
    first_b = min(cols.index(c) for c in b_present)

    if last_a >= first_b:
        raise ExportError(
            f"summary_wide column order violation: Block A column at position "
            f"{last_a} appears after Block B column at position {first_b}. "
            f"Run build_summary_wide before export."
        )


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

_PIPELINE_VERSION = "2.0-phaseA"

_DISCLAIMER_LINES = [
    "⚠  RESEARCH OUTPUT ONLY — READ BEFORE USE  ⚠",
    "",
    "This output is for RESEARCH purposes only.",
    "Do NOT use as a sole basis for real-money trading decisions.",
    "",
    "Tier 1 means: strong Phase A ok-only candidate (all WF steps passed basic quality gates).",
    "Tier 1 does NOT mean profitable, not a trading signal, not ready for paper trading.",
    "",
    "Score values are relative within this grid run only.",
    "They are NOT comparable across different runs or datasets.",
    "",
    "ok_ratio < 1.0 means some OOS segments were excluded.",
    "Aggregates may overstate performance (survivorship bias).",
    "",
    "Minimum recommended: paper trading for 3+ months",
    "before any real capital allocation.",
    "",
    "⚠  SURVIVORSHIP BIAS WARNING  ⚠",
    "",
    "Aggregates (Median, Min, Std, etc.) are computed using ONLY ok-status steps.",
    "Segments that failed quality gates are EXCLUDED from all aggregate statistics.",
    "This means reported metrics reflect only surviving segments, not the full history.",
    "",
    "ok_ratio < 0.9 → at least 10% of segments were excluded (check survivorship_warning column).",
    "ok_ratio < 0.7 → HIGH survivorship bias risk. Aggregate metrics are unreliable.",
    "ok_ratio < 0.5 → CRITICAL. Results should NOT inform any trading decision.",
    "",
    "Always check the survivorship_warning column in the summary sheet before",
    "drawing conclusions from aggregate performance metrics.",
    "",
    "⚠  WHAT SCORES / TIERS / RANKS MEAN  ⚠",
    "",
    "tester_seed_score is a RELATIVE ranking metric within THIS grid run ONLY.",
    "It is NOT a prediction of future profitability.",
    "It is NOT comparable across different runs, datasets, or timeframes.",
    "",
    "Tier 1 means: strong Phase A ok-only candidate (all WF steps passed basic quality gates).",
    "Tier 1 is NOT proof of edge, not a trading signal, not ready for paper trading or live deployment.",
    "Tier 2 means: passed some gates. Tier 3 means: failed core gates.",
    "",
    "Rank 1 means: best score among the combinations tested in THIS run.",
    "It does NOT mean the strategy will be profitable in live trading.",
    "",
    "Check score_discrimination_status and score_interpretation_note columns",
    "before drawing conclusions from scores or rankings.",
    "",
    "⚠  MULTIPLE COMPARISONS / DATA MINING WARNING  ⚠",
    "",
    "This grid search evaluates many parameter combinations simultaneously.",
    "Rank 1 is the best result among all tested combinations — NOT an independent OOS proof.",
    "With hundreds of grid points, at least one combination will appear profitable by chance alone.",
    "",
    "DO NOT make trading decisions based on a single grid run without:",
    "  - An independent forward test on fully unseen data",
    "  - Multiple testing correction (e.g. Bonferroni, FDR / Benjamini-Hochberg)",
    "  - Comparison with a random baseline drawn from the same grid",
    "",
    "One grid run = one hypothesis generation step. NOT standalone evidence.",
    "Rank 1 / Tier 1 in Phase A is a starting point for Phase B validation, nothing more.",
    "",
    "⚠  PIPELINE PHASE STATUS  ⚠",
    "",
    "Current: Phase A (grid search + walk-forward + ranking).",
    "Missing: Phase B (regime analysis, confidence intervals, benchmark comparison).",
    "Missing: Phase C (position sizing, portfolio risk management, live monitoring).",
    "",
    "Do NOT trade real money without at least Phase B completed.",
    "Phase A output is a STARTING POINT for research, not a trading signal.",
    "",
    "⚠  WF_TRADES / WF_TRAIN_TRADES SHEETS  ⚠",
    "",
    "WF_Trades and WF_Train_Trades contain trades for the top-ranked (rank-1) grid point ONLY.",
    "This is a limited debug export, NOT a full-grid audit of all parameter combinations.",
    "Do NOT treat this trade list as representative of the full grid or the strategy population.",
    "It is provided for inspection and manual sanity-checking of the best candidate only.",
]


def _write_disclaimer_sheet(writer: pd.ExcelWriter) -> None:
    """Write DISCLAIMER as the first sheet in the workbook."""
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = (
        [{"Text": line} for line in _DISCLAIMER_LINES]
        + [{"Text": ""}]
        + [{"Text": f"Generated: {generated_at}"}]
        + [{"Text": f"Pipeline version: {_PIPELINE_VERSION}"}]
    )
    df = pd.DataFrame(rows, columns=["Text"])
    df.to_excel(writer, sheet_name="DISCLAIMER", index=False)


_WARNING_ROWS = [
    {
        "Section": "WARNING",
        "Parameter": "survivorship_bias",
        "Value": (
            "Aggregates use ONLY ok-status steps. "
            "Segments that failed gates are excluded. "
            "Check ok_ratio and survivorship_warning column per candidate."
        ),
    },
    {
        "Section": "WARNING",
        "Parameter": "ok_ratio_interpretation",
        "Value": (
            "ok_ratio >= 0.9: acceptable. "
            "ok_ratio < 0.9: MODERATE bias risk. "
            "ok_ratio < 0.7: HIGH bias risk. "
            "ok_ratio < 0.5: CRITICAL — do not use aggregates for decisions."
        ),
    },
    {
        "Section": "WARNING",
        "Parameter": "multiple_comparisons",
        "Value": (
            "Grid search tests many parameter combinations simultaneously. "
            "Rank 1 is the best result among all tested — NOT an independent OOS proof. "
            "With a large grid, top-ranked results may be statistical artefacts. "
            "Require independent forward test + multiple testing correction (Bonferroni / FDR) "
            "before drawing any conclusions. Do NOT trade based on a single grid run."
        ),
    },
]


_PHASE_ROWS = [
    {
        "Section": "PHASE",
        "Parameter": "current_phase",
        "Value": "A — grid search + walk-forward + ranking",
    },
    {
        "Section": "PHASE",
        "Parameter": "missing_phases",
        "Value": "B (regime analysis, confidence intervals, benchmark comparison), C (position sizing, portfolio risk management)",
    },
    {
        "Section": "PHASE",
        "Parameter": "recommendation",
        "Value": "Do NOT trade real money without Phase B + C. Phase A is a research starting point only.",
    },
]


def _write_config_sheet(
    writer: pd.ExcelWriter,
    config_dict: dict,
    config: GridConfig,
) -> None:
    """Write WF_Config sheet (Section, Parameter, Value) — donor reuse."""
    rows = _flatten_config_to_rows(config_dict)
    # Append resolved_periods_per_year explicitly
    if config.resolved_periods_per_year is not None:
        rows.append({
            "Section": "data",
            "Parameter": "resolved_periods_per_year",
            "Value": str(config.resolved_periods_per_year),
        })
    # Append WARNING rows (FIX-5.3: survivorship bias)
    rows.extend(_WARNING_ROWS)
    # Append PHASE rows (FIX-5.5: phase marker)
    rows.extend(_PHASE_ROWS)
    df = pd.DataFrame(rows, columns=["Section", "Parameter", "Value"])
    if not df.empty:
        df = df.sort_values(["Section", "Parameter"], kind="mergesort")
    df.to_excel(writer, sheet_name="WF_Config", index=False)
    _apply_autofilter(writer, "WF_Config", df)


def _write_gates_sheet(writer: pd.ExcelWriter, gates_result: Any) -> None:
    """Write WF_Gates sheet from gates_result.checks."""
    import numpy as np

    checks = getattr(gates_result, "checks", None) or []
    rows = []
    for check in checks:
        rows.append({
            "Name": getattr(check, "name", ""),
            "Passed": "PASS" if getattr(check, "passed", False) else "FAIL",
            "Value": getattr(check, "value", None),
            "Threshold": getattr(check, "threshold", None),
            "Message": getattr(check, "message", "") or "",
        })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["Name", "Passed", "Value", "Threshold", "Message"]
    )
    df.to_excel(writer, sheet_name="WF_Gates", index=False)
    _apply_autofilter(writer, "WF_Gates", df)


def _write_step_sheet(
    writer: pd.ExcelWriter,
    step_oos_long: pd.DataFrame,
    step_idx: int,
    sheet_name: str,
) -> None:
    """Write WF_01..WF_N sheet: all grid points at this wf_step, key metrics.

    Core columns come first in a fixed order; filter summary columns (§8.4, exit-off)
    are appended after the core when present (no silent drop — plan §8.4 / §2).
    """
    core_cols = [
        "grid_point_id",
        "atr_period",
        "multiplier",
        "trade_mode",
        "step_status",
        "sum_pnl_pct",
        "sharpe",
        "sortino",
        "max_drawdown",
        "num_trades",
        "profit_factor",
        "effective_oos_bars",
        "used_prepend",
        "prepend_bars_applied",
    ]
    # §8.4: filter summary columns — included when present (exit-off modes + others).
    filter_summary_cols = [
        "filter_states_visited",
        "n_bars_in_off",
        "n_bars_in_wait_first_st_flip",
        "n_bars_in_freeze",
        "n_bars_in_monitoring",
        "n_bars_in_counting_zz_legs",
        "n_bars_in_stopping",
        "n_filter_blocked_entries",
        "lifecycle_starts_count",
        "median_stop_triggered_count",
        "zz_leg_stop_triggered_count",
        "exit_off_mode",
        "exit_off_zz_leg_count",
        # Plan v3 §8 / §6: immediate-off summary columns
        "exit_b_immediate_off",
        "exit_b_immediate_off_count",
        "filter_diagnostics_available",
        "trigger_count_candidate_threshold",
        "trigger_count_confirmed_median",
        "trigger_count_both",
        "stopping_started_count",
        "n_volume_blocked_start_attempts",
        "n_volume_blocked_start_attempts_long",
        "n_volume_blocked_start_attempts_short",
        "n_volume_blocked_start_attempts_unknown_direction",
        "n_volume_warmup_blocked_start_attempts",
        "n_volume_below_baseline_blocked_start_attempts",
        "n_volume_above_baseline_blocked_start_attempts",
        "n_volume_baseline_zero_blocked_start_attempts",
        "n_volume_direction_warmup_blocked_start_attempts",
        "n_volume_unknown_direction_blocked_start_attempts",
        "n_volume_trade_mode_disallowed_direction_blocked_start_attempts",
        "n_volume_low_regime_bars",
        "n_volume_normal_regime_bars",
        "n_volume_high_regime_bars",
        "avg_median_relative_volume",
        "n_volume_started_cycles",
    ]

    step_df = step_oos_long[step_oos_long["wf_step"] == step_idx].copy()
    step_df = _expand_grid_point_id(step_df)
    present_core = [c for c in core_cols if c in step_df.columns]
    present_filter = [c for c in filter_summary_cols if c in step_df.columns]
    # Extra columns not in either list (forward-compat): append at end
    known = set(core_cols) | set(filter_summary_cols)
    extra = [c for c in step_df.columns if c not in known]
    cols = present_core + present_filter + extra
    step_df = step_df[cols].reset_index(drop=True)

    format_excel_export_df(step_df).to_excel(writer, sheet_name=sheet_name, index=False)
    _apply_autofilter(writer, sheet_name, step_df)


def _write_trades_sheet(
    writer: pd.ExcelWriter,
    trades_df: pd.DataFrame,
    sheet_name: str,
) -> None:
    """Write WF_Trades or WF_Train_Trades with deterministic column order.

    WF_Trades uses _TRADES_COLUMN_ORDER (test_start_idx / test_end_idx).
    WF_Train_Trades uses _TRAIN_TRADES_COLUMN_ORDER (train_start_idx / train_end_idx).
    """
    column_order = (
        _TRAIN_TRADES_COLUMN_ORDER
        if sheet_name == "WF_Train_Trades"
        else _TRADES_COLUMN_ORDER
    )

    if trades_df is None or trades_df.empty:
        pd.DataFrame(columns=column_order).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        _apply_autofilter(writer, sheet_name, pd.DataFrame())
        return

    df = _expand_grid_point_id(trades_df)

    # Strip timezone from datetime columns (Excel incompatibility)
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            if hasattr(df[col].dtype, "tz") and df[col].dtype.tz is not None:
                df[col] = df[col].dt.tz_localize(None)

    # Enforce deterministic column order: known cols first, extras appended
    ordered = [c for c in column_order if c in df.columns]
    extras = [c for c in df.columns if c not in ordered]
    df = df[ordered + extras]

    # Sort for stable golden tests: (grid_point_id, wf_step, trade_id)
    sort_cols = [c for c in ["grid_point_id", "wf_step", "trade_id"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols).reset_index(drop=True)

    format_excel_export_df(df).to_excel(writer, sheet_name=sheet_name, index=False)
    _apply_autofilter(writer, sheet_name, df)


def _write_filter_diagnostics_sheet(
    writer: pd.ExcelWriter,
    step_results_oos: Optional[dict[str, list]],
) -> None:
    """Write retained per-bar filter diagnostics when explicitly requested."""
    if not step_results_oos:
        return

    rows: list[dict[str, Any]] = []
    for gp_id, step_results in step_results_oos.items():
        for sr in step_results:
            diag = getattr(sr, "filter_diagnostics_oos", None)
            if not diag:
                continue
            n = len(next(iter(diag.values()))) if diag else 0
            for bar_idx in range(n):
                row = {
                    "grid_point_id": gp_id,
                    "wf_step": getattr(sr, "wf_step", None),
                    "bar_index": bar_idx,
                }
                for key, arr in diag.items():
                    row[str(key)] = arr[bar_idx]
                rows.append(row)
                if len(rows) > _ROW_LIMIT:
                    raise ExportError(
                        "WF_FilterDiagnostics has more than Excel's row limit; "
                        "disable export.retain_per_bar_filter_diagnostics or "
                        "reduce grid/window size."
                    )

    if not rows:
        return
    df = format_excel_export_df(pd.DataFrame(rows))
    df.to_excel(writer, sheet_name="WF_FilterDiagnostics", index=False)
    _apply_autofilter(writer, "WF_FilterDiagnostics", df)


def _render_disabled_gate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Replace pd.NA in gate_ok_worst_segment with "disabled" for XLSX display.

    Plan §4.5 / G1: when worst_segment_pnl_threshold is null, the gate is
    disabled and gate_ok_worst_segment is stored as pd.NA.  In the XLSX the
    cell must show "disabled" (not blank, not True, not False) so the user
    can immediately see that the gate was intentionally turned off.

    Only gate_ok_worst_segment is treated this way; other gate columns store
    bool values and must not be touched.
    """
    col = "gate_ok_worst_segment"
    if col not in df.columns:
        return df
    # Identify pd.NA cells (works for both object-dtype and pandas NA)
    na_mask = df[col].isna()
    if na_mask.any():
        # Cast column to object to allow mixed True/False/"disabled" values
        df[col] = df[col].astype(object)
        df.loc[na_mask, col] = "disabled"
    return df


def _write_summary_sheet(
    writer: pd.ExcelWriter,
    summary_wide: pd.DataFrame,
    config: GridConfig,
) -> None:
    """Write summary sheet with auto-filter, freeze panes, and soft highlights."""
    df = format_excel_export_df(_render_disabled_gate_columns(summary_wide.copy()))
    df.to_excel(writer, sheet_name="summary", index=False)
    _apply_autofilter(writer, "summary", df)
    _apply_freeze_panes(writer, "summary", df)
    _apply_summary_conditional_formatting(writer, "summary", df)
    _apply_summary_segment_highlights(writer, "summary", df)
    _apply_pnl_sum_quartile_highlights(writer, "summary", df)


# ---------------------------------------------------------------------------
# Summary conditional formatting (soft leader/loser highlights)
# ---------------------------------------------------------------------------

# Columns where higher = better → green top, red bottom
_CF_HIGH_IS_GOOD: List[str] = [
    "tester_seed_score",
    "ok_ratio",
    "sum_pnl_pct_Median",
    "sum_pnl_pct_Min",
    "profit_factor_Median",
    "num_trades_Median",
]

# Columns where lower magnitude = better (drawdown: closer to 0 is good,
# more negative is bad) → invert: red top, green bottom
_CF_LOW_IS_GOOD: List[str] = [
    "max_drawdown_Min",
    "max_drawdown_Median",
]

# Pastel green / white / pastel red (3-color scale)
_CF_COLOR_BEST = "C6EFCE"   # soft green
_CF_COLOR_MID = "FFFFFF"    # white (neutral)
_CF_COLOR_WORST = "FFC7CE"  # soft red

# ---------------------------------------------------------------------------
# Segment column highlight constants
# ---------------------------------------------------------------------------

# sum_pnl_pct: negative → red, zero → white, positive → green
_SEG_PNL_COLOR_MIN = "FFC7CE"   # soft red  (most negative)
_SEG_PNL_COLOR_MID = "FFFFFF"   # white     (around zero)
_SEG_PNL_COLOR_MAX = "C6EFCE"   # soft green (most positive)

# max_drawdown: values are ≤ 0; closer to 0 is better (green), more negative is worse (red)
_SEG_DD_COLOR_MIN = "FFC7CE"    # soft red  (worst drawdown, most negative)
_SEG_DD_COLOR_MID = "FFEB9C"    # soft amber (mid drawdown)
_SEG_DD_COLOR_MAX = "C6EFCE"    # soft green (closest to 0, best)

_RE_SEG_PNL = re.compile(r"^S\d+_sum_pnl_pct$")
_RE_SEG_DD = re.compile(r"^S\d+_max_drawdown$")

# ---------------------------------------------------------------------------
# sum_pnl_pct_Sum quartile highlight constants
# ---------------------------------------------------------------------------

_QUARTILE_WINNER_FILL = "C6EFCE"  # light green — top 25%
_QUARTILE_LOSER_FILL = "FFC7CE"   # light red  — bottom 25%


def _apply_summary_conditional_formatting(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """
    Apply a soft 3-color scale to key decision columns on the summary sheet.

    Rules:
    - Only numeric data rows are covered (row 2 onward; row 1 is the header).
    - Empty sheet → no-op.
    - Uses openpyxl ColorScaleRule; no icons, data bars, or borders changed.
    """
    if df.empty:
        return

    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return

    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    n_rows = len(df)
    cols_list = list(df.columns)

    def _add_scale(col_name: str, best_color: str, worst_color: str) -> None:
        if col_name not in cols_list:
            return
        col_idx = cols_list.index(col_name) + 1  # 1-based
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
        rule = ColorScaleRule(
            start_type="min",
            start_color=worst_color,
            mid_type="percentile",
            mid_value=50,
            mid_color=_CF_COLOR_MID,
            end_type="max",
            end_color=best_color,
        )
        ws.conditional_formatting.add(data_range, rule)

    for col in _CF_HIGH_IS_GOOD:
        _add_scale(col, best_color=_CF_COLOR_BEST, worst_color=_CF_COLOR_WORST)

    for col in _CF_LOW_IS_GOOD:
        # Invert: min (most negative drawdown = worst) → red; max (closest to 0 = best) → green
        _add_scale(col, best_color=_CF_COLOR_WORST, worst_color=_CF_COLOR_BEST)


# ---------------------------------------------------------------------------
# Segment column highlights (S1_sum_pnl_pct, S2_max_drawdown, …)
# ---------------------------------------------------------------------------


def _apply_summary_segment_highlights(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """
    Apply 3-color scale conditional formatting to segment metric columns.

    Targets:
    - S*_sum_pnl_pct : red (min) → white (mid) → green (max)
    - S*_max_drawdown: red (min, worst drawdown) → amber (mid) → green (max, closest to 0)

    Rules:
    - Only data rows are covered (row 2 onward; row 1 is the header).
    - No-op when df is empty or no matching columns exist.
    - Does not modify cell values or column order.
    """
    if df.empty:
        return

    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return

    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    n_rows = len(df)
    cols_list = list(df.columns)

    def _add_segment_scale(
        col_name: str,
        min_color: str,
        mid_color: str,
        max_color: str,
    ) -> None:
        col_idx = cols_list.index(col_name) + 1  # 1-based
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{n_rows + 1}"
        rule = ColorScaleRule(
            start_type="min",
            start_color=min_color,
            mid_type="percentile",
            mid_value=50,
            mid_color=mid_color,
            end_type="max",
            end_color=max_color,
        )
        ws.conditional_formatting.add(data_range, rule)

    for col in cols_list:
        if _RE_SEG_PNL.match(col):
            _add_segment_scale(
                col,
                min_color=_SEG_PNL_COLOR_MIN,
                mid_color=_SEG_PNL_COLOR_MID,
                max_color=_SEG_PNL_COLOR_MAX,
            )
        elif _RE_SEG_DD.match(col):
            _add_segment_scale(
                col,
                min_color=_SEG_DD_COLOR_MIN,
                mid_color=_SEG_DD_COLOR_MID,
                max_color=_SEG_DD_COLOR_MAX,
            )


# ---------------------------------------------------------------------------
# sum_pnl_pct_Sum quartile fill highlight (openpyxl PatternFill, not CF formula)
# ---------------------------------------------------------------------------


def _apply_pnl_sum_quartile_highlights(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """
    Apply quartile-based cell fill to the sum_pnl_pct_Sum column.

    Rules (applied to data rows only, not the header):
      - value <= Q1 (25th percentile) → _QUARTILE_LOSER_FILL  (light red)
      - value >= Q3 (75th percentile) → _QUARTILE_WINNER_FILL (light green)
      - otherwise                     → no fill

    Edge cases:
      - Column absent or df empty → no-op.
      - Fewer than 2 numeric values → no fill applied (not enough data).
    """
    col_name = "sum_pnl_pct_Sum"
    if df.empty or col_name not in df.columns:
        return

    ws = writer.sheets.get(sheet_name)
    if ws is None:
        return

    import numpy as np
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    col_idx = list(df.columns).index(col_name) + 1  # 1-based
    col_letter = get_column_letter(col_idx)

    # Collect numeric non-NaN values to compute quartiles
    numeric_vals = pd.to_numeric(df[col_name], errors="coerce").dropna()
    if len(numeric_vals) < 2:
        return

    q1 = float(np.percentile(numeric_vals, 25))
    q3 = float(np.percentile(numeric_vals, 75))

    winner_fill = PatternFill(fill_type="solid", fgColor=_QUARTILE_WINNER_FILL)
    loser_fill = PatternFill(fill_type="solid", fgColor=_QUARTILE_LOSER_FILL)

    for row_idx, raw_val in enumerate(df[col_name], start=2):  # row 1 is header
        val = pd.to_numeric(raw_val, errors="coerce")
        if pd.isna(val):
            continue
        cell = ws[f"{col_letter}{row_idx}"]
        if val <= q1:
            cell.fill = loser_fill
        elif val >= q3:
            cell.fill = winner_fill


# ---------------------------------------------------------------------------
# Excel UX helpers
# ---------------------------------------------------------------------------

def _apply_autofilter(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """Set auto-filter on header row (§11.1)."""
    ws = writer.sheets.get(sheet_name)
    if ws is None or df.empty:
        return
    n_cols = len(df.columns)
    if n_cols == 0:
        return
    from openpyxl.utils import get_column_letter
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A1:{last_col}1"


def _apply_freeze_panes(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    """
    Freeze leftmost consecutive Block-A columns (§11.2).

    Freezes up to (but not including) the first column that is NOT in
    _FREEZE_COLS.  If all columns are in _FREEZE_COLS, no freeze is applied
    (edge case: nothing to scroll to).
    """
    ws = writer.sheets.get(sheet_name)
    if ws is None or df.empty:
        return

    cols = list(df.columns)
    freeze_count = 0
    for c in cols:
        if c in _FREEZE_COLS:
            freeze_count += 1
        else:
            break

    if freeze_count == 0 or freeze_count >= len(cols):
        return

    from openpyxl.utils import get_column_letter
    freeze_col = get_column_letter(freeze_count + 1)
    ws.freeze_panes = f"{freeze_col}2"


# ---------------------------------------------------------------------------
# Config serialization helper
# ---------------------------------------------------------------------------

def _config_to_dict(config: GridConfig) -> dict:
    """Convert GridConfig dataclass tree to nested dict for _flatten_config_to_rows."""
    def _to_dict(obj: Any) -> Any:
        if dataclasses.is_dataclass(obj) and not isinstance(obj, type):
            return {
                f.name: _to_dict(getattr(obj, f.name))
                for f in dataclasses.fields(obj)
                if f.name != "resolved_periods_per_year"
            }
        if isinstance(obj, list):
            return [_to_dict(i) for i in obj]
        if isinstance(obj, dict):
            return {k: _to_dict(v) for k, v in obj.items()}
        return obj

    config_dict = _to_dict(config)
    tf = getattr(config, "trade_filter", None)
    if not is_volume_enabled(tf):
        trade_filter = config_dict.get("trade_filter")
        if isinstance(trade_filter, dict):
            trade_filter.pop("volume", None)
    else:
        volume = tf.volume
        config_dict["filter_config_snapshot"] = {
            "volume_filter_enabled": True,
            "volume_filter_mode": volume.mode,
            "volume_aggregation": getattr(volume, "aggregation", "median"),
            "volume_short_window": volume.short_window,
            "volume_baseline_window": volume.baseline_window,
            "volume_baseline_session_enabled": bool(
                getattr(getattr(volume, "baseline_session", None), "enabled", False)
            ),
            "volume_baseline_session_window": getattr(
                getattr(volume, "baseline_session", None), "window", None
            ),
            "volume_threshold_ratio": volume.threshold_ratio,
            "volume_regime_low_ratio": volume.regime_low_ratio,
            "volume_regime_high_ratio": volume.regime_high_ratio,
            "volume_direction_lookback_bars": volume.direction_lookback_bars,
        }
    return config_dict
