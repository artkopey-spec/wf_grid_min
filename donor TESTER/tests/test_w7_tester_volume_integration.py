from __future__ import annotations

import ast
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from supertrend_optimizer.cli.tester import load_tester_config
from supertrend_optimizer.core.trade_filter_config import (
    build_trade_filter_config_from_raw,
    collect_raw_user_keys,
    resolve_exit_b_immediate_off_in_place,
    resolve_exit_off_mode_in_place,
    resolve_time_filter_in_place,
    resolve_trade_filter_mode_in_place,
    resolve_volume_defaults_in_place,
    resolve_volume_enabled_in_place,
    resolve_zigzag_enabled_in_place,
    validate_trade_filter,
)
from supertrend_optimizer.core.volume_metrics import build_volume_global_metrics
from supertrend_optimizer.engine.result import BacktestResult
from supertrend_optimizer.io.excel_tester import export_tester_results
from supertrend_optimizer.testing.signal_events import build_signal_events
from supertrend_optimizer.testing import runner
from supertrend_optimizer.testing.runner import run_all_periods, run_period
from supertrend_optimizer.utils.enums import ExecutionModel
from supertrend_optimizer.utils.exceptions import ConfigError


def test_run_batch_tester_passes_index_to_volume_metrics():
    path = Path(__file__).resolve().parents[1] / "run_batch_tester.py"
    tree = ast.parse(path.read_text(encoding="utf-8"))
    calls = [
        node
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Name)
        and node.func.id == "build_volume_global_metrics"
    ]

    assert calls
    assert all(any(keyword.arg == "index" for keyword in call.keywords) for call in calls)


def test_repo_config_tester_yaml_loads_volume_aggregation_and_baseline_session():
    path = Path(__file__).resolve().parents[2] / "config_tester.yaml"

    cfg = load_tester_config(str(path))

    volume = cfg["trade_filter"].volume
    assert volume.aggregation == "mean"
    assert volume.baseline_session.enabled is True
    assert volume.baseline_session.window == "09:00-19:00"
    assert volume.baseline_session._start_hour == 9
    assert volume.baseline_session._end_hour == 19


def _df(n: int = 80) -> pd.DataFrame:
    close = 100.0 + np.linspace(0.0, 12.0, n)
    volume = np.full(n, 100.0)
    return pd.DataFrame(
        {
            "open": close,
            "high": close + 1.0,
            "low": close - 1.0,
            "close": close,
            "volume": volume,
        },
        index=pd.date_range("2025-01-01", periods=n, freq="min"),
    )


def _volume_raw(enabled: bool = True) -> dict:
    if not enabled:
        return {"enabled": False}
    return {
        "enabled": True,
        "mode": "volume_A",
        "short_window": 2,
        "baseline_window": 3,
        "threshold_ratio": 1.0,
        "regime_low_ratio": 0.8,
        "regime_high_ratio": 1.2,
        "direction_lookback_bars": 1,
    }


def _tf_cfg():
    raw = {
        "enabled": True,
        "volume": _volume_raw(True),
    }
    raw_user_keys = collect_raw_user_keys({"trade_filter": raw})
    tf_cfg = build_trade_filter_config_from_raw(raw)
    resolve_zigzag_enabled_in_place(tf_cfg, raw_user_keys)
    resolve_volume_enabled_in_place(tf_cfg, raw_user_keys)
    errors: list[str] = []
    validate_trade_filter(tf_cfg, errors, raw_user_keys, caller_pipeline="tester")
    assert errors == []
    resolve_trade_filter_mode_in_place(tf_cfg, raw_user_keys)
    resolve_exit_off_mode_in_place(tf_cfg, raw_user_keys)
    resolve_exit_b_immediate_off_in_place(tf_cfg, raw_user_keys)
    resolve_time_filter_in_place(tf_cfg, raw_user_keys)
    resolve_volume_defaults_in_place(tf_cfg, raw_user_keys)
    return tf_cfg


def _write_config(path, *, enabled: bool) -> None:
    volume_lines = "\n".join(
        f"    {key}: {str(value).lower() if isinstance(value, bool) else value}"
        for key, value in _volume_raw(enabled).items()
    )
    path.write_text(
        f"""
supertrend:
  atr_period: 5
  multiplier: 2.0
trade_mode: revers
segmentation:
  mode: equal_blocks
  n_parts: 3
trade_filter:
  enabled: {str(enabled).lower()}
  volume:
{volume_lines}
""",
        encoding="utf-8",
    )


def test_equal_blocks_rejects_enabled_trade_filter(tmp_path):
    cfg_path = tmp_path / "enabled.yaml"
    _write_config(cfg_path, enabled=True)

    with pytest.raises(ConfigError, match="equal_blocks segmentation is not supported"):
        load_tester_config(str(cfg_path))


def test_equal_blocks_allows_disabled_trade_filter(tmp_path):
    cfg_path = tmp_path / "disabled.yaml"
    _write_config(cfg_path, enabled=False)

    cfg = load_tester_config(str(cfg_path))

    assert cfg["segmentation"]["mode"] == "equal_blocks"


def test_run_period_slices_volume_runtime(monkeypatch):
    df = _df(20)
    tf_cfg = _tf_cfg()
    runtime = build_volume_global_metrics(
        df["volume"].to_numpy(),
        df["close"].to_numpy(),
        tf_cfg.volume,
    )
    captured = {}

    def fake_run_single_backtest(**kwargs):
        sliced = kwargs["volume_runtime"]
        captured["absolute_offset"] = sliced.absolute_offset
        captured["reference_length"] = sliced.reference_length
        n = len(kwargs["close"])
        return BacktestResult(
            atr_period=kwargs["atr_period"],
            multiplier=kwargs["multiplier"],
            trade_mode=kwargs["trade_mode"],
            commission=kwargs["commission"],
            warmup=0,
            returns=np.zeros(n - 1, dtype=np.float64),
            equity_curve=np.ones(n, dtype=np.float64),
            positions=np.zeros(n, dtype=np.int8),
            trend=np.ones(n, dtype=np.int8),
            metrics={"num_trades": 0, "sum_pnl_pct": 0.0},
            early_exit=False,
            exit_bar=None,
            exit_drawdown=None,
            trades_df=pd.DataFrame(),
            n_bars_original=n,
            filter_config_snapshot=sliced.filter_config_snapshot,
        )

    monkeypatch.setattr(runner, "run_single_backtest", fake_run_single_backtest)

    run_period(
        df=df.iloc[5:15],
        atr_period=5,
        multiplier=2.0,
        trade_mode="revers",
        commission=0.0,
        trade_filter_config=tf_cfg,
        volume_runtime=runtime,
        global_offset=5,
    )

    assert captured == {"absolute_offset": 5, "reference_length": 10}


def test_standalone_volume_full_run_and_workbook_export(tmp_path):
    df = _df(80)
    tf_cfg = _tf_cfg()
    runtime = build_volume_global_metrics(
        df["volume"].to_numpy(),
        df["close"].to_numpy(),
        tf_cfg.volume,
    )

    results = run_all_periods(
        df=df,
        atr_period=5,
        multiplier=2.0,
        trade_mode="revers",
        commission=0.0,
        periods_per_year=365.0,
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        trade_filter_config=tf_cfg,
        volume_runtime=runtime,
    )

    assert np.any(results[0].result.positions != 0)
    assert results[0].filter_config_snapshot is runtime.filter_config_snapshot
    signals_df = build_signal_events(
        df=df,
        trend=results[0].result.trend,
        atr_period=5,
        trade_mode="revers",
        execution_model=ExecutionModel.OPEN_TO_OPEN,
        filter_diagnostics=results[0].filter_diagnostics,
    )
    for column in (
        "filter_state_at_signal",
        "filter_decision",
        "filter_block_reason",
        "filter_trigger_source",
    ):
        assert column in signals_df.columns

    actual = export_tester_results(
        results,
        str(tmp_path / "standalone_volume.xlsx"),
        signals_df=signals_df,
        trade_filter_config=tf_cfg,
        df=df,
    )
    wb = openpyxl.load_workbook(actual, read_only=True)
    try:
        assert "FilterDiagnostics_100" in wb.sheetnames
        assert "ZigZag_Trigger_Events" not in wb.sheetnames
        assert "filters_summary" not in wb.sheetnames
        # volume-only produces a cycle sheet (informational artifact)
        assert "cycle" in wb.sheetnames
        from supertrend_optimizer.io.excel_tester import VOLUME_CYCLE_SHEET_COLUMNS
        cycle_headers = [
            cell.value
            for cell in next(wb["cycle"].iter_rows(max_row=1))
            if cell.value is not None
        ]
        assert cycle_headers == list(VOLUME_CYCLE_SHEET_COLUMNS), (
            f"cycle sheet header contract mismatch.\n"
            f"  Expected: {list(VOLUME_CYCLE_SHEET_COLUMNS)}\n"
            f"  Got:      {cycle_headers}"
        )

        fd_headers = [cell.value for cell in next(wb["FilterDiagnostics_100"].iter_rows(max_row=1))]
        assert "Volume Regime" in fd_headers
        assert "Median Relative Volume" in fd_headers
        assert "ZigZag Mode" not in fd_headers

        cfg_rows = list(wb["Tester_Config"].iter_rows(values_only=True))
        assert (
            "run",
            "filter_config_snapshot.volume_filter_mode",
            "volume_A",
        ) in cfg_rows
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("snippet", "expected"),
    [
        ("  volume: true\n", "trade_filter.volume must be a YAML mapping"),
        ("  zigzag: false\n", "trade_filter.zigzag must be a YAML mapping"),
        (
            "  triggers:\n    candidate_threshold: true\n",
            "trade_filter.triggers.candidate_threshold must be a YAML mapping",
        ),
    ],
)
def test_tester_rejects_non_mapping_trade_filter_blocks(tmp_path, snippet, expected):
    cfg_path = tmp_path / "bad_nested.yaml"
    cfg_path.write_text(
        f"""
supertrend:
  atr_period: 5
  multiplier: 2.0
trade_mode: revers
trade_filter:
  enabled: true
{snippet}
""",
        encoding="utf-8",
    )

    with pytest.raises(ConfigError, match=expected):
        load_tester_config(str(cfg_path))
